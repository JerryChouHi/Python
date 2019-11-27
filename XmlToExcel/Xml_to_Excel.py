# encoding:utf-8
# @Time     : 2019/11/1 10:21
# @Author   : Jerry Chou
# @File     : Xml_to_Excel.py
# @Function :

from sys import argv, path
from os.path import abspath, join
from os import getcwd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.styles.colors import RED, YELLOW, GREEN, BLACK
from openpyxl.utils import quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation
from xml.dom.minidom import parse

path.append(abspath(join(getcwd(), '..')))
import Common

alignment = Alignment(horizontal='center', vertical='center')
thin = Side(border_style='thin', color=BLACK)
border = Border(top=thin, left=thin, right=thin, bottom=thin)


def read_project(project_folder):
    socketmap_path = join(project_folder, "Project.xml")
    dom = parse(socketmap_path)
    root = dom.documentElement
    socketmap = root.getElementsByTagName('SocketMap')
    datalog = root.getElementsByTagName('Datalog')
    projectsetup = root.getElementsByTagName('ProjectSetup')
    project_data = []
    for item in socketmap:
        project_data.append(Common.get_tree(item))
    for item in datalog:
        project_data.append(Common.get_tree(item))
    for item in projectsetup:
        project_data.append(Common.get_tree(item))
    return project_data


def read_socketmap(project_folder):
    socketmap_path = join(project_folder, "XML\\SocketMap.xml")
    dom = parse(socketmap_path)
    root = dom.documentElement
    itemlist = root.getElementsByTagName('SignalRef')
    socketmap_data = []
    for item in itemlist:
        socketmap_data.append(Common.get_tree(item))
    return socketmap_data


def read_signals(project_folder):
    signals_path = join(project_folder, "XML\\Signals.xml")
    dom = parse(signals_path)
    root = dom.documentElement
    itemlist = root.getElementsByTagName('Signal')
    signals_data = []
    for item in itemlist:
        signals_data.append(Common.get_tree(item))
    return signals_data


def read_limit(project_folder):
    limit_path = join(project_folder, "XML\\Limit.xml")
    dom = parse(limit_path)
    root = dom.documentElement
    itemlist = root.getElementsByTagName('Limit')
    limit_data = []
    for item in itemlist:
        limit_data.append(Common.get_tree(item))
    return limit_data


def read_signalgroups(project_folder):
    signalgroups_path = join(project_folder, "XML\\SignalGroups.xml")
    dom = parse(signalgroups_path)
    root = dom.documentElement
    itemlist = root.getElementsByTagName('Signalgroup')
    signalgroups_data = []
    for item in itemlist:
        signalgroups_data.append(Common.get_tree(item))
    return signalgroups_data


def read_bindefinition(project_folder):
    bindefinition_path = join(project_folder, "XML\\BinDefinition.xml")
    dom = parse(bindefinition_path)
    root = dom.documentElement
    itemlist = root.getElementsByTagName('BinGroup')
    bindefinition_data = []
    for item in itemlist:
        if item.getAttribute('name') == 'HardBins':
            bindefinition_data.append(Common.get_tree(item))
        elif item.getAttribute('name') == 'SoftBins':
            bindefinition_data.append(Common.get_tree(item))
    return bindefinition_data


def read_dcmeasure(project_folder):
    testblock_path = join(project_folder, "XML\\TestBlock.xml")
    dom = parse(testblock_path)
    root = dom.documentElement
    itemlist = root.getElementsByTagName('DCMeasure')
    dcmeasure_data = []
    for item in itemlist:
        dcmeasure_data.append(Common.get_tree(item))
    return dcmeasure_data


def read_tests(project_folder):
    testblock_path = join(project_folder, "XML\\TestBlock.xml")
    dom = parse(testblock_path)
    root = dom.documentElement
    itemlist = root.getElementsByTagName('Test')
    tests_data = []
    for item in itemlist:
        tests_data.append(Common.get_tree(item))
    return tests_data


def read_uservars(project_folder):
    uservars_path = join(project_folder, "XML\\UserVars.xml")
    dom = parse(uservars_path)
    root = dom.documentElement
    itemlist = root.getElementsByTagName('Variable')
    uservars_data = []
    for item in itemlist:
        uservars_data.append(Common.get_tree(item))
    return uservars_data


def read_levels(project_folder):
    levels_path = join(project_folder, "XML\\Levels.xml")
    dom = parse(levels_path)
    root = dom.documentElement
    itemlist = root.getElementsByTagName('Levels')
    levels_data = []
    for item in itemlist:
        levels_data.append(Common.get_tree(item))
    return levels_data


def write_excel(project_data, socketmap_data, signals_data, limit_data, signalgroups_data, bindefinition_data,
                dcmeasure_data, tests_data, uservars_data, levels_data):
    file_name = 'Project.xlsx'

    wb = Workbook()  # 创建文件对象

    dv_SignalType = DataValidation(type='list', formula1='"In,Out,InOut,Supply,Pseudo,System"', allow_blank=True)
    dv_Unit = DataValidation(type='list', formula1='"A,mA,uA,V,mV"', allow_blank=True)
    dv_HWBinPassFail = DataValidation(type='list', formula1='"Pass,Fail"', allow_blank=True)
    dv_MeasureModeType = DataValidation(type='list', formula1='"FVMI,FIMV,FVMV,FIMI"', allow_blank=True)
    dv_MeasureMethodType = DataValidation(type='list', formula1='"Parallel,Serial"', allow_blank=True)
    dv_VariableType = DataValidation(type='list',
                                     formula1='"Voltage,Current,Time,Frequency,Resistance,Capacitance,SingleValue"',
                                     allow_blank=True)
    dv_PatternExecModeType = DataValidation(type='list', formula1='"Burst,Individual"', allow_blank=True)
    dv_TestParamType = DataValidation(type='list',
                                      formula1='"SocketMap,PatternBurst,DCMeasure,Value_String,Value_Voltage,Value_Current,Value_Time,Value_Frequency,Value_Resistance,Value_Capacitance,Value_Single"',
                                      allow_blank=True)

    sigref_str = ''
    for i in range(len(signals_data)):
        if signals_data[i][1][0][1] in ('Supply,InOut'):
            sigref_str += (signals_data[i][0][1] + ',')
    for i in range(len(signalgroups_data)):
        if signalgroups_data[i][1][0][1] in ('Supply,InOut'):
            sigref_str += (signalgroups_data[i][0][1] + ',')
    sigref_str = sigref_str[:-1]
    dv_sigref = DataValidation(type='list', formula1='"' + sigref_str + '"', allow_blank=True)

    sheet_project = wb.create_sheet('Project')
    sheet_project.add_data_validation(dv_SignalType)
    sheet_project.freeze_panes = 'B1'
    irow = 1
    sheet_project.cell(row=irow, column=1).value = 'SocketMap'
    sheet_project.cell(row=irow, column=2).value = 'SocketMap'
    sheet_project.cell(row=irow, column=3).value = project_data[0][0][1][1]
    irow += 1
    if len(project_data[1][1]) > 0:
        sheet_project.cell(row=irow, column=1).value = 'Datalog'
        for i in range(len(project_data[1][1])):
            sheet_project.cell(row=irow, column=2).value = project_data[1][1][i][0]
            sheet_project.cell(row=irow, column=3).value = project_data[1][1][i][1]
            irow += 1
    sheet_project.merge_cells(start_row=2, end_row=irow - 1, start_column=1, end_column=1)
    if len(project_data[2][0]) > 0:
        start = irow
        sheet_project.cell(row=irow, column=1).value = 'ProjectSetup'
        for i in range(len(project_data[2][0])):
            sheet_project.cell(row=irow, column=2).value = project_data[2][0][i][0]
            sheet_project.cell(row=irow, column=3).value = project_data[2][0][i][1]
            irow += 1
        sheet_project.merge_cells(start_row=start, end_row=irow - 1, start_column=1, end_column=1)
    for row in sheet_project.rows:
        for cell in row:
            cell.border = border
            cell.alignment = alignment
            if cell.column in (1, 2):
                cell.fill = PatternFill(fill_type='solid', fgColor=YELLOW)
                if cell.row == 1:
                    cell.font = Font(color=RED, bold=True)
                else:
                    cell.font = Font(color=BLACK, bold=True)

    sheet_pinmap = wb.create_sheet('PinMap')
    sheet_pinmap.add_data_validation(dv_SignalType)
    dv_SignalType.add('G2:G1048576')
    sheet_pinmap.add_data_validation(dv_Unit)
    dv_Unit.add('K2:K1048576')
    dv_Unit.add('N2:N1048576')
    dv_Unit.add('Q2:Q1048576')
    sheet_pinmap.freeze_panes = 'B2'
    loadboard = [
        ['J21', 'J22', 'Seg 4_1/Site0', 'Seg 4_1/Site1'],
        ['J23', 'J24', 'Seg 5_1/Site2', 'Seg 5_1/Site3'],
        ['J9', 'J10', 'Seg 1_1/Site4', 'Seg 1_1/Site5'],
        ['J7', 'J8', 'Seg 0_1/Site6', 'Seg 0_1/Site7'],
        ['J17', 'J18', 'Seg 5_0/Site8', 'Seg 5_0/Site9'],
        ['J15', 'J16', 'Seg 4_0/Site10', 'Seg 4_0/Site11'],
        ['J1', 'J2', 'Seg 0_0/Site12', 'Seg 0_0/Site13'],
        ['J3', 'J4', 'Seg 1_0/Site14', 'Seg 1_0/Site15']
    ]
    LoadBoardPins = []
    for i in range(len(loadboard)):
        LoadBoardPins.append(
            [
                loadboard[i][2],
                ['PMU_0', loadboard[i][0] + ': H_LCH00(PMU 0-31)', 0],
                ['PMU_1', loadboard[i][0] + ': H_LCH01(PMU 0-31)', 1],
                ['PMU_2', loadboard[i][0] + ': H_LCH02(PMU 0-31)', 2],
                ['PMU_3', loadboard[i][0] + ': H_LCH03(PMU 0-31)', 3],
                ['PMU_4', loadboard[i][0] + ': H_LCH04(PMU 0-31)', 4],
                ['PMU_5', loadboard[i][0] + ': H_LCH05(PMU 0-31)', 5],
                ['PMU_6', loadboard[i][0] + ': H_LCH06(PMU 0-31)', 6],
                ['PMU_7', loadboard[i][0] + ': H_LCH07(PMU 0-31)', 7],
                ['PMU_8', loadboard[i][0] + ': H_LCH08(PMU 0-31)', 8],
                ['PMU_9', loadboard[i][0] + ': H_LCH09(PMU 0-31)', 9],
                ['PMU_10', loadboard[i][0] + ': H_LCH10(PMU 0-31)', 10],
                ['PMU_11', loadboard[i][0] + ': H_LCH11(PMU 0-31)', 11],
                ['PMU_12/Mipi clk n', loadboard[i][0] + ': H_LCH12(PMU 0-31)/CH1_CSI_A_n', 12],
                ['PMU_13/Mipi clk p', loadboard[i][0] + ': H_LCH13(PMU 0-31)/CH1_CSI_A_p', 13],
                ['PMU_14/Mipi data0 n', loadboard[i][0] + ': H_LCH14(PMU 0-31)/CH1_CSI_B_n', 14],
                ['PMU_15/Mipi data0 p', loadboard[i][0] + ': H_LCH15(PMU 0-31)/CH1_CSI_B_p', 15],
                ['PMU_16/Mipi data1 n', loadboard[i][0] + ': H_LCH16(PMU 0-31)/CH1_CSI_C_n', 16],
                ['PMU_17/Mipi data1 p', loadboard[i][0] + ': H_LCH17(PMU 0-31)/CH1_CSI_C_p', 17],
                ['PMU_18/Mipi data2 n', loadboard[i][0] + ': H_LCH18(PMU 0-31)/CH1_CSI_D_n', 18],
                ['PMU_19/Mipi data2 p', loadboard[i][0] + ': H_LCH19(PMU 0-31)/CH1_CSI_D_p', 19],
                ['PMU_20/Mipi data3 n', loadboard[i][0] + ': H_LCH20(PMU 0-31)/CH1_CSI_E_n', 20],
                ['PMU_21/Mipi data3 p', loadboard[i][0] + ': H_LCH21(PMU 0-31)/CH1_CSI_E_p', 21],
                ['PMU_22/Mipi clk n', loadboard[i][0] + ': H_LCH22(PMU 0-31)/CH2_CSI_A_n', 22],
                ['PMU_23/Mipi clk p', loadboard[i][0] + ': H_LCH23(PMU 0-31)/CH2_CSI_A_p', 23],
                ['PMU_24/Mipi data0 n', loadboard[i][0] + ': H_LCH24(PMU 0-31)/CH2_CSI_B_n', 24],
                ['PMU_25/Mipi data0 p', loadboard[i][0] + ': H_LCH25(PMU 0-31)/CH2_CSI_B_p', 25],
                ['PMU_26/Mipi data1 n', loadboard[i][0] + ': H_LCH26(PMU 0-31)/CH2_CSI_C_n', 26],
                ['PMU_27/Mipi data1 p', loadboard[i][0] + ': H_LCH27(PMU 0-31)/CH2_CSI_C_p', 27],
                ['PMU_28/Mipi data2 n', loadboard[i][0] + ': H_LCH28(PMU 0-31)/CH2_CSI_D_n', 28],
                ['PMU_29/Mipi data2 p', loadboard[i][0] + ': H_LCH29(PMU 0-31)/CH2_CSI_D_p', 29],
                ['PMU_30/Mipi data3 n', loadboard[i][0] + ': H_LCH30(PMU 0-31)/CH2_CSI_E_n', 30],
                ['PMU_31/Mipi data3 p', loadboard[i][0] + ': H_LCH31(PMU 0-31)/CH2_CSI_E_p', 31],
                ['UTP_0', loadboard[i][1] + ': PE_E_S_0', 'M0'],
                ['UTP_1', loadboard[i][1] + ': PE_E_S_1', 'M1'],
                ['UTP_2', loadboard[i][1] + ': PE_E_S_2', 'M2'],
                ['UTP_3', loadboard[i][1] + ': PE_E_S_3', 'M3'],
                ['UTP_4', loadboard[i][1] + ': PE_E_S_4', 'M4'],
                ['UTP_5', loadboard[i][1] + ': PE_E_S_5', 'M5'],
                ['UTP_6', loadboard[i][1] + ': PE_E_S_6', 'M6'],
                ['UTP_7', loadboard[i][1] + ': PE_E_S_7', 'M7'],
                ['UTP_8', loadboard[i][0] + ': PE_E_S_0_B', 'M8'],
                ['UTP_9', loadboard[i][0] + ': PE_E_S_1_B', 'M9'],
                ['UTP_10', loadboard[i][0] + ': PE_E_S_2_B', 'M10'],
                ['UTP_11', loadboard[i][0] + ': PE_E_S_3_B', 'M11'],
                ['UTP_12', loadboard[i][0] + ': PE_E_S_4_B', 'M12'],
                ['UTP_13', loadboard[i][0] + ': PE_E_S_5_B', 'M13'],
                ['UTP_14', loadboard[i][0] + ': PE_E_S_6_B', 'M14'],
                ['UTP_15', loadboard[i][0] + ': PE_E_S_7_B', 'M15'],
                ['DPS_0', loadboard[i][1] + ': LVLC_F+_0', 2000],
                ['DPS_1', loadboard[i][1] + ': LVLC_F+_1', 2001],
                ['DPS_2', loadboard[i][1] + ': LVLC_F+_2', 2002],
                ['DPS_3', loadboard[i][1] + ': LVLC_F+_3', 2003],
                ['DPS_4', loadboard[i][1] + ': LVLC_F+_4', 2004],
                ['DPS_5', loadboard[i][1] + ': LVLC_F+_5', 2005],
                ['DPS_6', loadboard[i][1] + ': LVLC_F+_6', 2006],
                ['DPS_7', loadboard[i][1] + ': LVLC_F+_7', 2007],
                ['DPS_8', loadboard[i][1] + ': LVLC_S+_0', 2008],
                ['DPS_9', loadboard[i][1] + ': LVLC_S+_1', 2009],
                ['DPS_10', loadboard[i][1] + ': LVLC_S+_2', 2010],
                ['DPS_11', loadboard[i][1] + ': LVLC_S+_3', 2011],
                ['DPS_12', loadboard[i][1] + ': LVLC_S+_4', 2012],
                ['DPS_13', loadboard[i][1] + ': LVLC_S+_5', 2013],
                ['DPS_14', loadboard[i][1] + ': LVLC_S+_6', 2014],
                ['DPS_15', loadboard[i][1] + ': LVLC_S+_7', 2015],
                ['GPIO_0', 'gpio_2V5_0_p(Transition Board)', 5048],
                ['GPIO_1', 'gpio_2V5_0_n(Transition Board)', 5049],
                ['GPIO_2', 'gpio_2V5_1_p(Transition Board)', 5050],
                ['GPIO_3', 'gpio_2V5_1_n(Transition Board)', 5051],
                ['GPIO_4', loadboard[i][0] + ': gpio_2v5_2_p', 5052],
                ['GPIO_5', loadboard[i][0] + ': gpio_2v5_2_n', 5053],
                ['GPIO_6', loadboard[i][0] + ': gpio_2v5_3_p', 5054],
                ['GPIO_7', loadboard[i][0] + ': gpio_2v5_3_n', 5055],
                ['GPIO_8', loadboard[i][0] + ': gpio_1v8_0_p', 5056],
                ['GPIO_9', loadboard[i][0] + ': gpio_1v8_0_n', 5057],
                ['GPIO_10', loadboard[i][0] + ': gpio_1v8_1_p', 5058],
                ['GPIO_11', loadboard[i][0] + ': gpio_1v8_1_n', 5059],
                ['GPIO_12', loadboard[i][0] + ': gpio_1v8_2_p', 5060],
                ['GPIO_13', loadboard[i][0] + ': gpio_1v8_2_n', 5061],
                ['GPIO_14', loadboard[i][0] + ': gpio_1v8_3_p', 5062],
                ['GPIO_15', loadboard[i][0] + ': gpio_1v8_3_n', 5063],
                ['GPIO_16', loadboard[i][0] + ': gpio_3v3_1', 5064],
                ['GPIO_17', loadboard[i][0] + ': gpio_3v3_2', 5065],
                ['GPIO_18', loadboard[i][0] + ': gpio_3v3_3', 5066],
                ['GPIO_19', loadboard[i][0] + ': gpio_3v3_4', 5067],
                ['GPIO_20', loadboard[i][0] + ': gpio_3v3_5', 5068],
                ['GPIO_21', loadboard[i][0] + ': gpio_3v3_6', 5069],
                ['GPIO_22', loadboard[i][0] + ': gpio_3v3_7', 5070],
                ['GPIO_23', loadboard[i][0] + ': gpio_3v3_8', 5071],
                ['GPIO_24', loadboard[i][0] + ': gpio_3v3_9', 5072],
                ['GPIO_25', loadboard[i][0] + ': gpio_3v3_10', 5073],
                ['GPIO_26', loadboard[i][0] + ': gpio_3v3_11', 5074],
                ['GPIO_27', loadboard[i][0] + ': gpio_3v3_12', 5075]
            ])
        LoadBoardPins.append(
            [
                loadboard[i][3],
                ['PMU_0', loadboard[i][0] + ': H_LCH00(PMU 0-31)', 0],
                ['PMU_1', loadboard[i][0] + ': H_LCH01(PMU 0-31)', 1],
                ['PMU_2', loadboard[i][0] + ': H_LCH02(PMU 0-31)', 2],
                ['PMU_3', loadboard[i][0] + ': H_LCH03(PMU 0-31)', 3],
                ['PMU_4', loadboard[i][0] + ': H_LCH04(PMU 0-31)', 4],
                ['PMU_5', loadboard[i][0] + ': H_LCH05(PMU 0-31)', 5],
                ['PMU_6', loadboard[i][0] + ': H_LCH06(PMU 0-31)', 6],
                ['PMU_7', loadboard[i][0] + ': H_LCH07(PMU 0-31)', 7],
                ['PMU_8', loadboard[i][0] + ': H_LCH08(PMU 0-31)', 8],
                ['PMU_9', loadboard[i][0] + ': H_LCH09(PMU 0-31)', 9],
                ['PMU_10', loadboard[i][0] + ': H_LCH10(PMU 0-31)', 10],
                ['PMU_11', loadboard[i][0] + ': H_LCH11(PMU 0-31)', 11],
                ['PMU_12/Mipi clk n', loadboard[i][0] + ': H_LCH12(PMU 0-31)/CH1_CSI_A_n', 12],
                ['PMU_13/Mipi clk p', loadboard[i][0] + ': H_LCH13(PMU 0-31)/CH1_CSI_A_p', 13],
                ['PMU_14/Mipi data0 n', loadboard[i][0] + ': H_LCH14(PMU 0-31)/CH1_CSI_B_n', 14],
                ['PMU_15/Mipi data0 p', loadboard[i][0] + ': H_LCH15(PMU 0-31)/CH1_CSI_B_p', 15],
                ['PMU_16/Mipi data1 n', loadboard[i][0] + ': H_LCH16(PMU 0-31)/CH1_CSI_C_n', 16],
                ['PMU_17/Mipi data1 p', loadboard[i][0] + ': H_LCH17(PMU 0-31)/CH1_CSI_C_p', 17],
                ['PMU_18/Mipi data2 n', loadboard[i][0] + ': H_LCH18(PMU 0-31)/CH1_CSI_D_n', 18],
                ['PMU_19/Mipi data2 p', loadboard[i][0] + ': H_LCH19(PMU 0-31)/CH1_CSI_D_p', 19],
                ['PMU_20/Mipi data3 n', loadboard[i][0] + ': H_LCH20(PMU 0-31)/CH1_CSI_E_n', 20],
                ['PMU_21/Mipi data3 p', loadboard[i][0] + ': H_LCH21(PMU 0-31)/CH1_CSI_E_p', 21],
                ['PMU_22/Mipi clk n', loadboard[i][0] + ': H_LCH22(PMU 0-31)/CH2_CSI_A_n', 22],
                ['PMU_23/Mipi clk p', loadboard[i][0] + ': H_LCH23(PMU 0-31)/CH2_CSI_A_p', 23],
                ['PMU_24/Mipi data0 n', loadboard[i][0] + ': H_LCH24(PMU 0-31)/CH2_CSI_B_n', 24],
                ['PMU_25/Mipi data0 p', loadboard[i][0] + ': H_LCH25(PMU 0-31)/CH2_CSI_B_p', 25],
                ['PMU_26/Mipi data1 n', loadboard[i][0] + ': H_LCH26(PMU 0-31)/CH2_CSI_C_n', 26],
                ['PMU_27/Mipi data1 p', loadboard[i][0] + ': H_LCH27(PMU 0-31)/CH2_CSI_C_p', 27],
                ['PMU_28/Mipi data2 n', loadboard[i][0] + ': H_LCH28(PMU 0-31)/CH2_CSI_D_n', 28],
                ['PMU_29/Mipi data2 p', loadboard[i][0] + ': H_LCH29(PMU 0-31)/CH2_CSI_D_p', 29],
                ['PMU_30/Mipi data3 n', loadboard[i][0] + ': H_LCH30(PMU 0-31)/CH2_CSI_E_n', 30],
                ['PMU_31/Mipi data3 p', loadboard[i][0] + ': H_LCH31(PMU 0-31)/CH2_CSI_E_p', 31],
                ['UTP_0', loadboard[i][1] + ': PE_E_S_0', 'M0'],
                ['UTP_1', loadboard[i][1] + ': PE_E_S_1', 'M1'],
                ['UTP_2', loadboard[i][1] + ': PE_E_S_2', 'M2'],
                ['UTP_3', loadboard[i][1] + ': PE_E_S_3', 'M3'],
                ['UTP_4', loadboard[i][1] + ': PE_E_S_4', 'M4'],
                ['UTP_5', loadboard[i][1] + ': PE_E_S_5', 'M5'],
                ['UTP_6', loadboard[i][1] + ': PE_E_S_6', 'M6'],
                ['UTP_7', loadboard[i][1] + ': PE_E_S_7', 'M7'],
                ['UTP_8', loadboard[i][0] + ': PE_E_S_0_B', 'M8'],
                ['UTP_9', loadboard[i][0] + ': PE_E_S_1_B', 'M9'],
                ['UTP_10', loadboard[i][0] + ': PE_E_S_2_B', 'M10'],
                ['UTP_11', loadboard[i][0] + ': PE_E_S_3_B', 'M11'],
                ['UTP_12', loadboard[i][0] + ': PE_E_S_4_B', 'M12'],
                ['UTP_13', loadboard[i][0] + ': PE_E_S_5_B', 'M13'],
                ['UTP_14', loadboard[i][0] + ': PE_E_S_6_B', 'M14'],
                ['UTP_15', loadboard[i][0] + ': PE_E_S_7_B', 'M15'],
                ['DPS_0', loadboard[i][1] + ': LVLC_F+_0', 2000],
                ['DPS_1', loadboard[i][1] + ': LVLC_F+_1', 2001],
                ['DPS_2', loadboard[i][1] + ': LVLC_F+_2', 2002],
                ['DPS_3', loadboard[i][1] + ': LVLC_F+_3', 2003],
                ['DPS_4', loadboard[i][1] + ': LVLC_F+_4', 2004],
                ['DPS_5', loadboard[i][1] + ': LVLC_F+_5', 2005],
                ['DPS_6', loadboard[i][1] + ': LVLC_F+_6', 2006],
                ['DPS_7', loadboard[i][1] + ': LVLC_F+_7', 2007],
                ['DPS_8', loadboard[i][1] + ': LVLC_S+_0', 2008],
                ['DPS_9', loadboard[i][1] + ': LVLC_S+_1', 2009],
                ['DPS_10', loadboard[i][1] + ': LVLC_S+_2', 2010],
                ['DPS_11', loadboard[i][1] + ': LVLC_S+_3', 2011],
                ['DPS_12', loadboard[i][1] + ': LVLC_S+_4', 2012],
                ['DPS_13', loadboard[i][1] + ': LVLC_S+_5', 2013],
                ['DPS_14', loadboard[i][1] + ': LVLC_S+_6', 2014],
                ['DPS_15', loadboard[i][1] + ': LVLC_S+_7', 2015],
                ['GPIO_0', 'gpio_2V5_0_p(Transition Board)', 5048],
                ['GPIO_1', 'gpio_2V5_0_n(Transition Board)', 5049],
                ['GPIO_2', 'gpio_2V5_1_p(Transition Board)', 5050],
                ['GPIO_3', 'gpio_2V5_1_n(Transition Board)', 5051],
                ['GPIO_4', loadboard[i][0] + ': gpio_2v5_2_p', 5052],
                ['GPIO_5', loadboard[i][0] + ': gpio_2v5_2_n', 5053],
                ['GPIO_6', loadboard[i][0] + ': gpio_2v5_3_p', 5054],
                ['GPIO_7', loadboard[i][0] + ': gpio_2v5_3_n', 5055],
                ['GPIO_8', loadboard[i][0] + ': gpio_1v8_0_p', 5056],
                ['GPIO_9', loadboard[i][0] + ': gpio_1v8_0_n', 5057],
                ['GPIO_10', loadboard[i][0] + ': gpio_1v8_1_p', 5058],
                ['GPIO_11', loadboard[i][0] + ': gpio_1v8_1_n', 5059],
                ['GPIO_12', loadboard[i][0] + ': gpio_1v8_2_p', 5060],
                ['GPIO_13', loadboard[i][0] + ': gpio_1v8_2_n', 5061],
                ['GPIO_14', loadboard[i][0] + ': gpio_1v8_3_p', 5062],
                ['GPIO_15', loadboard[i][0] + ': gpio_1v8_3_n', 5063],
                ['GPIO_16', loadboard[i][0] + ': gpio_3v3_1', 5064],
                ['GPIO_17', loadboard[i][0] + ': gpio_3v3_2', 5065],
                ['GPIO_18', loadboard[i][0] + ': gpio_3v3_3', 5066],
                ['GPIO_19', loadboard[i][0] + ': gpio_3v3_4', 5067],
                ['GPIO_20', loadboard[i][0] + ': gpio_3v3_5', 5068],
                ['GPIO_21', loadboard[i][0] + ': gpio_3v3_6', 5069],
                ['GPIO_22', loadboard[i][0] + ': gpio_3v3_7', 5070],
                ['GPIO_23', loadboard[i][0] + ': gpio_3v3_8', 5071],
                ['GPIO_24', loadboard[i][0] + ': gpio_3v3_9', 5072],
                ['GPIO_25', loadboard[i][0] + ': gpio_3v3_10', 5073],
                ['GPIO_26', loadboard[i][0] + ': gpio_3v3_11', 5074],
                ['GPIO_27', loadboard[i][0] + ': gpio_3v3_12', 5075]
            ]
        )

    pin_count = len(LoadBoardPins[0]) - 1
    irow = 1
    sheet_pinmap.cell(row=irow, column=1).value = 'Segment/Site'
    sheet_pinmap.cell(row=irow, column=2).value = 'Name'
    sheet_pinmap.cell(row=irow, column=3).value = 'LoadBoard Pin'
    sheet_pinmap.cell(row=irow, column=5).value = 'SocketBoard Pin'
    sheet_pinmap.cell(row=irow, column=6).value = 'PinName'
    sheet_pinmap.cell(row=irow, column=7).value = 'PinType'
    sheet_pinmap.merge_cells(start_row=1, end_row=1, start_column=9, end_column=11)
    sheet_pinmap.cell(row=irow, column=9).value = 'OS'
    sheet_pinmap.merge_cells(start_row=1, end_row=1, start_column=12, end_column=14)
    sheet_pinmap.cell(row=irow, column=12).value = 'IIL'
    sheet_pinmap.merge_cells(start_row=1, end_row=1, start_column=15, end_column=17)
    sheet_pinmap.cell(row=irow, column=15).value = 'IIH'
    irow += 1
    site_row = []
    used_limit = []
    for i in range(len(LoadBoardPins)):
        start = i * pin_count + i + 2
        end = (i + 1) * (pin_count + 1)
        site_row.append((start, end))
        sheet_pinmap.cell(row=irow, column=1).value = LoadBoardPins[i][0]
        for j in range(1, pin_count + 1):
            sheet_pinmap.cell(row=irow, column=2).value = LoadBoardPins[i][j][0]
            sheet_pinmap.cell(row=irow, column=3).value = LoadBoardPins[i][j][1]
            loadboard_channel_id = str(LoadBoardPins[i][j][2])
            for m in range(len(socketmap_data)):
                socket_list = socketmap_data[m][1][0][1].split(' ')
                socket_channel_id = socket_list[i].split('.')[3]
                if loadboard_channel_id == socket_channel_id:
                    pin_name = socketmap_data[m][0][1]
                    sheet_pinmap.cell(row=irow, column=6).value = pin_name
                    for x in range(len(signals_data)):
                        if signals_data[x][0][1] == pin_name:
                            sheet_pinmap.cell(row=irow, column=7).value = signals_data[x][1][0][1]
                            break
                    for y in range(len(limit_data)):
                        if limit_data[y][0][1].split('_')[0].upper() == pin_name.upper() and limit_data[y][0][1].split('_')[
                            1].upper() == 'OS':
                            used_limit.append(y)
                            sheet_pinmap.cell(row=irow, column=9).value = limit_data[y][1][2][1]
                            sheet_pinmap.cell(row=irow, column=10).value = limit_data[y][1][1][1]
                            sheet_pinmap.cell(row=irow, column=11).value = limit_data[y][1][3][1]
                        if limit_data[y][0][1].split('_')[0].upper() == pin_name.upper() and limit_data[y][0][1].split('_')[
                            1].upper() == 'IIL':
                            used_limit.append(y)
                            sheet_pinmap.cell(row=irow, column=12).value = limit_data[y][1][2][1]
                            sheet_pinmap.cell(row=irow, column=13).value = limit_data[y][1][1][1]
                            sheet_pinmap.cell(row=irow, column=14).value = limit_data[y][1][3][1]
                        if limit_data[y][0][1].split('_')[0].upper() == pin_name.upper() and limit_data[y][0][1].split('_')[
                            1].upper() == 'IIH':
                            used_limit.append(y)
                            sheet_pinmap.cell(row=irow, column=15).value = limit_data[y][1][2][1]
                            sheet_pinmap.cell(row=irow, column=16).value = limit_data[y][1][1][1]
                            sheet_pinmap.cell(row=irow, column=17).value = limit_data[y][1][3][1]
                    break
            irow += 1
        irow += 1
    used_limit = list(set(used_limit))
    for item in site_row:
        sheet_pinmap.merge_cells(start_row=item[0], end_row=item[1], start_column=1, end_column=1)

    for row in sheet_pinmap.rows:
        for cell in row:
            cell.border = border
            if cell.row == 1 and cell.column not in (4, 8):
                cell.alignment = alignment
                cell.font = Font(color=RED, bold=True)
                cell.fill = PatternFill(fill_type='solid', fgColor=YELLOW)
            if cell.column == 1:
                cell.alignment = alignment
            for item in site_row:
                if item[0] <= cell.row <= item[1]:
                    if cell.column == 2:
                        cell.font = Font(bold=True)
                    if 1 <= cell.column <= 3:
                        cell.fill = PatternFill(fill_type='solid', fgColor='CAE1FF')  # 浅蓝
                    if 5 <= cell.column <= 7 or cell.column >= 9:
                        cell.fill = PatternFill(fill_type='solid', fgColor='66CD00')  # 浅绿

    sheet_pingroup = wb.create_sheet('PinGroup')
    sheet_pingroup.add_data_validation(dv_SignalType)
    dv_SignalType.add('B2:B1048576')
    sheet_pingroup.freeze_panes = 'A2'
    irow = 1
    sheet_pingroup.cell(row=irow, column=1).value = 'PinGroup'
    sheet_pingroup.cell(row=irow, column=2).value = 'PinType'
    sheet_pingroup.cell(row=irow, column=3).value = 'Pins'
    irow += 1
    for i in range(len(signalgroups_data)):
        sheet_pingroup.cell(row=irow, column=1).value = signalgroups_data[i][0][1]
        sheet_pingroup.cell(row=irow, column=2).value = signalgroups_data[i][1][0][1]
        sheet_pingroup.cell(row=irow, column=3).value = signalgroups_data[i][1][1][1]
        irow += 1
    for row in sheet_pingroup.rows:
        for cell in row:
            cell.border = border
            if cell.row == 1:
                cell.alignment = alignment
                cell.fill = PatternFill(fill_type='solid', fgColor=YELLOW)
                cell.font = Font(color=RED, bold=True)

    sheet_binmap = wb.create_sheet('BinMap')
    sheet_binmap.add_data_validation(dv_HWBinPassFail)
    dv_HWBinPassFail.add('H2:H1048576')
    sheet_binmap.freeze_panes = 'A2'
    irow = 1
    sheet_binmap.cell(row=irow, column=1).value = 'SWBinID'
    sheet_binmap.cell(row=irow, column=2).value = 'SWBinName'
    sheet_binmap.cell(row=irow, column=3).value = 'Comment'
    sheet_binmap.cell(row=irow, column=4).value = 'HWBinID'
    sheet_binmap.cell(row=irow, column=6).value = 'HWBinID'
    sheet_binmap.cell(row=irow, column=7).value = 'HWBinName'
    sheet_binmap.cell(row=irow, column=8).value = 'Pass/Fail'
    sheet_binmap.cell(row=irow, column=9).value = 'Comment'
    irow += 1
    for i in range(2, len(bindefinition_data[1])):
        sheet_binmap.cell(row=irow, column=1).value = bindefinition_data[1][i][1][0][1]
        sheet_binmap.cell(row=irow, column=2).value = bindefinition_data[1][i][0][1]
        sheet_binmap.cell(row=irow, column=3).value = bindefinition_data[1][i][1][1][1]
        hwbinname = bindefinition_data[1][i][1][2][1]
        for j in range(2, len(bindefinition_data[0])):
            if bindefinition_data[0][j][0][1] == hwbinname:
                sheet_binmap.cell(row=irow, column=4).value = bindefinition_data[0][j][1][0][1]
                if bindefinition_data[0][j][1][2][1] == 'Pass':
                    sheet_binmap.cell(row=irow, column=4).fill = PatternFill(fill_type='solid', fgColor=GREEN)
                else:
                    sheet_binmap.cell(row=irow, column=4).fill = PatternFill(fill_type='solid', fgColor=RED)
        irow += 1
    dv_SWBinId = DataValidation(type='list', formula1="{0}!$A$2:$A${1}".format(quote_sheetname('BinMap'), irow))
    irow = 2
    for i in range(2, len(bindefinition_data[0])):
        sheet_binmap.cell(row=irow, column=6).value = bindefinition_data[0][i][1][0][1]
        sheet_binmap.cell(row=irow, column=7).value = bindefinition_data[0][i][0][1]
        sheet_binmap.cell(row=irow, column=8).value = bindefinition_data[0][i][1][2][1]
        if bindefinition_data[0][i][1][2][1] == 'Pass':
            sheet_binmap.cell(row=irow, column=8).fill = PatternFill(fill_type='solid', fgColor=GREEN)
        else:
            sheet_binmap.cell(row=irow, column=8).fill = PatternFill(fill_type='solid', fgColor=RED)
        sheet_binmap.cell(row=irow, column=9).value = bindefinition_data[0][i][1][1][1]
        irow += 1
    dv_HWBinId = DataValidation(type='list', formula1="{0}!$F$2:$F${1}".format(quote_sheetname('BinMap'), irow))
    sheet_binmap.add_data_validation(dv_HWBinId)
    dv_HWBinId.add('D2:D1048576')
    for row in sheet_binmap.rows:
        for cell in row:
            cell.border = border
            cell.alignment = alignment
            if cell.row == 1 and cell.column != 5:
                cell.fill = PatternFill(fill_type='solid', fgColor=YELLOW)
                if cell.column == 3 or cell.column == 9:
                    cell.font = Font(color=BLACK, bold=True)
                else:
                    cell.font = Font(color=RED, bold=True)

    sheet_limits = wb.create_sheet('Limits')
    sheet_limits.add_data_validation(dv_Unit)
    dv_Unit.add('D2:D1048576')
    sheet_limits.add_data_validation(dv_SWBinId)
    dv_SWBinId.add('E2:E1048576')
    sheet_limits.freeze_panes = 'A2'
    irow = 1
    sheet_limits.cell(row=irow, column=1).value = 'Name'
    sheet_limits.cell(row=irow, column=2).value = 'HiLimit'
    sheet_limits.cell(row=irow, column=3).value = 'LoLimit'
    sheet_limits.cell(row=irow, column=4).value = 'Unit'
    sheet_limits.cell(row=irow, column=5).value = 'BinNum'
    sheet_limits.cell(row=irow, column=6).value = 'Comment'
    irow += 1
    for i in range(len(limit_data)):
        if i not in used_limit:
            sheet_limits.cell(row=irow, column=1).value = limit_data[i][0][1]
            if len(limit_data[i][1]) >= 2:
                sheet_limits.cell(row=irow, column=2).value = limit_data[i][1][2][1]
            if len(limit_data[i][1]) >= 3:
                sheet_limits.cell(row=irow, column=3).value = limit_data[i][1][1][1]
            if len(limit_data[i][1]) >= 4:
                sheet_limits.cell(row=irow, column=4).value = limit_data[i][1][3][1]
            if len(limit_data[i][1]) >= 5:
                sheet_limits.cell(row=irow, column=5).value = limit_data[i][1][4][1]
            if len(limit_data[i][1]) >= 6:
                sheet_limits.cell(row=irow, column=6).value = limit_data[i][1][5][1]
            irow += 1
    for row in sheet_limits.rows:
        for cell in row:
            cell.border = border
            if cell.row == 1:
                cell.alignment = alignment
                cell.fill = PatternFill(fill_type='solid', fgColor=YELLOW)
                if cell.column == 1:
                    cell.font = Font(color=RED, bold=True)
                else:
                    cell.font = Font(color=BLACK, bold=True)

    sheet_dcmeasure = wb.create_sheet('DCMeasure')
    sheet_dcmeasure.add_data_validation(dv_sigref)
    dv_sigref.add('B2:B1048576')
    sheet_dcmeasure.add_data_validation(dv_MeasureModeType)
    dv_MeasureModeType.add('D2:D1048576')
    sheet_dcmeasure.add_data_validation(dv_MeasureMethodType)
    dv_MeasureMethodType.add('E2:E1048576')
    sheet_dcmeasure.freeze_panes = 'A2'
    irow = 1
    sheet_dcmeasure.cell(row=irow, column=1).value = 'DCMeasure'
    sheet_dcmeasure.cell(row=irow, column=2).value = 'Signal'
    sheet_dcmeasure.cell(row=irow, column=3).value = 'minorTestNum'
    sheet_dcmeasure.cell(row=irow, column=4).value = 'Mode'
    sheet_dcmeasure.cell(row=irow, column=5).value = 'Method'
    sheet_dcmeasure.cell(row=irow, column=6).value = 'ForceValue'
    sheet_dcmeasure.cell(row=irow, column=7).value = 'IRange'
    sheet_dcmeasure.cell(row=irow, column=8).value = 'HiLimit'
    sheet_dcmeasure.cell(row=irow, column=9).value = 'LoLimit'
    sheet_dcmeasure.cell(row=irow, column=10).value = 'HiClamp'
    sheet_dcmeasure.cell(row=irow, column=11).value = 'LoClamp'
    sheet_dcmeasure.cell(row=irow, column=12).value = 'SampleNum'
    sheet_dcmeasure.cell(row=irow, column=13).value = 'Delay'
    irow += 1
    for i in range(len(dcmeasure_data)):
        sheet_dcmeasure.cell(row=irow, column=1).value = dcmeasure_data[i][0][1]
        for j in range(1, len(dcmeasure_data[i])):
            sheet_dcmeasure.cell(row=irow, column=2).value = dcmeasure_data[i][j][0][1]
            sheet_dcmeasure.cell(row=irow, column=3).value = dcmeasure_data[i][j][1][1]
            sheet_dcmeasure.cell(row=irow, column=4).value = dcmeasure_data[i][j][2][0][1]
            sheet_dcmeasure.cell(row=irow, column=5).value = dcmeasure_data[i][j][2][1][1]
            sheet_dcmeasure.cell(row=irow, column=6).value = dcmeasure_data[i][j][2][2][1]
            sheet_dcmeasure.cell(row=irow, column=7).value = dcmeasure_data[i][j][2][3][1]
            sheet_dcmeasure.cell(row=irow, column=8).value = dcmeasure_data[i][j][2][4][1]
            sheet_dcmeasure.cell(row=irow, column=9).value = dcmeasure_data[i][j][2][5][1]
            sheet_dcmeasure.cell(row=irow, column=10).value = dcmeasure_data[i][j][2][6][1]
            sheet_dcmeasure.cell(row=irow, column=11).value = dcmeasure_data[i][j][2][7][1]
            sheet_dcmeasure.cell(row=irow, column=12).value = dcmeasure_data[i][j][2][8][1]
            sheet_dcmeasure.cell(row=irow, column=13).value = dcmeasure_data[i][j][2][9][1]
            irow += 1
        irow += 1
    for row in sheet_dcmeasure.rows:
        for cell in row:
            cell.border = border
            cell.alignment = alignment
            if cell.row == 1:
                cell.fill = PatternFill(fill_type='solid', fgColor=YELLOW)
                cell.font = Font(color=RED, bold=True)

    sheet_test = wb.create_sheet('Test')
    sheet_test.freeze_panes = 'A2'
    irow = 1
    sheet_test.cell(row=irow, column=1).value = 'Test Item'
    sheet_test.cell(row=irow, column=2).value = 'TestNum'
    sheet_test.cell(row=irow, column=3).value = 'ExecAPI'
    irow += 1
    param_count = 0
    for i in range(len(tests_data)):
        sheet_test.cell(row=irow, column=1).value = tests_data[i][0][1]
        sheet_test.cell(row=irow, column=2).value = tests_data[i][1][1]
        sheet_test.cell(row=irow, column=3).value = tests_data[i][-1][1][1]
        temp_count = len(tests_data[i]) - 3
        if param_count < temp_count:
            param_count = temp_count
        for j in range(temp_count):
            para = tests_data[i][2 + j][0][1] + ':' + tests_data[i][2 + j][1][1][1]
            sheet_test.cell(row=irow, column=(4 + j)).value = para
        irow += 1
    for i in range(param_count):
        sheet_test.cell(row=1, column=(4 + i)).value = 'Param' + str(i)
    for row in sheet_test.rows:
        for cell in row:
            cell.border = border
            cell.alignment = alignment
            if cell.row == 1:
                cell.fill = PatternFill(fill_type='solid', fgColor=YELLOW)
                if cell.column in (1, 2, 3):
                    cell.font = Font(color=RED, bold=True)
                else:
                    cell.font = Font(color=BLACK, bold=True)

    sheet_uservars = wb.create_sheet('UserVars')
    sheet_uservars.add_data_validation(dv_VariableType)
    dv_VariableType.add('B2:B1048576')
    sheet_uservars.freeze_panes = 'A2'
    irow = 1
    sheet_uservars.cell(row=irow, column=1).value = 'UserVars'
    sheet_uservars.cell(row=irow, column=2).value = 'Type'
    sheet_uservars.cell(row=irow, column=3).value = 'Val'
    irow += 1
    for i in range(len(uservars_data)):
        sheet_uservars.cell(row=irow, column=1).value = uservars_data[i][0][1]
        sheet_uservars.cell(row=irow, column=2).value = uservars_data[i][1][1]
        sheet_uservars.cell(row=irow, column=3).value = uservars_data[i][2][0][1]
        irow += 1
    for row in sheet_uservars.rows:
        for cell in row:
            cell.border = border
            cell.alignment = alignment
            if cell.row == 1:
                cell.fill = PatternFill(fill_type='solid', fgColor=YELLOW)
                cell.font = Font(color=RED, bold=True)

    sheet_level = wb.create_sheet('Level')
    sheet_level.add_data_validation(dv_sigref)
    dv_sigref.add('B2:B1048576')
    sheet_level.freeze_panes = 'A2'
    irow = 1
    level_title = ['LevelName', 'SigRef', 'VPS', 'IClamp', 'IRange', 'Delay', 'VIL', 'VIH', 'VTERM', 'VTERM_EN', 'VOL',
                   'VOH', 'VCH', 'VCL']
    for i in range(len(level_title)):
        sheet_level.cell(row=irow, column=(1 + i)).value = level_title[i]
    irow += 1
    for i in range(len(levels_data)):
        sheet_level.cell(row=irow, column=1).value = levels_data[i][0][1]
        for j in range(1, len(levels_data[i])):
            sheet_level.cell(row=irow, column=2).value = levels_data[i][j][0][1]
            for m in range(len(levels_data[i][j][1])):
                col = level_title.index(levels_data[i][j][1][m][0])
                sheet_level.cell(row=irow, column=(1 + col)).value = levels_data[i][j][1][m][1]
            irow += 1
        irow += 1
    for row in sheet_level.rows:
        for cell in row:
            cell.border = border
            cell.alignment = alignment
            if cell.row == 1:
                cell.fill = PatternFill(fill_type='solid', fgColor=YELLOW)
                if cell.column in (1, 2):
                    cell.font = Font(color=RED, bold=True)
                elif cell.column in (3, 4, 5, 6):
                    cell.font = Font(color='00CCCC', bold=True)  # 蓝色
                else:
                    cell.font = Font(color='666633', bold=True)  # 褐色

    for sheet_name in wb.sheetnames:
        if sheet_name == 'Sheet':
            del wb[sheet_name]
        else:
            Common.set_column_width(wb[sheet_name])

    print("保存数据开始-----------------")
    wb.save(file_name)
    print("保存数据结束-----------------")


def main():
    # Project folder path
    if argv.count('-p') == 0:
        print("Error：Project folder path 格式：“-p D:\Project” 。")
        exit()
    else:
        project_folder = argv[argv.index('-p') + 1]
        project_data = read_project(project_folder)
        socketmap_data = read_socketmap(project_folder)
        signals_data = read_signals(project_folder)
        limit_data = read_limit(project_folder)
        signalgroups_data = read_signalgroups(project_folder)
        bindefinition_data = read_bindefinition(project_folder)
        dcmeasure_data = read_dcmeasure(project_folder)
        tests_data = read_tests(project_folder)
        uservars_data = read_uservars(project_folder)
        levels_data = read_levels(project_folder)
    write_excel(project_data, socketmap_data, signals_data, limit_data, signalgroups_data, bindefinition_data,
                dcmeasure_data, tests_data, uservars_data, levels_data)


if __name__ == '__main__':
    main()
