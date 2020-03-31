# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'Datalog_Analysis.ui'
#
# Created by: PyQt5 UI code generator 5.11.3
#
# WARNING! All changes made in this file will be lost!

import Datalog_Analysis_UI
from sys import argv, exit, maxsize
from os import listdir, makedirs, getcwd
from csv import reader, field_size_limit
from PyQt5.QtCore import QThread, pyqtSignal, QRegExp
from PyQt5.QtGui import QIntValidator, QRegExpValidator
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QErrorMessage
from os.path import join, isdir, exists, basename, dirname
from math import isnan
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Side, Border
from openpyxl.styles.colors import BLACK, WHITE, RED, GREEN
from openpyxl.utils import column_index_from_string, get_column_letter
from datetime import datetime
from numpy import array

thin = Side(border_style='thin', color=BLACK)
border = Border(top=thin, left=thin, right=thin, bottom=thin)

hiLimitRowNum = 0
loLimitRowNum = 0
testDataRowNum = 0
testDataColLetter = 'A'
allTestData = False
beginColLetter = 'A'
endColLetter = 'A'
totalRowCount = 0
currentRowCount = 0
nowTime = 'Unknown'


def GetFileList(folder, postfix=None):
    """
    find a list of files with a postfix
    """
    fullNameList = []
    if isdir(folder):
        files = listdir(folder)
        for filename in files:
            fullNameList.append(join(folder, filename))
        if postfix:
            targetFileList = []
            for fullname in fullNameList:
                if fullname.endswith(postfix):
                    targetFileList.append(fullname)
            return targetFileList
        else:
            return fullNameList
    else:
        print("Error：Not a folder!")
        return False


def MkDir(path):
    """
    create a folder
    """
    path = path.strip()
    path = path.rstrip("\\")
    isExists = exists(path)
    if not isExists:
        makedirs(path)
        return True
    else:
        return False


def SearchString(data, target):
    """
    find out if target exists in data
    """
    for i in range(len(data)):
        try:
            colNum = data[i].index(target)
            rowNum = i
            return rowNum, colNum
        except:
            pass
    print("Can't find " + target + " !")
    return False


def ParseAndSave(file, _signal, analysisFolder):
    """
    parse and save file
    """
    global currentRowCount
    data = []
    # get file data
    with open(file, encoding='unicode_escape') as f:
        csvReader = reader(f)
        for row in csvReader:
            data.append(row)

    # data check----------------------------
    rowCount = len(data)

    dataCheckResult = []
    for rowNum in range(rowCount):
        rowData = []
        overLimitCount = 0
        for colNum in range(len(data[rowNum])):
            value = data[rowNum][colNum]
            highLimitData = data[hiLimitRowNum - 1][colNum]
            lowLimitData = data[loLimitRowNum - 1][colNum]
            if allTestData:
                beginColNum = column_index_from_string(testDataColLetter)
                endColNum = len(data[rowNum])
            else:
                beginColNum = column_index_from_string(beginColLetter)
                endColNum = column_index_from_string(endColLetter)
            if colNum in range(beginColNum - 1, endColNum) and rowNum in (hiLimitRowNum - 1, loLimitRowNum - 1):
                # fill color of limit value is green
                rowData.append([value, GREEN])
            elif colNum in range(beginColNum - 1, endColNum) and rowNum >= testDataRowNum - 1:
                try:
                    valueConvert = float(value)
                    try:
                        highLimit = float(highLimitData)
                    except:
                        highLimit = highLimitData
                    try:
                        lowLimit = float(lowLimitData)
                    except:
                        lowLimit = lowLimitData
                    if valueConvert == int(valueConvert):
                        valueConvert = int(valueConvert)
                    if highLimit == 'N' or (highLimit != 'N' and lowLimit <= valueConvert <= highLimit):
                        # fill color of following scenario value is white
                        # 1.limit is N
                        # 2.limit is not N and data wihtin limit
                        rowData.append([valueConvert, WHITE])
                    else:
                        # fill color of over limit is red
                        rowData.append([valueConvert, RED])
                        overLimitCount += 1
                except:
                    if not value.strip():
                        # fill color of '' is purple
                        rowData.append([value, 'A020F0'])
                    elif not highLimitData.strip():
                        # fill color is white
                        rowData.append([value, WHITE])
            else:
                # set '' when value is nan
                if isinstance(value, float) and isnan(value):
                    value = ''
                # fill color is white
                rowData.append([value, WHITE])
        if overLimitCount > 0:
            rowData[0][0] = rowData[0][0] + '(' + str(overLimitCount) + ')'
            rowData[0][1] = RED
        dataCheckResult.append(rowData)
        currentRowCount += 1
        if currentRowCount != totalRowCount:
            _signal.emit(str(currentRowCount * 100 // totalRowCount))

    fileName = basename(file).split('.')[0]
    dataCheckFile = join(analysisFolder, fileName + '_DataCheck_Analysis' + nowTime + '.xlsx')
    dataCheckWb = Workbook()  # create file object
    dataCheckSheet = dataCheckWb.active  # get first sheet
    freezePanes = testDataColLetter + str(testDataRowNum)
    dataCheckSheet.freezePanes = freezePanes  # set freeze panes

    irow = 1
    for i in range(len(dataCheckResult)):
        for j in range(len(dataCheckResult[i])):
            dataCheckSheet.cell(row=irow, column=j + 1).value = dataCheckResult[i][j][0]
            dataCheckSheet.cell(row=irow, column=j + 1).fill = PatternFill(fill_type='solid',
                                                                           fgColor=dataCheckResult[i][j][1])
            dataCheckSheet.cell(row=irow, column=j + 1).border = border
        irow += 1
        currentRowCount += 1
        if currentRowCount != totalRowCount:
            _signal.emit(str(currentRowCount * 100 // totalRowCount))
    dataCheckWb.save(dataCheckFile)


class Runthread(QThread):
    _signal = pyqtSignal(str)

    def __init__(self, openPath, chooseRadio):
        super(Runthread, self).__init__()
        self.openPath = openPath
        self.chooseRadio = chooseRadio

    def __del__(self):
        self.wait()

    def run(self):
        if self.chooseRadio == 'DateFolder':
            lotnoNames = listdir(self.openPath)
            for name in lotnoNames:
                folder = join(self.openPath, name)
                if isdir(folder):
                    # get CSV file under the folder
                    fileList = GetFileList(folder, '.csv')

                    # create analysis folder
                    analysisFolder = folder + '\Analysis'
                    MkDir(analysisFolder)

                    for file in fileList:
                        ParseAndSave(file, self._signal, analysisFolder)
        elif self.chooseRadio == 'LotFolder':
            # get CSV file under the folder
            fileList = GetFileList(self.openPath, '.csv')

            # create analysis folder
            analysisFolder = self.openPath + '\Analysis'
            MkDir(analysisFolder)

            for file in fileList:
                ParseAndSave(file, self._signal, analysisFolder)
        else:
            # create analysis folder
            analysisFolder = dirname(self.openPath) + '\Analysis'
            MkDir(analysisFolder)
            ParseAndSave(self.openPath, self._signal, analysisFolder)
        self._signal.emit(str(100))


class MainWindow(QMainWindow, Datalog_Analysis_UI.Ui_MainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        self.setupUi(self)
        self.cwd = getcwd()
        self.qe = QErrorMessage(self)
        self.DateFolder_radioButton.setChecked(True)
        self.Open_pushButton.clicked.connect(self.Open)
        self.HiLimitRowNum_lineEdit.setValidator(QIntValidator())
        self.LoLimitRowNum_lineEdit.setValidator(QIntValidator())
        self.TestDataRowNum_lineEdit.setValidator(QIntValidator())
        reg = QRegExp('[A-Z]+')
        pValidator = QRegExpValidator(self)
        pValidator.setRegExp(reg)
        self.TestDataColLetter_lineEdit.setValidator(pValidator)
        self.AllTestData_radioButton.setChecked(True)
        self.AllTestData_radioButton.toggled.connect(self.BtnState)
        self.BeginColLetter_lineEdit.setValidator(pValidator)
        self.EndColLetter_lineEdit.setValidator(pValidator)
        if self.AllTestData_radioButton.isChecked():
            self.BeginColLetter_lineEdit.setEnabled(False)
            self.EndColLetter_lineEdit.setEnabled(False)
        self.Analysis_pushButton.clicked.connect(self.Analysis)

    def BtnState(self):
        if self.AllTestData_radioButton.isChecked():
            self.BeginColLetter_lineEdit.setEnabled(False)
            self.EndColLetter_lineEdit.setEnabled(False)
        else:
            self.BeginColLetter_lineEdit.setEnabled(True)
            self.EndColLetter_lineEdit.setEnabled(True)

    def Open(self):
        global totalRowCount
        totalRowCount = 0
        if self.SingleFile_radioButton.isChecked():
            fileNameChoose, fileType = QFileDialog.getOpenFileName(self, 'Select CSV File', self.cwd,
                                                                   'CSV Files(*.csv);;All Files(*)')
            if not fileNameChoose:
                return
            self.OpenPath_lineEdit.setText(fileNameChoose)
            with open(fileNameChoose, encoding='unicode_escape') as f:
                csvReader = reader(f)
                totalRowCount += (array(list(csvReader)).shape[0] * 2)
            getParaFile = fileNameChoose
        else:
            dirChoose = QFileDialog.getExistingDirectory(self, 'Select Directory', self.cwd)
            if not dirChoose:
                return
            self.OpenPath_lineEdit.setText(dirChoose)
            if self.DateFolder_radioButton.isChecked():
                lotnoNames = listdir(self.OpenPath_lineEdit.text())
                for name in lotnoNames:
                    folder = join(self.OpenPath_lineEdit.text(), name)
                    if isdir(folder):
                        # get CSV file under the folder
                        fileList = GetFileList(folder, '.csv')
                        if not fileList:
                            return
                        for file in fileList:
                            getParaFile = file
                            with open(file, encoding='unicode_escape') as f:
                                csvReader = reader(f)
                                totalRowCount += (array(list(csvReader)).shape[0] * 2)
            elif self.LotFolder_radioButton.isChecked():
                fileList = GetFileList(self.OpenPath_lineEdit.text(), '.csv')
                if not fileList:
                    return
                for file in fileList:
                    getParaFile = file
                    with open(file, encoding='unicode_escape') as f:
                        csvReader = reader(f)
                        totalRowCount += (array(list(csvReader)).shape[0] * 2)
        data = []
        with open(getParaFile, encoding='unicode_escape') as f:
            csvReader = reader(f)
            for row in csvReader:
                data.append(row)
        for i in range(len(data)):
            if data[i][0] == 'MAX':
                self.HiLimitRowNum_lineEdit.setText(str(i + 1))
            elif data[i][0] == 'MIN':
                self.LoLimitRowNum_lineEdit.setText(str(i + 1))
            else:
                try:
                    int(data[i][0])
                    self.TestDataRowNum_lineEdit.setText(str(i + 1))
                    for j in range(len(data[i])):
                        if type(eval(data[i][j])) == float:
                            self.TestDataColLetter_lineEdit.setText(get_column_letter(j + 1))
                            return
                except:
                    pass

    def CallBackLog(self, msg):
        self.progressBar.setValue(int(msg))  # pass the thread's parameters to progressBar
        if msg == '100':
            self.Analysis_pushButton.setEnabled(True)

    def Analysis(self):
        if not self.OpenPath_lineEdit.text():
            self.qe.showMessage('Path cannot be empty!')
            return
        elif not self.HiLimitRowNum_lineEdit.text():
            self.qe.showMessage('hiLimitRowNum cannot be empty!Input limit:Integer only.')
            return
        elif not self.LoLimitRowNum_lineEdit.text():
            self.qe.showMessage('loLimitRowNum cannot be empty!Input limit:Integer only.')
            return
        elif not self.TestDataRowNum_lineEdit.text():
            self.qe.showMessage('testDataRowNum cannot be empty!Input limit:Integer only.')
            return
        elif not self.TestDataColLetter_lineEdit.text():
            self.qe.showMessage('testDataColLetter cannot be empty!Input limit:Uppercase only.')
            return
        else:
            if totalRowCount == 0:
                self.qe.showMessage('The test data in open path is illegal.')
                return
            global nowTime
            nowTime = datetime.now().strftime("%Y%m%d%H%M%S")
            global currentRowCount
            currentRowCount = 0
            global hiLimitRowNum
            hiLimitRowNum = int(self.HiLimitRowNum_lineEdit.text())
            global loLimitRowNum
            loLimitRowNum = int(self.LoLimitRowNum_lineEdit.text())
            global testDataRowNum
            testDataRowNum = int(self.TestDataRowNum_lineEdit.text())
            global testDataColLetter
            testDataColLetter = self.TestDataColLetter_lineEdit.text()
            global allTestData
            if self.AllTestData_radioButton.isChecked():
                allTestData = True
            else:
                global beginColLetter
                global endColLetter
                beginColLetter = self.BeginColLetter_lineEdit.text()
                endColLetter = self.EndColLetter_lineEdit.text()
            if self.DateFolder_radioButton.isChecked():
                chooseRadio = self.DateFolder_radioButton.text()
            elif self.LotFolder_radioButton.isChecked():
                chooseRadio = self.LotFolder_radioButton.text()
            else:
                chooseRadio = self.SingleFile_radioButton.text()
            # create thread
            self.Analysis_pushButton.setEnabled(False)
            self.thread = Runthread(self.OpenPath_lineEdit.text(), chooseRadio)
            # connect signal
            self.thread._signal.connect(self.CallBackLog)
            self.thread.start()


if __name__ == '__main__':
    # _csv.Error:field larger than field limit(131072)
    maxInt = maxsize
    while True:
        try:
            field_size_limit(maxInt)
            break
        except OverflowError:
            maxInt = int(maxInt / 10)
    app = QApplication(argv)
    myMainWindow = MainWindow()
    myMainWindow.show()
    exit(app.exec_())
