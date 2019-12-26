# encoding:utf-8
# @Time     : 2019/9/17
# @Author   : Jerry Chou
# @File     :
# @Function : 多文件HWBIN、SWBIN分析

from csv import reader
from os.path import basename, dirname, abspath, join
from os import getcwd
from datetime import datetime
from sys import argv, path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.styles.colors import YELLOW, GREEN, BLACK, WHITE, RED
from progressbar import *

path.append(abspath(join(getcwd(), '..')))
import Common

alignment = Alignment(horizontal='center', vertical='center')
thin = Side(border_style='thin', color=BLACK)
border = Border(top=thin, left=thin, right=thin, bottom=thin)

row_offset = 5
col_offset = 4


def parse_file(file, group_by_id, project_id):
    """
    解析文件
    :param project_id: 项目id
    :param file: datalog csv文件
    :param group_by_id: 列序号（1：Site，2：SW_BIN，3：hW_BIN）
    :return: 
    """
    data = []
    file_name = basename(file).split('.')[0]
    with open(file) as f:
        csv_reader = reader(f)
        for row in csv_reader:
            data.append(row)
    if project_id == 0:
        search_last_test_item = Common.search_string(data, 'Full_Error')
    elif project_id == 1:
        search_last_test_item = Common.search_string(data, 'ChipVer')
    if not search_last_test_item:
        exit()
    else:
        last_test_item_row_num = search_last_test_item[0]
        last_test_item_col_num = search_last_test_item[1]

    first_register_row_num = len(data) - 1
    for i in range(last_test_item_row_num + row_offset, len(data)):
        try:
            int(data[i][0])
        except:
            first_register_row_num = i
            break
    group_list = []
    for i in range(last_test_item_row_num + row_offset, first_register_row_num):
        group_list.append(int(data[i][group_by_id]))
    format_group_list = list(set(group_list))
    format_group_list.sort()

    group_index = []
    for i in format_group_list:
        temp = [i]
        data_list = Common.find_item(group_list, i)
        if group_by_id == 1:
            temp_list = []
            for j in data_list:
                temp_list.append(data[last_test_item_row_num + row_offset + j][2])
            format_temp_list = list(set(temp_list))
            temp_index = []
            for m in format_temp_list:
                temp_swbin = [m]
                temp_data_list = Common.find_item(temp_list, m)
                temp_swbin.append(temp_data_list)
                temp_index.append(temp_swbin)
            temp.append(temp_index)
        else:
            temp.append(data_list)
        testitem_fail_count = []
        for col_num in range(col_offset, last_test_item_col_num + 1):
            test_item_name = data[last_test_item_row_num][col_num].strip()
            if test_item_name == 'AVDD_O/S':  # JX828测试项特殊处理
                high_limit_data = -0.2
                low_limit_data = -0.6
            else:
                high_limit_data = data[last_test_item_row_num + 2][col_num]
                low_limit_data = data[last_test_item_row_num + 3][col_num]
            temp_list = [data[last_test_item_row_num][col_num], 0]
            try:
                high_limit = float(high_limit_data)
                low_limit = float(low_limit_data)
                for j in data_list:
                    try:
                        value = data[last_test_item_row_num + row_offset + j][col_num]
                        value_convert = float(value)
                        if value_convert == int(value_convert):
                            value_convert = int(value_convert)
                        if value_convert < low_limit or high_limit < value_convert:
                            temp_list[1] += 1
                    except:
                        temp_list[1] += 1
            except:
                continue
            testitem_fail_count.append(temp_list)

        temp.append(testitem_fail_count)
        group_index.append(temp)
    chip_count = first_register_row_num - last_test_item_row_num - row_offset
    lotno = data[5][1]
    return file_name, group_index, chip_count, lotno


def save_data(analysis_folder, site_data, softbin_data, hardbin_data, project_id):
    """
    写数据
    :param project_id: 项目id
    :param analysis_folder: 保存分析文件夹路径
    :param site_data: site数据
    :param softbin_data: softbin数据
    :param hardbin_data: hardbin数据
    :return: 
    """
    date = datetime.now().strftime("%Y%m%d%H%M")
    test_file_name = analysis_folder + '/Summary_Analysis_' + date + '.xlsx'
    wb = Workbook()

    hwbin_to_swbin_list = [
        [
            [1, [1, 2]],
            [2, [41, 42, 43, 44, 45, 53, 54, 55]],
            [4, [23, 24, 25, 29, 30, 33, 34, 36, 39, 40, 70, 71, 72]],
            [5, [5, 6, 7, 8, 9, 12, 96, 97, 98, 99]],
            [6, [13, 14, 15, 35]],
            [8, [26, 27, 31, 32]]  # 专门分出来测试良率
        ],
        [
            [3, [1, 2, 3]],
            [1, [63, 64, 65, 89, 90, 94]],
            [2, [53, 54, 73, 74]],
            [4, [23, 24, 25, 26, 27, 31, 32, 36, 56, 57, 58, 75]],
            [5, [5, 6, 7, 8, 9, 12, 93, 96, 98, 99]],
            [6, [13, 14, 15, 46, 47, 48, 51, 60]],
            [8, [29, 30, 33, 34, 36, 39, 40]]
        ]
    ]
    if project_id == 0:
        ok_hwbin_count = 2
    elif project_id == 1:
        ok_hwbin_count = 3
    begin_fail_swbin = 0
    for i in range(ok_hwbin_count):
        begin_fail_swbin += len(hwbin_to_swbin_list[project_id][i][1])

    color_list = ['99FFFF', '33FF00', 'FFFFCC', 'FFFF33', 'FF9900', '9900CC', 'FF0000']
    swbin_list = []
    for i in range(len(hwbin_to_swbin_list[project_id])):
        for j in range(len(hwbin_to_swbin_list[project_id][i][1])):
            swbin_list.append([hwbin_to_swbin_list[project_id][i][1][j], color_list[i]])

    hardbin_sheet = wb.create_sheet('HWBin')
    hardbin_sheet.freeze_panes = 'B2'
    summary_data = []
    for i in range(len(hardbin_data)):
        temp_list = []
        if i == 0:
            temp_list.append('FT')
        else:
            temp_list.append('RT' + str(i))
        test_count = 0
        for hw_bin in hwbin_to_swbin_list[project_id]:
            for j in range(len(hardbin_data[i][1])):
                if hw_bin[0] == hardbin_data[i][1][j][0]:
                    bin_count = len(hardbin_data[i][1][j][1])
                    break
                else:
                    bin_count = 0
            test_count += bin_count
            temp_list.append(bin_count)
        temp_list.append(test_count)
        pass_count = 0
        for m in range(ok_hwbin_count):
            pass_count += temp_list[1 + m]
        pass_percent = '{:.2%}'.format(pass_count / test_count)
        temp_list.append(pass_percent)
        summary_data.append(temp_list)
    irow = 1
    icol = 1
    hardbin_sheet.cell(row=irow, column=icol).value = softbin_data[0][3]
    icol += 1
    for i in range(len(hwbin_to_swbin_list[project_id])):
        hardbin_sheet.cell(row=irow, column=icol).value = 'HWBin' + str(hwbin_to_swbin_list[project_id][i][0])
        icol += 1
    hardbin_sheet.cell(row=irow, column=icol).value = 'TestCount'
    icol += 1
    hardbin_sheet.cell(row=irow, column=icol).value = 'PassPercent'
    irow += 1
    for i in range(len(summary_data)):
        for j in range(len(summary_data[i])):
            hardbin_sheet.cell(row=irow, column=(j + 1)).value = summary_data[i][j]
            if j == 0:
                hardbin_sheet.cell(row=irow, column=(j + 1)).fill = PatternFill(fill_type='solid', fgColor=GREEN)
            elif 1 <= j < len(summary_data[i]) - 2:
                hardbin_sheet.cell(row=irow + 1, column=(j + 1)).value = '{:.2%}'.format(
                    summary_data[i][j] / summary_data[i][-2])
                hardbin_sheet.cell(row=irow + 1, column=(j + 1)).fill = PatternFill(fill_type='solid', fgColor=GREEN)
            elif j == len(summary_data[i]) - 2 and i > 0:
                compare_count = 0
                for m in range(ok_hwbin_count + 1, len(summary_data[i]) - 2):
                    compare_count += summary_data[i - 1][m]
                if summary_data[i][j] != compare_count:
                    hardbin_sheet.cell(row=irow, column=(j + 1)).fill = PatternFill(fill_type='solid', fgColor=RED)
        irow += 2
    hardbin_sheet.cell(row=irow, column=1).value = 'Summary'
    hardbin_sheet.cell(row=irow, column=1).fill = PatternFill(fill_type='solid', fgColor=GREEN)
    summary_bin_count = [0] * ok_hwbin_count
    for i in range(len(summary_data)):
        for j in range(len(summary_bin_count)):
            summary_bin_count[j] += summary_data[i][j + 1]
    pass_bin_count = 0
    for i in range(len(summary_bin_count)):
        hardbin_sheet.cell(row=irow, column=2 + i).value = summary_bin_count[i]
        pass_bin_count += summary_bin_count[i]
    for i in range(len(summary_bin_count) + 1, len(summary_data[-1]) - 2):
        hardbin_sheet.cell(row=irow, column=(i + 1)).value = summary_data[-1][i]
    hardbin_sheet.cell(row=irow, column=icol - 1).value = summary_data[0][-2]
    hardbin_sheet.cell(row=irow, column=icol).value = '{:.2%}'.format(pass_bin_count / summary_data[0][-2])

    for row in hardbin_sheet.rows:
        for cell in row:
            cell.border = border
            if cell.row == 1:
                cell.fill = PatternFill(fill_type='solid', fgColor=YELLOW)
                cell.font = Font(bold=True)
                cell.alignment = alignment
            if cell.row == 2 and cell.column == hardbin_sheet.max_column:
                cell.fill = PatternFill(fill_type='solid', fgColor=GREEN)
            if cell.row == hardbin_sheet.max_row and cell.column == hardbin_sheet.max_column:
                cell.fill = PatternFill(fill_type='solid', fgColor=GREEN)

    hardsoftbin_sheet = wb.create_sheet('HWBin-SWBin')
    hardsoftbin_sheet.freeze_panes = 'C2'
    irow = 1
    hardsoftbin_sheet.cell(row=irow, column=1).value = softbin_data[0][3]
    hardsoftbin_sheet.cell(row=irow, column=1).fill = PatternFill(fill_type='solid', fgColor=YELLOW)
    hardsoftbin_sheet.cell(row=irow, column=3).value = 'FT'
    hardsoftbin_sheet.cell(row=irow, column=3).fill = PatternFill(fill_type='solid', fgColor=YELLOW)
    hardsoftbin_sheet.merge_cells(start_row=irow, end_row=irow, start_column=3, end_column=4)
    for i in range(1, len(softbin_data)):
        hardsoftbin_sheet.cell(row=irow, column=3 + 2 * i).value = 'RT' + str(i)
        hardsoftbin_sheet.cell(row=irow, column=3 + 2 * i).fill = PatternFill(fill_type='solid', fgColor=YELLOW)
        hardsoftbin_sheet.merge_cells(start_row=irow, end_row=irow, start_column=3 + 2 * i, end_column=4 + 2 * i)
    irow += 1

    swbin_sort_by_FT_list = []
    for i in range(len(hwbin_to_swbin_list[project_id])):
        temp_list = []
        for j in range(len(hwbin_to_swbin_list[project_id][i][1])):
            find_softbin = False
            for x in range(len(softbin_data[0][1])):
                if hwbin_to_swbin_list[project_id][i][1][j] == softbin_data[0][1][x][0]:
                    find_softbin = True
                    swbin_count = len(softbin_data[0][1][x][1])
                    temp_list.append((j, swbin_count))
            if not find_softbin:
                temp_list.append((j, 0))
        for m in range(len(temp_list) - 1):
            for n in range(m + 1, len(temp_list)):
                if temp_list[m][1] < temp_list[n][1]:
                    temp_list[m], temp_list[n] = temp_list[n], temp_list[m]
        temp_count_list = [hwbin_to_swbin_list[project_id][i][0]]
        for y in range(len(temp_list)):
            temp_count_list.append(hwbin_to_swbin_list[project_id][i][1][temp_list[y][0]])
        swbin_sort_by_FT_list.append(temp_count_list)

    summary_soft_count = []
    for i in range(len(swbin_sort_by_FT_list)):
        hardsoftbin_sheet.cell(row=irow, column=1).value = 'HWBin' + str(swbin_sort_by_FT_list[i][0])
        hardsoftbin_sheet.cell(row=irow, column=1).fill = PatternFill(fill_type='solid', fgColor=GREEN)
        hardsoftbin_sheet.cell(row=irow, column=1).font = Font(bold=True)
        for j in range(1, len(swbin_sort_by_FT_list[i])):
            hardsoftbin_sheet.cell(row=irow, column=2).value = 'SWBin' + str(swbin_sort_by_FT_list[i][j])
            temp_count = 0
            for m in range(len(softbin_data)):
                find_softbin = False
                for n in range(len(softbin_data[m][1])):
                    if swbin_sort_by_FT_list[i][j] == softbin_data[m][1][n][0]:
                        find_softbin = True
                        swbin_count = len(softbin_data[m][1][n][1])
                if not find_softbin:
                    swbin_count = 0
                if i in (0, 1):
                    temp_count += swbin_count
                elif m == len(softbin_data) - 1:
                    temp_count = swbin_count
                hardsoftbin_sheet.cell(row=irow, column=(2 * m + 3)).value = swbin_count
                hardsoftbin_sheet.cell(row=irow, column=(2 * m + 4)).value = '{:.2%}'.format(
                    swbin_count / summary_data[0][-2])  # 当次测试总数 summary_data[m][6])
                hardsoftbin_sheet.cell(row=irow, column=(2 * m + 4)).fill = PatternFill(fill_type='solid',
                                                                                        fgColor=GREEN)
            irow += 1
            summary_soft_count.append(temp_count)
        summary_soft_count.append('')
        irow += 1

    irow = 1
    icol = 4 + 2 * len(softbin_data)
    hardsoftbin_sheet.cell(row=irow, column=icol).value = 'Summary'
    hardsoftbin_sheet.cell(row=irow, column=icol).fill = PatternFill(fill_type='solid', fgColor='FFA500')
    hardsoftbin_sheet.merge_cells(start_row=irow, end_row=irow, start_column=icol, end_column=icol + 1)
    irow += 1
    for i in range(len(summary_soft_count)):
        if isinstance(summary_soft_count[i], int):
            hardsoftbin_sheet.cell(row=irow, column=icol).value = summary_soft_count[i]
            hardsoftbin_sheet.cell(row=irow, column=icol + 1).value = '{:.2%}'.format(
                summary_soft_count[i] / summary_data[0][-2])
            hardsoftbin_sheet.cell(row=irow, column=icol + 1).fill = PatternFill(fill_type='solid', fgColor='FFA500')
        irow += 1

    for row in hardsoftbin_sheet.rows:
        for cell in row:
            cell.border = border
            if cell.row == 1:
                cell.font = Font(bold=True)
                cell.alignment = alignment

    sitesoftbin_sheet = wb.create_sheet('Site-SWBin')
    sitesoftbin_sheet.freeze_panes = 'B2'
    irow = 1
    sitesoftbin_sheet.cell(row=irow, column=1).value = site_data[0][3]
    sitesoftbin_sheet.cell(row=irow, column=1).fill = PatternFill(fill_type='solid', fgColor=YELLOW)
    for i in range(len(site_data[0][1])):
        sitesoftbin_sheet.cell(row=irow, column=2 + i).value = 'Site' + str(site_data[0][1][i][0])
        sitesoftbin_sheet.cell(row=irow, column=2 + i).fill = PatternFill(fill_type='solid', fgColor=YELLOW)
    sitesoftbin_sheet.merge_cells(start_row=irow, end_row=irow, start_column=3 + len(site_data[0][1]),
                                  end_column=4 + len(site_data[0][1]))
    sitesoftbin_sheet.cell(row=irow, column=3 + len(site_data[0][1])).value = 'Summary'
    sitesoftbin_sheet.cell(row=irow, column=3 + len(site_data[0][1])).fill = PatternFill(fill_type='solid',
                                                                                         fgColor='FFA500')
    irow += 1

    site_swbin_count = []
    for x in range(len(swbin_list)):
        temp_total_count = 0
        temp_list = []
        for i in range(len(site_data[0][1])):
            find_softbin = False
            for j in range(len(site_data[0][1][i][1])):
                if swbin_list[x][0] == int(site_data[0][1][i][1][j][0]):
                    temp_swbin_count = len(site_data[0][1][i][1][j][1])
                    find_softbin = True
            if not find_softbin:
                temp_swbin_count = 0
            temp_total_count += temp_swbin_count
            temp_list.append([temp_swbin_count, WHITE])
            if i == len(site_data[0][1]) - 1:
                temp_list.append([temp_total_count, 'FFA500'])
        site_swbin_count.append(temp_list)
    site_fail_total_list = [0] * 16
    for i in range(begin_fail_swbin, len(site_swbin_count)):
        min_value = site_swbin_count[i][0][0]
        max_value = site_swbin_count[i][0][0]
        for m in range(1, len(site_swbin_count[i]) - 1):
            if min_value > site_swbin_count[i][m][0]:
                min_value = site_swbin_count[i][m][0]
            if max_value < site_swbin_count[i][m][0]:
                max_value = site_swbin_count[i][m][0]
        min_index = []
        max_index = []
        for j in range(len(site_swbin_count[i]) - 1):
            site_fail_total_list[j] += site_swbin_count[i][j][0]
            if site_swbin_count[i][j][0] == min_value:
                min_index.append(j)
            if site_swbin_count[i][j][0] == max_value:
                max_index.append(j)
        if min_value == 0 and len(min_index) == 1:
            site_swbin_count[i][min_index[0]][1] = GREEN
            for y in range(len(max_index)):
                site_swbin_count[i][max_index[y]][1] = RED
        elif min_value == 0 and max_value > 0:
            for y in range(len(max_index)):
                site_swbin_count[i][max_index[y]][1] = RED
        elif min_value > 0:
            for x in range(len(min_index)):
                site_swbin_count[i][min_index[x]][1] = GREEN
            for y in range(len(max_index)):
                site_swbin_count[i][max_index[y]][1] = RED

    for x in range(len(swbin_list)):
        sitesoftbin_sheet.cell(row=irow, column=1).value = 'SWBin' + str(swbin_list[x][0])
        sitesoftbin_sheet.cell(row=irow, column=1).fill = PatternFill(fill_type='solid', fgColor=swbin_list[x][1])
        for y in range(len(site_swbin_count[x])):
            if y < len(site_swbin_count[x]) - 1:
                if site_swbin_count[x][y][0] > 0:
                    sitesoftbin_sheet.cell(row=irow, column=2 + y).value = '{:.4%}'.format(
                        site_swbin_count[x][y][0] / summary_data[0][-2])
                sitesoftbin_sheet.cell(row=irow, column=2 + y).fill = PatternFill(fill_type='solid',
                                                                                  fgColor=site_swbin_count[x][y][1])
            else:
                sitesoftbin_sheet.cell(row=irow, column=3 + y).value = site_swbin_count[x][y][0]
                sitesoftbin_sheet.cell(row=irow, column=4 + y).value = '{:.4%}'.format(
                    site_swbin_count[x][y][0] / summary_data[0][-2])
                sitesoftbin_sheet.cell(row=irow, column=4 + y).fill = PatternFill(fill_type='solid',
                                                                                  fgColor=site_swbin_count[x][y][1])
        irow += 1
    irow += 1
    sitesoftbin_sheet.merge_cells(start_row=irow, end_row=irow + 1, start_column=1, end_column=1)
    sitesoftbin_sheet.cell(row=irow, column=1).value = 'FailPercent'
    sitesoftbin_sheet.cell(row=irow, column=1).font = Font(bold=True)
    sitesoftbin_sheet.cell(row=irow, column=1).alignment = alignment
    sitesoftbin_sheet.cell(row=irow, column=1).fill = PatternFill(fill_type='solid', fgColor='FFA500')
    for i in range(len(site_fail_total_list)):
        sitesoftbin_sheet.cell(row=irow, column=2 + i).value = site_fail_total_list[i]
        sitesoftbin_sheet.cell(row=irow + 1, column=2 + i).value = '{:.4%}'.format(
            site_fail_total_list[i] / summary_data[0][-2])
        sitesoftbin_sheet.cell(row=irow + 1, column=2 + i).fill = PatternFill(fill_type='solid', fgColor='FFA500')

    for row in sitesoftbin_sheet.rows:
        for cell in row:
            cell.border = border
            if cell.row == 1:
                cell.font = Font(bold=True)
                cell.alignment = alignment

    hwbin_testitem_sheet = wb.create_sheet('HWBin-TestItem')
    hwbin_testitem_sheet.freeze_panes = 'B2'
    irow = 1
    hwbin_testitem_sheet.cell(row=irow, column=1).value = hardbin_data[0][2]
    for i in range(len(hardbin_data[0][1][0][2])):
        hwbin_testitem_sheet.cell(row=irow, column=i + 2).value = hardbin_data[0][1][0][2][i][0]
    irow += 1
    for i in range(len(hardbin_data[0][1])):
        hwbin_testitem_sheet.cell(row=irow, column=1).value = 'HWBin' + str(hardbin_data[0][1][i][0])
        hwbin_testitem_sheet.cell(row=irow, column=1).fill = PatternFill(fill_type='solid', fgColor=GREEN)
        for j in range(len(hardbin_data[0][1][i][2])):
            hwbin_testitem_sheet.cell(row=irow, column=2 + j).value = hardbin_data[0][1][i][2][j][1]
            hwbin_testitem_sheet.cell(row=irow + 1, column=2 + j).value = '{:.2%}'.format(
                hardbin_data[0][1][i][2][j][1] / hardbin_data[0][2])
            if hardbin_data[0][1][i][2][j][1] > 0:
                hwbin_testitem_sheet.cell(row=irow + 1, column=2 + j).fill = PatternFill(fill_type='solid',
                                                                                         fgColor=RED)
        irow += 2

    for row in hwbin_testitem_sheet.rows:
        for cell in row:
            cell.border = border
            if cell.row == 1 and cell.column > 1:
                cell.fill = PatternFill(fill_type='solid', fgColor=YELLOW)
                cell.font = Font(bold=True)
                cell.alignment = alignment

    swbin_testitem_sheet = wb.create_sheet('SWBin-TestItem')
    swbin_testitem_sheet.freeze_panes = 'B2'
    irow = 1
    swbin_testitem_sheet.cell(row=irow, column=1).value = softbin_data[0][2]
    for i in range(len(softbin_data[0][1][0][2])):
        swbin_testitem_sheet.cell(row=irow, column=i + 2).value = softbin_data[0][1][0][2][i][0]
    irow += 1
    for x in range(len(swbin_list)):
        swbin_testitem_sheet.cell(row=irow, column=1).value = 'SWBin' + str(swbin_list[x][0])
        swbin_testitem_sheet.cell(row=irow, column=1).fill = PatternFill(fill_type='solid', fgColor=swbin_list[x][1])
        for i in range(len(softbin_data[0][1])):
            if swbin_list[x][0] == softbin_data[0][1][i][0]:
                for j in range(len(softbin_data[0][1][i][2])):
                    swbin_testitem_sheet.cell(row=irow, column=2 + j).value = softbin_data[0][1][i][2][j][1]
                    swbin_testitem_sheet.cell(row=irow + 1, column=2 + j).value = '{:.2%}'.format(
                        softbin_data[0][1][i][2][j][1] / softbin_data[0][2])
                    if softbin_data[0][1][i][2][j][1] > 0:
                        swbin_testitem_sheet.cell(row=irow + 1, column=2 + j).fill = PatternFill(fill_type='solid',
                                                                                                 fgColor=RED)
        irow += 2

    for row in swbin_testitem_sheet.rows:
        for cell in row:
            cell.border = border
            if cell.row == 1 and cell.column > 1:
                cell.fill = PatternFill(fill_type='solid', fgColor=YELLOW)
                cell.font = Font(bold=True)
                cell.alignment = alignment

    site_testitem_sheet = wb.create_sheet('Site-TestItem')
    site_testitem_sheet.freeze_panes = 'B2'
    irow = 1
    site_testitem_sheet.cell(row=irow, column=1).value = site_data[0][2]
    for i in range(len(site_data[0][1][0][2])):
        site_testitem_sheet.cell(row=irow, column=i + 2).value = site_data[0][1][0][2][i][0]
    irow += 1
    for i in range(len(site_data[0][1])):
        site_testitem_sheet.cell(row=irow, column=1).value = 'Site' + str(site_data[0][1][i][0])
        site_testitem_sheet.cell(row=irow, column=1).fill = PatternFill(fill_type='solid',
                                                                        fgColor=GREEN)
        for j in range(len(site_data[0][1][i][2])):
            site_testitem_sheet.cell(row=irow, column=2 + j).value = site_data[0][1][i][2][j][1]
            site_testitem_sheet.cell(row=irow + 1, column=2 + j).value = '{:.2%}'.format(
                site_data[0][1][i][2][j][1] / site_data[0][2])
            if site_data[0][1][i][2][j][1] > 0:
                site_testitem_sheet.cell(row=irow + 1, column=2 + j).fill = PatternFill(fill_type='solid',
                                                                                        fgColor=RED)
        irow += 2

    for row in site_testitem_sheet.rows:
        for cell in row:
            cell.border = border
            if cell.row == 1 and cell.column > 1:
                cell.fill = PatternFill(fill_type='solid', fgColor=YELLOW)
                cell.font = Font(bold=True)
                cell.alignment = alignment

    for sheet_name in wb.sheetnames:
        if sheet_name == 'Sheet':
            del wb[sheet_name]
        else:
            Common.set_column_width(wb[sheet_name])

    wb.save(test_file_name)


def main():
    # Sourcefile folder path
    if argv.count('-s') == 0 and argv.count('-f') == 0:
        print(
            "Error：Sourcefile folder path 或 single file path为必填项，格式：“-s D:\sourcefile” 或 “-f D:\sourcefile\ST5-440mm-out1.csv”。")
        exit()

    if argv.count('-s') != 0:
        sourcefile_folder = argv[argv.index('-s') + 1]
        for i in range(argv.index('-s') + 2, len(argv)):
            if not argv[i].startswith('-'):
                sourcefile_folder += (' ' + argv[i])
            else:
                break
        file_list = Common.get_filelist(sourcefile_folder, '.csv')
        if not file_list:
            exit()

    if argv.count('-f') != 0:
        single_file = argv[argv.index('-f') + 1]
        for i in range(argv.index('-f') + 2, len(argv)):
            if not argv[i].startswith('-'):
                single_file += (' ' + argv[i])
            else:
                break
        sourcefile_folder = dirname(single_file)
        file_list = [single_file]

    # Analysis folder path
    if argv.count('-a') == 0:
        analysis_folder = sourcefile_folder + '\Analysis'
    else:
        analysis_folder = argv[argv.index('-a') + 1]

    Common.mkdir(analysis_folder)

    # 项目 0:F28,1:JX828
    if argv.count('-p') != 0:
        project_id = int(argv[argv.index('-p') + 1])
    else:
        project_id = 0  # 默认F28项目

    hardbin_data = []
    softbin_data = []
    site_data = []
    parse_file_widgets = ['ParseFile: ', Percentage(), ' ', Bar('#'), ' ', Timer(), ' ', ETA(), ' ',
                          FileTransferSpeed()]
    parse_file_pbar = ProgressBar(widgets=parse_file_widgets, maxval=len(file_list)).start()
    for i in range(len(file_list)):
        # parse file
        site_data.append(parse_file(file_list[i], 1, project_id))
        softbin_data.append(parse_file(file_list[i], 2, project_id))
        hardbin_data.append(parse_file(file_list[i], 3, project_id))
        parse_file_pbar.update(i + 1)
    parse_file_pbar.finish()
    for data in softbin_data:
        Common.sort_data(data[1])

    # save data
    save_data(analysis_folder, site_data, softbin_data, hardbin_data, project_id)


if __name__ == '__main__':
    main()
