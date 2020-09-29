# encoding:utf-8
# @Time     : 2019/12/2
# @Author   : Jerry Chou
# @File     :
# @Function : analysis all data

import Datalog_Total_Analysis_UI
from csv import reader, field_size_limit
from os.path import basename, dirname, exists, join, isdir
from os import listdir, makedirs, getcwd
from datetime import datetime
from sys import argv, exit, maxsize
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.styles.colors import YELLOW, GREEN, BLACK, WHITE, RED
from openpyxl.utils import get_column_letter
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QErrorMessage
from PyQt5.QtCore import QThread, pyqtSignal
from PyQt5.QtGui import QBrush, QPixmap, QPalette


def FindItem(itemList, value):
    """
    find value in item list
    """
    return [i for i, v in enumerate(itemList) if v == value]


def SearchString(data, target):
    """
    find out if target exists in data
    """
    for i in range(len(data)):
        try:
            colNum = data[i].index(target)
            rowNum = i
            return rowNum, colNum
        except:
            pass
    print("Can't find " + target + " !")
    return False


def SetColumnWidth(sheet):
    """
    set column width
    """
    # get the maximum width of each column
    colWidth = [0.5] * sheet.max_column
    for row in range(sheet.max_row):
        for col in range(sheet.max_column):
            value = sheet.cell(row=row + 1, column=col + 1).value
            if value:
                width = len(str(value))
                if width > colWidth[col]:
                    colWidth[col] = width
    # set column width
    for i in range(len(colWidth)):
        colLettert = get_column_letter(i + 1)
        if colWidth[i] > 100:
            # set to 100 if col_width greater than 100
            sheet.column_dimensions[colLettert].width = 100
        else:
            sheet.column_dimensions[colLettert].width = colWidth[i] + 4


def GetFileList(folder, postfix=None):
    """
    find a list of files with a postfix
    """
    fullNameList = []
    if isdir(folder):
        files = listdir(folder)
        for filename in files:
            fullNameList.append(join(folder, filename))
        if postfix:
            targetFileList = []
            for fullname in fullNameList:
                if fullname.endswith(postfix.upper()) or fullname.endswith(postfix.lower()):
                    targetFileList.append(fullname)
            return targetFileList
        else:
            return fullNameList
    else:
        print("Error：Not a folder!")
        return False


def MkDir(path):
    """
    create a folder
    """
    path = path.strip()
    path = path.rstrip("\\")
    isExists = exists(path)
    if not isExists:
        makedirs(path)
        return True
    else:
        return False


def GetLotno(file):
    """
    get lotno
    """
    data = []
    # get file data
    with open(file, encoding='unicode_escape') as f:
        csvReader = reader((line.replace('\0', '') for line in f))
        for row in csvReader:
            data.append(row)
    # get lotno
    lotno = data[5][1]
    return lotno


alignment = Alignment(horizontal='center', vertical='center')
thin = Side(border_style='thin', color=BLACK)
border = Border(top=thin, left=thin, right=thin, bottom=thin)

rowOffset = 5
project = 'Unknown'
totalLotCount = 0
nowTime = 'Unknown'

hwbinToSwbin = {
    'F28': {
        3: {'SWBin': (1, 2), 'isPassBin': True},
        1: {'SWBin': (37, 38, 61, 62, 64, 65, 66, 90, 91, 92, 94), 'isPassBin': False},
        2: {'SWBin': (41, 42, 43, 44, 45, 53, 54, 55), 'isPassBin': False},
        4: {'SWBin': (23, 24, 25, 26, 27, 29, 30, 31, 32, 33, 34, 36, 39, 40), 'isPassBin': False},
        5: {'SWBin': (5, 6, 7, 8, 9, 12, 96, 97, 98, 99), 'isPassBin': False},
        6: {'SWBin': (13, 14, 15, 35), 'isPassBin': False}
    },
    'JX828': {
        3: {'SWBin': (1, 2, 3), 'isPassBin': True},
        1: {'SWBin': (63, 64, 65, 89, 90, 94), 'isPassBin': True},
        2: {'SWBin': (53, 54, 73, 74), 'isPassBin': True},
        4: {'SWBin': (23, 24, 25, 26, 27, 29, 30, 31, 32, 33, 34, 36, 39, 40, 56, 57, 58, 75), 'isPassBin': False},
        5: {'SWBin': (5, 6, 7, 8, 9, 12, 93, 96, 98, 99), 'isPassBin': False},
        6: {'SWBin': (13, 14, 15, 35, 46, 47, 48, 51, 60), 'isPassBin': False}
    },
    'JX825': {
        3: {'SWBin': (2, 255), 'isPassBin': True},
        4: {'SWBin': (23, 24, 25, 26, 27, 29, 30, 31, 32, 33, 34, 36, 39, 40, 63, 64, 65), 'isPassBin': False},
        5: {'SWBin': (5, 6, 7, 8, 9, 12, 89), 'isPassBin': False},
        6: {'SWBin': (13, 14, 15, 35, 46, 48, 53, 54, 60), 'isPassBin': False},
        8: {'SWBin': (94, 96, 97, 98, 99), 'isPassBin': False}
    },
    'JX832': {
        3: {'SWBin': (1, 2, 3), 'isPassBin': True},
        1: {'SWBin': (56, 57, 58, 66, 67, 68, 95, 96), 'isPassBin': True},
        2: {'SWBin': (53, 54, 63, 64, 65, 73, 74, 89, 90, 94), 'isPassBin': True},
        4: {'SWBin': (23, 24, 25, 26, 27, 29, 30, 31, 32, 33, 34, 36, 39, 40, 78), 'isPassBin': False},
        5: {'SWBin': (5, 6, 7, 8, 9, 12, 93, 97, 98, 99), 'isPassBin': False},
        6: {'SWBin': (13, 14, 15, 35, 46, 47, 48, 51, 60), 'isPassBin': False}
    }
}


def ParseFile(file, groupById):
    """
    get basic information and group data
    """
    parseResult = {}
    data = []
    # get file data
    with open(file, encoding='unicode_escape') as f:
        csvReader = reader((line.replace('\0', '') for line in f))
        for row in csvReader:
            data.append(row)

    # get row of ChipNo
    searchChipno = SearchString(data, 'ChipNo')
    if not searchChipno:
        exit()
    else:
        chipnoRowNum = searchChipno[0]

    # get row of first register
    firstRegisterRowNum = len(data) - 1
    for i in range(chipnoRowNum + rowOffset, len(data)):
        try:
            int(data[i][0])
        except:
            firstRegisterRowNum = i
            break

    groupList = []
    for i in range(chipnoRowNum + rowOffset, firstRegisterRowNum):
        groupList.append(int(data[i][groupById]))
    formatGroupList = list(set(groupList))
    formatGroupList.sort()

    groupIndex = {}
    for i in formatGroupList:
        temp = []
        dataList = FindItem(groupList, i)
        if groupById == 1:
            tempList = []
            for j in dataList:
                # get the corresponding SW_BIN of Site
                tempList.append(data[chipnoRowNum + rowOffset + j][2])
            # remove duplicate value
            formatTempList = list(set(tempList))
            tempIndex = {}
            for m in formatTempList:
                # find m in tempList
                tempDataList = FindItem(tempList, m)
                tempIndex[m] = tempDataList
            temp.append(tempIndex)
        else:
            temp.append(dataList)
        groupIndex[i] = temp
    # calculate chip count
    chipCount = firstRegisterRowNum - chipnoRowNum - rowOffset
    parseResult['group index'] = groupIndex
    parseResult['chip count'] = chipCount
    return parseResult


def SaveData(analysisFile, parseData):
    """
    save data to file
    """
    wb = Workbook()

    okHwbinCount = 0
    beginFailSwbin = 0
    for hwBinKey in hwbinToSwbin[project].keys():
        if hwbinToSwbin[project][hwBinKey]['isPassBin']:
            okHwbinCount += 1
            beginFailSwbin += len(hwbinToSwbin[project][hwBinKey]['SWBin'])

    colorList = ['99FFFF', '33FF00', 'FFFFCC', 'FFFF33', 'FF9900', 'FF0099', 'FF0000']
    swbinList = []
    keyIndex = 0
    for hwbinKey in hwbinToSwbin[project].keys():
        for swbin in hwbinToSwbin[project][hwbinKey]['SWBin']:
            swbinList.append([swbin, colorList[keyIndex]])
        keyIndex += 1

    siteSoftbinSheet = wb.create_sheet('Site-SWBin')
    siteSoftbinSheet.freeze_panes = 'B2'
    irow = 1
    totalCount = 0
    for i in range(len(parseData)):
        totalCount += parseData[i]['site data'][0]['chip count']
    siteSoftbinSheet.cell(row=irow, column=1).value = totalCount
    for i in range(16):
        siteSoftbinSheet.cell(row=irow, column=2 + i).value = 'Site' + str(i)
        siteSoftbinSheet.cell(row=irow, column=2 + i).fill = PatternFill(fill_type='solid', fgColor=YELLOW)
    siteSoftbinSheet.merge_cells(start_row=irow, end_row=irow, start_column=19, end_column=20)
    siteSoftbinSheet.cell(row=irow, column=19).value = 'Summary'
    siteSoftbinSheet.cell(row=irow, column=19).fill = PatternFill(fill_type='solid', fgColor='FFA500')
    irow += 1

    lotnoSiteSwbinCount = []
    for i in range(len(parseData)):
        lotnoTempList = [parseData[i]['lotno']]
        for x in range(len(swbinList)):
            swbin = str(swbinList[x][0])
            fileTempList = [swbinList[x][0]]
            siteCountList = [[0, 0], [1, 0], [2, 0], [3, 0], [4, 0], [5, 0], [6, 0], [7, 0], [8, 0], [9, 0], [10, 0],
                             [11, 0], [12, 0], [13, 0], [14, 0], [15, 0]]
            for site in range(16):
                if x < beginFailSwbin:
                    for m in range(len(parseData[i]['site data'])):
                        if (site in parseData[i]['site data'][m]['group index'].keys()) and (
                                    swbin in parseData[i]['site data'][m]['group index'][site][0].keys()):
                            siteCountList[site][1] += len(
                                parseData[i]['site data'][m]['group index'][site][0][swbin])
                else:
                    if (site in parseData[i]['site data'][-1]['group index'].keys()) and (
                                swbin in parseData[i]['site data'][-1]['group index'][site][0].keys()):
                        siteCountList[site][1] += len(parseData[i]['site data'][-1]['group index'][site][0][swbin])
            fileTempList.append(siteCountList)
            lotnoTempList.append(fileTempList)
        lotnoSiteSwbinCount.append(lotnoTempList)
    siteSwbinCount = []
    actualTotalCount = 0
    for x in range(len(swbinList)):
        tempTotalCount = 0
        tempList = [0] * 16
        tempSwbinCount = []
        for i in range(len(tempList)):
            for y in range(len(lotnoSiteSwbinCount)):
                tempList[i] += lotnoSiteSwbinCount[y][x + 1][1][i][1]
                if y == len(lotnoSiteSwbinCount) - 1:
                    tempSwbinCount.append([tempList[i], WHITE])
                    tempTotalCount += tempList[i]
                    actualTotalCount += tempList[i]
            if i == len(tempList) - 1:
                tempSwbinCount.append([tempTotalCount, 'FFA500'])
        siteSwbinCount.append(tempSwbinCount)
    # add the number of fail swbin by all site
    # fill color of max value in fail swbin is green(all values are 0 do not fill green)
    # fill color of min value in fail swbin is red(multiple values of 0 do not fill red)
    sitePassTotalList = [0] * 16
    for i in range(beginFailSwbin):
        for j in range(len(siteSwbinCount[i]) - 1):
            sitePassTotalList[j] += siteSwbinCount[i][j][0]
    siteFailTotalList = [0] * 16
    for i in range(beginFailSwbin, len(siteSwbinCount)):
        minValue = siteSwbinCount[i][0][0]
        maxValue = siteSwbinCount[i][0][0]
        for m in range(1, len(siteSwbinCount[i]) - 1):
            if minValue > siteSwbinCount[i][m][0]:
                minValue = siteSwbinCount[i][m][0]
            if maxValue < siteSwbinCount[i][m][0]:
                maxValue = siteSwbinCount[i][m][0]
        minIndex = []
        maxIndex = []
        for j in range(len(siteSwbinCount[i]) - 1):
            siteFailTotalList[j] += siteSwbinCount[i][j][0]
            if siteSwbinCount[i][j][0] == minValue:
                minIndex.append(j)
            if siteSwbinCount[i][j][0] == maxValue:
                maxIndex.append(j)
        if minValue == 0 and len(minIndex) == 1:
            siteSwbinCount[i][minIndex[0]][1] = GREEN
            for y in range(len(maxIndex)):
                siteSwbinCount[i][maxIndex[y]][1] = RED
        elif minValue == 0 and maxValue > 0:
            for y in range(len(maxIndex)):
                siteSwbinCount[i][maxIndex[y]][1] = RED
        elif minValue > 0:
            for x in range(len(minIndex)):
                siteSwbinCount[i][minIndex[x]][1] = GREEN
            for y in range(len(maxIndex)):
                siteSwbinCount[i][maxIndex[y]][1] = RED

    for x in range(len(swbinList)):
        siteSoftbinSheet.cell(row=irow, column=1).value = 'SWBin' + str(swbinList[x][0])
        siteSoftbinSheet.cell(row=irow, column=1).fill = PatternFill(fill_type='solid', fgColor=swbinList[x][1])
        for y in range(len(siteSwbinCount[x])):
            if y < len(siteSwbinCount[x]) - 1:
                if siteSwbinCount[x][y][0] > 0:
                    siteSoftbinSheet.cell(row=irow, column=2 + y).value = '{:.4%}'.format(
                        siteSwbinCount[x][y][0] / totalCount)
                siteSoftbinSheet.cell(row=irow, column=2 + y).fill = PatternFill(fill_type='solid',
                                                                                 fgColor=siteSwbinCount[x][y][1])
            else:
                siteSoftbinSheet.cell(row=irow, column=3 + y).value = siteSwbinCount[x][y][0]
                siteSoftbinSheet.cell(row=irow, column=4 + y).value = '{:.4%}'.format(
                    siteSwbinCount[x][y][0] / totalCount)
                siteSoftbinSheet.cell(row=irow, column=4 + y).fill = PatternFill(fill_type='solid',
                                                                                 fgColor=siteSwbinCount[x][y][1])
        irow += 1
    irow += 1
    siteSoftbinSheet.merge_cells(start_row=irow, end_row=irow + 1, start_column=1, end_column=1)
    siteSoftbinSheet.cell(row=irow, column=1).value = 'FailPercent'
    siteSoftbinSheet.cell(row=irow, column=1).font = Font(bold=True)
    siteSoftbinSheet.cell(row=irow, column=1).alignment = alignment
    siteSoftbinSheet.cell(row=irow, column=1).fill = PatternFill(fill_type='solid', fgColor=RED)
    for i in range(len(siteFailTotalList)):
        siteSoftbinSheet.cell(row=irow, column=2 + i).value = siteFailTotalList[i]
        siteSoftbinSheet.cell(row=irow + 1, column=2 + i).value = '{:.4%}'.format(
            siteFailTotalList[i] / totalCount)
        siteSoftbinSheet.cell(row=irow + 1, column=2 + i).fill = PatternFill(fill_type='solid', fgColor=RED)
    siteSoftbinSheet.cell(row=irow, column=3 + len(siteFailTotalList)).value = sum(siteFailTotalList)
    siteSoftbinSheet.cell(row=irow, column=3 + len(siteFailTotalList)).fill = PatternFill(fill_type='solid',
                                                                                          fgColor=RED)
    siteSoftbinSheet.cell(row=irow + 1, column=3 + len(siteFailTotalList)).value = '{:.4%}'.format(
        sum(siteFailTotalList) / totalCount)
    siteSoftbinSheet.cell(row=irow + 1, column=3 + len(siteFailTotalList)).fill = PatternFill(fill_type='solid',
                                                                                              fgColor=RED)
    irow += 2
    siteSoftbinSheet.merge_cells(start_row=irow, end_row=irow + 1, start_column=1, end_column=1)
    siteSoftbinSheet.cell(row=irow, column=1).value = 'PassPercent'
    siteSoftbinSheet.cell(row=irow, column=1).font = Font(bold=True)
    siteSoftbinSheet.cell(row=irow, column=1).alignment = alignment
    siteSoftbinSheet.cell(row=irow, column=1).fill = PatternFill(fill_type='solid', fgColor=GREEN)
    for i in range(len(sitePassTotalList)):
        siteSoftbinSheet.cell(row=irow, column=2 + i).value = sitePassTotalList[i]
        siteSoftbinSheet.cell(row=irow + 1, column=2 + i).value = '{:.4%}'.format(
            sitePassTotalList[i] / totalCount)
        siteSoftbinSheet.cell(row=irow + 1, column=2 + i).fill = PatternFill(fill_type='solid', fgColor=GREEN)
    siteSoftbinSheet.cell(row=irow, column=3 + len(sitePassTotalList)).value = sum(sitePassTotalList)
    siteSoftbinSheet.cell(row=irow, column=3 + len(sitePassTotalList)).fill = PatternFill(fill_type='solid',
                                                                                          fgColor=GREEN)
    siteSoftbinSheet.cell(row=irow + 1, column=3 + len(sitePassTotalList)).value = '{:.4%}'.format(
        sum(sitePassTotalList) / totalCount)
    siteSoftbinSheet.cell(row=irow + 1, column=3 + len(sitePassTotalList)).fill = PatternFill(fill_type='solid',
                                                                                              fgColor=GREEN)

    for row in siteSoftbinSheet.rows:
        for cell in row:
            cell.border = border
            if cell.row == 1:
                cell.font = Font(bold=True)
                cell.alignment = alignment

    LotnoSoftbinSheet = wb.create_sheet('LotNo-SWBin')
    LotnoSoftbinSheet.freeze_panes = 'D2'
    LotnoSoftbinSheet.cell(row=1, column=1).value = len(parseData)
    LotnoSoftbinSheet.cell(row=1, column=2).value = totalCount
    LotnoSoftbinSheet.cell(row=1, column=3).value = 'PassPercent'
    for i in range(len(parseData)):
        LotnoSoftbinSheet.cell(row=2 + i, column=1).value = parseData[i]['date']
        LotnoSoftbinSheet.cell(row=2 + i, column=2).value = parseData[i]['lotno']
        LotnoSoftbinSheet.cell(row=2 + i, column=2).fill = PatternFill(fill_type='solid', fgColor=YELLOW)

    lotnoSwbinCount = []

    for i in range(len(parseData)):
        lotnoTempList = []
        for x in range(len(swbinList)):
            swbin = swbinList[x][0]
            fileTempList = [swbin, 0]
            if x < beginFailSwbin:
                for m in range(len(parseData[i]['swbin data'])):
                    if swbin in parseData[i]['swbin data'][m]['group index'].keys():
                        fileTempList[1] += len(parseData[i]['swbin data'][m]['group index'][swbin][0])
            else:
                if swbin in parseData[i]['swbin data'][-1]['group index'].keys():
                    fileTempList[1] += len(parseData[i]['swbin data'][-1]['group index'][swbin][0])
            lotnoTempList.append(fileTempList)
        lotnoSwbinCount.append(lotnoTempList)
    for x in range(len(swbinList)):
        LotnoSoftbinSheet.merge_cells(start_row=1, end_row=1, start_column=4 + 2 * x, end_column=5 + 2 * x)
        LotnoSoftbinSheet.cell(row=1, column=4 + 2 * x).value = 'SWBin' + str(swbinList[x][0])
        LotnoSoftbinSheet.cell(row=1, column=4 + 2 * x).fill = PatternFill(fill_type='solid', fgColor=swbinList[x][1])
        for i in range(len(lotnoSwbinCount)):
            passCount = 0
            for j in range(beginFailSwbin):
                passCount += lotnoSwbinCount[i][j][1]
            LotnoSoftbinSheet.cell(row=2 + i, column=3).value = '{:.2%}'.format(
                passCount / parseData[i]['swbin data'][0]['chip count'])
            if passCount >= parseData[i]['swbin data'][0]['chip count']:
                LotnoSoftbinSheet.cell(row=2 + i, column=3).fill = PatternFill(fill_type='solid', fgColor=RED)
            LotnoSoftbinSheet.cell(row=2 + i, column=4 + 2 * x).value = lotnoSwbinCount[i][x][1]
            LotnoSoftbinSheet.cell(row=2 + i, column=5 + 2 * x).value = '{:.2%}'.format(
                lotnoSwbinCount[i][x][1] / parseData[i]['swbin data'][0]['chip count'])
            LotnoSoftbinSheet.cell(row=2 + i, column=5 + 2 * x).fill = PatternFill(fill_type='solid',
                                                                                   fgColor=swbinList[x][1])

    for row in LotnoSoftbinSheet.rows:
        for cell in row:
            cell.border = border
            if cell.row == 1 or cell.column in (1, 2):
                cell.font = Font(bold=True)
                cell.alignment = alignment

    for sheet_name in wb.sheetnames:
        if sheet_name == 'Sheet':
            del wb[sheet_name]
        else:
            SetColumnWidth(wb[sheet_name])

    wb.save(analysisFile)


class Runthread(QThread):
    _signal = pyqtSignal(str)

    def __init__(self, openPath, chooseRadio):
        super(Runthread, self).__init__()
        self.openPath = openPath
        self.chooseRadio = chooseRadio
        self.currentLotCount = 0

    def __del__(self):
        self.wait()

    def run(self):
        if self.chooseRadio == 'ProjectFolder':
            handlerNames = listdir(self.openPath)
            handlerFolders = []
            for handlerName in handlerNames:
                if isdir(join(self.openPath, handlerName)):
                    handlerFolders.append(join(self.openPath, handlerName))

            for handlerFolder in handlerFolders:
                dateFolders = []
                dateNames = listdir(handlerFolder)
                for dateName in dateNames:
                    if dateName != 'Analysis' and isdir(join(handlerFolder, dateName)):
                        dateFolders.append(join(handlerFolder, dateName))

                lotnoFolders = []
                for dateFolder in dateFolders:
                    lotNames = listdir(dateFolder)
                    for lotName in lotNames:
                        if isdir(join(dateFolder, lotName)):
                            lotnoFolders.append(join(dateFolder, lotName))
                fileList = []
                for lotFolder in lotnoFolders:
                    # get CSV file under the folder
                    fileList.append(GetFileList(lotFolder, '.csv'))
                    if not fileList:
                        exit()

                # analysis folder path
                if argv.count('-a') == 0:
                    analysisFolder = handlerFolder + '\Analysis'
                else:
                    analysisFolder = argv[argv.index('-a') + 1]

                MkDir(analysisFolder)

                parseData = []

                for i in range(len(fileList)):
                    lotnoData = {}
                    tempSiteData = []
                    tempSoftbinData = []
                    for file in fileList[i]:
                        # parse file
                        tempSiteData.append(ParseFile(file, 1))
                        tempSoftbinData.append(ParseFile(file, 2))
                    # get Date
                    date = basename(dirname(dirname(fileList[i][0])))
                    lotnoData['site data'] = tempSiteData
                    lotnoData['swbin data'] = tempSoftbinData
                    lotnoData['date'] = date
                    lotnoData['lotno'] = GetLotno(fileList[i][0])
                    parseData.append(lotnoData)
                    self.currentLotCount += 1
                    self._signal.emit(str(self.currentLotCount * 100 // totalLotCount))
                handler = basename(handlerFolder)
                analysisFile = join(analysisFolder, handler + '_Total_Analysis' + nowTime + '.xlsx')
                # save data
                SaveData(analysisFile, parseData)
        else:
            dateFolders = []
            dateNames = listdir(self.openPath)
            for dateName in dateNames:
                if dateName != 'Analysis' and isdir(join(self.openPath, dateName)):
                    dateFolders.append(join(self.openPath, dateName))

            lotnoFolders = []
            for dateFolder in dateFolders:
                lotNames = listdir(dateFolder)
                for lotName in lotNames:
                    if isdir(join(dateFolder, lotName)):
                        lotnoFolders.append(join(dateFolder, lotName))
            fileList = []
            for lotFolder in lotnoFolders:
                # get CSV file under the folder
                fileList.append(GetFileList(lotFolder, '.csv'))
                if not fileList:
                    exit()

            # analysis folder path
            if argv.count('-a') == 0:
                analysisFolder = self.openPath + '\Analysis'
            else:
                analysisFolder = argv[argv.index('-a') + 1]

            MkDir(analysisFolder)

            parseData = []

            for i in range(len(fileList)):
                lotnoData = {}
                tempSiteData = []
                tempSoftbinData = []
                for file in fileList[i]:
                    # parse file
                    tempSiteData.append(ParseFile(file, 1))
                    tempSoftbinData.append(ParseFile(file, 2))
                # get Date
                date = basename(dirname(dirname(fileList[i][0])))
                lotnoData['site data'] = tempSiteData
                lotnoData['swbin data'] = tempSoftbinData
                lotnoData['date'] = date
                lotnoData['lotno'] = GetLotno(fileList[i][0])
                parseData.append(lotnoData)
                self.currentLotCount += 1
                self._signal.emit(str(self.currentLotCount * 100 // totalLotCount))
            handler = basename(self.openPath)
            analysisFile = join(analysisFolder, handler + '_Total_Analysis' + nowTime + '.xlsx')
            # save data
            SaveData(analysisFile, parseData)
        self._signal.emit(str(100))


class MainWindow(QMainWindow, Datalog_Total_Analysis_UI.Ui_MainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        self.setupUi(self)
        self.cwd = getcwd()
        self.qe = QErrorMessage(self)
        self.pale = QPalette()
        # self.pale.setBrush(self.backgroundRole(), QBrush(QPixmap('./images/kobe3.jpg')))
        self.setPalette(self.pale)
        self.Open_pushButton.clicked.connect(self.Open)
        self.Project_comboBox.insertItem(0, self.tr('F28'))
        self.Project_comboBox.insertItem(1, self.tr('JX828'))
        self.Project_comboBox.insertItem(2, self.tr('JX825'))
        self.Project_comboBox.insertItem(2, self.tr('JX832'))
        self.Analysis_pushButton.clicked.connect(self.Analysis)

    def Open(self):
        if self.ProjectFolder_radioButton.isChecked():
            dirChoose = QFileDialog.getExistingDirectory(self, 'Select ProjectFolder Directory', self.cwd)
        else:
            dirChoose = QFileDialog.getExistingDirectory(self, 'Select HandlerFolder Directory', self.cwd)
        if not dirChoose:
            return
        self.Open_lineEdit.setText(dirChoose)

    def CallBackLog(self, msg):
        self.progressBar.setValue(int(msg))  # pass the thread's parameters to progressBar
        if msg == '100':
            self.Analysis_pushButton.setEnabled(True)

    def Analysis(self):
        if not self.Open_lineEdit.text():
            self.qe.showMessage('Path cannot be empty!')
            return
        else:
            global totalLotCount
            totalLotCount = 0
            if self.ProjectFolder_radioButton.isChecked():
                handlerNames = listdir(self.Open_lineEdit.text())
                handlerFolders = []
                for handlerName in handlerNames:
                    if isdir(join(self.Open_lineEdit.text(), handlerName)):
                        handlerFolders.append(join(self.Open_lineEdit.text(), handlerName))
                if not handlerFolders:
                    self.qe.showMessage('Path is incorrect,please check!')
                    return
                for handlerFolder in handlerFolders:
                    dateFolders = []
                    dateNames = listdir(handlerFolder)
                    for dateName in dateNames:
                        if dateName != 'Analysis' and isdir(join(handlerFolder, dateName)):
                            dateFolders.append(join(handlerFolder, dateName))
                    if not dateFolders:
                        self.qe.showMessage('Path is incorrect,please check!')
                        return
                    lotnoFolders = []
                    for dateFolder in dateFolders:
                        lotNames = listdir(dateFolder)
                        for lotName in lotNames:
                            if isdir(join(dateFolder, lotName)):
                                lotnoFolders.append(join(dateFolder, lotName))
                    if not lotnoFolders:
                        self.qe.showMessage('Path is incorrect,please check!')
                        return
                    totalLotCount += len(lotnoFolders)
            else:
                dateFolders = []
                dateNames = listdir(self.Open_lineEdit.text())
                for dateName in dateNames:
                    if dateName != 'Analysis' and isdir(join(self.Open_lineEdit.text(), dateName)):
                        dateFolders.append(join(self.Open_lineEdit.text(), dateName))
                if not dateFolders:
                    self.qe.showMessage('Path is incorrect,please check!')
                    return
                lotnoFolders = []
                for dateFolder in dateFolders:
                    lotNames = listdir(dateFolder)
                    for lotName in lotNames:
                        if isdir(join(dateFolder, lotName)):
                            lotnoFolders.append(join(dateFolder, lotName))
                if not lotnoFolders:
                    self.qe.showMessage('Path is incorrect,please check!')
                    return
                totalLotCount += len(lotnoFolders)
            if totalLotCount == 0:
                return
            global nowTime
            nowTime = datetime.now().strftime("%Y%m%d%H%M%S")
            self.progressBar.setValue(0)
            global project
            project = self.Project_comboBox.currentText()
            if self.ProjectFolder_radioButton.isChecked():
                chooseRadio = self.ProjectFolder_radioButton.text()
            else:
                chooseRadio = self.HandlerFolder_radioButton.text()
            # create thread
            self.Analysis_pushButton.setEnabled(False)
            self.thread = Runthread(self.Open_lineEdit.text(), chooseRadio)
            # connect signal
            self.thread._signal.connect(self.CallBackLog)
            self.thread.start()


if __name__ == '__main__':
    # _csv.Error:field larger than field limit(131072)
    maxInt = maxsize
    while True:
        try:
            field_size_limit(maxInt)
            break
        except OverflowError:
            maxInt = int(maxInt / 10)
    app = QApplication(argv)
    myMainWindow = MainWindow()
    myMainWindow.show()
    exit(app.exec_())
