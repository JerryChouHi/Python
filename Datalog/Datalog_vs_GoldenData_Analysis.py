# encoding:utf-8
# @Time     : 2019/9/9 13:32
# @Author   : Jerry Chou
# @File     :
# @Function :

from csv import reader
from os.path import basename, dirname, join, abspath
from os import getcwd
from datetime import datetime
from math import isnan
from sys import argv, path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.styles.colors import BLACK

path.append(abspath(join(getcwd(), '..')))
import Common

row_offset = 5
col_offset = 4

golden_data_color = '00FFFF'  # 水绿色
alignment = Alignment(horizontal='center', vertical='center')
thin = Side(border_style='thin', color=BLACK)
border = Border(top=thin, left=thin, right=thin, bottom=thin)


def get_golden_data(project_id):
    """
    golden数据列表
    :param project_id: 0:F28  1:JX828
    :return: golden高阈值和低阈值
    """
    golden_data_list = [[('PCLK_O/S', -0.38, -0.5),
                         ('HSYNC_O/S', -0.38, -0.5),
                         ('VSYNC_O/S', -0.38, -0.5),
                         ('D9_O/S', -0.38, -0.5),
                         ('D8_O/S', -0.38, -0.5),
                         ('D7_O/S', -0.38, -0.5),
                         ('D6_O/S', -0.38, -0.5),
                         ('D5_O/S', -0.38, -0.5),
                         ('D4_O/S', -0.38, -0.5),
                         ('D3_O/S', -0.38, -0.5),
                         ('D2_O/S', -0.38, -0.5),
                         ('D1_O/S', -0.38, -0.5),
                         ('D0_O/S', -0.38, -0.5),
                         ('SCL_O/S', -0.38, -0.5),
                         ('SDA_O/S', -0.38, -0.5),
                         ('RSTB_O/S', -0.38, -0.5),
                         ('PWDN_O/S', -0.38, -0.5),
                         ('DVDD_O/S', -0.32, -0.34),
                         ('VRamp_O/S', -0.4, -0.5),
                         ('VH_O/S', -0.4, -0.5),
                         ('VN1_O/S', 0.5, 0.4),
                         ('EXCLK_O/S', -0.4, -0.5),
                         ('PCLK_Leakage/iiL', 0.005, -0.02),
                         ('HSYNC_Leakage/iiL', 0.005, -0.02),
                         ('VSYNC_Leakage/iiL', 0.005, -0.02),
                         ('D9_Leakage/iiL', 0.005, -0.02),
                         ('D8_Leakage/iiL', 0.005, -0.02),
                         ('D7_Leakage/iiL', 0.005, -0.02),
                         ('D6_Leakage/iiL', 0.005, -0.02),
                         ('D5_Leakage/iiL', 0.005, -0.02),
                         ('D4_Leakage/iiL', 0.005, -0.02),
                         ('D3_Leakage/iiL', 0.005, -0.02),
                         ('D2_Leakage/iiL', 0.005, -0.02),
                         ('D1_Leakage/iiL', 0.02, -0.02),
                         ('D0_Leakage/iiL', 0.02, -0.02),
                         ('SCL_Leakage/iiL', 0.02, -0.02),
                         ('SDA_Leakage/iiL', 0.02, -0.02),
                         ('RSTB_Leakage/iiL', -0.15, -0.35),
                         ('PWDN_Leakage/iiL', 0.1, -0.1),
                         ('EXCLK_Leakage/iiL', 0.02, -0.02),
                         ('PCLK_Leakage/iiH', 0.02, -0.02),
                         ('HSYNC_Leakage/iiH', 0.02, -0.02),
                         ('VSYNC_Leakage/iiH', 0.02, -0.02),
                         ('D9_Leakage/iiH', 0.005, -0.02),
                         ('D8_Leakage/iiH', 0.005, -0.02),
                         ('D7_Leakage/iiH', 0.005, -0.02),
                         ('D6_Leakage/iiH', 0.005, -0.02),
                         ('D5_Leakage/iiH', 0.005, -0.02),
                         ('D4_Leakage/iiH', 0.005, -0.02),
                         ('D3_Leakage/iiH', 0.005, -0.02),
                         ('D2_Leakage/iiH', 0.005, -0.02),
                         ('D1_Leakage/iiH', 32, 27),
                         ('D0_Leakage/iiH', 32, 27),
                         ('SCL_Leakage/iiH', 0.005, -0.02),
                         ('SDA_Leakage/iiH', 0.005, -0.02),
                         ('RSTB_Leakage/iiH', 0.06, -0.02),
                         ('PWDN_Leakage/iiH', 1.5, 0.5),
                         ('EXCLK_Leakage/iiH', 0.005, -0.02),
                         ('iic_test', 1, 1),
                         ('DVDD_voltage', 1.65, 1.45, 9),
                         ('VH_voltage', 4.4, 4.1, 9),
                         ('VN1_voltage', -1.3, -1.6, 9),
                         ('Active_AVDD', 27, 23, 12),
                         ('Active_DOVDD', 70, 50, 12),
                         ('PWDN_AVDD', 10, 0, 8),
                         ('PWDN_DOVDD', 50, 0, 8),
                         ('PWDN_Total', 60, 10, 8),
                         ('BLC_R', 100, 90),
                         ('BLC_Gr', 100, 90),
                         ('BLC_Gb', 100, 90),
                         ('BLC_B', 100, 90),
                         ('PLCK_Freq', 86.5, 86.3, 9),
                         ('BK_DeadRowExBPix_R', 0, 0, 15),
                         ('BK_DeadRowExBPix_Gr', 0, 0, 15),
                         ('BK_DeadRowExBPix_Gb', 0, 0, 15),
                         ('BK_DeadRowExBPix_B', 0, 0, 15),
                         ('BK_DeadColExBPix_R', 0, 0, 14),
                         ('BK_DeadColExBPix_Gr', 0, 0, 14),
                         ('BK_DeadColExBPix_Gb', 0, 0, 14),
                         ('BK_DeadColExBPix_B', 0, 0, 14),
                         ('BK_Mean_R', 18, 14, 13),
                         ('BK_Mean_Gr', 18, 14, 13),
                         ('BK_Mean_Gb', 18, 14, 13),
                         ('BK_Mean_B', 18, 14, 13),
                         ('BK_StdDEV_R', 4, 2, 13),
                         ('BK_StdDEV_Gr', 4, 2, 13),
                         ('BK_StdDEV_Gb', 4, 2, 13),
                         ('BK_StdDEV_B', 4, 2, 13),
                         ('LT_DRow_R', 0, 0, 25),
                         ('LT_DRow_Gr', 0, 0, 25),
                         ('LT_DRow_Gb', 0, 0, 25),
                         ('LT_DRow_B', 0, 0, 25),
                         ('LT_DCol_R', 0, 0, 24),
                         ('LT_DCol_Gr', 0, 0, 24),
                         ('LT_DCol_Gb', 0, 0, 24),
                         ('LT_DCol_B', 0, 0, 24),
                         ('LT_DRow_Color', 0, 0, 25),
                         ('LT_DCol_Color', 0, 0, 24),
                         ('LT_Mean_R', 500, 400, 23),
                         ('LT_Mean_Gr', 700, 600, 23),
                         ('LT_Mean_Gb', 700, 600, 23),
                         ('LT_Mean_B', 500, 420, 23),
                         ('LT_StdDEV_R', 7, 4, 23),
                         ('LT_StdDEV_Gr', 7, 4, 23),
                         ('LT_StdDEV_Gb', 7, 4, 23),
                         ('LT_StdDEV_B', 7, 4, 23),
                         ('LT_RI_R', 95, 85, 23),
                         ('LT_RI_Gr', 95, 85, 23),
                         ('LT_RI_Gb', 95, 85, 23),
                         ('LT_RI_B', 95, 85, 23),
                         ('LT_Ratio_GrR', 1.6, 1.4, 23),
                         ('LT_Ratio_GbR', 1.6, 1.4, 23),
                         ('LT_Ratio_GrB', 1.6, 1.4, 23),
                         ('LT_Ratio_GbB', 1.6, 1.4, 23),
                         ('LT_Ratio_GbGr', 1.1, 1, 23),
                         ('LT_LostBit', 0, 0, 23),
                         ('LT_LostBitSNR_SumDIFF_R', 2, 0, 23),
                         ('LT_LostBitSNR_SumDIFF_Gr', 2, 0, 23),
                         ('LT_LostBitSNR_SumDIFF_Gb', 2, 0, 23),
                         ('LT_LostBitSNR_SumDIFF_B', 2, 0, 23),
                         ('LB_LT_Mean_R', 680, 580),
                         ('LB_LT_Mean_Gr', 980, 840),
                         ('LB_LT_Mean_Gb', 980, 840),
                         ('LB_LT_Mean_B', 680, 580),
                         ('LB_LT_LostBit', 1, 0, 23),
                         ('LB_LT_LostBitSNR_SumDIFF_R', 3, 0, 23),
                         ('LB_LT_LostBitSNR_SumDIFF_Gr', 3, 0, 23),
                         ('LB_LT_LostBitSNR_SumDIFF_Gb', 3, 0, 23),
                         ('LB_LT_LostBitSNR_SumDIFF_B', 3, 0, 23),
                         ('FW_LT_DRow_R', 0, 0, 36),
                         ('FW_LT_DRow_Gr', 0, 0, 36),
                         ('FW_LT_DRow_Gb', 0, 0, 36),
                         ('FW_LT_DRow_B', 0, 0, 36),
                         ('FW_LT_DCol_R', 0, 0, 36),
                         ('FW_LT_DCol_Gr', 0, 0, 36),
                         ('FW_LT_DCol_Gb', 0, 0, 36),
                         ('FW_LT_DCol_B', 0, 0, 36),
                         ('FW_LT_DRow_Color', 0, 0, 36),
                         ('FW_LT_DCol_Color', 0, 0, 36),
                         ('FW_LT_Mean_R', 1023, 1023, 36),
                         ('FW_LT_Mean_Gr', 1023, 1023, 36),
                         ('FW_LT_Mean_Gb', 1023, 1023, 36),
                         ('FW_LT_Mean_B', 1023, 1023, 36),
                         ('FW_LT_StdDEV_R', 0, 0, 36),
                         ('FW_LT_StdDEV_Gr', 0, 0, 36),
                         ('FW_LT_StdDEV_Gb', 0, 0, 36),
                         ('FW_LT_StdDEV_B', 0, 0, 36),
                         ('FW_LT_RI_R', 100, 100),
                         ('FW_LT_RI_Gr', 100, 100),
                         ('FW_LT_RI_Gb', 100, 100),
                         ('FW_LT_RI_B', 100, 100),
                         ('FW_LT_Ratio_GrR', 1, 1),
                         ('FW_LT_Ratio_GbR', 1, 1),
                         ('FW_LT_Ratio_GrB', 1, 1),
                         ('FW_LT_Ratio_GbB', 1, 1),
                         ('FW_LT_Ratio_GbGr', 1, 1),
                         ('LT_CornerLine', 0, 0, 26),
                         ('LT_ScratchLine', 0, 0, 26),
                         ('LT_Blemish', 0, 0, 31),
                         ('LT_LineStripe', 0, 0, 27),
                         ('LT_Particle', 0, 0, 32),
                         ('BK_Cluster2', 0, 0, 30),
                         ('LT_Cluster2', 0, 0, 30),
                         ('BK_Cluster1', 0, 0, 29),
                         ('LT_Cluster1', 0, 0, 29),
                         ('WP_Count', 0, 0, 35),
                         ('BK_Cluster3GrGb', 0, 0, 39),
                         ('LT_Cluster3GrGb', 0, 0, 40),
                         ('BK_Cluster3SubtractGrGb', 1, 0, 33),
                         ('LT_Cluster3SubtractGrGb', 1, 0, 34),
                         ('LB_BK_DeadRowExBPix_R', 0, 0, 45),
                         ('LB_BK_DeadRowExBPix_Gr', 0, 0, 45),
                         ('LB_BK_DeadRowExBPix_Gb', 0, 0, 45),
                         ('LB_BK_DeadRowExBPix_B', 0, 0, 45),
                         ('LB_BK_DeadColExBPix_R', 0, 0, 44),
                         ('LB_BK_DeadColExBPix_Gr', 0, 0, 44),
                         ('LB_BK_DeadColExBPix_Gb', 0, 0, 44),
                         ('LB_BK_DeadColExBPix_B', 0, 0, 44),
                         ('SR_LT_DRow_R', 0, 0, 55),
                         ('SR_LT_DRow_Gr', 0, 0, 55),
                         ('SR_LT_DRow_Gb', 0, 0, 55),
                         ('SR_LT_DRow_B', 0, 0, 55),
                         ('SR_LT_DCol_R', 0, 0, 54),
                         ('SR_LT_DCol_Gr', 0, 0, 54),
                         ('SR_LT_DCol_Gb', 0, 0, 54),
                         ('SR_LT_DCol_B', 0, 0, 54),
                         ('SR_LT_DRow_Color', 0, 0, 55),
                         ('SR_LT_DCol_Color', 0, 0, 54),
                         ('SR_LT_Mean_R', 480, 400, 53),
                         ('SR_LT_Mean_Gr', 700, 600, 53),
                         ('SR_LT_Mean_Gb', 700, 600, 53),
                         ('SR_LT_Mean_B', 480, 400, 53),
                         ('SR_LT_StdDEV_R', 8, 4, 53),
                         ('SR_LT_StdDEV_Gr', 8, 4, 53),
                         ('SR_LT_StdDEV_Gb', 8, 4, 53),
                         ('SR_LT_StdDEV_B', 8, 4, 53),
                         ('SR_LT_RI_R', 100, 90, 53),
                         ('SR_LT_RI_Gr', 100, 90, 53),
                         ('SR_LT_RI_Gb', 100, 90, 53),
                         ('SR_LT_RI_B', 100, 90, 53),
                         ('SR_LT_Ratio_GrR', 1.6, 1.4, 53),
                         ('SR_LT_Ratio_GbR', 1.6, 1.4, 53),
                         ('SR_LT_Ratio_GrB', 1.6, 1.4, 53),
                         ('SR_LT_Ratio_GbB', 1.6, 1.4, 53),
                         ('SR_LT_Ratio_GbGr', 1.1, 1, 53),
                         ('SR_LT_LostBit', 1, 0, 53),
                         ('SR_BK_DeadRowExBPix_R', 0, 0, 42),
                         ('SR_BK_DeadRowExBPix_Gr', 0, 0, 42),
                         ('SR_BK_DeadRowExBPix_Gb', 0, 0, 42),
                         ('SR_BK_DeadRowExBPix_B', 0, 0, 42),
                         ('SR_BK_DeadColExBPix_R', 0, 0, 41),
                         ('SR_BK_DeadColExBPix_Gr', 0, 0, 41),
                         ('SR_BK_DeadColExBPix_Gb', 0, 0, 41),
                         ('SR_BK_DeadColExBPix_B', 0, 0, 41),
                         ('SR_BK_Mean_R', 70, 60, 43),
                         ('SR_BK_Mean_Gr', 70, 60, 43),
                         ('SR_BK_Mean_Gb', 70, 60, 43),
                         ('SR_BK_Mean_B', 70, 60, 43),
                         ('SR_BK_StdDEV_R', 2, 0, 43),
                         ('SR_BK_StdDEV_Gr', 2, 0, 43),
                         ('SR_BK_StdDEV_Gb', 2, 0, 43),
                         ('SR_BK_StdDEV_B', 2, 0, 43),
                         # ('Diffuser2_LT_WeakLineRow_R', 0, 0, 72),
                         # ('Diffuser2_LT_WeakLineRow_Gr', 0, 0, 72),
                         # ('Diffuser2_LT_WeakLineRow_Gb', 0, 0, 72),
                         # ('Diffuser2_LT_WeakLineRow_B', 0, 0, 72),
                         # ('Diffuser2_LT_WeakLineCol_R', 0, 0, 71),
                         # ('Diffuser2_LT_WeakLineCol_Gr', 0, 0, 71),
                         # ('Diffuser2_LT_WeakLineCol_Gb', 0, 0, 71),
                         # ('Diffuser2_LT_WeakLineCol_B', 0, 0, 71),
                         ('BK_Z1_BT_DP_R', 4, 0),
                         ('BK_Z1_BT_DP_Gr', 4, 0),
                         ('BK_Z1_BT_DP_Gb', 4, 0),
                         ('BK_Z1_BT_DP_B', 4, 0),
                         ('BK_Z2_BT_DP_R', 4, 0),
                         ('BK_Z2_BT_DP_Gr', 4, 0),
                         ('BK_Z2_BT_DP_Gb', 4, 0),
                         ('BK_Z2_BT_DP_B', 4, 0),
                         ('BK_Z1_BT_WP_R', 4, 0),
                         ('BK_Z1_BT_WP_Gr', 4, 0),
                         ('BK_Z1_BT_WP_Gb', 4, 0),
                         ('BK_Z1_BT_WP_B', 4, 0),
                         ('BK_Z2_BT_WP_R', 4, 0),
                         ('BK_Z2_BT_WP_Gr', 4, 0),
                         ('BK_Z2_BT_WP_Gb', 4, 0),
                         ('BK_Z2_BT_WP_B', 4, 0),
                         ('LT_Z1_BT_DP_R', 0, 0),
                         ('LT_Z1_BT_DP_Gr', 0, 0),
                         ('LT_Z1_BT_DP_Gb', 0, 0),
                         ('LT_Z1_BT_DP_B', 0, 0),
                         ('LT_Z2_BT_DP_R', 0, 0),
                         ('LT_Z2_BT_DP_Gr', 0, 0),
                         ('LT_Z2_BT_DP_Gb', 0, 0),
                         ('LT_Z2_BT_DP_B', 0, 0),
                         ('LT_Z1_BT_WP_R', 0, 0),
                         ('LT_Z1_BT_WP_Gr', 0, 0),
                         ('LT_Z1_BT_WP_Gb', 0, 0),
                         ('LT_Z1_BT_WP_B', 0, 0),
                         ('LT_Z2_BT_WP_R', 0, 0),
                         ('LT_Z2_BT_WP_Gr', 0, 0),
                         ('LT_Z2_BT_WP_Gb', 0, 0),
                         ('LT_Z2_BT_WP_B', 0, 0),
                         ('LT_Z1_DK_WP_R', 0, 0),
                         ('LT_Z1_DK_WP_Gr', 0, 0),
                         ('LT_Z1_DK_WP_Gb', 0, 0),
                         ('LT_Z1_DK_WP_B', 0, 0),
                         ('LT_Z2_DK_WP_R', 0, 0),
                         ('LT_Z2_DK_WP_Gr', 0, 0),
                         ('LT_Z2_DK_WP_Gb', 0, 0),
                         ('LT_Z2_DK_WP_B', 0, 0),
                         ('LT_Z1_DK_DP_R', 0, 0),
                         ('LT_Z1_DK_DP_Gr', 0, 0),
                         ('LT_Z1_DK_DP_Gb', 0, 0),
                         ('LT_Z1_DK_DP_B', 0, 0),
                         ('LT_Z2_DK_DP_R', 0, 0),
                         ('LT_Z2_DK_DP_Gr', 0, 0),
                         ('LT_Z2_DK_DP_Gb', 0, 0),
                         ('LT_Z2_DK_DP_B', 0, 0),
                         ('BK_Z1_BT_DP', 9, 0),
                         ('BK_Z2_BT_DP', 9, 0),
                         ('BK_Z1_BT_WP', 9, 0),
                         ('BK_Z2_BT_WP', 9, 0),
                         ('LT_Z1_BT_DP', 0, 0),
                         ('LT_Z2_BT_DP', 0, 0),
                         ('LT_Z1_BT_WP', 0, 0),
                         ('LT_Z2_BT_WP', 0, 0),
                         ('LT_Z1_DK_DP', 0, 0),
                         ('LT_Z2_DK_DP', 0, 0),
                         ('LT_Z1_DK_WP', 0, 0),
                         ('LT_Z2_DK_WP', 0, 0),
                         ('Binning', 0, 0, 2)
                         ],
                        [('VSYNC_O/S', -0.4, -0.5),
                         ('HSYNC_O/S', -0.4, -0.5),
                         ('PCLK_O/S', -0.4, -0.5),
                         ('EXCLK_O/S', -0.4, -0.5),
                         ('RSTB_O/S', -0.4, -0.5),
                         ('PWDN_O/S', -0.4, -0.5),
                         ('SDA_O/S', -0.4, -0.5),
                         ('SCL_O/S', -0.4, -0.5),
                         ('D0_O/S', -0.4, -0.5),
                         ('D1_O/S', -0.4, -0.5),
                         ('D2_O/S', -0.4, -0.5),
                         ('D3_O/S', -0.4, -0.5),
                         ('D4_O/S', -0.4, -0.5),
                         ('D5_O/S', -0.4, -0.5),
                         ('D6_O/S', -0.4, -0.5),
                         ('D7_O/S', -0.4, -0.5),
                         ('D8_O/S', -0.4, -0.5),
                         ('D9_O/S', -0.4, -0.5),
                         ('MCP_O/S', -0.4, -0.5),
                         ('MCN_O/S', -0.4, -0.5),
                         ('MDP0_O/S', -0.4, -0.5),
                         ('MDN0_O/S', -0.4, -0.5),
                         ('MDP1_O/S', -0.4, -0.5),
                         ('MDN1_O/S', -0.4, -0.5),
                         ('AVDD_O/S', -0.4, -0.5),
                         ('DOVDD_O/S', -0.4, -0.5),
                         ('VRamp_O/S', -0.4, -0.5),
                         ('VH_O/S', -0.4, -0.5),
                         ('VN1_O/S', 0.4, 0.3),

                         ('VSYNC_Leakage/iiL', 0.01, -0.1),
                         ('HSYNC_Leakage/iiL', 0.01, -0.1),
                         ('PCLK_Leakage/iiL', 0.01, -0.1),
                         ('EXCLK_Leakage/iiL', 0.01, -0.1),
                         ('RSTB_Leakage/iiL', 0.01, -0.1),
                         ('PWDN_Leakage/iiL', 0.01, -0.1),
                         ('SDA_Leakage/iiL', 0.01, -0.1),
                         ('SCL_Leakage/iiL', 0.01, -0.1),
                         ('D0_Leakage/iiL', 0.01, -0.1),
                         ('D1_Leakage/iiL', 0.01, -0.1),
                         ('D2_Leakage/iiL', 0.01, -0.1),
                         ('D4_Leakage/iiL', 0.01, -0.1),
                         ('D5_Leakage/iiL', 0.01, -0.1),
                         ('D3_Leakage/iiL', 0.01, -0.1),
                         ('D6_Leakage/iiL', 0.01, -0.1),
                         ('D7_Leakage/iiL', 0.01, -0.1),
                         ('D8_Leakage/iiL', 0.01, -0.1),
                         ('D9_Leakage/iiL', 0.01, -0.1),
                         ('MCP_Leakage/iiL', 0.01, -0.1),
                         ('MCN_Leakage/iiL', 0.01, -0.1),
                         ('MDP0_Leakage/iiL', 0.01, -0.1),
                         ('MDN0_Leakage/iiL', 0.01, -0.1),
                         ('MDP1_Leakage/iiL', 0.01, -0.1),
                         ('MDN1_Leakage/iiL', 0.01, -0.1),

                         ('VSYNC_Leakage/iiH', 0.05, -0.01),
                         ('HSYNC_Leakage/iiH', 0.05, -0.01),
                         ('PCLK_Leakage/iiH', 0.05, -0.01),
                         ('EXCLK_Leakage/iiH', 0.05, -0.01),
                         ('RSTB_Leakage/iiH', 0.05, -0.01),
                         ('PWDN_Leakage/iiH', 0.5, 0.1),
                         ('SDA_Leakage/iiH', 0.01, -0.01),
                         ('SCL_Leakage/iiH', 0.01, -0.01),
                         ('D0_Leakage/iiH', 19, 18),
                         ('D1_Leakage/iiH', 19, 18),
                         ('D2_Leakage/iiH', 0.05, -0.01),
                         ('D4_Leakage/iiH', 0.05, -0.01),
                         ('D5_Leakage/iiH', 0.05, -0.01),
                         ('D3_Leakage/iiH', 0.05, -0.01),
                         ('D6_Leakage/iiH', 0.05, -0.01),
                         ('D7_Leakage/iiH', 0.05, -0.01),
                         ('D8_Leakage/iiH', 0.05, -0.01),
                         ('D9_Leakage/iiH', 0.05, -0.01),
                         ('MCP_Leakage/iiH', 0.05, -0.01),
                         ('MCN_Leakage/iiH', 0.05, -0.01),
                         ('MDP0_Leakage/iiH', 0.05, -0.01),
                         ('MDN0_Leakage/iiH', 0.05, -0.01),
                         ('MDP1_Leakage/iiH', 0.05, -0.01),
                         ('MDN1_Leakage/iiH', 0.05, -0.01),

                         ('iic_test', 1, 1),
                         ('DVDD_voltage', 1.6, 1.5, 9),
                         ('VH_voltage', 3.3, 3.1, 9),
                         ('VN1_voltage', -1.35, -1.45, 9),
                         ('Active_AVDD', 43, 40, 12),
                         ('Active_DOVDD', 51, 47, 12),
                         ('PWDN_AVDD', 40, 37, 8),
                         ('PWDN_DOVDD', 22, 19, 8),
                         ('PWDN_Total', 61, 57, 8),

                         ('BLC_R', 395, 380),
                         ('BLC_Gr', 395, 380),
                         ('BLC_Gb', 395, 380),
                         ('BLC_B', 395, 380),
                         ('PLCK_Freq', 86.5, 86.3, 93),

                         ('BK_DeadRowExBPix_R', 0, 0, 15),
                         ('BK_DeadRowExBPix_Gr', 0, 0, 15),
                         ('BK_DeadRowExBPix_Gb', 0, 0, 15),
                         ('BK_DeadRowExBPix_B', 0, 0, 15),
                         ('BK_DeadColExBPix_R', 0, 0, 14),
                         ('BK_DeadColExBPix_Gr', 0, 0, 14),
                         ('BK_DeadColExBPix_Gb', 0, 0, 14),
                         ('BK_DeadColExBPix_B', 0, 0, 14),

                         ('BK_MixDCol_VfpnQty_R', 0, 0),
                         ('BK_MixDCol_VfpnQty_Gr', 0, 0),
                         ('BK_MixDCol_VfpnQty_Gb', 0, 0),
                         ('BK_MixDCol_VfpnQty_B', 0, 0),

                         ('BK_MixDCol_VFPN_MaxValue0_R', 1, -1),
                         ('BK_MixDCol_VFPN_MaxValue0_Gr', 1, -1),
                         ('BK_MixDCol_VFPN_MaxValue0_Gb', 1, -1),
                         ('BK_MixDCol_VFPN_MaxValue0_B', 1, -1),
                         ('BK_MixDCol_VFPN_MaxValue1_R', 1, -1),
                         ('BK_MixDCol_VFPN_MaxValue1_Gr', 1, -1),
                         ('BK_MixDCol_VFPN_MaxValue1_Gb', 1, -1),
                         ('BK_MixDCol_VFPN_MaxValue1_B', 1, -1),
                         ('BK_MixDCol_VFPN_MaxValue2_R', 1, -1),
                         ('BK_MixDCol_VFPN_MaxValue2_Gr', 1, -1),
                         ('BK_MixDCol_VFPN_MaxValue2_Gb', 1, -1),
                         ('BK_MixDCol_VFPN_MaxValue2_B', 1, -1),
                         ('BK_MixDCol_VFPN_MaxValue3_R', 1, -1),
                         ('BK_MixDCol_VFPN_MaxValue3_Gr', 1, -1),
                         ('BK_MixDCol_VFPN_MaxValue3_Gb', 1, -1),
                         ('BK_MixDCol_VFPN_MaxValue3_B', 1, -1),
                         ('BK_MixDCol_VFPN_MaxValue4_R', 1, -1),
                         ('BK_MixDCol_VFPN_MaxValue4_Gr', 1, -1),
                         ('BK_MixDCol_VFPN_MaxValue4_Gb', 1, -1),
                         ('BK_MixDCol_VFPN_MaxValue4_B', 1, -1),
                         ('BK_Mean_R', 18, 14, 13),
                         ('BK_Mean_Gr', 18, 14, 13),
                         ('BK_Mean_Gb', 18, 14, 13),
                         ('BK_Mean_B', 18, 14, 13),
                         ('BK_StdDEV_R', 1.3, 1.1, 13),
                         ('BK_StdDEV_Gr', 1.3, 1.1, 13),
                         ('BK_StdDEV_Gb', 1.3, 1.1, 13),
                         ('BK_StdDEV_B', 1.3, 1.1, 13),

                         ('LT_DRow_R', 0, 0, 25),
                         ('LT_DRow_Gr', 0, 0, 25),
                         ('LT_DRow_Gb', 0, 0, 25),
                         ('LT_DRow_B', 0, 0, 25),
                         ('LT_DCol_R', 0, 0, 24),
                         ('LT_DCol_Gr', 0, 0, 24),
                         ('LT_DCol_Gb', 0, 0, 24),
                         ('LT_DCol_B', 0, 0, 24),
                         ('LT_DRow_Color', 0, 0, 24),
                         ('LT_DCol_Color', 0, 0, 24),
                         ('LT_WeakLineRow_R', 0, 0, 25),
                         ('LT_WeakLineRow_Gr', 0, 0, 25),
                         ('LT_WeakLineRow_Gb', 0, 0, 25),
                         ('LT_WeakLineRow_B', 0, 0, 25),
                         ('LT_WeakLineCol_R', 0, 0, 24),
                         ('LT_WeakLineCol_Gr', 0, 0, 24),
                         ('LT_WeakLineCol_Gb', 0, 0, 24),
                         ('LT_WeakLineCol_B', 0, 0, 24),
                         ('LT_Mean_R', 405, 390, 23),
                         ('LT_Mean_Gr', 685, 650, 23),
                         ('LT_Mean_Gb', 685, 650, 23),
                         ('LT_Mean_B', 450, 410, 23),
                         ('LT_StdDEV_R', 7, 4, 23),
                         ('LT_StdDEV_Gr', 10, 6, 23),
                         ('LT_StdDEV_Gb', 10, 6, 23),
                         ('LT_StdDEV_B', 7, 4, 23),
                         ('LT_RI_R', 72, 63, 23),
                         ('LT_RI_Gr', 68, 60, 23),
                         ('LT_RI_Gb', 68, 60, 23),
                         ('LT_RI_B', 68, 59, 23),
                         ('LT_Ratio_GrR', 1.72, 1.65, 23),
                         ('LT_Ratio_GbR', 1.72, 1.65, 23),
                         ('LT_Ratio_GrB', 1.62, 1.5, 23),
                         ('LT_Ratio_GbB', 1.62, 1.5, 23),
                         ('LT_Ratio_GbGr', 1, 0.99, 23),
                         ('LT_LostBit', 0, 0, 23),
                         ('LT_LostBitSNR_SumDIFF_R', 2, 0, 23),
                         ('LT_LostBitSNR_SumDIFF_Gr', 2, 0, 23),
                         ('LT_LostBitSNR_SumDIFF_Gb', 2, 0, 23),
                         ('LT_LostBitSNR_SumDIFF_B', 2, 0, 23),

                         ('LBIT_LT_Mean_R', 560, 530),
                         ('LBIT_LT_Mean_Gr', 950, 910),
                         ('LBIT_LT_Mean_Gb', 950, 900),
                         ('LBIT_LT_Mean_B', 620, 560),
                         ('LBIT_LT_LostBit', 0, 0, 23),
                         ('LBIT_LT_LostBitSNR_SumDIFF_R', 4, 1, 23),
                         ('LBIT_LT_LostBitSNR_SumDIFF_Gr', 4, 1, 23),
                         ('LBIT_LT_LostBitSNR_SumDIFF_Gb', 3, 1, 23),
                         ('LBIT_LT_LostBitSNR_SumDIFF_B', 3, 1, 23),

                         ('FW_LT_DRow_R', 0, 0, 36),
                         ('FW_LT_DRow_Gr', 0, 0, 36),
                         ('FW_LT_DRow_Gb', 0, 0, 36),
                         ('FW_LT_DRow_B', 0, 0, 36),
                         ('FW_LT_DCol_R', 0, 0, 36),
                         ('FW_LT_DCol_Gr', 0, 0, 36),
                         ('FW_LT_DCol_Gb', 0, 0, 36),
                         ('FW_LT_DCol_B', 0, 0, 36),
                         ('FW_LT_DRow_Color', 0, 0, 36),
                         ('FW_LT_DCol_Color', 0, 0, 36),
                         ('FW_LT_Mean_R', 1023, 1023, 36),
                         ('FW_LT_Mean_Gr', 1023, 1023, 36),
                         ('FW_LT_Mean_Gb', 1023, 1023, 36),
                         ('FW_LT_Mean_B', 1023, 1023, 36),
                         ('FW_LT_StdDEV_R', 0, 0, 36),
                         ('FW_LT_StdDEV_Gr', 0, 0, 36),
                         ('FW_LT_StdDEV_Gb', 0, 0, 36),
                         ('FW_LT_StdDEV_B', 0, 0, 36),
                         ('FW_LT_RI_R', 100, 98),
                         ('FW_LT_RI_Gr', 100, 100),
                         ('FW_LT_RI_Gb', 100, 100),
                         ('FW_LT_RI_B', 100, 98),
                         ('FW_LT_Ratio_GrR', 1, 1),
                         ('FW_LT_Ratio_GbR', 1, 1),
                         ('FW_LT_Ratio_GrB', 1, 1),
                         ('FW_LT_Ratio_GbB', 1, 1),
                         ('FW_LT_Ratio_GbGr', 1, 1),
                         ('FW_LT_LostBit', 1, 1),

                         ('BSun_BK_DeadRowExBPix_R', 0, 0, 48),
                         ('BSun_BK_DeadRowExBPix_Gr', 0, 0, 48),
                         ('BSun_BK_DeadRowExBPix_Gb', 0, 0, 48),
                         ('BSun_BK_DeadRowExBPix_B', 0, 0, 48),
                         ('BSun_BK_DeadColExBPix_R', 0, 0, 47),
                         ('BSun_BK_DeadColExBPix_Gr', 0, 0, 47),
                         ('BSun_BK_DeadColExBPix_Gb', 0, 0, 47),
                         ('BSun_BK_DeadColExBPix_B', 0, 0, 47),
                         ('BSun_BK_Mean_R', 68, 64),
                         ('BSun_BK_Mean_Gr', 68, 64),
                         ('BSun_BK_Mean_Gb', 68, 64),
                         ('BSun_BK_Mean_B', 68, 64),
                         ('BSun_BK_StdDEV_R', 1.9, 1.6),
                         ('BSun_BK_StdDEV_Gr', 1.9, 1.6),
                         ('BSun_BK_StdDEV_Gb', 1.9, 1.6),
                         ('BSun_BK_StdDEV_B', 1.9, 1.6),

                         ('LMFlip_LT_DRow_R', 0, 0, 58),
                         ('LMFlip_LT_DRow_Gr', 0, 0, 58),
                         ('LMFlip_LT_DRow_Gb', 0, 0, 58),
                         ('LMFlip_LT_DRow_B', 0, 0, 58),
                         ('LMFlip_LT_DCol_R', 0, 0, 57),
                         ('LMFlip_LT_DCol_Gr', 0, 0, 57),
                         ('LMFlip_LT_DCol_Gb', 0, 0, 57),
                         ('LMFlip_LT_DCol_B', 0, 0, 57),
                         ('LMFlip_LT_DRow_Color', 0, 0, 58),
                         ('LMFlip_LT_DCol_Color', 0, 0, 57),
                         ('LMFlip_LT_WeakLineRow_R', 0, 0, 58),
                         ('LMFlip_LT_WeakLineRow_Gr', 0, 0, 58),
                         ('LMFlip_LT_WeakLineRow_Gb', 0, 0, 58),
                         ('LMFlip_LT_WeakLineRow_B', 0, 0, 58),
                         ('LMFlip_LT_WeakLineCol_R', 0, 0, 57),
                         ('LMFlip_LT_WeakLineCol_Gr', 0, 0, 57),
                         ('LMFlip_LT_WeakLineCol_Gb', 0, 0, 57),
                         ('LMFlip_LT_WeakLineCol_B', 0, 0, 57),
                         ('LMFlip_LT_Mean_R', 400, 380, 56),
                         ('LMFlip_LT_Mean_Gr', 660, 640, 56),
                         ('LMFlip_LT_Mean_Gb', 660, 640, 56),
                         ('LMFlip_LT_Mean_B', 440, 400, 56),
                         ('LMFlip_LT_StdDEV_R', 6, 4, 56),
                         ('LMFlip_LT_StdDEV_Gr', 10, 6, 56),
                         ('LMFlip_LT_StdDEV_Gb', 10, 6, 56),
                         ('LMFlip_LT_StdDEV_B', 7, 4, 56),
                         ('LMFlip_LT_RI_R', 71, 64, 56),
                         ('LMFlip_LT_RI_Gr', 68, 60, 56),
                         ('LMFlip_LT_RI_Gb', 68, 60, 56),
                         ('LMFlip_LT_RI_B', 66, 59, 56),
                         ('LMFlip_LT_Ratio_GrR', 1.7, 1.6, 56),
                         ('LMFlip_LT_Ratio_GbR', 1.7, 1.6, 56),
                         ('LMFlip_LT_Ratio_GrB', 1.6, 1.5, 56),
                         ('LMFlip_LT_Ratio_GbB', 1.6, 1.5, 56),
                         ('LMFlip_LT_Ratio_GbGr', 1, 0.99, 56),
                         ('LMFlip_LT_LostBit', 0, 0, 56),
                         ('LMFlip_LT_LostBitSNR_SumDIFF_R', 2, 0, 56),
                         ('LMFlip_LT_LostBitSNR_SumDIFF_Gr', 2, 0, 56),
                         ('LMFlip_LT_LostBitSNR_SumDIFF_Gb', 2, 0, 56),
                         ('LMFlip_LT_LostBitSNR_SumDIFF_B', 2, 0, 56),

                         ('LT_CornerLine', 0, 0, 26),
                         ('LT_ScratchLine', 0, 0, 26),
                         ('LT_Blemish', 0, 0, 31),
                         ('LT_LineStripe', 0, 0, 27),
                         ('LT_Particle', 0, 0, 32),
                         ('BK_Cluster2', 0, 0, 30),
                         ('LT_Cluster2', 0, 0, 30),
                         ('BK_Cluster1', 0, 0, 29),
                         ('LT_Cluster1', 0, 0, 29),
                         ('BK_Cluster3GrGb', 0, 0, 39),
                         ('LT_Cluster3GrGb', 0, 0, 40),
                         ('BK_Cluster3SubtractGrGb', 0, 0, 33),
                         ('LT_Cluster3SubtractGrGb', 0, 0, 34),
                         ('WP_Count', 0, 0, 35),

                         ('BK32X_BK_BadPixel_Area', 0, 0, 75),
                         ('BK32X_BK_MixDCol_DashQty_R', 0, 0, 51),
                         ('BK32X_BK_MixDCol_DashQty_Gr', 0, 0, 51),
                         ('BK32X_BK_MixDCol_DashQty_Gb', 0, 0, 51),
                         ('BK32X_BK_MixDCol_DashQty_B', 0, 0, 51),
                         ('BK32X_SP_BK_MixDCol_DashQty_R', 0, 0, 54),
                         ('BK32X_SP_BK_MixDCol_DashQty_Gr', 0, 0, 54),
                         ('BK32X_SP_BK_MixDCol_DashQty_Gb', 0, 0, 54),
                         ('BK32X_SP_BK_MixDCol_DashQty_B', 0, 0, 54),
                         ('BK32X_BK_MixDCol_DASH_DeltaAvg_R', 15, 5),
                         ('BK32X_BK_MixDCol_DASH_DeltaAvg_Gr', 15, 5),
                         ('BK32X_BK_MixDCol_DASH_DeltaAvg_Gb', 15, 5),
                         ('BK32X_BK_MixDCol_DASH_DeltaAvg_B', 15, 5),
                         ('BK32X_BK_MixDCol_DASH_MaxValue0_R', 20, 8),
                         ('BK32X_BK_MixDCol_DASH_MaxValue0_Gr', 20, 7),
                         ('BK32X_BK_MixDCol_DASH_MaxValue0_Gb', 20, 7),
                         ('BK32X_BK_MixDCol_DASH_MaxValue0_B', 20, 7),
                         ('BK32X_BK_MixDCol_DASH_MaxValue1_R', 20, 7),
                         ('BK32X_BK_MixDCol_DASH_MaxValue1_Gr', 20, 7),
                         ('BK32X_BK_MixDCol_DASH_MaxValue1_Gb', 20, 7),
                         ('BK32X_BK_MixDCol_DASH_MaxValue1_B', 20, 7),
                         ('BK32X_BK_MixDCol_DASH_MaxValue2_R', 20, 7),
                         ('BK32X_BK_MixDCol_DASH_MaxValue2_Gr', 20, 7),
                         ('BK32X_BK_MixDCol_DASH_MaxValue2_Gb', 20, 7),
                         ('BK32X_BK_MixDCol_DASH_MaxValue2_B', 20, 7),
                         ('BK32X_BK_MixDCol_DASH_MaxValue3_R', 20, 7),
                         ('BK32X_BK_MixDCol_DASH_MaxValue3_Gr', 20, 7),
                         ('BK32X_BK_MixDCol_DASH_MaxValue3_Gb', 20, 7),
                         ('BK32X_BK_MixDCol_DASH_MaxValue3_B', 20, 7),
                         ('BK32X_BK_MixDCol_DASH_MaxValue4_R', 20, 7),
                         ('BK32X_BK_MixDCol_DASH_MaxValue4_Gr', 20, 7),
                         ('BK32X_BK_MixDCol_DASH_MaxValue4_Gb', 20, 7),
                         ('BK32X_BK_MixDCol_DASH_MaxValue4_B', 20, 7),
                         ('BK32X_BK_Mean_R', 30, 15),
                         ('BK32X_BK_Mean_Gr', 30, 15),
                         ('BK32X_BK_Mean_Gb', 30, 15),
                         ('BK32X_BK_Mean_B', 30, 15),
                         ('BK32X_BK_StdDEV_R', 4, 1),
                         ('BK32X_BK_StdDEV_Gr', 4, 1),
                         ('BK32X_BK_StdDEV_Gb', 4, 1),
                         ('BK32X_BK_StdDEV_B', 4, 1),
                         ('BK_Cluster3_2', 0, 0),
                         ('LT_Cluster3_2', 0, 0),

                         ('MIPI_One_W1_FixedPattern', 0, 0, 94),

                         ('MIPI_LT_DRow_R', 0, 0, 65),
                         ('MIPI_LT_DRow_Gr', 0, 0, 65),
                         ('MIPI_LT_DRow_Gb', 0, 0, 65),
                         ('MIPI_LT_DRow_B', 0, 0, 65),
                         ('MIPI_LT_DCol_R', 0, 0, 64),
                         ('MIPI_LT_DCol_Gr', 0, 0, 64),
                         ('MIPI_LT_DCol_Gb', 0, 0, 64),
                         ('MIPI_LT_DCol_B', 0, 0, 64),
                         ('MIPI_LT_DRow_Color', 0, 0, 65),
                         ('MIPI_LT_DCol_Color', 0, 0, 64),
                         ('MIPI_LT_WeakLineRow_R', 0, 0),
                         ('MIPI_LT_WeakLineRow_Gr', 0, 0),
                         ('MIPI_LT_WeakLineRow_Gb', 0, 0),
                         ('MIPI_LT_WeakLineRow_B', 0, 0),
                         ('MIPI_LT_WeakLineCol_R', 0, 0),
                         ('MIPI_LT_WeakLineCol_Gr', 0, 0),
                         ('MIPI_LT_WeakLineCol_Gb', 0, 0),
                         ('MIPI_LT_WeakLineCol_B', 0, 0),
                         ('MIPI_LT_Mean_R', 400, 380, 63),
                         ('MIPI_LT_Mean_Gr', 670, 640, 63),
                         ('MIPI_LT_Mean_Gb', 670, 640, 63),
                         ('MIPI_LT_Mean_B', 440, 400, 63),
                         ('MIPI_LT_StdDEV_R', 7, 4, 63),
                         ('MIPI_LT_StdDEV_Gr', 9, 6, 63),
                         ('MIPI_LT_StdDEV_Gb', 9, 6, 63),
                         ('MIPI_LT_StdDEV_B', 7, 4, 63),
                         ('MIPI_LT_RI_R', 70, 60, 63),
                         ('MIPI_LT_RI_Gr', 68, 59, 63),
                         ('MIPI_LT_RI_Gb', 68, 59, 63),
                         ('MIPI_LT_RI_B', 68, 59, 63),
                         ('MIPI_LT_Ratio_GrR', 1.7, 1.6, 63),
                         ('MIPI_LT_Ratio_GbR', 1.7, 1.6, 63),
                         ('MIPI_LT_Ratio_GrB', 1.6, 1.5, 63),
                         ('MIPI_LT_Ratio_GbB', 1.6, 1.5, 63),
                         ('MIPI_LT_Ratio_GbGr', 1, 0.99, 63),
                         ('MIPI_LT_LostBit', 0, 0, 63),
                         ('MIPI_LT_LostBitSNR_SumDIFF_R', 3, 0, 63),
                         ('MIPI_LT_LostBitSNR_SumDIFF_Gr', 3, 0, 63),
                         ('MIPI_LT_LostBitSNR_SumDIFF_Gb', 3, 0, 63),
                         ('MIPI_LT_LostBitSNR_SumDIFF_B', 3, 0, 63),

                         ('MIPI_LBIT_LT_Mean_R', 550, 530),
                         ('MIPI_LBIT_LT_Mean_Gr', 920, 890),
                         ('MIPI_LBIT_LT_Mean_Gb', 920, 890),
                         ('MIPI_LBIT_LT_Mean_B', 610, 560),
                         ('MIPI_LBIT_LT_LostBit', 0, 0, 63),
                         ('MIPI_LBIT_LT_LostBitSNR_SumDIFF_R', 4, 1, 63),
                         ('MIPI_LBIT_LT_LostBitSNR_SumDIFF_Gr', 4, 1, 63),
                         ('MIPI_LBIT_LT_LostBitSNR_SumDIFF_Gb', 4, 1, 63),
                         ('MIPI_LBIT_LT_LostBitSNR_SumDIFF_B', 4, 1, 63),

                         ('BK_Z1_BT_DP_R', 0, 0),
                         ('BK_Z1_BT_DP_Gr', 0, 0),
                         ('BK_Z1_BT_DP_Gb', 0, 0),
                         ('BK_Z1_BT_DP_B', 0, 0),
                         ('BK_Z2_BT_DP_R', 0, 0),
                         ('BK_Z2_BT_DP_Gr', 0, 0),
                         ('BK_Z2_BT_DP_Gb', 0, 0),
                         ('BK_Z2_BT_DP_B', 0, 0),
                         ('BK_Z1_BT_WP_R', 0, 0),
                         ('BK_Z1_BT_WP_Gr', 0, 0),
                         ('BK_Z1_BT_WP_Gb', 0, 0),
                         ('BK_Z1_BT_WP_B', 0, 0),
                         ('BK_Z2_BT_WP_R', 0, 0),
                         ('BK_Z2_BT_WP_Gr', 0, 0),
                         ('BK_Z2_BT_WP_Gb', 0, 0),
                         ('BK_Z2_BT_WP_B', 0, 0),
                         ('LT_Z1_BT_DP_R', 0, 0),
                         ('LT_Z1_BT_DP_Gr', 0, 0),
                         ('LT_Z1_BT_DP_Gb', 0, 0),
                         ('LT_Z1_BT_DP_B', 0, 0),
                         ('LT_Z2_BT_DP_R', 0, 0),
                         ('LT_Z2_BT_DP_Gr', 0, 0),
                         ('LT_Z2_BT_DP_Gb', 0, 0),
                         ('LT_Z2_BT_DP_B', 0, 0),
                         ('LT_Z1_BT_WP_R', 0, 0),
                         ('LT_Z1_BT_WP_Gr', 0, 0),
                         ('LT_Z1_BT_WP_Gb', 0, 0),
                         ('LT_Z1_BT_WP_B', 0, 0),
                         ('LT_Z2_BT_WP_R', 0, 0),
                         ('LT_Z2_BT_WP_Gr', 0, 0),
                         ('LT_Z2_BT_WP_Gb', 0, 0),
                         ('LT_Z2_BT_WP_B', 0, 0),
                         ('LT_Z1_DK_WP_R', 0, 0),
                         ('LT_Z1_DK_WP_Gr', 0, 0),
                         ('LT_Z1_DK_WP_Gb', 0, 0),
                         ('LT_Z1_DK_WP_B', 0, 0),
                         ('LT_Z2_DK_WP_R', 0, 0),
                         ('LT_Z2_DK_WP_Gr', 0, 0),
                         ('LT_Z2_DK_WP_Gb', 0, 0),
                         ('LT_Z2_DK_WP_B', 0, 0),
                         ('LT_Z1_DK_DP_R', 0, 0),
                         ('LT_Z1_DK_DP_Gr', 0, 0),
                         ('LT_Z1_DK_DP_Gb', 0, 0),
                         ('LT_Z1_DK_DP_B', 0, 0),
                         ('LT_Z2_DK_DP_R', 0, 0),
                         ('LT_Z2_DK_DP_Gr', 0, 0),
                         ('LT_Z2_DK_DP_Gb', 0, 0),
                         ('LT_Z2_DK_DP_B', 0, 0),
                         ('BK_Z1_BT_DP', 0, 0),
                         ('BK_Z2_BT_DP', 0, 0),
                         ('BK_Z1_BT_WP', 0, 0),
                         ('BK_Z2_BT_WP', 0, 0),
                         ('LT_Z1_BT_DP', 0, 0),
                         ('LT_Z2_BT_DP', 0, 0),
                         ('LT_Z1_BT_WP', 0, 0),
                         ('LT_Z2_BT_WP', 0, 0),
                         ('LT_Z1_DK_DP', 0, 0),
                         ('LT_Z2_DK_DP', 0, 0),
                         ('LT_Z1_DK_WP', 0, 0),
                         ('LT_Z2_DK_WP', 0, 0),

                         ('Binning', 0, 0, 2),
                         ]]
    golden_high_list = ['GoldenHigh', '', '', '']
    golden_low_list = ['GoldenLow', '', '', '']
    for i in range(len(golden_data_list[project_id])):
        golden_high_list.append(golden_data_list[project_id][i][1])
        golden_low_list.append(golden_data_list[project_id][i][2])
    return golden_high_list, golden_low_list


def parse_file(file, golden_data, project_id):
    """
    解析文件
    :param project_id: 项目id
    :param file: 待分析文件
    :param golden_data: golden数据
    :return: 分析结果
    """
    data = []
    with open(file) as f:
        csv_reader = reader(f)
        for row in csv_reader:
            data.append(row)
    if project_id == 0:
        search_last_test_item = Common.search_string(data, 'Full_Error')
    elif project_id == 1:
        search_last_test_item = Common.search_string(data, 'ChipVer')
    if not search_last_test_item:
        exit()
    else:
        last_test_item_row_num = search_last_test_item[0]
        last_test_item_col_num = search_last_test_item[1]

    search_binning = Common.search_string(data, 'Binning')
    if not search_binning:
        exit()
    else:
        binning_col_num = search_binning[1]

    search_iic_test = Common.search_string(data, 'iic_test')
    if not search_iic_test:
        exit()
    else:
        iic_test_col_num = search_iic_test[1]

    first_register_row_num = len(data)
    for i in range(last_test_item_row_num + row_offset, len(data)):
        try:
            int(data[i][0])
        except:
            first_register_row_num = i
            break

    for i in range(0, len(data)):
        add_count = last_test_item_col_num + 1 - len(data[i])
        if add_count > 0:
            for j in range(add_count):
                data[i].append('')

    col_count = last_test_item_col_num + 1
    row_count = len(data)
    exist_nan_row_list = []
    for i in range(last_test_item_row_num + row_offset, first_register_row_num):
        for j in range(last_test_item_col_num - 1):
            if len(data[i][j].strip()) == 0:
                exist_nan_row_list.append(i)
                break

    result = []

    for row_num in range(row_count):
        row_data = []
        for col_num in range(col_count):
            value = data[row_num][col_num]
            if col_offset <= col_num <= binning_col_num and last_test_item_row_num + 2 <= row_num <= last_test_item_row_num + 3:
                row_data.append((value, '008000'))  # 纯绿
            elif col_num == 0 and row_num in exist_nan_row_list:
                row_data.append((value, 'A020F0'))  # purple
            elif col_offset <= col_num <= binning_col_num and last_test_item_row_num + row_offset <= row_num < first_register_row_num:
                try:
                    value_convert = float(value)
                    high_limit_data = data[last_test_item_row_num + 2][col_num]
                    try:
                        high_limit = float(high_limit_data)
                    except:
                        high_limit = high_limit_data
                    low_limit_data = data[last_test_item_row_num + 3][col_num]
                    try:
                        low_limit = float(low_limit_data)
                    except:
                        low_limit = low_limit_data
                    if value_convert == int(value_convert):
                        value_convert = int(value_convert)
                    if golden_data[1][col_num] <= value_convert <= golden_data[0][col_num]:
                        row_data.append((value_convert, 'FFFFFF'))  # 白色
                    elif high_limit == 'N' or (high_limit != 'N' and (
                                        low_limit <= value_convert < golden_data[1][col_num] or golden_data[0][
                                col_num] < value_convert <= high_limit)):
                        row_data.append((value_convert, 'FFFF00'))  # 纯黄
                    else:
                        row_data.append((value_convert, 'FF0000'))  # 纯红
                except:
                    if not value.strip():
                        row_data.append((value, 'A020F0'))  # purple
                    elif col_num == iic_test_col_num and value != '1':
                        row_data.append((value, 'FFFF00'))  # 纯黄
            else:
                if isinstance(value, float) and isnan(value):
                    value = ''
                row_data.append((value, 'FFFFFF'))
        result.append(row_data)
        if row_num % 100 == 0 or row_num == row_count - 1:
            print("解析文件：第 " + str(row_num + 1) + " 行")
    return result


def save_data(analysis_folder, file, parse_data, golden_data):
    """
    保存数据
    :param analysis_folder: 分析文件夹
    :param file: 文件名
    :param parse_data: 解析数据
    :param golden_data: golden数据
    :return: 
    """
    analysis_file = join(analysis_folder, basename(file))
    file_name = analysis_file.split('.')[0]
    date = datetime.now().strftime("%Y%m%d%H%M")
    test_file_name = file_name + ' vs GoldenData_' + date + '.xlsx'
    print(test_file_name)

    wb = Workbook()  # 创建文件对象
    ws = wb.active  # 获取第一个sheet
    ws.freeze_panes = 'E19'

    irow = 1
    for i in range(len(parse_data)):
        if i == 14:
            for m in range(len(golden_data[0])):
                ws.cell(row=irow, column=m + 1).value = golden_data[0][m]
                ws.cell(row=irow, column=m + 1).fill = PatternFill(fill_type='solid', fgColor=golden_data_color)
                ws.cell(row=irow, column=m + 1).border = border
                if m == len(golden_data[0]) - 1:
                    for x in range(m + 1, len(parse_data[0])):
                        ws.cell(row=irow, column=x + 1).border = border
            irow += 1
            for n in range(len(golden_data[1])):
                ws.cell(row=irow, column=n + 1).value = golden_data[1][n]
                ws.cell(row=irow, column=n + 1).fill = PatternFill(fill_type='solid', fgColor=golden_data_color)
                ws.cell(row=irow, column=n + 1).border = border
                if n == len(golden_data[0]) - 1:
                    for x in range(n + 1, len(parse_data[0])):
                        ws.cell(row=irow, column=x + 1).border = border
            irow += 1
        for j in range(len(parse_data[i])):
            ws.cell(row=irow, column=j + 1).value = parse_data[i][j][0]
            ws.cell(row=irow, column=j + 1).fill = PatternFill(fill_type='solid', fgColor=parse_data[i][j][1])
            ws.cell(row=irow, column=j + 1).border = border
        if i % 100 == 0 or i == len(parse_data) - 1:
            print("组装Excel数据：第 " + str(i + 1) + " 行")
        irow += 1

    print("保存数据开始-----------------")
    wb.save(test_file_name)
    print("保存数据结束-----------------")


def main():
    # Sourcefile folder path
    if argv.count('-s') == 0 and argv.count('-f') == 0:
        print(
            "Error：Sourcefile folder path 或 single file path为必填项，格式：“-s D:\sourcefile” 或 “-f D:\sourcefile\ST5-440mm-out1.csv”。")
        exit()

    if argv.count('-s') != 0:
        sourcefile_folder = argv[argv.index('-s') + 1]
        for i in range(argv.index('-s') + 2, len(argv)):
            if not argv[i].startswith('-'):
                sourcefile_folder += (' ' + argv[i])
            else:
                break
        file_list = Common.get_filelist(sourcefile_folder, '.csv')
        if not file_list:
            exit()

    if argv.count('-f') != 0:
        single_file = argv[argv.index('-f') + 1]
        for i in range(argv.index('-f') + 2, len(argv)):
            if not argv[i].startswith('-'):
                single_file += (' ' + argv[i])
            else:
                break
        sourcefile_folder = dirname(single_file)
        file_list = [single_file]

    # Analysis folder path
    if argv.count('-a') == 0:
        analysis_folder = sourcefile_folder + '\Analysis'
    else:
        analysis_folder = argv[argv.index('-a') + 1]

    # 项目 0:F28,1:JX828
    if argv.count('-p') != 0:
        project_id = int(argv[argv.index('-p') + 1])
    else:
        project_id = 0  # 默认F28项目

    golden_data = get_golden_data(project_id)
    Common.mkdir(analysis_folder)

    for file in file_list:
        print(file)
        # parse file
        parse_data = parse_file(file, golden_data, project_id)
        # save data
        save_data(analysis_folder, file, parse_data, golden_data)


if __name__ == '__main__':
    main()
