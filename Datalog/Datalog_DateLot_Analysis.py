# encoding:utf-8
# @Time     : 2019/9/9 13:32
# @Author   : Jerry Chou
# @File     :
# @Function : Date data analysis

import Datalog_DateLot_Analysis_UI
from csv import reader, field_size_limit
from os.path import basename, isdir, join, exists
from os import listdir, makedirs, getcwd
from datetime import datetime
from math import isnan
from sys import argv, maxsize, exit
from PyQt5.QtCore import QThread, pyqtSignal
from PyQt5.QtGui import QPalette, QBrush, QPixmap
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QErrorMessage
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.styles.colors import YELLOW, GREEN, BLACK, WHITE, RED
from openpyxl.utils import get_column_letter
from numpy import array
import time
from copy import deepcopy


def time_calc(func):
    """
    calculate function execution time and write in file
    """

    def wrapper(*args, **kargs):
        start_time = time.time()
        f = func(*args, **kargs)
        exec_time = time.time() - start_time
        if exec_time > 0:
            with open(join(open_path, 'log.log'), 'a') as log:
                log.write('{} function cost time : {}s\n'.format(func.__name__, exec_time))
        return f

    return wrapper


def find_item(item_list, item):
    """
    find value in item list
    :return: a list of item index
    """
    return [i for i, v in enumerate(item_list) if v == item]


def search_string(data, target):
    """
    find out if target exists in data
    :return: return index if find target,else return False
    """
    for i in range(len(data)):
        try:
            column_num = data[i].index(target)
            row_num = i
            return row_num, column_num
        except:
            pass
    print("Can't find " + str(target) + " !")
    return False


def set_column_width(sheet):
    """
    set column width
    """
    # get the maximum width of each column
    column_width = [0.5] * sheet.max_column
    for row in range(sheet.max_row):
        for col in range(sheet.max_column):
            value = sheet.cell(row=row + 1, column=col + 1).value
            if value:
                width = len(str(value))
                if width > column_width[col]:
                    column_width[col] = width
    # set column width
    for i in range(len(column_width)):
        column_letter = get_column_letter(i + 1)
        if column_width[i] > 100:
            # set to 100 if col_width greater than 100
            sheet.column_dimensions[column_letter].width = 100
        else:
            sheet.column_dimensions[column_letter].width = column_width[i] + 4


def get_file_list(folder, postfix=None):
    """
    find a list of files with a postfix
    """
    full_name_list = []
    if isdir(folder):
        files = listdir(folder)
        for filename in files:
            full_name_list.append(join(folder, filename))
        if postfix:
            target_file_list = []
            for full_name in full_name_list:
                if full_name.endswith(postfix.upper()) or full_name.endswith(postfix.lower()):
                    target_file_list.append(full_name)
            return target_file_list
        else:
            return full_name_list
    else:
        print('{} is not a folder.'.format(folder))
        return False


def create_directory(path):
    """
    create a folder
    """
    path = path.strip()
    path = path.rstrip("\\")
    is_exists = exists(path)
    if not is_exists:
        makedirs(path)
        return True
    else:
        print('{} already exists.'.format(path))
        return False


row_offset = 5
column_offset = 4
high_limit_row_num = 0
# default project
project = 'Unknown'
analysis_item = 'Unknown'
# get current time
now_time = 'Unknown'
alignment = Alignment(horizontal='center', vertical='center')
thin = Side(border_style='thin', color=BLACK)
border = Border(top=thin, left=thin, right=thin, bottom=thin)
total_row_count = 0
current_row_count = 0
open_path = ''

HWBIN_TO_SWBIN = {
    'F28': {
        3: {'SWBin': (1, 2), 'isPassBin': True},
        1: {'SWBin': (37, 38, 61, 62, 64, 65, 66, 90, 91, 92, 94), 'isPassBin': False},
        2: {'SWBin': (41, 42, 43, 44, 45, 53, 54, 55), 'isPassBin': False},
        4: {'SWBin': (23, 24, 25, 26, 27, 29, 30, 31, 32, 33, 34, 36, 39, 40), 'isPassBin': False},
        5: {'SWBin': (5, 6, 7, 8, 9, 12, 96, 97, 98, 99), 'isPassBin': False},
        6: {'SWBin': (13, 14, 15, 35), 'isPassBin': False}
    },
    'JX828': {
        3: {'SWBin': (1, 2, 3), 'isPassBin': True},
        1: {'SWBin': (63, 64, 65, 89, 90, 94), 'isPassBin': True},
        2: {'SWBin': (53, 54, 73, 74), 'isPassBin': True},
        4: {'SWBin': (23, 24, 25, 26, 27, 29, 30, 31, 32, 33, 34, 36, 39, 40, 56, 57, 58, 75), 'isPassBin': False},
        5: {'SWBin': (5, 6, 7, 8, 9, 12, 93, 96, 98, 99), 'isPassBin': False},
        6: {'SWBin': (13, 14, 15, 35, 46, 47, 48, 51, 60), 'isPassBin': False}
    },
    'JX825': {
        3: {'SWBin': (2, 255), 'isPassBin': True},
        4: {'SWBin': (23, 24, 25, 26, 27, 29, 30, 31, 32, 33, 34, 36, 39, 40, 63, 64, 65), 'isPassBin': False},
        5: {'SWBin': (5, 6, 7, 8, 9, 12, 89), 'isPassBin': False},
        6: {'SWBin': (13, 14, 15, 35, 46, 48, 53, 54, 60), 'isPassBin': False},
        8: {'SWBin': (94, 96, 97, 98, 99), 'isPassBin': False}
    },
    'JX832': {
        3: {'SWBin': (1, 2, 3), 'isPassBin': True},
        1: {'SWBin': (56, 57, 58, 66, 67, 68, 95, 96), 'isPassBin': True},
        2: {'SWBin': (53, 54, 63, 64, 65, 73, 74, 89, 90, 94), 'isPassBin': True},
        4: {'SWBin': (23, 24, 25, 26, 27, 29, 30, 31, 32, 33, 34, 36, 39, 40, 78), 'isPassBin': False},
        5: {'SWBin': (5, 6, 7, 8, 9, 12, 93, 97, 98, 99), 'isPassBin': False},
        6: {'SWBin': (13, 14, 15, 35, 46, 47, 48, 51, 60), 'isPassBin': False}
    }
}

BIN_DEFINITION = {
    'F28':
        {
            'PCLK_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 1},
            'HSYNC_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 2},
            'VSYNC_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 3},
            'D9_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 4},
            'D8_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 5},
            'D7_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 6},
            'D6_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 7},
            'D5_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 8},
            'D4_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 9},
            'D3_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 10},
            'D2_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 11},
            'D1_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 12},
            'D0_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 13},
            'SCL_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 14},
            'SDA_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 15},
            'RSTB_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 16},
            'PWDN_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 17},
            'DVDD_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 18},
            'VRamp_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 19},
            'VH_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 20},
            'VN1_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 21},
            'EXCLK_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 22},
            'PCLK_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 23},
            'HSYNC_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 24},
            'VSYNC_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 25},
            'D9_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 26},
            'D8_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 27},
            'D7_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 28},
            'D6_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 29},
            'D5_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 30},
            'D4_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 31},
            'D3_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 32},
            'D2_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 33},
            'D1_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 34},
            'D0_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 35},
            'SCL_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 36},
            'SDA_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 37},
            'RSTB_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 38},
            'PWDN_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 39},
            'EXCLK_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 40},
            'PCLK_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 41},
            'HSYNC_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 42},
            'VSYNC_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 43},
            'D9_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 44},
            'D8_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 45},
            'D7_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 46},
            'D6_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 47},
            'D5_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 48},
            'D4_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 49},
            'D3_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 50},
            'D2_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 51},
            'D1_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 52},
            'D0_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 53},
            'SCL_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 54},
            'SDA_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 55},
            'RSTB_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 56},
            'PWDN_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 57},
            'EXCLK_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 58},
            'iic_test': {'SWBin': 7, 'HWBin': 5, 'Priority': 59},
            'DVDD_voltage': {'SWBin': 9, 'HWBin': 5, 'Priority': 60},
            'VH_voltage': {'SWBin': 9, 'HWBin': 5, 'Priority': 61},
            'VN1_voltage': {'SWBin': 9, 'HWBin': 5, 'Priority': 62},
            'Active_AVDD': {'SWBin': 12, 'HWBin': 5, 'Priority': 63},
            'Active_DOVDD': {'SWBin': 12, 'HWBin': 5, 'Priority': 64},
            'PWDN_AVDD': {'SWBin': 8, 'HWBin': 5, 'Priority': 65},
            'PWDN_DOVDD': {'SWBin': 8, 'HWBin': 5, 'Priority': 66},
            'PWDN_Total': {'SWBin': 8, 'HWBin': 5, 'Priority': 67},
            'PLCK_Freq': {'SWBin': 9, 'HWBin': 5, 'Priority': 68},
            'BK_ImageCapture': {'SWBin': 98, 'HWBin': 5, 'Priority': 69, 'Status': []},
            'BK_ImageCalculation': {'SWBin': 97, 'HWBin': 5, 'Priority': 70},
            'BLC_R': {'SWBin': 7, 'HWBin': 5, 'Priority': 71},
            'BLC_Gr': {'SWBin': 7, 'HWBin': 5, 'Priority': 72},
            'BLC_Gb': {'SWBin': 7, 'HWBin': 5, 'Priority': 73},
            'BLC_B': {'SWBin': 7, 'HWBin': 5, 'Priority': 74},
            'BK_DeadRowExBPix_R': {'SWBin': 15, 'HWBin': 6, 'Priority': 75, 'IsNan': 'BK_ImageCapture'},
            'BK_DeadRowExBPix_Gr': {'SWBin': 15, 'HWBin': 6, 'Priority': 76, 'IsNan': 'BK_ImageCapture'},
            'BK_DeadRowExBPix_Gb': {'SWBin': 15, 'HWBin': 6, 'Priority': 77, 'IsNan': 'BK_ImageCapture'},
            'BK_DeadRowExBPix_B': {'SWBin': 15, 'HWBin': 6, 'Priority': 78, 'IsNan': 'BK_ImageCapture'},
            'BK_DeadColExBPix_R': {'SWBin': 14, 'HWBin': 6, 'Priority': 79, 'IsNan': 'BK_ImageCapture'},
            'BK_DeadColExBPix_Gr': {'SWBin': 14, 'HWBin': 6, 'Priority': 80, 'IsNan': 'BK_ImageCapture'},
            'BK_DeadColExBPix_Gb': {'SWBin': 14, 'HWBin': 6, 'Priority': 81, 'IsNan': 'BK_ImageCapture'},
            'BK_DeadColExBPix_B': {'SWBin': 14, 'HWBin': 6, 'Priority': 82, 'IsNan': 'BK_ImageCapture'},
            'BK_Mean_R': {'SWBin': 13, 'HWBin': 6, 'Priority': 83, 'IsNan': 'BK_ImageCapture'},
            'BK_Mean_Gr': {'SWBin': 13, 'HWBin': 6, 'Priority': 84, 'IsNan': 'BK_ImageCapture'},
            'BK_Mean_Gb': {'SWBin': 13, 'HWBin': 6, 'Priority': 85, 'IsNan': 'BK_ImageCapture'},
            'BK_Mean_B': {'SWBin': 13, 'HWBin': 6, 'Priority': 86, 'IsNan': 'BK_ImageCapture'},
            'BK_StdDEV_R': {'SWBin': 13, 'HWBin': 6, 'Priority': 87, 'IsNan': 'BK_ImageCapture'},
            'BK_StdDEV_Gr': {'SWBin': 13, 'HWBin': 6, 'Priority': 88, 'IsNan': 'BK_ImageCapture'},
            'BK_StdDEV_Gb': {'SWBin': 13, 'HWBin': 6, 'Priority': 89, 'IsNan': 'BK_ImageCapture'},
            'BK_StdDEV_B': {'SWBin': 13, 'HWBin': 6, 'Priority': 90, 'IsNan': 'BK_ImageCapture','IsLast': ''},
            'LT_ImageCapture': {'SWBin': 99, 'HWBin': 5, 'Priority': 91, 'Status': []},
            'LT_ImageCalculation': {'SWBin': 97, 'HWBin': 5, 'Priority': 92},
            'LT_DRow_R': {'SWBin': 25, 'HWBin': 4, 'Priority': 93, 'IsNan': 'LT_ImageCapture'},
            'LT_DRow_Gr': {'SWBin': 25, 'HWBin': 4, 'Priority': 94, 'IsNan': 'LT_ImageCapture'},
            'LT_DRow_Gb': {'SWBin': 25, 'HWBin': 4, 'Priority': 95, 'IsNan': 'LT_ImageCapture'},
            'LT_DRow_B': {'SWBin': 25, 'HWBin': 4, 'Priority': 96, 'IsNan': 'LT_ImageCapture'},
            'LT_DCol_R': {'SWBin': 24, 'HWBin': 4, 'Priority': 97, 'IsNan': 'LT_ImageCapture'},
            'LT_DCol_Gr': {'SWBin': 24, 'HWBin': 4, 'Priority': 98, 'IsNan': 'LT_ImageCapture'},
            'LT_DCol_Gb': {'SWBin': 24, 'HWBin': 4, 'Priority': 99, 'IsNan': 'LT_ImageCapture'},
            'LT_DCol_B': {'SWBin': 24, 'HWBin': 4, 'Priority': 100, 'IsNan': 'LT_ImageCapture'},
            'LT_DRow_Color': {'SWBin': 25, 'HWBin': 4, 'Priority': 101, 'IsNan': 'LT_ImageCapture'},
            'LT_DCol_Color': {'SWBin': 24, 'HWBin': 4, 'Priority': 102, 'IsNan': 'LT_ImageCapture'},
            'LT_Mean_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 103, 'IsNan': 'LT_ImageCapture'},
            'LT_Mean_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 104, 'IsNan': 'LT_ImageCapture'},
            'LT_Mean_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 105, 'IsNan': 'LT_ImageCapture'},
            'LT_Mean_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 106, 'IsNan': 'LT_ImageCapture'},
            'LT_StdDEV_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 107, 'IsNan': 'LT_ImageCapture'},
            'LT_StdDEV_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 108, 'IsNan': 'LT_ImageCapture'},
            'LT_StdDEV_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 109, 'IsNan': 'LT_ImageCapture'},
            'LT_StdDEV_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 110, 'IsNan': 'LT_ImageCapture'},
            'LT_RI_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 111, 'IsNan': 'LT_ImageCapture'},
            'LT_RI_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 112, 'IsNan': 'LT_ImageCapture'},
            'LT_RI_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 113, 'IsNan': 'LT_ImageCapture'},
            'LT_RI_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 114, 'IsNan': 'LT_ImageCapture'},
            'LT_Ratio_GrR': {'SWBin': 23, 'HWBin': 4, 'Priority': 115, 'IsNan': 'LT_ImageCapture'},
            'LT_Ratio_GbR': {'SWBin': 23, 'HWBin': 4, 'Priority': 116, 'IsNan': 'LT_ImageCapture'},
            'LT_Ratio_GrB': {'SWBin': 23, 'HWBin': 4, 'Priority': 117, 'IsNan': 'LT_ImageCapture'},
            'LT_Ratio_GbB': {'SWBin': 23, 'HWBin': 4, 'Priority': 118, 'IsNan': 'LT_ImageCapture'},
            'LT_Ratio_GbGr': {'SWBin': 23, 'HWBin': 4, 'Priority': 119, 'IsNan': 'LT_ImageCapture'},
            'LT_LostBit': {'SWBin': 23, 'HWBin': 4, 'Priority': 120, 'IsNan': 'LT_ImageCapture'},
            'LT_LostBitSNR_SumDIFF_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 121, 'IsNan': 'LT_ImageCapture'},
            'LT_LostBitSNR_SumDIFF_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 122, 'IsNan': 'LT_ImageCapture'},
            'LT_LostBitSNR_SumDIFF_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 123, 'IsNan': 'LT_ImageCapture'},
            'LT_LostBitSNR_SumDIFF_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 124, 'IsNan': 'LT_ImageCapture','IsLast': 'LT_ImageCapture'},
            'LB_LT_ImageCapture': {'SWBin': 99, 'HWBin': 5, 'Priority': 125, 'Status': []},
            'LB_LT_ImageCalculation': {'SWBin': 97, 'HWBin': 5, 'Priority': 126},
            'LB_LT_Mean_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 127, 'IsNan': 'LB_LT_ImageCapture'},
            'LB_LT_Mean_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 128, 'IsNan': 'LB_LT_ImageCapture'},
            'LB_LT_Mean_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 129, 'IsNan': 'LB_LT_ImageCapture'},
            'LB_LT_Mean_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 130, 'IsNan': 'LB_LT_ImageCapture'},
            'LB_LT_LostBit': {'SWBin': 23, 'HWBin': 4, 'Priority': 131, 'IsNan': 'LB_LT_ImageCapture'},
            'LB_LT_LostBitSNR_SumDIFF_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 132, 'IsNan': 'LB_LT_ImageCapture'},
            'LB_LT_LostBitSNR_SumDIFF_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 133, 'IsNan': 'LB_LT_ImageCapture'},
            'LB_LT_LostBitSNR_SumDIFF_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 134, 'IsNan': 'LB_LT_ImageCapture'},
            'LB_LT_LostBitSNR_SumDIFF_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 135, 'IsNan': 'LB_LT_ImageCapture','IsLast': 'LB_LT_ImageCapture'},
            'FW_LT_ImageCapture': {'SWBin': 99, 'HWBin': 5, 'Priority': 136, 'Status': []},
            'FW_LT_Calculation': {'SWBin': 97, 'HWBin': 5, 'Priority': 137},
            'FW_LT_DRow_R': {'SWBin': 36, 'HWBin': 4, 'Priority': 138, 'IsNan': 'FW_LT_ImageCapture'},
            'FW_LT_DRow_Gr': {'SWBin': 36, 'HWBin': 4, 'Priority': 139, 'IsNan': 'FW_LT_ImageCapture'},
            'FW_LT_DRow_Gb': {'SWBin': 36, 'HWBin': 4, 'Priority': 140, 'IsNan': 'FW_LT_ImageCapture'},
            'FW_LT_DRow_B': {'SWBin': 36, 'HWBin': 4, 'Priority': 141, 'IsNan': 'FW_LT_ImageCapture'},
            'FW_LT_DCol_R': {'SWBin': 36, 'HWBin': 4, 'Priority': 142, 'IsNan': 'FW_LT_ImageCapture'},
            'FW_LT_DCol_Gr': {'SWBin': 36, 'HWBin': 4, 'Priority': 143, 'IsNan': 'FW_LT_ImageCapture'},
            'FW_LT_DCol_Gb': {'SWBin': 36, 'HWBin': 4, 'Priority': 144, 'IsNan': 'FW_LT_ImageCapture'},
            'FW_LT_DCol_B': {'SWBin': 36, 'HWBin': 4, 'Priority': 145, 'IsNan': 'FW_LT_ImageCapture'},
            'FW_LT_DRow_Color': {'SWBin': 36, 'HWBin': 4, 'Priority': 146, 'IsNan': 'FW_LT_ImageCapture'},
            'FW_LT_DCol_Color': {'SWBin': 36, 'HWBin': 4, 'Priority': 147, 'IsNan': 'FW_LT_ImageCapture'},
            'FW_LT_Mean_R': {'SWBin': 36, 'HWBin': 4, 'Priority': 148, 'IsNan': 'FW_LT_ImageCapture'},
            'FW_LT_Mean_Gr': {'SWBin': 36, 'HWBin': 4, 'Priority': 149, 'IsNan': 'FW_LT_ImageCapture'},
            'FW_LT_Mean_Gb': {'SWBin': 36, 'HWBin': 4, 'Priority': 150, 'IsNan': 'FW_LT_ImageCapture'},
            'FW_LT_Mean_B': {'SWBin': 36, 'HWBin': 4, 'Priority': 151, 'IsNan': 'FW_LT_ImageCapture'},
            'FW_LT_StdDEV_R': {'SWBin': 36, 'HWBin': 4, 'Priority': 152, 'IsNan': 'FW_LT_ImageCapture'},
            'FW_LT_StdDEV_Gr': {'SWBin': 36, 'HWBin': 4, 'Priority': 153, 'IsNan': 'FW_LT_ImageCapture'},
            'FW_LT_StdDEV_Gb': {'SWBin': 36, 'HWBin': 4, 'Priority': 154, 'IsNan': 'FW_LT_ImageCapture'},
            'FW_LT_StdDEV_B': {'SWBin': 36, 'HWBin': 4, 'Priority': 155, 'IsNan': 'FW_LT_ImageCapture'},
            'FW_LT_RI_R': {'SWBin': 36, 'HWBin': 4, 'Priority': 156, 'IsNan': 'FW_LT_ImageCapture'},
            'FW_LT_RI_Gr': {'SWBin': 36, 'HWBin': 4, 'Priority': 157, 'IsNan': 'FW_LT_ImageCapture'},
            'FW_LT_RI_Gb': {'SWBin': 36, 'HWBin': 4, 'Priority': 158, 'IsNan': 'FW_LT_ImageCapture'},
            'FW_LT_RI_B': {'SWBin': 36, 'HWBin': 4, 'Priority': 159, 'IsNan': 'FW_LT_ImageCapture'},
            'FW_LT_Ratio_GrR': {'SWBin': 36, 'HWBin': 4, 'Priority': 160, 'IsNan': 'FW_LT_ImageCapture'},
            'FW_LT_Ratio_GbR': {'SWBin': 36, 'HWBin': 4, 'Priority': 161, 'IsNan': 'FW_LT_ImageCapture'},
            'FW_LT_Ratio_GrB': {'SWBin': 36, 'HWBin': 4, 'Priority': 162, 'IsNan': 'FW_LT_ImageCapture'},
            'FW_LT_Ratio_GbB': {'SWBin': 36, 'HWBin': 4, 'Priority': 163, 'IsNan': 'FW_LT_ImageCapture'},
            'FW_LT_Ratio_GbGr': {'SWBin': 36, 'HWBin': 4, 'Priority': 164, 'IsNan': 'FW_LT_ImageCapture','IsLast': 'FW_LT_ImageCapture'},
            'LT_CornerLine': {'SWBin': 26, 'HWBin': 4, 'Priority': 165},
            'LT_ScratchLine': {'SWBin': 26, 'HWBin': 4, 'Priority': 166},
            'LT_Blemish': {'SWBin': 31, 'HWBin': 4, 'Priority': 167},
            'LT_LineStripe': {'SWBin': 27, 'HWBin': 4, 'Priority': 168},
            'LT_Particle': {'SWBin': 32, 'HWBin': 4, 'Priority': 169},
            'BK_Cluster2': {'SWBin': 30, 'HWBin': 4, 'Priority': 170},
            'LT_Cluster2': {'SWBin': 30, 'HWBin': 4, 'Priority': 171},
            'BK_Cluster1': {'SWBin': 29, 'HWBin': 4, 'Priority': 172},
            'LT_Cluster1': {'SWBin': 29, 'HWBin': 4, 'Priority': 173},
            'WP_Count': {'SWBin': 35, 'HWBin': 6, 'Priority': 174},
            'BK_Cluster3GrGb': {'SWBin': 39, 'HWBin': 4, 'Priority': 175},
            'LT_Cluster3GrGb': {'SWBin': 40, 'HWBin': 4, 'Priority': 176},
            'BK_Cluster3SubtractGrGb': {'SWBin': 33, 'HWBin': 4, 'Priority': 177},
            'LT_Cluster3SubtractGrGb': {'SWBin': 34, 'HWBin': 4, 'Priority': 178},
            'LB_BK_DeadRowExBPix_R': {'SWBin': 45, 'HWBin': 2, 'Priority': 179},
            'LB_BK_DeadRowExBPix_Gr': {'SWBin': 45, 'HWBin': 2, 'Priority': 180},
            'LB_BK_DeadRowExBPix_Gb': {'SWBin': 45, 'HWBin': 2, 'Priority': 181},
            'LB_BK_DeadRowExBPix_B': {'SWBin': 45, 'HWBin': 2, 'Priority': 182},
            'LB_BK_DeadColExBPix_R': {'SWBin': 44, 'HWBin': 2, 'Priority': 183},
            'LB_BK_DeadColExBPix_Gr': {'SWBin': 44, 'HWBin': 2, 'Priority': 184},
            'LB_BK_DeadColExBPix_Gb': {'SWBin': 44, 'HWBin': 2, 'Priority': 185},
            'LB_BK_DeadColExBPix_B': {'SWBin': 44, 'HWBin': 2, 'Priority': 186},
            'SR_LT_ImageCapture': {'SWBin': 96, 'HWBin': 5, 'Priority': 187, 'Status': []},
            'SR_LT_Calculation': {'SWBin': 97, 'HWBin': 5, 'Priority': 188},
            'SR_LT_DRow_R': {'SWBin': 55, 'HWBin': 2, 'Priority': 189, 'IsNan': 'SR_LT_ImageCapture'},
            'SR_LT_DRow_Gr': {'SWBin': 55, 'HWBin': 2, 'Priority': 190, 'IsNan': 'SR_LT_ImageCapture'},
            'SR_LT_DRow_Gb': {'SWBin': 55, 'HWBin': 2, 'Priority': 191, 'IsNan': 'SR_LT_ImageCapture'},
            'SR_LT_DRow_B': {'SWBin': 55, 'HWBin': 2, 'Priority': 192, 'IsNan': 'SR_LT_ImageCapture'},
            'SR_LT_DCol_R': {'SWBin': 54, 'HWBin': 2, 'Priority': 193, 'IsNan': 'SR_LT_ImageCapture'},
            'SR_LT_DCol_Gr': {'SWBin': 54, 'HWBin': 2, 'Priority': 194, 'IsNan': 'SR_LT_ImageCapture'},
            'SR_LT_DCol_Gb': {'SWBin': 54, 'HWBin': 2, 'Priority': 195, 'IsNan': 'SR_LT_ImageCapture'},
            'SR_LT_DCol_B': {'SWBin': 54, 'HWBin': 2, 'Priority': 196, 'IsNan': 'SR_LT_ImageCapture'},
            'SR_LT_DRow_Color': {'SWBin': 55, 'HWBin': 2, 'Priority': 197, 'IsNan': 'SR_LT_ImageCapture'},
            'SR_LT_DCol_Color': {'SWBin': 54, 'HWBin': 2, 'Priority': 198, 'IsNan': 'SR_LT_ImageCapture'},
            'SR_LT_Mean_R': {'SWBin': 53, 'HWBin': 2, 'Priority': 199, 'IsNan': 'SR_LT_ImageCapture'},
            'SR_LT_Mean_Gr': {'SWBin': 53, 'HWBin': 2, 'Priority': 200, 'IsNan': 'SR_LT_ImageCapture'},
            'SR_LT_Mean_Gb': {'SWBin': 53, 'HWBin': 2, 'Priority': 201, 'IsNan': 'SR_LT_ImageCapture'},
            'SR_LT_Mean_B': {'SWBin': 53, 'HWBin': 2, 'Priority': 202, 'IsNan': 'SR_LT_ImageCapture'},
            'SR_LT_StdDEV_R': {'SWBin': 53, 'HWBin': 2, 'Priority': 203, 'IsNan': 'SR_LT_ImageCapture'},
            'SR_LT_StdDEV_Gr': {'SWBin': 53, 'HWBin': 2, 'Priority': 204, 'IsNan': 'SR_LT_ImageCapture'},
            'SR_LT_StdDEV_Gb': {'SWBin': 53, 'HWBin': 2, 'Priority': 205, 'IsNan': 'SR_LT_ImageCapture'},
            'SR_LT_StdDEV_B': {'SWBin': 53, 'HWBin': 2, 'Priority': 206, 'IsNan': 'SR_LT_ImageCapture'},
            'SR_LT_RI_R': {'SWBin': 53, 'HWBin': 2, 'Priority': 207, 'IsNan': 'SR_LT_ImageCapture'},
            'SR_LT_RI_Gr': {'SWBin': 53, 'HWBin': 2, 'Priority': 208, 'IsNan': 'SR_LT_ImageCapture'},
            'SR_LT_RI_Gb': {'SWBin': 53, 'HWBin': 2, 'Priority': 209, 'IsNan': 'SR_LT_ImageCapture'},
            'SR_LT_RI_B': {'SWBin': 53, 'HWBin': 2, 'Priority': 210, 'IsNan': 'SR_LT_ImageCapture'},
            'SR_LT_Ratio_GrR': {'SWBin': 53, 'HWBin': 2, 'Priority': 211, 'IsNan': 'SR_LT_ImageCapture'},
            'SR_LT_Ratio_GbR': {'SWBin': 53, 'HWBin': 2, 'Priority': 212, 'IsNan': 'SR_LT_ImageCapture'},
            'SR_LT_Ratio_GrB': {'SWBin': 53, 'HWBin': 2, 'Priority': 213, 'IsNan': 'SR_LT_ImageCapture'},
            'SR_LT_Ratio_GbB': {'SWBin': 53, 'HWBin': 2, 'Priority': 214, 'IsNan': 'SR_LT_ImageCapture'},
            'SR_LT_Ratio_GbGr': {'SWBin': 53, 'HWBin': 2, 'Priority': 215, 'IsNan': 'SR_LT_ImageCapture'},
            'SR_LT_LostBit': {'SWBin': 53, 'HWBin': 2, 'Priority': 216, 'IsNan': 'SR_LT_ImageCapture','IsLast': 'SR_LT_ImageCapture'},
            'SR_BK_ImageCapture': {'SWBin': 96, 'HWBin': 5, 'Priority': 217, 'Status': []},
            'SR_BK_Calculation': {'SWBin': 97, 'HWBin': 5, 'Priority': 218},
            'SR_BK_DeadRowExBPix_R': {'SWBin': 42, 'HWBin': 2, 'Priority': 219, 'IsNan': 'SR_BK_ImageCapture'},
            'SR_BK_DeadRowExBPix_Gr': {'SWBin': 42, 'HWBin': 2, 'Priority': 220, 'IsNan': 'SR_BK_ImageCapture'},
            'SR_BK_DeadRowExBPix_Gb': {'SWBin': 42, 'HWBin': 2, 'Priority': 221, 'IsNan': 'SR_BK_ImageCapture'},
            'SR_BK_DeadRowExBPix_B': {'SWBin': 42, 'HWBin': 2, 'Priority': 222, 'IsNan': 'SR_BK_ImageCapture'},
            'SR_BK_DeadColExBPix_R': {'SWBin': 41, 'HWBin': 2, 'Priority': 223, 'IsNan': 'SR_BK_ImageCapture'},
            'SR_BK_DeadColExBPix_Gr': {'SWBin': 41, 'HWBin': 2, 'Priority': 224, 'IsNan': 'SR_BK_ImageCapture'},
            'SR_BK_DeadColExBPix_Gb': {'SWBin': 41, 'HWBin': 2, 'Priority': 225, 'IsNan': 'SR_BK_ImageCapture'},
            'SR_BK_DeadColExBPix_B': {'SWBin': 41, 'HWBin': 2, 'Priority': 226, 'IsNan': 'SR_BK_ImageCapture'},
            'SR_BK_Mean_R': {'SWBin': 43, 'HWBin': 2, 'Priority': 227, 'IsNan': 'SR_BK_ImageCapture'},
            'SR_BK_Mean_Gr': {'SWBin': 43, 'HWBin': 2, 'Priority': 228, 'IsNan': 'SR_BK_ImageCapture'},
            'SR_BK_Mean_Gb': {'SWBin': 43, 'HWBin': 2, 'Priority': 229, 'IsNan': 'SR_BK_ImageCapture'},
            'SR_BK_Mean_B': {'SWBin': 43, 'HWBin': 2, 'Priority': 230, 'IsNan': 'SR_BK_ImageCapture'},
            'SR_BK_StdDEV_R': {'SWBin': 43, 'HWBin': 2, 'Priority': 231, 'IsNan': 'SR_BK_ImageCapture'},
            'SR_BK_StdDEV_Gr': {'SWBin': 43, 'HWBin': 2, 'Priority': 232, 'IsNan': 'SR_BK_ImageCapture'},
            'SR_BK_StdDEV_Gb': {'SWBin': 43, 'HWBin': 2, 'Priority': 233, 'IsNan': 'SR_BK_ImageCapture'},
            'SR_BK_StdDEV_B': {'SWBin': 43, 'HWBin': 2, 'Priority': 234, 'IsNan': 'SR_BK_ImageCapture','IsLast': 'SR_BK_ImageCapture'},
            'MIPI2L30F_BK_ImageCapture': {'SWBin': 91, 'HWBin': 1, 'Priority': 235, 'Status': []},
            'MIPI2L30F_BK_Calculation': {'SWBin': 92, 'HWBin': 1, 'Priority': 236},
            'MIPI2L30F_BK_DeadRowExBPix_R': {'SWBin': 62, 'HWBin': 1, 'Priority': 237,
                                             'IsNan': 'MIPI2L30F_BK_ImageCapture'},
            'MIPI2L30F_BK_DeadRowExBPix_Gr': {'SWBin': 62, 'HWBin': 1, 'Priority': 238,
                                              'IsNan': 'MIPI2L30F_BK_ImageCapture'},
            'MIPI2L30F_BK_DeadRowExBPix_Gb': {'SWBin': 62, 'HWBin': 1, 'Priority': 239,
                                              'IsNan': 'MIPI2L30F_BK_ImageCapture'},
            'MIPI2L30F_BK_DeadRowExBPix_B': {'SWBin': 62, 'HWBin': 1, 'Priority': 240,
                                             'IsNan': 'MIPI2L30F_BK_ImageCapture'},
            'MIPI2L30F_BK_DeadColExBPix_R': {'SWBin': 61, 'HWBin': 1, 'Priority': 241,
                                             'IsNan': 'MIPI2L30F_BK_ImageCapture'},
            'MIPI2L30F_BK_DeadColExBPix_Gr': {'SWBin': 61, 'HWBin': 1, 'Priority': 242,
                                              'IsNan': 'MIPI2L30F_BK_ImageCapture'},
            'MIPI2L30F_BK_DeadColExBPix_Gb': {'SWBin': 61, 'HWBin': 1, 'Priority': 243,
                                              'IsNan': 'MIPI2L30F_BK_ImageCapture'},
            'MIPI2L30F_BK_DeadColExBPix_B': {'SWBin': 61, 'HWBin': 1, 'Priority': 244,
                                             'IsNan': 'MIPI2L30F_BK_ImageCapture'},
            'MIPI2L30F_BK_Mean_R': {'SWBin': 66, 'HWBin': 1, 'Priority': 245, 'IsNan': 'MIPI2L30F_BK_ImageCapture'},
            'MIPI2L30F_BK_Mean_Gr': {'SWBin': 66, 'HWBin': 1, 'Priority': 246, 'IsNan': 'MIPI2L30F_BK_ImageCapture'},
            'MIPI2L30F_BK_Mean_Gb': {'SWBin': 66, 'HWBin': 1, 'Priority': 247, 'IsNan': 'MIPI2L30F_BK_ImageCapture'},
            'MIPI2L30F_BK_Mean_B': {'SWBin': 66, 'HWBin': 1, 'Priority': 248, 'IsNan': 'MIPI2L30F_BK_ImageCapture'},
            'MIPI2L30F_BK_StdDEV_R': {'SWBin': 66, 'HWBin': 1, 'Priority': 249, 'IsNan': 'MIPI2L30F_BK_ImageCapture'},
            'MIPI2L30F_BK_StdDEV_Gr': {'SWBin': 66, 'HWBin': 1, 'Priority': 250, 'IsNan': 'MIPI2L30F_BK_ImageCapture'},
            'MIPI2L30F_BK_StdDEV_Gb': {'SWBin': 66, 'HWBin': 1, 'Priority': 251, 'IsNan': 'MIPI2L30F_BK_ImageCapture'},
            'MIPI2L30F_BK_StdDEV_B': {'SWBin': 66, 'HWBin': 1, 'Priority': 252, 'IsNan': 'MIPI2L30F_BK_ImageCapture','IsLast': 'MIPI2L30F_BK_ImageCapture'},
            'MIPI2L30F_LT_ImageCapture': {'SWBin': 90, 'HWBin': 1, 'Priority': 253, 'Status': []},
            'MIPI2L30F_LT_Calculation': {'SWBin': 92, 'HWBin': 1, 'Priority': 254},
            'MIPI2L30F_LT_DRow_R': {'SWBin': 65, 'HWBin': 1, 'Priority': 255, 'IsNan': 'MIPI2L30F_LT_ImageCapture'},
            'MIPI2L30F_LT_DRow_Gr': {'SWBin': 65, 'HWBin': 1, 'Priority': 256, 'IsNan': 'MIPI2L30F_LT_ImageCapture'},
            'MIPI2L30F_LT_DRow_Gb': {'SWBin': 65, 'HWBin': 1, 'Priority': 257, 'IsNan': 'MIPI2L30F_LT_ImageCapture'},
            'MIPI2L30F_LT_DRow_B': {'SWBin': 65, 'HWBin': 1, 'Priority': 258, 'IsNan': 'MIPI2L30F_LT_ImageCapture'},
            'MIPI2L30F_LT_DCol_R': {'SWBin': 64, 'HWBin': 1, 'Priority': 259, 'IsNan': 'MIPI2L30F_LT_ImageCapture'},
            'MIPI2L30F_LT_DCol_Gr': {'SWBin': 64, 'HWBin': 1, 'Priority': 260, 'IsNan': 'MIPI2L30F_LT_ImageCapture'},
            'MIPI2L30F_LT_DCol_Gb': {'SWBin': 64, 'HWBin': 1, 'Priority': 261, 'IsNan': 'MIPI2L30F_LT_ImageCapture'},
            'MIPI2L30F_LT_DCol_B': {'SWBin': 64, 'HWBin': 1, 'Priority': 262, 'IsNan': 'MIPI2L30F_LT_ImageCapture'},
            'MIPI2L30F_LT_DRow_Color': {'SWBin': 65, 'HWBin': 1, 'Priority': 263, 'IsNan': 'MIPI2L30F_LT_ImageCapture'},
            'MIPI2L30F_LT_DCol_Color': {'SWBin': 64, 'HWBin': 1, 'Priority': 264, 'IsNan': 'MIPI2L30F_LT_ImageCapture'},
            'MIPI2L30F_LT_Mean_R': {'SWBin': 63, 'HWBin': 1, 'Priority': 265, 'IsNan': 'MIPI2L30F_LT_ImageCapture'},
            'MIPI2L30F_LT_Mean_Gr': {'SWBin': 63, 'HWBin': 1, 'Priority': 266, 'IsNan': 'MIPI2L30F_LT_ImageCapture'},
            'MIPI2L30F_LT_Mean_Gb': {'SWBin': 63, 'HWBin': 1, 'Priority': 267, 'IsNan': 'MIPI2L30F_LT_ImageCapture'},
            'MIPI2L30F_LT_Mean_B': {'SWBin': 63, 'HWBin': 1, 'Priority': 268, 'IsNan': 'MIPI2L30F_LT_ImageCapture'},
            'MIPI2L30F_LT_StdDEV_R': {'SWBin': 63, 'HWBin': 1, 'Priority': 269, 'IsNan': 'MIPI2L30F_LT_ImageCapture'},
            'MIPI2L30F_LT_StdDEV_Gr': {'SWBin': 63, 'HWBin': 1, 'Priority': 270, 'IsNan': 'MIPI2L30F_LT_ImageCapture'},
            'MIPI2L30F_LT_StdDEV_Gb': {'SWBin': 63, 'HWBin': 1, 'Priority': 271, 'IsNan': 'MIPI2L30F_LT_ImageCapture'},
            'MIPI2L30F_LT_StdDEV_B': {'SWBin': 63, 'HWBin': 1, 'Priority': 272, 'IsNan': 'MIPI2L30F_LT_ImageCapture'},
            'MIPI2L30F_LT_RI_R': {'SWBin': 63, 'HWBin': 1, 'Priority': 273, 'IsNan': 'MIPI2L30F_LT_ImageCapture'},
            'MIPI2L30F_LT_RI_Gr': {'SWBin': 63, 'HWBin': 1, 'Priority': 274, 'IsNan': 'MIPI2L30F_LT_ImageCapture'},
            'MIPI2L30F_LT_RI_Gb': {'SWBin': 63, 'HWBin': 1, 'Priority': 275, 'IsNan': 'MIPI2L30F_LT_ImageCapture'},
            'MIPI2L30F_LT_RI_B': {'SWBin': 63, 'HWBin': 1, 'Priority': 276, 'IsNan': 'MIPI2L30F_LT_ImageCapture'},
            'MIPI2L30F_LT_Ratio_GrR': {'SWBin': 63, 'HWBin': 1, 'Priority': 277, 'IsNan': 'MIPI2L30F_LT_ImageCapture'},
            'MIPI2L30F_LT_Ratio_GbR': {'SWBin': 63, 'HWBin': 1, 'Priority': 278, 'IsNan': 'MIPI2L30F_LT_ImageCapture'},
            'MIPI2L30F_LT_Ratio_GrB': {'SWBin': 63, 'HWBin': 1, 'Priority': 279, 'IsNan': 'MIPI2L30F_LT_ImageCapture'},
            'MIPI2L30F_LT_Ratio_GbB': {'SWBin': 63, 'HWBin': 1, 'Priority': 280, 'IsNan': 'MIPI2L30F_LT_ImageCapture'},
            'MIPI2L30F_LT_Ratio_GbGr': {'SWBin': 63, 'HWBin': 1, 'Priority': 281, 'IsNan': 'MIPI2L30F_LT_ImageCapture'},
            'MIPI2L30F_LT_LostBit': {'SWBin': 63, 'HWBin': 1, 'Priority': 282, 'IsNan': 'MIPI2L30F_LT_ImageCapture'},
            'MIPI2L30F_LT_LostBitSNR_SumDIFF_R': {'SWBin': 63, 'HWBin': 1, 'Priority': 283,
                                                  'IsNan': 'MIPI2L30F_LT_ImageCapture'},
            'MIPI2L30F_LT_LostBitSNR_SumDIFF_Gr': {'SWBin': 63, 'HWBin': 1, 'Priority': 284,
                                                   'IsNan': 'MIPI2L30F_LT_ImageCapture'},
            'MIPI2L30F_LT_LostBitSNR_SumDIFF_Gb': {'SWBin': 63, 'HWBin': 1, 'Priority': 285,
                                                   'IsNan': 'MIPI2L30F_LT_ImageCapture'},
            'MIPI2L30F_LT_LostBitSNR_SumDIFF_B': {'SWBin': 63, 'HWBin': 1, 'Priority': 286,
                                                  'IsNan': 'MIPI2L30F_LT_ImageCapture','IsLast': 'MIPI2L30F_LT_ImageCapture'},
            'MIPI2L30F_LB_LT_ImageCapture': {'SWBin': 90, 'HWBin': 1, 'Priority': 287, 'Status': []},
            'MIPI2L30F_LB_LT_Calculation': {'SWBin': 92, 'HWBin': 1, 'Priority': 288},
            'MIPI2L30F_LB_LT_Mean_R': {'SWBin': 63, 'HWBin': 1, 'Priority': 289,
                                       'IsNan': 'MIPI2L30F_LB_LT_ImageCapture'},
            'MIPI2L30F_LB_LT_Mean_Gr': {'SWBin': 63, 'HWBin': 1, 'Priority': 290,
                                        'IsNan': 'MIPI2L30F_LB_LT_ImageCapture'},
            'MIPI2L30F_LB_LT_Mean_Gb': {'SWBin': 63, 'HWBin': 1, 'Priority': 291,
                                        'IsNan': 'MIPI2L30F_LB_LT_ImageCapture'},
            'MIPI2L30F_LB_LT_Mean_B': {'SWBin': 63, 'HWBin': 1, 'Priority': 292,
                                       'IsNan': 'MIPI2L30F_LB_LT_ImageCapture'},
            'MIPI2L30F_LB_LT_LostBit': {'SWBin': 63, 'HWBin': 1, 'Priority': 293,
                                        'IsNan': 'MIPI2L30F_LB_LT_ImageCapture'},
            'MIPI2L30F_LB_LT_LostBitSNR_SumDIFF_R': {'SWBin': 63, 'HWBin': 1, 'Priority': 294,
                                                     'IsNan': 'MIPI2L30F_LB_LT_ImageCapture'},
            'MIPI2L30F_LB_LT_LostBitSNR_SumDIFF_Gr': {'SWBin': 63, 'HWBin': 1, 'Priority': 295,
                                                      'IsNan': 'MIPI2L30F_LB_LT_ImageCapture'},
            'MIPI2L30F_LB_LT_LostBitSNR_SumDIFF_Gb': {'SWBin': 63, 'HWBin': 1, 'Priority': 296,
                                                      'IsNan': 'MIPI2L30F_LB_LT_ImageCapture'},
            'MIPI2L30F_LB_LT_LostBitSNR_SumDIFF_B': {'SWBin': 63, 'HWBin': 1, 'Priority': 297,
                                                     'IsNan': 'MIPI2L30F_LB_LT_ImageCapture','IsLast': 'MIPI2L30F_LB_LT_ImageCapture'},
            'MIPI2L30F_FW_LT_ImageCapture': {'SWBin': 90, 'HWBin': 1, 'Priority': 298, 'Status': []},
            'MIPI2L30F_FW_LT_Calculation': {'SWBin': 92, 'HWBin': 1, 'Priority': 299},
            'MIPI2L30F_FW_LT_DRow_R': {'SWBin': 37, 'HWBin': 1, 'Priority': 300,
                                       'IsNan': 'MIPI2L30F_FW_LT_ImageCapture'},
            'MIPI2L30F_FW_LT_DRow_Gr': {'SWBin': 37, 'HWBin': 1, 'Priority': 301,
                                        'IsNan': 'MIPI2L30F_FW_LT_ImageCapture'},
            'MIPI2L30F_FW_LT_DRow_Gb': {'SWBin': 37, 'HWBin': 1, 'Priority': 302,
                                        'IsNan': 'MIPI2L30F_FW_LT_ImageCapture'},
            'MIPI2L30F_FW_LT_DRow_B': {'SWBin': 37, 'HWBin': 1, 'Priority': 303,
                                       'IsNan': 'MIPI2L30F_FW_LT_ImageCapture'},
            'MIPI2L30F_FW_LT_DCol_R': {'SWBin': 37, 'HWBin': 1, 'Priority': 304,
                                       'IsNan': 'MIPI2L30F_FW_LT_ImageCapture'},
            'MIPI2L30F_FW_LT_DCol_Gr': {'SWBin': 37, 'HWBin': 1, 'Priority': 305,
                                        'IsNan': 'MIPI2L30F_FW_LT_ImageCapture'},
            'MIPI2L30F_FW_LT_DCol_Gb': {'SWBin': 37, 'HWBin': 1, 'Priority': 306,
                                        'IsNan': 'MIPI2L30F_FW_LT_ImageCapture'},
            'MIPI2L30F_FW_LT_DCol_B': {'SWBin': 37, 'HWBin': 1, 'Priority': 307,
                                       'IsNan': 'MIPI2L30F_FW_LT_ImageCapture'},
            'MIPI2L30F_FW_LT_DRow_Color': {'SWBin': 37, 'HWBin': 1, 'Priority': 308,
                                           'IsNan': 'MIPI2L30F_FW_LT_ImageCapture'},
            'MIPI2L30F_FW_LT_DCol_Color': {'SWBin': 37, 'HWBin': 1, 'Priority': 309,
                                           'IsNan': 'MIPI2L30F_FW_LT_ImageCapture'},
            'MIPI2L30F_FW_LT_Mean_R': {'SWBin': 37, 'HWBin': 1, 'Priority': 310,
                                       'IsNan': 'MIPI2L30F_FW_LT_ImageCapture'},
            'MIPI2L30F_FW_LT_Mean_Gr': {'SWBin': 37, 'HWBin': 1, 'Priority': 311,
                                        'IsNan': 'MIPI2L30F_FW_LT_ImageCapture'},
            'MIPI2L30F_FW_LT_Mean_Gb': {'SWBin': 37, 'HWBin': 1, 'Priority': 312,
                                        'IsNan': 'MIPI2L30F_FW_LT_ImageCapture'},
            'MIPI2L30F_FW_LT_Mean_B': {'SWBin': 37, 'HWBin': 1, 'Priority': 313,
                                       'IsNan': 'MIPI2L30F_FW_LT_ImageCapture'},
            'MIPI2L30F_FW_LT_StdDEV_R': {'SWBin': 37, 'HWBin': 1, 'Priority': 314,
                                         'IsNan': 'MIPI2L30F_FW_LT_ImageCapture'},
            'MIPI2L30F_FW_LT_StdDEV_Gr': {'SWBin': 37, 'HWBin': 1, 'Priority': 315,
                                          'IsNan': 'MIPI2L30F_FW_LT_ImageCapture'},
            'MIPI2L30F_FW_LT_StdDEV_Gb': {'SWBin': 37, 'HWBin': 1, 'Priority': 316,
                                          'IsNan': 'MIPI2L30F_FW_LT_ImageCapture'},
            'MIPI2L30F_FW_LT_StdDEV_B': {'SWBin': 37, 'HWBin': 1, 'Priority': 317,
                                         'IsNan': 'MIPI2L30F_FW_LT_ImageCapture'},
            'MIPI2L30F_FW_LT_RI_R': {'SWBin': 37, 'HWBin': 1, 'Priority': 318, 'IsNan': 'MIPI2L30F_FW_LT_ImageCapture'},
            'MIPI2L30F_FW_LT_RI_Gr': {'SWBin': 37, 'HWBin': 1, 'Priority': 319,
                                      'IsNan': 'MIPI2L30F_FW_LT_ImageCapture'},
            'MIPI2L30F_FW_LT_RI_Gb': {'SWBin': 37, 'HWBin': 1, 'Priority': 320,
                                      'IsNan': 'MIPI2L30F_FW_LT_ImageCapture'},
            'MIPI2L30F_FW_LT_RI_B': {'SWBin': 37, 'HWBin': 1, 'Priority': 321, 'IsNan': 'MIPI2L30F_FW_LT_ImageCapture'},
            'MIPI2L30F_FW_LT_Ratio_GrR': {'SWBin': 37, 'HWBin': 1, 'Priority': 322,
                                          'IsNan': 'MIPI2L30F_FW_LT_ImageCapture'},
            'MIPI2L30F_FW_LT_Ratio_GbR': {'SWBin': 37, 'HWBin': 1, 'Priority': 323,
                                          'IsNan': 'MIPI2L30F_FW_LT_ImageCapture'},
            'MIPI2L30F_FW_LT_Ratio_GrB': {'SWBin': 37, 'HWBin': 1, 'Priority': 324,
                                          'IsNan': 'MIPI2L30F_FW_LT_ImageCapture'},
            'MIPI2L30F_FW_LT_Ratio_GbB': {'SWBin': 37, 'HWBin': 1, 'Priority': 325,
                                          'IsNan': 'MIPI2L30F_FW_LT_ImageCapture'},
            'MIPI2L30F_FW_LT_Ratio_GbGr': {'SWBin': 37, 'HWBin': 1, 'Priority': 326,
                                           'IsNan': 'MIPI2L30F_FW_LT_ImageCapture','IsLast': 'MIPI2L30F_FW_LT_ImageCapture'},
            'MIPI2L30F_BK_Cluster2': {'SWBin': 38, 'HWBin': 1, 'Priority': 327},
            'MIPI2L30F_LT_Cluster2': {'SWBin': 38, 'HWBin': 1, 'Priority': 328},
            'MIPI2L30F_BK_Cluster1': {'SWBin': 38, 'HWBin': 1, 'Priority': 329},
            'MIPI2L30F_LT_Cluster1': {'SWBin': 38, 'HWBin': 1, 'Priority': 330},
            'Binning': {'SWBin': 2, 'HWBin': 3, 'Priority': 331},
            'All Pass': {'SWBin': 1, 'HWBin': 3, 'Priority': 332}
        },
    'JX828':
        {
            'VSYNC_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 1},
            'HSYNC_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 2},
            'PCLK_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 3},
            'EXCLK_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 4},
            'RSTB_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 5},
            'PWDN_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 6},
            'SDA_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 7},
            'SCL_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 8},
            'D0_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 9},
            'D1_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 10},
            'D2_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 11},
            'D3_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 12},
            'D4_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 13},
            'D5_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 14},
            'D6_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 15},
            'D7_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 16},
            'D8_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 17},
            'D9_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 18},
            'MCP_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 19},
            'MCN_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 20},
            'MDP0_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 21},
            'MDN0_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 22},
            'MDP1_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 23},
            'MDN1_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 24},
            'AVDD_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 25, 'Limit': {'L': -2, 'H': 2}},
            'DOVDD_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 26},
            'VRamp_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 27},
            'VH_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 28},
            'VN1_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 29},
            'VSYNC_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 30},
            'HSYNC_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 31},
            'PCLK_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 32},
            'EXCLK_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 33},
            'RSTB_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 34},
            'PWDN_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 35},
            'SDA_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 36},
            'SCL_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 37},
            'D0_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 38},
            'D1_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 39},
            'D2_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 40},
            'D4_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 41},
            'D5_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 42},
            'D3_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 43},
            'D6_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 44},
            'D7_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 45},
            'D8_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 46},
            'D9_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 47},
            'MCP_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 48},
            'MCN_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 49},
            'MDP0_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 50},
            'MDN0_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 51},
            'MDP1_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 52},
            'MDN1_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 53},
            'VSYNC_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 54},
            'HSYNC_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 55},
            'PCLK_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 56},
            'EXCLK_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 57},
            'RSTB_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 58},
            'PWDN_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 59},
            'SDA_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 60},
            'SCL_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 61},
            'D0_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 62},
            'D1_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 63},
            'D2_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 64},
            'D4_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 65},
            'D5_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 66},
            'D3_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 67},
            'D6_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 68},
            'D7_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 69},
            'D8_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 70},
            'D9_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 71},
            'MCP_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 72},
            'MCN_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 73},
            'MDP0_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 74},
            'MDN0_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 75},
            'MDP1_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 76},
            'MDN1_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 77},
            'iic_test': {'SWBin': 7, 'HWBin': 5, 'Priority': 78},
            'DVDD_voltage': {'SWBin': 9, 'HWBin': 5, 'Priority': 79},
            'VH_voltage': {'SWBin': 9, 'HWBin': 5, 'Priority': 80},
            'VN1_voltage': {'SWBin': 9, 'HWBin': 5, 'Priority': 81},
            'Active_AVDD': {'SWBin': 12, 'HWBin': 5, 'Priority': 82},
            'Active_DOVDD': {'SWBin': 12, 'HWBin': 5, 'Priority': 83},
            'PWDN_AVDD': {'SWBin': 8, 'HWBin': 5, 'Priority': 84},
            'PWDN_DOVDD': {'SWBin': 8, 'HWBin': 5, 'Priority': 85},
            'PWDN_Total': {'SWBin': 8, 'HWBin': 5, 'Priority': 86},
            'DVP27M30F_BK_Cap': {'SWBin': 98, 'HWBin': 5, 'Priority': 87, 'Status': []},
            'DVP27M30F_BK_Calc': {'SWBin': 97, 'HWBin': 5, 'Priority': 88},
            'DVP27M30F_BK_REG_VRamp': {'SWBin': 7, 'HWBin': 5, 'Priority': 89},
            'BLC_R': {'SWBin': 7, 'HWBin': 5, 'Priority': 90},
            'BLC_Gr': {'SWBin': 7, 'HWBin': 5, 'Priority': 91},
            'BLC_Gb': {'SWBin': 7, 'HWBin': 5, 'Priority': 92},
            'BLC_B': {'SWBin': 7, 'HWBin': 5, 'Priority': 93},
            'BK_DeadRowExBPix_R': {'SWBin': 15, 'HWBin': 6, 'Priority': 94, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_DeadRowExBPix_Gr': {'SWBin': 15, 'HWBin': 6, 'Priority': 95, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_DeadRowExBPix_Gb': {'SWBin': 15, 'HWBin': 6, 'Priority': 96, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_DeadRowExBPix_B': {'SWBin': 15, 'HWBin': 6, 'Priority': 97, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_DeadColExBPix_R': {'SWBin': 14, 'HWBin': 6, 'Priority': 98, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_DeadColExBPix_Gr': {'SWBin': 14, 'HWBin': 6, 'Priority': 99, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_DeadColExBPix_Gb': {'SWBin': 14, 'HWBin': 6, 'Priority': 100, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_DeadColExBPix_B': {'SWBin': 14, 'HWBin': 6, 'Priority': 101, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VfpnQty_R': {'SWBin': 14, 'HWBin': 6, 'Priority': 102, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VfpnQty_Gr': {'SWBin': 14, 'HWBin': 6, 'Priority': 103, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VfpnQty_Gb': {'SWBin': 14, 'HWBin': 6, 'Priority': 104, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VfpnQty_B': {'SWBin': 14, 'HWBin': 6, 'Priority': 105, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue0_R': {'SWBin': 14, 'HWBin': 6, 'Priority': 106, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue0_Gr': {'SWBin': 14, 'HWBin': 6, 'Priority': 107, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue0_Gb': {'SWBin': 14, 'HWBin': 6, 'Priority': 108, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue0_B': {'SWBin': 14, 'HWBin': 6, 'Priority': 109, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue1_R': {'SWBin': 14, 'HWBin': 6, 'Priority': 110, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue1_Gr': {'SWBin': 14, 'HWBin': 6, 'Priority': 111, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue1_Gb': {'SWBin': 14, 'HWBin': 6, 'Priority': 112, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue1_B': {'SWBin': 14, 'HWBin': 6, 'Priority': 113, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue2_R': {'SWBin': 14, 'HWBin': 6, 'Priority': 114, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue2_Gr': {'SWBin': 14, 'HWBin': 6, 'Priority': 115, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue2_Gb': {'SWBin': 14, 'HWBin': 6, 'Priority': 116, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue2_B': {'SWBin': 14, 'HWBin': 6, 'Priority': 117, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue3_R': {'SWBin': 14, 'HWBin': 6, 'Priority': 118, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue3_Gr': {'SWBin': 14, 'HWBin': 6, 'Priority': 119, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue3_Gb': {'SWBin': 14, 'HWBin': 6, 'Priority': 120, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue3_B': {'SWBin': 14, 'HWBin': 6, 'Priority': 121, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue4_R': {'SWBin': 14, 'HWBin': 6, 'Priority': 122, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue4_Gr': {'SWBin': 14, 'HWBin': 6, 'Priority': 123, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue4_Gb': {'SWBin': 14, 'HWBin': 6, 'Priority': 124, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue4_B': {'SWBin': 14, 'HWBin': 6, 'Priority': 125, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_Mean_R': {'SWBin': 13, 'HWBin': 6, 'Priority': 126, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_Mean_Gr': {'SWBin': 13, 'HWBin': 6, 'Priority': 127, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_Mean_Gb': {'SWBin': 13, 'HWBin': 6, 'Priority': 128, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_Mean_B': {'SWBin': 13, 'HWBin': 6, 'Priority': 129, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_StdDEV_R': {'SWBin': 13, 'HWBin': 6, 'Priority': 130, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_StdDEV_Gr': {'SWBin': 13, 'HWBin': 6, 'Priority': 131, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_StdDEV_Gb': {'SWBin': 13, 'HWBin': 6, 'Priority': 132, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_StdDEV_B': {'SWBin': 13, 'HWBin': 6, 'Priority': 133, 'IsNan': 'DVP27M30F_BK_Cap','IsLast': 'DVP27M30F_BK_Cap'},
            'DVP27M30F_LT_Cap': {'SWBin': 99, 'HWBin': 5, 'Priority': 134, 'Status': []},
            'DVP27M30F_LT_Calc': {'SWBin': 97, 'HWBin': 5, 'Priority': 135},
            'PLCK_Freq': {'SWBin': 93, 'HWBin': 5, 'Priority': 136},
            'LT_DRow_R': {'SWBin': 25, 'HWBin': 4, 'Priority': 137, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_DRow_Gr': {'SWBin': 25, 'HWBin': 4, 'Priority': 138, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_DRow_Gb': {'SWBin': 25, 'HWBin': 4, 'Priority': 139, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_DRow_B': {'SWBin': 25, 'HWBin': 4, 'Priority': 140, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_DCol_R': {'SWBin': 24, 'HWBin': 4, 'Priority': 141, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_DCol_Gr': {'SWBin': 24, 'HWBin': 4, 'Priority': 142, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_DCol_Gb': {'SWBin': 24, 'HWBin': 4, 'Priority': 143, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_DCol_B': {'SWBin': 24, 'HWBin': 4, 'Priority': 144, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_DRow_Color': {'SWBin': 25, 'HWBin': 4, 'Priority': 145, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_DCol_Color': {'SWBin': 24, 'HWBin': 4, 'Priority': 146, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_WeakLineRow_R': {'SWBin': 25, 'HWBin': 4, 'Priority': 147, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_WeakLineRow_Gr': {'SWBin': 25, 'HWBin': 4, 'Priority': 148, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_WeakLineRow_Gb': {'SWBin': 25, 'HWBin': 4, 'Priority': 149, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_WeakLineRow_B': {'SWBin': 25, 'HWBin': 4, 'Priority': 150, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_WeakLineCol_R': {'SWBin': 24, 'HWBin': 4, 'Priority': 151, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_WeakLineCol_Gr': {'SWBin': 24, 'HWBin': 4, 'Priority': 152, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_WeakLineCol_Gb': {'SWBin': 24, 'HWBin': 4, 'Priority': 153, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_WeakLineCol_B': {'SWBin': 24, 'HWBin': 4, 'Priority': 154, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_Mean_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 155, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_Mean_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 156, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_Mean_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 157, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_Mean_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 158, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_StdDEV_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 159, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_StdDEV_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 160, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_StdDEV_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 161, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_StdDEV_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 162, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_RI_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 163, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_RI_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 164, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_RI_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 165, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_RI_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 166, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_Ratio_GrR': {'SWBin': 23, 'HWBin': 4, 'Priority': 167, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_Ratio_GbR': {'SWBin': 23, 'HWBin': 4, 'Priority': 168, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_Ratio_GrB': {'SWBin': 23, 'HWBin': 4, 'Priority': 169, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_Ratio_GbB': {'SWBin': 23, 'HWBin': 4, 'Priority': 170, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_Ratio_GbGr': {'SWBin': 23, 'HWBin': 4, 'Priority': 171, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_LostBit': {'SWBin': 23, 'HWBin': 4, 'Priority': 172, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_LostBitSNR_SumDIFF_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 173, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_LostBitSNR_SumDIFF_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 174, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_LostBitSNR_SumDIFF_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 175, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_LostBitSNR_SumDIFF_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 176, 'IsNan': 'DVP27M30F_LT_Cap','IsLast': 'DVP27M30F_LT_Cap'},
            'DVP27M30F_LBIT_Cap': {'SWBin': 99, 'HWBin': 5, 'Priority': 177, 'Status': []},
            'DVP27M30F_LBIT_Calc': {'SWBin': 97, 'HWBin': 5, 'Priority': 178},
            'LBIT_LT_Mean_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 179, 'IsNan': 'DVP27M30F_LBIT_Cap'},
            'LBIT_LT_Mean_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 180, 'IsNan': 'DVP27M30F_LBIT_Cap'},
            'LBIT_LT_Mean_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 181, 'IsNan': 'DVP27M30F_LBIT_Cap'},
            'LBIT_LT_Mean_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 182, 'IsNan': 'DVP27M30F_LBIT_Cap'},
            'LBIT_LT_LostBit': {'SWBin': 23, 'HWBin': 4, 'Priority': 183, 'IsNan': 'DVP27M30F_LBIT_Cap'},
            'LBIT_LT_LostBitSNR_SumDIFF_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 184, 'IsNan': 'DVP27M30F_LBIT_Cap'},
            'LBIT_LT_LostBitSNR_SumDIFF_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 185, 'IsNan': 'DVP27M30F_LBIT_Cap'},
            'LBIT_LT_LostBitSNR_SumDIFF_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 186, 'IsNan': 'DVP27M30F_LBIT_Cap'},
            'LBIT_LT_LostBitSNR_SumDIFF_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 187, 'IsNan': 'DVP27M30F_LBIT_Cap','IsLast': 'DVP27M30F_LBIT_Cap'},
            'DVP27M30F_FW_Cap': {'SWBin': 99, 'HWBin': 5, 'Priority': 188, 'Status': []},
            'DVP27M30F_FW_Calc': {'SWBin': 97, 'HWBin': 5, 'Priority': 189},
            'FW_LT_DRow_R': {'SWBin': 36, 'HWBin': 4, 'Priority': 190, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_DRow_Gr': {'SWBin': 36, 'HWBin': 4, 'Priority': 191, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_DRow_Gb': {'SWBin': 36, 'HWBin': 4, 'Priority': 192, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_DRow_B': {'SWBin': 36, 'HWBin': 4, 'Priority': 193, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_DCol_R': {'SWBin': 36, 'HWBin': 4, 'Priority': 194, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_DCol_Gr': {'SWBin': 36, 'HWBin': 4, 'Priority': 195, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_DCol_Gb': {'SWBin': 36, 'HWBin': 4, 'Priority': 196, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_DCol_B': {'SWBin': 36, 'HWBin': 4, 'Priority': 197, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_DRow_Color': {'SWBin': 36, 'HWBin': 4, 'Priority': 198, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_DCol_Color': {'SWBin': 36, 'HWBin': 4, 'Priority': 199, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_Mean_R': {'SWBin': 36, 'HWBin': 4, 'Priority': 200, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_Mean_Gr': {'SWBin': 36, 'HWBin': 4, 'Priority': 201, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_Mean_Gb': {'SWBin': 36, 'HWBin': 4, 'Priority': 202, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_Mean_B': {'SWBin': 36, 'HWBin': 4, 'Priority': 203, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_StdDEV_R': {'SWBin': 36, 'HWBin': 4, 'Priority': 204, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_StdDEV_Gr': {'SWBin': 36, 'HWBin': 4, 'Priority': 205, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_StdDEV_Gb': {'SWBin': 36, 'HWBin': 4, 'Priority': 206, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_StdDEV_B': {'SWBin': 36, 'HWBin': 4, 'Priority': 207, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_RI_R': {'SWBin': 36, 'HWBin': 4, 'Priority': 208, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_RI_Gr': {'SWBin': 36, 'HWBin': 4, 'Priority': 209, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_RI_Gb': {'SWBin': 36, 'HWBin': 4, 'Priority': 210, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_RI_B': {'SWBin': 36, 'HWBin': 4, 'Priority': 211, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_Ratio_GrR': {'SWBin': 36, 'HWBin': 4, 'Priority': 212, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_Ratio_GbR': {'SWBin': 36, 'HWBin': 4, 'Priority': 213, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_Ratio_GrB': {'SWBin': 36, 'HWBin': 4, 'Priority': 214, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_Ratio_GbB': {'SWBin': 36, 'HWBin': 4, 'Priority': 215, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_Ratio_GbGr': {'SWBin': 36, 'HWBin': 4, 'Priority': 216, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_LostBit': {'SWBin': 36, 'HWBin': 4, 'Priority': 217, 'IsNan': 'DVP27M30F_FW_Cap','IsLast': 'DVP27M30F_FW_Cap'},
            'DVP27M30F_BSun_Cap': {'SWBin': 96, 'HWBin': 5, 'Priority': 218, 'Status': []},
            'DVP27M30F_BSun_Calc': {'SWBin': 97, 'HWBin': 5, 'Priority': 219},
            'BSun_BK_DeadRowExBPix_R': {'SWBin': 48, 'HWBin': 6, 'Priority': 220, 'IsNan': 'DVP27M30F_BSun_Cap'},
            'BSun_BK_DeadRowExBPix_Gr': {'SWBin': 48, 'HWBin': 6, 'Priority': 221, 'IsNan': 'DVP27M30F_BSun_Cap'},
            'BSun_BK_DeadRowExBPix_Gb': {'SWBin': 48, 'HWBin': 6, 'Priority': 222, 'IsNan': 'DVP27M30F_BSun_Cap'},
            'BSun_BK_DeadRowExBPix_B': {'SWBin': 48, 'HWBin': 6, 'Priority': 223, 'IsNan': 'DVP27M30F_BSun_Cap'},
            'BSun_BK_DeadColExBPix_R': {'SWBin': 47, 'HWBin': 6, 'Priority': 224, 'IsNan': 'DVP27M30F_BSun_Cap'},
            'BSun_BK_DeadColExBPix_Gr': {'SWBin': 47, 'HWBin': 6, 'Priority': 225, 'IsNan': 'DVP27M30F_BSun_Cap'},
            'BSun_BK_DeadColExBPix_Gb': {'SWBin': 47, 'HWBin': 6, 'Priority': 226, 'IsNan': 'DVP27M30F_BSun_Cap'},
            'BSun_BK_DeadColExBPix_B': {'SWBin': 47, 'HWBin': 6, 'Priority': 227, 'IsNan': 'DVP27M30F_BSun_Cap'},
            'BSun_BK_Mean_R': {'SWBin': 46, 'HWBin': 6, 'Priority': 228, 'IsNan': 'DVP27M30F_BSun_Cap'},
            'BSun_BK_Mean_Gr': {'SWBin': 46, 'HWBin': 6, 'Priority': 229, 'IsNan': 'DVP27M30F_BSun_Cap'},
            'BSun_BK_Mean_Gb': {'SWBin': 46, 'HWBin': 6, 'Priority': 230, 'IsNan': 'DVP27M30F_BSun_Cap'},
            'BSun_BK_Mean_B': {'SWBin': 46, 'HWBin': 6, 'Priority': 231, 'IsNan': 'DVP27M30F_BSun_Cap'},
            'BSun_BK_StdDEV_R': {'SWBin': 46, 'HWBin': 6, 'Priority': 232, 'IsNan': 'DVP27M30F_BSun_Cap'},
            'BSun_BK_StdDEV_Gr': {'SWBin': 46, 'HWBin': 6, 'Priority': 233, 'IsNan': 'DVP27M30F_BSun_Cap'},
            'BSun_BK_StdDEV_Gb': {'SWBin': 46, 'HWBin': 6, 'Priority': 234, 'IsNan': 'DVP27M30F_BSun_Cap'},
            'BSun_BK_StdDEV_B': {'SWBin': 46, 'HWBin': 6, 'Priority': 235, 'IsNan': 'DVP27M30F_BSun_Cap','IsLast': 'DVP27M30F_BSun_Cap'},
            'DVP27M30F_LMFlip_Cap': {'SWBin': 96, 'HWBin': 5, 'Priority': 236, 'Status': []},
            'DVP27M30F_LMFlip_Calc': {'SWBin': 97, 'HWBin': 5, 'Priority': 237},
            'LMFlip_LT_DRow_R': {'SWBin': 58, 'HWBin': 4, 'Priority': 238, 'IsNan': 'DVP27M30F_LMFlip_Cap'},
            'LMFlip_LT_DRow_Gr': {'SWBin': 58, 'HWBin': 4, 'Priority': 239, 'IsNan': 'DVP27M30F_LMFlip_Cap'},
            'LMFlip_LT_DRow_Gb': {'SWBin': 58, 'HWBin': 4, 'Priority': 240, 'IsNan': 'DVP27M30F_LMFlip_Cap'},
            'LMFlip_LT_DRow_B': {'SWBin': 58, 'HWBin': 4, 'Priority': 241, 'IsNan': 'DVP27M30F_LMFlip_Cap'},
            'LMFlip_LT_DCol_R': {'SWBin': 57, 'HWBin': 4, 'Priority': 242, 'IsNan': 'DVP27M30F_LMFlip_Cap'},
            'LMFlip_LT_DCol_Gr': {'SWBin': 57, 'HWBin': 4, 'Priority': 243, 'IsNan': 'DVP27M30F_LMFlip_Cap'},
            'LMFlip_LT_DCol_Gb': {'SWBin': 57, 'HWBin': 4, 'Priority': 244, 'IsNan': 'DVP27M30F_LMFlip_Cap'},
            'LMFlip_LT_DCol_B': {'SWBin': 57, 'HWBin': 4, 'Priority': 245, 'IsNan': 'DVP27M30F_LMFlip_Cap'},
            'LMFlip_LT_DRow_Color': {'SWBin': 58, 'HWBin': 4, 'Priority': 246, 'IsNan': 'DVP27M30F_LMFlip_Cap'},
            'LMFlip_LT_DCol_Color': {'SWBin': 57, 'HWBin': 4, 'Priority': 247, 'IsNan': 'DVP27M30F_LMFlip_Cap'},
            'LMFlip_LT_WeakLineRow_R': {'SWBin': 58, 'HWBin': 4, 'Priority': 248, 'IsNan': 'DVP27M30F_LMFlip_Cap'},
            'LMFlip_LT_WeakLineRow_Gr': {'SWBin': 58, 'HWBin': 4, 'Priority': 249, 'IsNan': 'DVP27M30F_LMFlip_Cap'},
            'LMFlip_LT_WeakLineRow_Gb': {'SWBin': 58, 'HWBin': 4, 'Priority': 250, 'IsNan': 'DVP27M30F_LMFlip_Cap'},
            'LMFlip_LT_WeakLineRow_B': {'SWBin': 58, 'HWBin': 4, 'Priority': 251, 'IsNan': 'DVP27M30F_LMFlip_Cap'},
            'LMFlip_LT_WeakLineCol_R': {'SWBin': 57, 'HWBin': 4, 'Priority': 252, 'IsNan': 'DVP27M30F_LMFlip_Cap'},
            'LMFlip_LT_WeakLineCol_Gr': {'SWBin': 57, 'HWBin': 4, 'Priority': 253, 'IsNan': 'DVP27M30F_LMFlip_Cap'},
            'LMFlip_LT_WeakLineCol_Gb': {'SWBin': 57, 'HWBin': 4, 'Priority': 254, 'IsNan': 'DVP27M30F_LMFlip_Cap'},
            'LMFlip_LT_WeakLineCol_B': {'SWBin': 57, 'HWBin': 4, 'Priority': 255, 'IsNan': 'DVP27M30F_LMFlip_Cap'},
            'LMFlip_LT_Mean_R': {'SWBin': 56, 'HWBin': 4, 'Priority': 256, 'IsNan': 'DVP27M30F_LMFlip_Cap'},
            'LMFlip_LT_Mean_Gr': {'SWBin': 56, 'HWBin': 4, 'Priority': 257, 'IsNan': 'DVP27M30F_LMFlip_Cap'},
            'LMFlip_LT_Mean_Gb': {'SWBin': 56, 'HWBin': 4, 'Priority': 258, 'IsNan': 'DVP27M30F_LMFlip_Cap'},
            'LMFlip_LT_Mean_B': {'SWBin': 56, 'HWBin': 4, 'Priority': 259, 'IsNan': 'DVP27M30F_LMFlip_Cap'},
            'LMFlip_LT_StdDEV_R': {'SWBin': 56, 'HWBin': 4, 'Priority': 260, 'IsNan': 'DVP27M30F_LMFlip_Cap'},
            'LMFlip_LT_StdDEV_Gr': {'SWBin': 56, 'HWBin': 4, 'Priority': 261, 'IsNan': 'DVP27M30F_LMFlip_Cap'},
            'LMFlip_LT_StdDEV_Gb': {'SWBin': 56, 'HWBin': 4, 'Priority': 262, 'IsNan': 'DVP27M30F_LMFlip_Cap'},
            'LMFlip_LT_StdDEV_B': {'SWBin': 56, 'HWBin': 4, 'Priority': 263, 'IsNan': 'DVP27M30F_LMFlip_Cap'},
            'LMFlip_LT_RI_R': {'SWBin': 56, 'HWBin': 4, 'Priority': 264, 'IsNan': 'DVP27M30F_LMFlip_Cap'},
            'LMFlip_LT_RI_Gr': {'SWBin': 56, 'HWBin': 4, 'Priority': 265, 'IsNan': 'DVP27M30F_LMFlip_Cap'},
            'LMFlip_LT_RI_Gb': {'SWBin': 56, 'HWBin': 4, 'Priority': 266, 'IsNan': 'DVP27M30F_LMFlip_Cap'},
            'LMFlip_LT_RI_B': {'SWBin': 56, 'HWBin': 4, 'Priority': 267, 'IsNan': 'DVP27M30F_LMFlip_Cap'},
            'LMFlip_LT_Ratio_GrR': {'SWBin': 56, 'HWBin': 4, 'Priority': 268, 'IsNan': 'DVP27M30F_LMFlip_Cap'},
            'LMFlip_LT_Ratio_GbR': {'SWBin': 56, 'HWBin': 4, 'Priority': 269, 'IsNan': 'DVP27M30F_LMFlip_Cap'},
            'LMFlip_LT_Ratio_GrB': {'SWBin': 56, 'HWBin': 4, 'Priority': 270, 'IsNan': 'DVP27M30F_LMFlip_Cap'},
            'LMFlip_LT_Ratio_GbB': {'SWBin': 56, 'HWBin': 4, 'Priority': 271, 'IsNan': 'DVP27M30F_LMFlip_Cap'},
            'LMFlip_LT_Ratio_GbGr': {'SWBin': 56, 'HWBin': 4, 'Priority': 272, 'IsNan': 'DVP27M30F_LMFlip_Cap'},
            'LMFlip_LT_LostBit': {'SWBin': 56, 'HWBin': 4, 'Priority': 273, 'IsNan': 'DVP27M30F_LMFlip_Cap'},
            'LMFlip_LT_LostBitSNR_SumDIFF_R': {'SWBin': 56, 'HWBin': 4, 'Priority': 274,
                                               'IsNan': 'DVP27M30F_LMFlip_Cap'},
            'LMFlip_LT_LostBitSNR_SumDIFF_Gr': {'SWBin': 56, 'HWBin': 4, 'Priority': 275,
                                                'IsNan': 'DVP27M30F_LMFlip_Cap'},
            'LMFlip_LT_LostBitSNR_SumDIFF_Gb': {'SWBin': 56, 'HWBin': 4, 'Priority': 276,
                                                'IsNan': 'DVP27M30F_LMFlip_Cap'},
            'LMFlip_LT_LostBitSNR_SumDIFF_B': {'SWBin': 56, 'HWBin': 4, 'Priority': 277,
                                               'IsNan': 'DVP27M30F_LMFlip_Cap','IsLast': 'DVP27M30F_LMFlip_Cap'},
            'LT_CornerLine': {'SWBin': 26, 'HWBin': 4, 'Priority': 278},
            'LT_ScratchLine': {'SWBin': 26, 'HWBin': 4, 'Priority': 279},
            'LT_Blemish': {'SWBin': 31, 'HWBin': 4, 'Priority': 280},
            'LT_LineStripe': {'SWBin': 27, 'HWBin': 4, 'Priority': 281},
            'LT_Particle': {'SWBin': 32, 'HWBin': 4, 'Priority': 282},
            'BK_Cluster2': {'SWBin': 30, 'HWBin': 4, 'Priority': 283},  # 4/8
            'LT_Cluster2': {'SWBin': 30, 'HWBin': 4, 'Priority': 284},  # 4/8
            'BK_Cluster1': {'SWBin': 29, 'HWBin': 4, 'Priority': 285},  # 4/8
            'LT_Cluster1': {'SWBin': 29, 'HWBin': 4, 'Priority': 286},  # 4/8
            'BK_Cluster3GrGb': {'SWBin': 39, 'HWBin': 4, 'Priority': 287},  # 4/8
            'LT_Cluster3GrGb': {'SWBin': 40, 'HWBin': 4, 'Priority': 288},  # 4/8
            'BK_Cluster3SubtractGrGb': {'SWBin': 33, 'HWBin': 4, 'Priority': 289},  # 4/8
            'LT_Cluster3SubtractGrGb': {'SWBin': 34, 'HWBin': 4, 'Priority': 290},  # 4/8
            'WP_Count': {'SWBin': 35, 'HWBin': 4, 'Priority': 291},  # 6/8
            'BK_Z1_BT_DP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 292},
            'BK_Z1_BT_DP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 293},
            'BK_Z1_BT_DP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 294},
            'BK_Z1_BT_DP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 295},
            'BK_Z2_BT_DP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 296},
            'BK_Z2_BT_DP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 297},
            'BK_Z2_BT_DP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 298},
            'BK_Z2_BT_DP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 299},
            'BK_Z1_BT_WP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 300},
            'BK_Z1_BT_WP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 301},
            'BK_Z1_BT_WP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 302},
            'BK_Z1_BT_WP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 303},
            'BK_Z2_BT_WP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 304},
            'BK_Z2_BT_WP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 305},
            'BK_Z2_BT_WP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 306},
            'BK_Z2_BT_WP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 307},
            'LT_Z1_BT_DP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 308},
            'LT_Z1_BT_DP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 309},
            'LT_Z1_BT_DP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 310},
            'LT_Z1_BT_DP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 311},
            'LT_Z2_BT_DP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 312},
            'LT_Z2_BT_DP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 313},
            'LT_Z2_BT_DP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 314},
            'LT_Z2_BT_DP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 315},
            'LT_Z1_BT_WP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 316},
            'LT_Z1_BT_WP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 317},
            'LT_Z1_BT_WP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 318},
            'LT_Z1_BT_WP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 319},
            'LT_Z2_BT_WP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 320},
            'LT_Z2_BT_WP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 321},
            'LT_Z2_BT_WP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 322},
            'LT_Z2_BT_WP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 323},
            'LT_Z1_DK_WP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 324},
            'LT_Z1_DK_WP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 325},
            'LT_Z1_DK_WP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 326},
            'LT_Z1_DK_WP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 327},
            'LT_Z2_DK_WP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 328},
            'LT_Z2_DK_WP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 329},
            'LT_Z2_DK_WP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 330},
            'LT_Z2_DK_WP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 331},
            'LT_Z1_DK_DP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 332},
            'LT_Z1_DK_DP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 333},
            'LT_Z1_DK_DP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 334},
            'LT_Z1_DK_DP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 335},
            'LT_Z2_DK_DP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 336},
            'LT_Z2_DK_DP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 337},
            'LT_Z2_DK_DP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 338},
            'LT_Z2_DK_DP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 339},
            'BK_Z1_BT_DP': {'SWBin': 60, 'HWBin': 6, 'Priority': 340},
            'BK_Z2_BT_DP': {'SWBin': 60, 'HWBin': 6, 'Priority': 341},
            'BK_Z1_BT_WP': {'SWBin': 60, 'HWBin': 6, 'Priority': 342},
            'BK_Z2_BT_WP': {'SWBin': 60, 'HWBin': 6, 'Priority': 343},
            'LT_Z1_BT_DP': {'SWBin': 60, 'HWBin': 6, 'Priority': 344},
            'LT_Z2_BT_DP': {'SWBin': 60, 'HWBin': 6, 'Priority': 345},
            'LT_Z1_BT_WP': {'SWBin': 60, 'HWBin': 6, 'Priority': 346},
            'LT_Z2_BT_WP': {'SWBin': 60, 'HWBin': 6, 'Priority': 347},
            'LT_Z1_DK_DP': {'SWBin': 60, 'HWBin': 6, 'Priority': 348},
            'LT_Z2_DK_DP': {'SWBin': 60, 'HWBin': 6, 'Priority': 349},
            'LT_Z1_DK_WP': {'SWBin': 60, 'HWBin': 6, 'Priority': 350},
            'LT_Z2_DK_WP': {'SWBin': 60, 'HWBin': 6, 'Priority': 351},
            'DVP27M30F_BK32X_Cap': {'SWBin': 98, 'HWBin': 5, 'Priority': 352, 'Status': []},
            'DVP27M30F_BK32X_Calc': {'SWBin': 97, 'HWBin': 5, 'Priority': 353},
            'BK32X_BK_BadPixel_Area': {'SWBin': 75, 'HWBin': 4, 'Priority': 354, 'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DashQty_R': {'SWBin': 51, 'HWBin': 6, 'Priority': 355, 'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DashQty_Gr': {'SWBin': 51, 'HWBin': 6, 'Priority': 356, 'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DashQty_Gb': {'SWBin': 51, 'HWBin': 6, 'Priority': 357, 'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DashQty_B': {'SWBin': 51, 'HWBin': 6, 'Priority': 358, 'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_SP_BK_MixDCol_DashQty_R': {'SWBin': 54, 'HWBin': 2, 'Priority': 359, 'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_SP_BK_MixDCol_DashQty_Gr': {'SWBin': 54, 'HWBin': 2, 'Priority': 360,
                                               'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_SP_BK_MixDCol_DashQty_Gb': {'SWBin': 54, 'HWBin': 2, 'Priority': 361,
                                               'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_SP_BK_MixDCol_DashQty_B': {'SWBin': 54, 'HWBin': 2, 'Priority': 362, 'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_DeltaAvg_R': {'SWBin': 54, 'HWBin': 2, 'Priority': 363,
                                                 'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_DeltaAvg_Gr': {'SWBin': 54, 'HWBin': 2, 'Priority': 364,
                                                  'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_DeltaAvg_Gb': {'SWBin': 54, 'HWBin': 2, 'Priority': 365,
                                                  'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_DeltaAvg_B': {'SWBin': 54, 'HWBin': 2, 'Priority': 366,
                                                 'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue0_R': {'SWBin': 54, 'HWBin': 2, 'Priority': 367,
                                                  'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue0_Gr': {'SWBin': 54, 'HWBin': 2, 'Priority': 368,
                                                   'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue0_Gb': {'SWBin': 54, 'HWBin': 2, 'Priority': 369,
                                                   'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue0_B': {'SWBin': 54, 'HWBin': 2, 'Priority': 370,
                                                  'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue1_R': {'SWBin': 54, 'HWBin': 2, 'Priority': 371,
                                                  'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue1_Gr': {'SWBin': 54, 'HWBin': 2, 'Priority': 372,
                                                   'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue1_Gb': {'SWBin': 54, 'HWBin': 2, 'Priority': 373,
                                                   'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue1_B': {'SWBin': 54, 'HWBin': 2, 'Priority': 374,
                                                  'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue2_R': {'SWBin': 54, 'HWBin': 2, 'Priority': 375,
                                                  'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue2_Gr': {'SWBin': 54, 'HWBin': 2, 'Priority': 376,
                                                   'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue2_Gb': {'SWBin': 54, 'HWBin': 2, 'Priority': 377,
                                                   'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue2_B': {'SWBin': 54, 'HWBin': 2, 'Priority': 378,
                                                  'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue3_R': {'SWBin': 54, 'HWBin': 2, 'Priority': 379,
                                                  'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue3_Gr': {'SWBin': 54, 'HWBin': 2, 'Priority': 380,
                                                   'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue3_Gb': {'SWBin': 54, 'HWBin': 2, 'Priority': 381,
                                                   'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue3_B': {'SWBin': 54, 'HWBin': 2, 'Priority': 382,
                                                  'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue4_R': {'SWBin': 54, 'HWBin': 2, 'Priority': 383,
                                                  'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue4_Gr': {'SWBin': 54, 'HWBin': 2, 'Priority': 384,
                                                   'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue4_Gb': {'SWBin': 54, 'HWBin': 2, 'Priority': 385,
                                                   'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue4_B': {'SWBin': 54, 'HWBin': 2, 'Priority': 386,
                                                  'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_Mean_R': {'SWBin': 53, 'HWBin': 2, 'Priority': 387, 'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_Mean_Gr': {'SWBin': 53, 'HWBin': 2, 'Priority': 388, 'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_Mean_Gb': {'SWBin': 53, 'HWBin': 2, 'Priority': 389, 'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_Mean_B': {'SWBin': 53, 'HWBin': 2, 'Priority': 390, 'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_StdDEV_R': {'SWBin': 53, 'HWBin': 2, 'Priority': 391, 'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_StdDEV_Gr': {'SWBin': 53, 'HWBin': 2, 'Priority': 392, 'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_StdDEV_Gb': {'SWBin': 53, 'HWBin': 2, 'Priority': 393, 'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_StdDEV_B': {'SWBin': 53, 'HWBin': 2, 'Priority': 394, 'IsNan': 'DVP27M30F_BK32X_Cap','IsLast': 'DVP27M30F_BK32X_Cap'},
            'BK_Cluster3_2': {'SWBin': 73, 'HWBin': 2, 'Priority': 395},
            'LT_Cluster3_2': {'SWBin': 74, 'HWBin': 2, 'Priority': 396},
            'MIPI2L24M30F_WK1_Cap': {'SWBin': 89, 'HWBin': 1, 'Priority': 397, 'Status': []},
            'MIPI2L24M30F_WK1_Calc': {'SWBin': 97, 'HWBin': 5, 'Priority': 398},
            'MIPI_One_W1_FixedPattern': {'SWBin': 94, 'HWBin': 1, 'Priority': 399, 'IsNan': 'MIPI2L24M30F_WK1_Cap','IsLast': 'MIPI2L24M30F_WK1_Cap'},
            'MIPI2L24M30F_LT_Cap': {'SWBin': 90, 'HWBin': 1, 'Priority': 400, 'Status': []},
            'MIPI2L24M30F_LT_Calc': {'SWBin': 97, 'HWBin': 5, 'Priority': 401},
            'MIPI_LT_DRow_R': {'SWBin': 65, 'HWBin': 1, 'Priority': 402, 'IsNan': 'MIPI2L24M30F_LT_Cap'},
            'MIPI_LT_DRow_Gr': {'SWBin': 65, 'HWBin': 1, 'Priority': 403, 'IsNan': 'MIPI2L24M30F_LT_Cap'},
            'MIPI_LT_DRow_Gb': {'SWBin': 65, 'HWBin': 1, 'Priority': 404, 'IsNan': 'MIPI2L24M30F_LT_Cap'},
            'MIPI_LT_DRow_B': {'SWBin': 65, 'HWBin': 1, 'Priority': 405, 'IsNan': 'MIPI2L24M30F_LT_Cap'},
            'MIPI_LT_DCol_R': {'SWBin': 64, 'HWBin': 1, 'Priority': 406, 'IsNan': 'MIPI2L24M30F_LT_Cap'},
            'MIPI_LT_DCol_Gr': {'SWBin': 64, 'HWBin': 1, 'Priority': 407, 'IsNan': 'MIPI2L24M30F_LT_Cap'},
            'MIPI_LT_DCol_Gb': {'SWBin': 64, 'HWBin': 1, 'Priority': 408, 'IsNan': 'MIPI2L24M30F_LT_Cap'},
            'MIPI_LT_DCol_B': {'SWBin': 64, 'HWBin': 1, 'Priority': 409, 'IsNan': 'MIPI2L24M30F_LT_Cap'},
            'MIPI_LT_DRow_Color': {'SWBin': 65, 'HWBin': 1, 'Priority': 410, 'IsNan': 'MIPI2L24M30F_LT_Cap'},
            'MIPI_LT_DCol_Color': {'SWBin': 64, 'HWBin': 1, 'Priority': 411, 'IsNan': 'MIPI2L24M30F_LT_Cap'},
            'MIPI_LT_WeakLineRow_R': {'SWBin': 65, 'HWBin': 1, 'Priority': 412, 'IsNan': 'MIPI2L24M30F_LT_Cap'},
            'MIPI_LT_WeakLineRow_Gr': {'SWBin': 65, 'HWBin': 1, 'Priority': 413, 'IsNan': 'MIPI2L24M30F_LT_Cap'},
            'MIPI_LT_WeakLineRow_Gb': {'SWBin': 65, 'HWBin': 1, 'Priority': 414, 'IsNan': 'MIPI2L24M30F_LT_Cap'},
            'MIPI_LT_WeakLineRow_B': {'SWBin': 65, 'HWBin': 1, 'Priority': 415, 'IsNan': 'MIPI2L24M30F_LT_Cap'},
            'MIPI_LT_WeakLineCol_R': {'SWBin': 64, 'HWBin': 1, 'Priority': 416, 'IsNan': 'MIPI2L24M30F_LT_Cap'},
            'MIPI_LT_WeakLineCol_Gr': {'SWBin': 64, 'HWBin': 1, 'Priority': 417, 'IsNan': 'MIPI2L24M30F_LT_Cap'},
            'MIPI_LT_WeakLineCol_Gb': {'SWBin': 64, 'HWBin': 1, 'Priority': 418, 'IsNan': 'MIPI2L24M30F_LT_Cap'},
            'MIPI_LT_WeakLineCol_B': {'SWBin': 64, 'HWBin': 1, 'Priority': 419, 'IsNan': 'MIPI2L24M30F_LT_Cap'},
            'MIPI_LT_Mean_R': {'SWBin': 63, 'HWBin': 1, 'Priority': 420, 'IsNan': 'MIPI2L24M30F_LT_Cap'},
            'MIPI_LT_Mean_Gr': {'SWBin': 63, 'HWBin': 1, 'Priority': 421, 'IsNan': 'MIPI2L24M30F_LT_Cap'},
            'MIPI_LT_Mean_Gb': {'SWBin': 63, 'HWBin': 1, 'Priority': 422, 'IsNan': 'MIPI2L24M30F_LT_Cap'},
            'MIPI_LT_Mean_B': {'SWBin': 63, 'HWBin': 1, 'Priority': 423, 'IsNan': 'MIPI2L24M30F_LT_Cap'},
            'MIPI_LT_StdDEV_R': {'SWBin': 63, 'HWBin': 1, 'Priority': 424, 'IsNan': 'MIPI2L24M30F_LT_Cap'},
            'MIPI_LT_StdDEV_Gr': {'SWBin': 63, 'HWBin': 1, 'Priority': 425, 'IsNan': 'MIPI2L24M30F_LT_Cap'},
            'MIPI_LT_StdDEV_Gb': {'SWBin': 63, 'HWBin': 1, 'Priority': 426, 'IsNan': 'MIPI2L24M30F_LT_Cap'},
            'MIPI_LT_StdDEV_B': {'SWBin': 63, 'HWBin': 1, 'Priority': 427, 'IsNan': 'MIPI2L24M30F_LT_Cap'},
            'MIPI_LT_RI_R': {'SWBin': 63, 'HWBin': 1, 'Priority': 428, 'IsNan': 'MIPI2L24M30F_LT_Cap'},
            'MIPI_LT_RI_Gr': {'SWBin': 63, 'HWBin': 1, 'Priority': 429, 'IsNan': 'MIPI2L24M30F_LT_Cap'},
            'MIPI_LT_RI_Gb': {'SWBin': 63, 'HWBin': 1, 'Priority': 430, 'IsNan': 'MIPI2L24M30F_LT_Cap'},
            'MIPI_LT_RI_B': {'SWBin': 63, 'HWBin': 1, 'Priority': 431, 'IsNan': 'MIPI2L24M30F_LT_Cap'},
            'MIPI_LT_Ratio_GrR': {'SWBin': 63, 'HWBin': 1, 'Priority': 432, 'IsNan': 'MIPI2L24M30F_LT_Cap'},
            'MIPI_LT_Ratio_GbR': {'SWBin': 63, 'HWBin': 1, 'Priority': 433, 'IsNan': 'MIPI2L24M30F_LT_Cap'},
            'MIPI_LT_Ratio_GrB': {'SWBin': 63, 'HWBin': 1, 'Priority': 434, 'IsNan': 'MIPI2L24M30F_LT_Cap'},
            'MIPI_LT_Ratio_GbB': {'SWBin': 63, 'HWBin': 1, 'Priority': 435, 'IsNan': 'MIPI2L24M30F_LT_Cap'},
            'MIPI_LT_Ratio_GbGr': {'SWBin': 63, 'HWBin': 1, 'Priority': 436, 'IsNan': 'MIPI2L24M30F_LT_Cap'},
            'MIPI_LT_LostBit': {'SWBin': 63, 'HWBin': 1, 'Priority': 437, 'IsNan': 'MIPI2L24M30F_LT_Cap'},
            'MIPI_LT_LostBitSNR_SumDIFF_R': {'SWBin': 63, 'HWBin': 1, 'Priority': 438, 'IsNan': 'MIPI2L24M30F_LT_Cap'},
            'MIPI_LT_LostBitSNR_SumDIFF_Gr': {'SWBin': 63, 'HWBin': 1, 'Priority': 439, 'IsNan': 'MIPI2L24M30F_LT_Cap'},
            'MIPI_LT_LostBitSNR_SumDIFF_Gb': {'SWBin': 63, 'HWBin': 1, 'Priority': 440, 'IsNan': 'MIPI2L24M30F_LT_Cap'},
            'MIPI_LT_LostBitSNR_SumDIFF_B': {'SWBin': 63, 'HWBin': 1, 'Priority': 441, 'IsNan': 'MIPI2L24M30F_LT_Cap','IsLast': 'MIPI2L24M30F_LT_Cap'},
            'MIPI2L24M30F_LBIT_Cap': {'SWBin': 90, 'HWBin': 1, 'Priority': 442, 'Status': []},
            'MIPI2L24M30F_LBIT_Calc': {'SWBin': 97, 'HWBin': 5, 'Priority': 443},
            'MIPI_LBIT_LT_Mean_R': {'SWBin': 63, 'HWBin': 1, 'Priority': 444, 'IsNan': 'MIPI2L24M30F_LBIT_Cap'},
            'MIPI_LBIT_LT_Mean_Gr': {'SWBin': 63, 'HWBin': 1, 'Priority': 445, 'IsNan': 'MIPI2L24M30F_LBIT_Cap'},
            'MIPI_LBIT_LT_Mean_Gb': {'SWBin': 63, 'HWBin': 1, 'Priority': 446, 'IsNan': 'MIPI2L24M30F_LBIT_Cap'},
            'MIPI_LBIT_LT_Mean_B': {'SWBin': 63, 'HWBin': 1, 'Priority': 447, 'IsNan': 'MIPI2L24M30F_LBIT_Cap'},
            'MIPI_LBIT_LT_LostBit': {'SWBin': 63, 'HWBin': 1, 'Priority': 448, 'IsNan': 'MIPI2L24M30F_LBIT_Cap'},
            'MIPI_LBIT_LT_LostBitSNR_SumDIFF_R': {'SWBin': 63, 'HWBin': 1, 'Priority': 449,
                                                  'IsNan': 'MIPI2L24M30F_LBIT_Cap'},
            'MIPI_LBIT_LT_LostBitSNR_SumDIFF_Gr': {'SWBin': 63, 'HWBin': 1, 'Priority': 450,
                                                   'IsNan': 'MIPI2L24M30F_LBIT_Cap'},
            'MIPI_LBIT_LT_LostBitSNR_SumDIFF_Gb': {'SWBin': 63, 'HWBin': 1, 'Priority': 451,
                                                   'IsNan': 'MIPI2L24M30F_LBIT_Cap'},
            'MIPI_LBIT_LT_LostBitSNR_SumDIFF_B': {'SWBin': 63, 'HWBin': 1, 'Priority': 452,
                                                  'IsNan': 'MIPI2L24M30F_LBIT_Cap','IsLast': 'MIPI2L24M30F_LBIT_Cap'},
            'Binning': {'SWBin': 2, 'HWBin': 3, 'Priority': 453},
            'All Pass': {'SWBin': 1, 'HWBin': 3, 'Priority': 454}
        },
    'JX825':
        {
            'VSYNC_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 1},
            'HSYNC_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 2},
            'PCLK_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 3},
            'EXCLK_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 4},
            'RSTB_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 5},
            'PWDN_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 6},
            'SDA_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 7},
            'SCL_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 8},
            'VH_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 9},
            'VN1_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 10},
            'VN2_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 11},
            'D0_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 12},
            'D1_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 13},
            'D2_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 14},
            'D3_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 15},
            'D4_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 16},
            'D5_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 17},
            'D6_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 18},
            'D7_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 19},
            'D8_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 20},
            'D9_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 21},
            'MCP_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 22},
            'MCN_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 23},
            'MDP0_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 24},
            'MDN0_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 25},
            'MDP1_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 26},
            'MDN1_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 27},
            'MDP2_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 28},
            'MDN2_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 29},
            'MDP3_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 30},
            'MDN3_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 31},
            'DVDD_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 32},
            'AVDD_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 33},
            'DOVDD_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 34},
            'VSYNC_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 35},
            'HSYNC_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 36},
            'PCLK_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 37},
            'EXCLK_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 38},
            'RSTB_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 39},
            'PWDN_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 40},
            'SDA_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 41},
            'SCL_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 42},
            'D0_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 43},
            'D1_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 44},
            'D2_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 45},
            'D3_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 46},
            'D4_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 47},
            'D5_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 48},
            'D6_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 49},
            'D7_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 50},
            'D8_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 51},
            'D9_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 52},
            'MCP_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 53},
            'MCN_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 54},
            'MDP0_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 55},
            'MDN0_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 56},
            'MDP1_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 57},
            'MDN1_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 58},
            'MDP2_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 59},
            'MDN2_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 60},
            'MDP3_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 61},
            'MDN3_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 62},
            'VSYNC_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 63},
            'HSYNC_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 64},
            'PCLK_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 65},
            'EXCLK_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 66},
            'RSTB_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 67},
            'PWDN_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 68},
            'SDA_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 69},
            'SCL_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 70},
            'D0_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 71},
            'D1_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 72},
            'D2_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 73},
            'D3_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 74},
            'D4_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 75},
            'D5_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 76},
            'D6_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 77},
            'D7_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 78},
            'D8_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 79},
            'D9_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 80},
            'MCP_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 81},
            'MCN_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 82},
            'MDP0_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 83},
            'MDN0_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 84},
            'MDP1_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 85},
            'MDN1_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 86},
            'MDP2_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 87},
            'MDN2_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 88},
            'MDP3_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 89},
            'MDN3_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 90},
            'iic_test': {'SWBin': 7, 'HWBin': 5, 'Priority': 91},
            'MIPI4L27M30F_BK_Cap': {'SWBin': 98, 'HWBin': 8, 'Priority': 92, 'Status': []},
            'MIPI4L27M30F_BK_Calc': {'SWBin': 97, 'HWBin': 8, 'Priority': 93},
            'BLC_R': {'SWBin': 7, 'HWBin': 5, 'Priority': 94},
            'BLC_Gr': {'SWBin': 7, 'HWBin': 5, 'Priority': 95},
            'BLC_Gb': {'SWBin': 7, 'HWBin': 5, 'Priority': 96},
            'BLC_B': {'SWBin': 7, 'HWBin': 5, 'Priority': 97},
            'BK_DeadRowExBPix_R': {'SWBin': 15, 'HWBin': 6, 'Priority': 98, 'IsNan': 'MIPI4L27M30F_BK_Cap'},
            'BK_DeadRowExBPix_Gr': {'SWBin': 15, 'HWBin': 6, 'Priority': 99, 'IsNan': 'MIPI4L27M30F_BK_Cap'},
            'BK_DeadRowExBPix_Gb': {'SWBin': 15, 'HWBin': 6, 'Priority': 100, 'IsNan': 'MIPI4L27M30F_BK_Cap'},
            'BK_DeadRowExBPix_B': {'SWBin': 15, 'HWBin': 6, 'Priority': 101, 'IsNan': 'MIPI4L27M30F_BK_Cap'},
            'BK_DeadColExBPix_R': {'SWBin': 14, 'HWBin': 6, 'Priority': 102, 'IsNan': 'MIPI4L27M30F_BK_Cap'},
            'BK_DeadColExBPix_Gr': {'SWBin': 14, 'HWBin': 6, 'Priority': 103, 'IsNan': 'MIPI4L27M30F_BK_Cap'},
            'BK_DeadColExBPix_Gb': {'SWBin': 14, 'HWBin': 6, 'Priority': 104, 'IsNan': 'MIPI4L27M30F_BK_Cap'},
            'BK_DeadColExBPix_B': {'SWBin': 14, 'HWBin': 6, 'Priority': 105, 'IsNan': 'MIPI4L27M30F_BK_Cap'},
            'BK_MixDCol_VfpnQty_R': {'SWBin': 14, 'HWBin': 6, 'Priority': 106, 'IsNan': 'MIPI4L27M30F_BK_Cap'},
            'BK_MixDCol_VfpnQty_Gr': {'SWBin': 14, 'HWBin': 6, 'Priority': 107, 'IsNan': 'MIPI4L27M30F_BK_Cap'},
            'BK_MixDCol_VfpnQty_Gb': {'SWBin': 14, 'HWBin': 6, 'Priority': 108, 'IsNan': 'MIPI4L27M30F_BK_Cap'},
            'BK_MixDCol_VfpnQty_B': {'SWBin': 14, 'HWBin': 6, 'Priority': 109, 'IsNan': 'MIPI4L27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue0_R': {'SWBin': 13, 'HWBin': 6, 'Priority': 110, 'IsNan': 'MIPI4L27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue0_Gr': {'SWBin': 13, 'HWBin': 6, 'Priority': 111, 'IsNan': 'MIPI4L27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue0_Gb': {'SWBin': 13, 'HWBin': 6, 'Priority': 112, 'IsNan': 'MIPI4L27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue0_B': {'SWBin': 13, 'HWBin': 6, 'Priority': 113, 'IsNan': 'MIPI4L27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue1_R': {'SWBin': 13, 'HWBin': 6, 'Priority': 114, 'IsNan': 'MIPI4L27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue1_Gr': {'SWBin': 13, 'HWBin': 6, 'Priority': 115, 'IsNan': 'MIPI4L27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue1_Gb': {'SWBin': 13, 'HWBin': 6, 'Priority': 116, 'IsNan': 'MIPI4L27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue1_B': {'SWBin': 13, 'HWBin': 6, 'Priority': 117, 'IsNan': 'MIPI4L27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue2_R': {'SWBin': 13, 'HWBin': 6, 'Priority': 118, 'IsNan': 'MIPI4L27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue2_Gr': {'SWBin': 13, 'HWBin': 6, 'Priority': 119, 'IsNan': 'MIPI4L27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue2_Gb': {'SWBin': 13, 'HWBin': 6, 'Priority': 120, 'IsNan': 'MIPI4L27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue2_B': {'SWBin': 13, 'HWBin': 6, 'Priority': 121, 'IsNan': 'MIPI4L27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue3_R': {'SWBin': 13, 'HWBin': 6, 'Priority': 122, 'IsNan': 'MIPI4L27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue3_Gr': {'SWBin': 13, 'HWBin': 6, 'Priority': 123, 'IsNan': 'MIPI4L27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue3_Gb': {'SWBin': 13, 'HWBin': 6, 'Priority': 124, 'IsNan': 'MIPI4L27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue3_B': {'SWBin': 13, 'HWBin': 6, 'Priority': 125, 'IsNan': 'MIPI4L27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue4_R': {'SWBin': 13, 'HWBin': 6, 'Priority': 126, 'IsNan': 'MIPI4L27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue4_Gr': {'SWBin': 13, 'HWBin': 6, 'Priority': 127, 'IsNan': 'MIPI4L27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue4_Gb': {'SWBin': 13, 'HWBin': 6, 'Priority': 128, 'IsNan': 'MIPI4L27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue4_B': {'SWBin': 13, 'HWBin': 6, 'Priority': 129, 'IsNan': 'MIPI4L27M30F_BK_Cap'},
            'BK_Mean_R': {'SWBin': 13, 'HWBin': 6, 'Priority': 130, 'IsNan': 'MIPI4L27M30F_BK_Cap'},
            'BK_Mean_Gr': {'SWBin': 13, 'HWBin': 6, 'Priority': 131, 'IsNan': 'MIPI4L27M30F_BK_Cap'},
            'BK_Mean_Gb': {'SWBin': 13, 'HWBin': 6, 'Priority': 132, 'IsNan': 'MIPI4L27M30F_BK_Cap'},
            'BK_Mean_B': {'SWBin': 13, 'HWBin': 6, 'Priority': 133, 'IsNan': 'MIPI4L27M30F_BK_Cap'},
            'BK_StdDEV_R': {'SWBin': 13, 'HWBin': 6, 'Priority': 134, 'IsNan': 'MIPI4L27M30F_BK_Cap'},
            'BK_StdDEV_Gr': {'SWBin': 13, 'HWBin': 6, 'Priority': 135, 'IsNan': 'MIPI4L27M30F_BK_Cap'},
            'BK_StdDEV_Gb': {'SWBin': 13, 'HWBin': 6, 'Priority': 136, 'IsNan': 'MIPI4L27M30F_BK_Cap'},
            'BK_StdDEV_B': {'SWBin': 13, 'HWBin': 6, 'Priority': 137, 'IsNan': 'MIPI4L27M30F_BK_Cap','IsLast': 'MIPI4L27M30F_BK_Cap'},
            'VH_voltage': {'SWBin': 9, 'HWBin': 5, 'Priority': 138},
            'VN1_voltage': {'SWBin': 9, 'HWBin': 5, 'Priority': 139},
            'VN2_voltage': {'SWBin': 9, 'HWBin': 5, 'Priority': 140},
            'DVDD_voltage': {'SWBin': 9, 'HWBin': 5, 'Priority': 141},
            'MIPI4L27M30F_LT_Cap': {'SWBin': 99, 'HWBin': 8, 'Priority': 142, 'Status': []},
            'MIPI4L27M30F_LT_Calc': {'SWBin': 99, 'HWBin': 8, 'Priority': 143},
            'LT_DRow_R': {'SWBin': 25, 'HWBin': 4, 'Priority': 144, 'IsNan': 'MIPI4L27M30F_LT_Cap'},
            'LT_DRow_Gr': {'SWBin': 25, 'HWBin': 4, 'Priority': 145, 'IsNan': 'MIPI4L27M30F_LT_Cap'},
            'LT_DRow_Gb': {'SWBin': 25, 'HWBin': 4, 'Priority': 146, 'IsNan': 'MIPI4L27M30F_LT_Cap'},
            'LT_DRow_B': {'SWBin': 25, 'HWBin': 4, 'Priority': 147, 'IsNan': 'MIPI4L27M30F_LT_Cap'},
            'LT_DCol_R': {'SWBin': 24, 'HWBin': 4, 'Priority': 148, 'IsNan': 'MIPI4L27M30F_LT_Cap'},
            'LT_DCol_Gr': {'SWBin': 24, 'HWBin': 4, 'Priority': 149, 'IsNan': 'MIPI4L27M30F_LT_Cap'},
            'LT_DCol_Gb': {'SWBin': 24, 'HWBin': 4, 'Priority': 150, 'IsNan': 'MIPI4L27M30F_LT_Cap'},
            'LT_DCol_B': {'SWBin': 24, 'HWBin': 4, 'Priority': 151, 'IsNan': 'MIPI4L27M30F_LT_Cap'},
            'LT_DRow_Color': {'SWBin': 25, 'HWBin': 4, 'Priority': 152, 'IsNan': 'MIPI4L27M30F_LT_Cap'},
            'LT_DCol_Color': {'SWBin': 24, 'HWBin': 4, 'Priority': 153, 'IsNan': 'MIPI4L27M30F_LT_Cap'},
            'LT_WeakLineRow_R': {'SWBin': 25, 'HWBin': 4, 'Priority': 154, 'IsNan': 'MIPI4L27M30F_LT_Cap'},
            'LT_WeakLineRow_Gr': {'SWBin': 25, 'HWBin': 4, 'Priority': 155, 'IsNan': 'MIPI4L27M30F_LT_Cap'},
            'LT_WeakLineRow_Gb': {'SWBin': 25, 'HWBin': 4, 'Priority': 156, 'IsNan': 'MIPI4L27M30F_LT_Cap'},
            'LT_WeakLineRow_B': {'SWBin': 25, 'HWBin': 4, 'Priority': 157, 'IsNan': 'MIPI4L27M30F_LT_Cap'},
            'LT_WeakLineCol_R': {'SWBin': 24, 'HWBin': 4, 'Priority': 158, 'IsNan': 'MIPI4L27M30F_LT_Cap'},
            'LT_WeakLineCol_Gr': {'SWBin': 24, 'HWBin': 4, 'Priority': 159, 'IsNan': 'MIPI4L27M30F_LT_Cap'},
            'LT_WeakLineCol_Gb': {'SWBin': 24, 'HWBin': 4, 'Priority': 160, 'IsNan': 'MIPI4L27M30F_LT_Cap'},
            'LT_WeakLineCol_B': {'SWBin': 24, 'HWBin': 4, 'Priority': 161, 'IsNan': 'MIPI4L27M30F_LT_Cap'},
            'LT_Mean_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 162, 'IsNan': 'MIPI4L27M30F_LT_Cap'},
            'LT_Mean_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 163, 'IsNan': 'MIPI4L27M30F_LT_Cap'},
            'LT_Mean_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 164, 'IsNan': 'MIPI4L27M30F_LT_Cap'},
            'LT_Mean_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 165, 'IsNan': 'MIPI4L27M30F_LT_Cap'},
            'LT_StdDEV_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 166, 'IsNan': 'MIPI4L27M30F_LT_Cap'},
            'LT_StdDEV_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 167, 'IsNan': 'MIPI4L27M30F_LT_Cap'},
            'LT_StdDEV_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 168, 'IsNan': 'MIPI4L27M30F_LT_Cap'},
            'LT_StdDEV_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 169, 'IsNan': 'MIPI4L27M30F_LT_Cap'},
            'LT_RI_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 170, 'IsNan': 'MIPI4L27M30F_LT_Cap'},
            'LT_RI_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 171, 'IsNan': 'MIPI4L27M30F_LT_Cap'},
            'LT_RI_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 172, 'IsNan': 'MIPI4L27M30F_LT_Cap'},
            'LT_RI_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 173, 'IsNan': 'MIPI4L27M30F_LT_Cap'},
            'LT_Ratio_GrR': {'SWBin': 23, 'HWBin': 4, 'Priority': 174, 'IsNan': 'MIPI4L27M30F_LT_Cap'},
            'LT_Ratio_GbR': {'SWBin': 23, 'HWBin': 4, 'Priority': 175, 'IsNan': 'MIPI4L27M30F_LT_Cap'},
            'LT_Ratio_GrB': {'SWBin': 23, 'HWBin': 4, 'Priority': 176, 'IsNan': 'MIPI4L27M30F_LT_Cap'},
            'LT_Ratio_GbB': {'SWBin': 23, 'HWBin': 4, 'Priority': 177, 'IsNan': 'MIPI4L27M30F_LT_Cap'},
            'LT_Ratio_GbGr': {'SWBin': 23, 'HWBin': 4, 'Priority': 178, 'IsNan': 'MIPI4L27M30F_LT_Cap'},
            'LT_LostBit': {'SWBin': 23, 'HWBin': 4, 'Priority': 179, 'IsNan': 'MIPI4L27M30F_LT_Cap'},
            'LT_LostBitSNR_SumDIFF_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 180, 'IsNan': 'MIPI4L27M30F_LT_Cap'},
            'LT_LostBitSNR_SumDIFF_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 181, 'IsNan': 'MIPI4L27M30F_LT_Cap'},
            'LT_LostBitSNR_SumDIFF_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 182, 'IsNan': 'MIPI4L27M30F_LT_Cap'},
            'LT_LostBitSNR_SumDIFF_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 183, 'IsNan': 'MIPI4L27M30F_LT_Cap','IsLast': 'MIPI4L27M30F_LT_Cap'},
            'MIPI4L27M30F_LBIT_Cap': {'SWBin': 99, 'HWBin': 8, 'Priority': 184, 'Status': []},
            'MIPI4L27M30F_LBIT_Calc': {'SWBin': 97, 'HWBin': 8, 'Priority': 185},
            'LBIT_LT_Mean_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 186, 'IsNan': 'MIPI4L27M30F_LBIT_Cap'},
            'LBIT_LT_Mean_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 187, 'IsNan': 'MIPI4L27M30F_LBIT_Cap'},
            'LBIT_LT_Mean_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 188, 'IsNan': 'MIPI4L27M30F_LBIT_Cap'},
            'LBIT_LT_Mean_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 189, 'IsNan': 'MIPI4L27M30F_LBIT_Cap'},
            'LBIT_LT_DRow_R': {'SWBin': 25, 'HWBin': 4, 'Priority': 190, 'IsNan': 'MIPI4L27M30F_LBIT_Cap'},
            'LBIT_LT_DRow_Gr': {'SWBin': 25, 'HWBin': 4, 'Priority': 191, 'IsNan': 'MIPI4L27M30F_LBIT_Cap'},
            'LBIT_LT_DRow_Gb': {'SWBin': 25, 'HWBin': 4, 'Priority': 192, 'IsNan': 'MIPI4L27M30F_LBIT_Cap'},
            'LBIT_LT_DRow_B': {'SWBin': 25, 'HWBin': 4, 'Priority': 193, 'IsNan': 'MIPI4L27M30F_LBIT_Cap'},
            'LBIT_LT_DCol_R': {'SWBin': 24, 'HWBin': 4, 'Priority': 194, 'IsNan': 'MIPI4L27M30F_LBIT_Cap'},
            'LBIT_LT_DCol_Gr': {'SWBin': 24, 'HWBin': 4, 'Priority': 195, 'IsNan': 'MIPI4L27M30F_LBIT_Cap'},
            'LBIT_LT_DCol_Gb': {'SWBin': 24, 'HWBin': 4, 'Priority': 196, 'IsNan': 'MIPI4L27M30F_LBIT_Cap'},
            'LBIT_LT_DCol_B': {'SWBin': 24, 'HWBin': 4, 'Priority': 197, 'IsNan': 'MIPI4L27M30F_LBIT_Cap'},
            'LBIT_LT_DRow_Color': {'SWBin': 25, 'HWBin': 4, 'Priority': 198, 'IsNan': 'MIPI4L27M30F_LBIT_Cap'},
            'LBIT_LT_DCol_Color': {'SWBin': 24, 'HWBin': 4, 'Priority': 199, 'IsNan': 'MIPI4L27M30F_LBIT_Cap'},
            'LBIT_LT_WeakLineRow_R': {'SWBin': 25, 'HWBin': 4, 'Priority': 200, 'IsNan': 'MIPI4L27M30F_LBIT_Cap'},
            'LBIT_LT_WeakLineRow_Gr': {'SWBin': 25, 'HWBin': 4, 'Priority': 201, 'IsNan': 'MIPI4L27M30F_LBIT_Cap'},
            'LBIT_LT_WeakLineRow_Gb': {'SWBin': 25, 'HWBin': 4, 'Priority': 202, 'IsNan': 'MIPI4L27M30F_LBIT_Cap'},
            'LBIT_LT_WeakLineRow_B': {'SWBin': 25, 'HWBin': 4, 'Priority': 203, 'IsNan': 'MIPI4L27M30F_LBIT_Cap'},
            'LBIT_LT_WeakLineCol_R': {'SWBin': 24, 'HWBin': 4, 'Priority': 204, 'IsNan': 'MIPI4L27M30F_LBIT_Cap'},
            'LBIT_LT_WeakLineCol_Gr': {'SWBin': 24, 'HWBin': 4, 'Priority': 205, 'IsNan': 'MIPI4L27M30F_LBIT_Cap'},
            'LBIT_LT_WeakLineCol_Gb': {'SWBin': 24, 'HWBin': 4, 'Priority': 206, 'IsNan': 'MIPI4L27M30F_LBIT_Cap'},
            'LBIT_LT_WeakLineCol_B': {'SWBin': 24, 'HWBin': 4, 'Priority': 207, 'IsNan': 'MIPI4L27M30F_LBIT_Cap'},
            'LBIT_LT_Cluster2': {'SWBin': 23, 'HWBin': 4, 'Priority': 208, 'IsNan': 'MIPI4L27M30F_LBIT_Cap'},
            'LBIT_LT_Cluster1': {'SWBin': 23, 'HWBin': 4, 'Priority': 209, 'IsNan': 'MIPI4L27M30F_LBIT_Cap'},
            'LBIT_LT_LostBit': {'SWBin': 23, 'HWBin': 4, 'Priority': 210, 'IsNan': 'MIPI4L27M30F_LBIT_Cap'},
            'LBIT_LT_LostBitSNR_SumDIFF_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 211,
                                             'IsNan': 'MIPI4L27M30F_LBIT_Cap'},
            'LBIT_LT_LostBitSNR_SumDIFF_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 212,
                                              'IsNan': 'MIPI4L27M30F_LBIT_Cap'},
            'LBIT_LT_LostBitSNR_SumDIFF_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 213,
                                              'IsNan': 'MIPI4L27M30F_LBIT_Cap'},
            'LBIT_LT_LostBitSNR_SumDIFF_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 214,
                                             'IsNan': 'MIPI4L27M30F_LBIT_Cap','IsLast': 'MIPI4L27M30F_LBIT_Cap'},
            'LT_CornerLine': {'SWBin': 26, 'HWBin': 4, 'Priority': 215},
            'LT_ScratchLine': {'SWBin': 26, 'HWBin': 4, 'Priority': 216},
            'LT_Blemish': {'SWBin': 31, 'HWBin': 4, 'Priority': 217},
            'LT_LineStripe': {'SWBin': 27, 'HWBin': 4, 'Priority': 218},
            'LT_Particle': {'SWBin': 32, 'HWBin': 4, 'Priority': 219},
            'BK_Cluster2': {'SWBin': 30, 'HWBin': 4, 'Priority': 220},
            'LT_Cluster2': {'SWBin': 30, 'HWBin': 4, 'Priority': 221},
            'BK_Cluster1': {'SWBin': 29, 'HWBin': 4, 'Priority': 222},
            'LT_Cluster1': {'SWBin': 29, 'HWBin': 4, 'Priority': 223},
            'BK_BadPixel_Area': {'SWBin': 33, 'HWBin': 4, 'Priority': 224},
            'LT_BadPixel_Area': {'SWBin': 34, 'HWBin': 4, 'Priority': 225},
            'WP_Count': {'SWBin': 35, 'HWBin': 6, 'Priority': 226},
            'BK_Z1_BT_DP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 227},
            'BK_Z1_BT_DP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 228},
            'BK_Z1_BT_DP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 229},
            'BK_Z1_BT_DP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 230},
            'BK_Z2_BT_DP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 231},
            'BK_Z2_BT_DP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 232},
            'BK_Z2_BT_DP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 233},
            'BK_Z2_BT_DP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 234},
            'BK_Z1_BT_WP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 235},
            'BK_Z1_BT_WP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 236},
            'BK_Z1_BT_WP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 237},
            'BK_Z1_BT_WP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 238},
            'BK_Z2_BT_WP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 239},
            'BK_Z2_BT_WP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 240},
            'BK_Z2_BT_WP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 241},
            'BK_Z2_BT_WP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 242},
            'LT_Z1_BT_DP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 243},
            'LT_Z1_BT_DP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 244},
            'LT_Z1_BT_DP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 245},
            'LT_Z1_BT_DP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 246},
            'LT_Z2_BT_DP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 247},
            'LT_Z2_BT_DP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 248},
            'LT_Z2_BT_DP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 249},
            'LT_Z2_BT_DP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 250},
            'LT_Z1_BT_WP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 251},
            'LT_Z1_BT_WP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 252},
            'LT_Z1_BT_WP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 253},
            'LT_Z1_BT_WP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 254},
            'LT_Z2_BT_WP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 255},
            'LT_Z2_BT_WP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 256},
            'LT_Z2_BT_WP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 257},
            'LT_Z2_BT_WP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 258},
            'LT_Z1_DK_WP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 259},
            'LT_Z1_DK_WP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 260},
            'LT_Z1_DK_WP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 261},
            'LT_Z1_DK_WP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 262},
            'LT_Z2_DK_WP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 263},
            'LT_Z2_DK_WP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 264},
            'LT_Z2_DK_WP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 265},
            'LT_Z2_DK_WP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 266},
            'LT_Z1_DK_DP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 267},
            'LT_Z1_DK_DP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 268},
            'LT_Z1_DK_DP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 269},
            'LT_Z1_DK_DP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 270},
            'LT_Z2_DK_DP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 271},
            'LT_Z2_DK_DP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 272},
            'LT_Z2_DK_DP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 273},
            'LT_Z2_DK_DP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 274},
            'BK_Z1_BT_DP': {'SWBin': 60, 'HWBin': 6, 'Priority': 275},
            'BK_Z2_BT_DP': {'SWBin': 60, 'HWBin': 6, 'Priority': 276},
            'BK_Z1_BT_WP': {'SWBin': 60, 'HWBin': 6, 'Priority': 277},
            'BK_Z2_BT_WP': {'SWBin': 60, 'HWBin': 6, 'Priority': 278},
            'LT_Z1_BT_DP': {'SWBin': 60, 'HWBin': 6, 'Priority': 279},
            'LT_Z2_BT_DP': {'SWBin': 60, 'HWBin': 6, 'Priority': 280},
            'LT_Z1_BT_WP': {'SWBin': 60, 'HWBin': 6, 'Priority': 281},
            'LT_Z2_BT_WP': {'SWBin': 60, 'HWBin': 6, 'Priority': 282},
            'LT_Z1_DK_DP': {'SWBin': 60, 'HWBin': 6, 'Priority': 283},
            'LT_Z2_DK_DP': {'SWBin': 60, 'HWBin': 6, 'Priority': 284},
            'LT_Z1_DK_WP': {'SWBin': 60, 'HWBin': 6, 'Priority': 285},
            'LT_Z2_DK_WP': {'SWBin': 60, 'HWBin': 6, 'Priority': 286},
            'Active_AVDD': {'SWBin': 12, 'HWBin': 5, 'Priority': 287},
            'Active_DOVDD': {'SWBin': 12, 'HWBin': 5, 'Priority': 288},
            'PWDN_AVDD': {'SWBin': 8, 'HWBin': 5, 'Priority': 289},
            'PWDN_DOVDD': {'SWBin': 8, 'HWBin': 5, 'Priority': 290},
            'PWDN_Total': {'SWBin': 8, 'HWBin': 5, 'Priority': 291},
            'BK_Cluster3GrGb': {'SWBin': 39, 'HWBin': 4, 'Priority': 292},
            'LT_Cluster3GrGb': {'SWBin': 40, 'HWBin': 4, 'Priority': 293},
            'MIPI4L27M30F_FW_Cap': {'SWBin': 99, 'HWBin': 8, 'Priority': 294, 'Status': []},
            'MIPI4L27M30F_FW_Calc': {'SWBin': 97, 'HWBin': 8, 'Priority': 295},
            'FW_LT_DRow_R': {'SWBin': 36, 'HWBin': 4, 'Priority': 296, 'IsNan': 'MIPI4L27M30F_FW_Cap'},
            'FW_LT_DRow_Gr': {'SWBin': 36, 'HWBin': 4, 'Priority': 297, 'IsNan': 'MIPI4L27M30F_FW_Cap'},
            'FW_LT_DRow_Gb': {'SWBin': 36, 'HWBin': 4, 'Priority': 298, 'IsNan': 'MIPI4L27M30F_FW_Cap'},
            'FW_LT_DRow_B': {'SWBin': 36, 'HWBin': 4, 'Priority': 299, 'IsNan': 'MIPI4L27M30F_FW_Cap'},
            'FW_LT_DCol_R': {'SWBin': 36, 'HWBin': 4, 'Priority': 300, 'IsNan': 'MIPI4L27M30F_FW_Cap'},
            'FW_LT_DCol_Gr': {'SWBin': 36, 'HWBin': 4, 'Priority': 301, 'IsNan': 'MIPI4L27M30F_FW_Cap'},
            'FW_LT_DCol_Gb': {'SWBin': 36, 'HWBin': 4, 'Priority': 302, 'IsNan': 'MIPI4L27M30F_FW_Cap'},
            'FW_LT_DCol_B': {'SWBin': 36, 'HWBin': 4, 'Priority': 303, 'IsNan': 'MIPI4L27M30F_FW_Cap'},
            'FW_LT_DRow_Color': {'SWBin': 36, 'HWBin': 4, 'Priority': 304, 'IsNan': 'MIPI4L27M30F_FW_Cap'},
            'FW_LT_DCol_Color': {'SWBin': 36, 'HWBin': 4, 'Priority': 305, 'IsNan': 'MIPI4L27M30F_FW_Cap'},
            'FW_LT_Mean_R': {'SWBin': 36, 'HWBin': 4, 'Priority': 306, 'IsNan': 'MIPI4L27M30F_FW_Cap'},
            'FW_LT_Mean_Gr': {'SWBin': 36, 'HWBin': 4, 'Priority': 307, 'IsNan': 'MIPI4L27M30F_FW_Cap'},
            'FW_LT_Mean_Gb': {'SWBin': 36, 'HWBin': 4, 'Priority': 308, 'IsNan': 'MIPI4L27M30F_FW_Cap'},
            'FW_LT_Mean_B': {'SWBin': 36, 'HWBin': 4, 'Priority': 309, 'IsNan': 'MIPI4L27M30F_FW_Cap'},
            'FW_LT_StdDEV_R': {'SWBin': 36, 'HWBin': 4, 'Priority': 310, 'IsNan': 'MIPI4L27M30F_FW_Cap'},
            'FW_LT_StdDEV_Gr': {'SWBin': 36, 'HWBin': 4, 'Priority': 311, 'IsNan': 'MIPI4L27M30F_FW_Cap'},
            'FW_LT_StdDEV_Gb': {'SWBin': 36, 'HWBin': 4, 'Priority': 312, 'IsNan': 'MIPI4L27M30F_FW_Cap'},
            'FW_LT_StdDEV_B': {'SWBin': 36, 'HWBin': 4, 'Priority': 313, 'IsNan': 'MIPI4L27M30F_FW_Cap'},
            'FW_LT_RI_R': {'SWBin': 36, 'HWBin': 4, 'Priority': 314, 'IsNan': 'MIPI4L27M30F_FW_Cap'},
            'FW_LT_RI_Gr': {'SWBin': 36, 'HWBin': 4, 'Priority': 315, 'IsNan': 'MIPI4L27M30F_FW_Cap'},
            'FW_LT_RI_Gb': {'SWBin': 36, 'HWBin': 4, 'Priority': 316, 'IsNan': 'MIPI4L27M30F_FW_Cap'},
            'FW_LT_RI_B': {'SWBin': 36, 'HWBin': 4, 'Priority': 317, 'IsNan': 'MIPI4L27M30F_FW_Cap'},
            'FW_LT_Ratio_GrR': {'SWBin': 36, 'HWBin': 4, 'Priority': 318, 'IsNan': 'MIPI4L27M30F_FW_Cap'},
            'FW_LT_Ratio_GbR': {'SWBin': 36, 'HWBin': 4, 'Priority': 319, 'IsNan': 'MIPI4L27M30F_FW_Cap'},
            'FW_LT_Ratio_GrB': {'SWBin': 36, 'HWBin': 4, 'Priority': 320, 'IsNan': 'MIPI4L27M30F_FW_Cap'},
            'FW_LT_Ratio_GbB': {'SWBin': 36, 'HWBin': 4, 'Priority': 321, 'IsNan': 'MIPI4L27M30F_FW_Cap'},
            'FW_LT_Ratio_GbGr': {'SWBin': 36, 'HWBin': 4, 'Priority': 322, 'IsNan': 'MIPI4L27M30F_FW_Cap'},
            'FW_LT_LostBit': {'SWBin': 36, 'HWBin': 4, 'Priority': 323, 'IsNan': 'MIPI4L27M30F_FW_Cap','IsLast': 'MIPI4L27M30F_FW_Cap'},
            'MIPI2L_LT_Cap': {'SWBin': 90, 'HWBin': 8, 'Priority': 324, 'Status': []},
            'MIPI2L_LT_Calc': {'SWBin': 97, 'HWBin': 8, 'Priority': 325},
            'MIPI2L_LT_DRow_R': {'SWBin': 65, 'HWBin': 4, 'Priority': 326, 'IsNan': 'MIPI2L_LT_Cap'},
            'MIPI2L_LT_DRow_Gr': {'SWBin': 65, 'HWBin': 4, 'Priority': 327, 'IsNan': 'MIPI2L_LT_Cap'},
            'MIPI2L_LT_DRow_Gb': {'SWBin': 65, 'HWBin': 4, 'Priority': 328, 'IsNan': 'MIPI2L_LT_Cap'},
            'MIPI2L_LT_DRow_B': {'SWBin': 65, 'HWBin': 4, 'Priority': 329, 'IsNan': 'MIPI2L_LT_Cap'},
            'MIPI2L_LT_DCol_R': {'SWBin': 64, 'HWBin': 4, 'Priority': 330, 'IsNan': 'MIPI2L_LT_Cap'},
            'MIPI2L_LT_DCol_Gr': {'SWBin': 64, 'HWBin': 4, 'Priority': 331, 'IsNan': 'MIPI2L_LT_Cap'},
            'MIPI2L_LT_DCol_Gb': {'SWBin': 64, 'HWBin': 4, 'Priority': 332, 'IsNan': 'MIPI2L_LT_Cap'},
            'MIPI2L_LT_DCol_B': {'SWBin': 64, 'HWBin': 4, 'Priority': 333, 'IsNan': 'MIPI2L_LT_Cap'},
            'MIPI2L_LT_DRow_Color': {'SWBin': 65, 'HWBin': 4, 'Priority': 334, 'IsNan': 'MIPI2L_LT_Cap'},
            'MIPI2L_LT_DCol_Color': {'SWBin': 64, 'HWBin': 4, 'Priority': 335, 'IsNan': 'MIPI2L_LT_Cap'},
            'MIPI2L_LT_WeakLineRow_R': {'SWBin': 65, 'HWBin': 4, 'Priority': 336, 'IsNan': 'MIPI2L_LT_Cap'},
            'MIPI2L_LT_WeakLineRow_Gr': {'SWBin': 65, 'HWBin': 4, 'Priority': 337, 'IsNan': 'MIPI2L_LT_Cap'},
            'MIPI2L_LT_WeakLineRow_Gb': {'SWBin': 65, 'HWBin': 4, 'Priority': 338, 'IsNan': 'MIPI2L_LT_Cap'},
            'MIPI2L_LT_WeakLineRow_B': {'SWBin': 65, 'HWBin': 4, 'Priority': 339, 'IsNan': 'MIPI2L_LT_Cap'},
            'MIPI2L_LT_WeakLineCol_R': {'SWBin': 64, 'HWBin': 4, 'Priority': 340, 'IsNan': 'MIPI2L_LT_Cap'},
            'MIPI2L_LT_WeakLineCol_Gr': {'SWBin': 64, 'HWBin': 4, 'Priority': 341, 'IsNan': 'MIPI2L_LT_Cap'},
            'MIPI2L_LT_WeakLineCol_Gb': {'SWBin': 64, 'HWBin': 4, 'Priority': 342, 'IsNan': 'MIPI2L_LT_Cap'},
            'MIPI2L_LT_WeakLineCol_B': {'SWBin': 64, 'HWBin': 4, 'Priority': 343, 'IsNan': 'MIPI2L_LT_Cap'},
            'MIPI2L_LT_Mean_R': {'SWBin': 63, 'HWBin': 4, 'Priority': 344, 'IsNan': 'MIPI2L_LT_Cap'},
            'MIPI2L_LT_Mean_Gr': {'SWBin': 63, 'HWBin': 4, 'Priority': 345, 'IsNan': 'MIPI2L_LT_Cap'},
            'MIPI2L_LT_Mean_Gb': {'SWBin': 63, 'HWBin': 4, 'Priority': 346, 'IsNan': 'MIPI2L_LT_Cap'},
            'MIPI2L_LT_Mean_B': {'SWBin': 63, 'HWBin': 4, 'Priority': 347, 'IsNan': 'MIPI2L_LT_Cap'},
            'MIPI2L_LT_StdDEV_R': {'SWBin': 63, 'HWBin': 4, 'Priority': 348, 'IsNan': 'MIPI2L_LT_Cap'},
            'MIPI2L_LT_StdDEV_Gr': {'SWBin': 63, 'HWBin': 4, 'Priority': 349, 'IsNan': 'MIPI2L_LT_Cap'},
            'MIPI2L_LT_StdDEV_Gb': {'SWBin': 63, 'HWBin': 4, 'Priority': 350, 'IsNan': 'MIPI2L_LT_Cap'},
            'MIPI2L_LT_StdDEV_B': {'SWBin': 63, 'HWBin': 4, 'Priority': 351, 'IsNan': 'MIPI2L_LT_Cap'},
            'MIPI2L_LT_RI_R': {'SWBin': 63, 'HWBin': 4, 'Priority': 352, 'IsNan': 'MIPI2L_LT_Cap'},
            'MIPI2L_LT_RI_Gr': {'SWBin': 63, 'HWBin': 4, 'Priority': 353, 'IsNan': 'MIPI2L_LT_Cap'},
            'MIPI2L_LT_RI_Gb': {'SWBin': 63, 'HWBin': 4, 'Priority': 354, 'IsNan': 'MIPI2L_LT_Cap'},
            'MIPI2L_LT_RI_B': {'SWBin': 63, 'HWBin': 4, 'Priority': 355, 'IsNan': 'MIPI2L_LT_Cap'},
            'MIPI2L_LT_Ratio_GrR': {'SWBin': 63, 'HWBin': 4, 'Priority': 356, 'IsNan': 'MIPI2L_LT_Cap'},
            'MIPI2L_LT_Ratio_GbR': {'SWBin': 63, 'HWBin': 4, 'Priority': 357, 'IsNan': 'MIPI2L_LT_Cap'},
            'MIPI2L_LT_Ratio_GrB': {'SWBin': 63, 'HWBin': 4, 'Priority': 358, 'IsNan': 'MIPI2L_LT_Cap'},
            'MIPI2L_LT_Ratio_GbB': {'SWBin': 63, 'HWBin': 4, 'Priority': 359, 'IsNan': 'MIPI2L_LT_Cap'},
            'MIPI2L_LT_Ratio_GbGr': {'SWBin': 63, 'HWBin': 4, 'Priority': 360, 'IsNan': 'MIPI2L_LT_Cap'},
            'MIPI2L_LT_LostBit': {'SWBin': 63, 'HWBin': 4, 'Priority': 361, 'IsNan': 'MIPI2L_LT_Cap'},
            'MIPI2L_LT_LostBitSNR_SumDIFF_R': {'SWBin': 63, 'HWBin': 4, 'Priority': 362, 'IsNan': 'MIPI2L_LT_Cap'},
            'MIPI2L_LT_LostBitSNR_SumDIFF_Gr': {'SWBin': 63, 'HWBin': 4, 'Priority': 363, 'IsNan': 'MIPI2L_LT_Cap'},
            'MIPI2L_LT_LostBitSNR_SumDIFF_Gb': {'SWBin': 63, 'HWBin': 4, 'Priority': 364, 'IsNan': 'MIPI2L_LT_Cap'},
            'MIPI2L_LT_LostBitSNR_SumDIFF_B': {'SWBin': 63, 'HWBin': 4, 'Priority': 365, 'IsNan': 'MIPI2L_LT_Cap','IsLast': 'MIPI2L_LT_Cap'},
            'MIPI4L27M30F_BK1X_Cap': {'SWBin': 96, 'HWBin': 8, 'Priority': 366, 'Status': []},
            'MIPI4L27M30F_BK1X_Calc': {'SWBin': 97, 'HWBin': 8, 'Priority': 367},
            'BK1X_BK_DeadRowExBPix_R': {'SWBin': 48, 'HWBin': 6, 'Priority': 368, 'IsNan': 'MIPI4L27M30F_BK1X_Cap'},
            'BK1X_BK_DeadRowExBPix_Gr': {'SWBin': 48, 'HWBin': 6, 'Priority': 369, 'IsNan': 'MIPI4L27M30F_BK1X_Cap'},
            'BK1X_BK_DeadRowExBPix_Gb': {'SWBin': 48, 'HWBin': 6, 'Priority': 370, 'IsNan': 'MIPI4L27M30F_BK1X_Cap'},
            'BK1X_BK_DeadRowExBPix_B': {'SWBin': 48, 'HWBin': 6, 'Priority': 371, 'IsNan': 'MIPI4L27M30F_BK1X_Cap'},
            'BK1X_BK_DeadColExBPix_R': {'SWBin': 47, 'HWBin': 6, 'Priority': 372, 'IsNan': 'MIPI4L27M30F_BK1X_Cap'},
            'BK1X_BK_DeadColExBPix_Gr': {'SWBin': 47, 'HWBin': 6, 'Priority': 373, 'IsNan': 'MIPI4L27M30F_BK1X_Cap'},
            'BK1X_BK_DeadColExBPix_Gb': {'SWBin': 47, 'HWBin': 6, 'Priority': 374, 'IsNan': 'MIPI4L27M30F_BK1X_Cap'},
            'BK1X_BK_DeadColExBPix_B': {'SWBin': 47, 'HWBin': 6, 'Priority': 375, 'IsNan': 'MIPI4L27M30F_BK1X_Cap'},
            'BK1X_BK_Mean_R': {'SWBin': 46, 'HWBin': 6, 'Priority': 376, 'IsNan': 'MIPI4L27M30F_BK1X_Cap'},
            'BK1X_BK_Mean_Gr': {'SWBin': 46, 'HWBin': 6, 'Priority': 377, 'IsNan': 'MIPI4L27M30F_BK1X_Cap'},
            'BK1X_BK_Mean_Gb': {'SWBin': 46, 'HWBin': 6, 'Priority': 378, 'IsNan': 'MIPI4L27M30F_BK1X_Cap'},
            'BK1X_BK_Mean_B': {'SWBin': 46, 'HWBin': 6, 'Priority': 379, 'IsNan': 'MIPI4L27M30F_BK1X_Cap'},
            'BK1X_BK_StdDEV_R': {'SWBin': 46, 'HWBin': 6, 'Priority': 380, 'IsNan': 'MIPI4L27M30F_BK1X_Cap'},
            'BK1X_BK_StdDEV_Gr': {'SWBin': 46, 'HWBin': 6, 'Priority': 381, 'IsNan': 'MIPI4L27M30F_BK1X_Cap'},
            'BK1X_BK_StdDEV_Gb': {'SWBin': 46, 'HWBin': 6, 'Priority': 382, 'IsNan': 'MIPI4L27M30F_BK1X_Cap'},
            'BK1X_BK_StdDEV_B': {'SWBin': 46, 'HWBin': 6, 'Priority': 383, 'IsNan': 'MIPI4L27M30F_BK1X_Cap','IsLast': 'MIPI4L27M30F_BK1X_Cap'},
            'MIPI4L27M30F_BK32X_Cap': {'SWBin': 96, 'HWBin': 8, 'Priority': 384, 'Status': []},
            'MIPI4L27M30F_BK32X_Calc': {'SWBin': 97, 'HWBin': 8, 'Priority': 385},
            'BK32X_BK_MixDCol_DashQty_R': {'SWBin': 54, 'HWBin': 6, 'Priority': 386, 'IsNan': 'MIPI4L27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DashQty_Gr': {'SWBin': 54, 'HWBin': 6, 'Priority': 387,
                                            'IsNan': 'MIPI4L27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DashQty_Gb': {'SWBin': 54, 'HWBin': 6, 'Priority': 388,
                                            'IsNan': 'MIPI4L27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DashQty_B': {'SWBin': 54, 'HWBin': 6, 'Priority': 389, 'IsNan': 'MIPI4L27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_DeltaAvg_R': {'SWBin': 54, 'HWBin': 6, 'Priority': 390,
                                                 'IsNan': 'MIPI4L27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_DeltaAvg_Gr': {'SWBin': 54, 'HWBin': 6, 'Priority': 391,
                                                  'IsNan': 'MIPI4L27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_DeltaAvg_Gb': {'SWBin': 54, 'HWBin': 6, 'Priority': 392,
                                                  'IsNan': 'MIPI4L27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_DeltaAvg_B': {'SWBin': 54, 'HWBin': 6, 'Priority': 393,
                                                 'IsNan': 'MIPI4L27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue0_R': {'SWBin': 54, 'HWBin': 6, 'Priority': 394,
                                                  'IsNan': 'MIPI4L27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue0_Gr': {'SWBin': 54, 'HWBin': 6, 'Priority': 395,
                                                   'IsNan': 'MIPI4L27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue0_Gb': {'SWBin': 54, 'HWBin': 6, 'Priority': 396,
                                                   'IsNan': 'MIPI4L27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue0_B': {'SWBin': 54, 'HWBin': 6, 'Priority': 397,
                                                  'IsNan': 'MIPI4L27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue1_R': {'SWBin': 54, 'HWBin': 6, 'Priority': 398,
                                                  'IsNan': 'MIPI4L27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue1_Gr': {'SWBin': 54, 'HWBin': 6, 'Priority': 399,
                                                   'IsNan': 'MIPI4L27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue1_Gb': {'SWBin': 54, 'HWBin': 6, 'Priority': 400,
                                                   'IsNan': 'MIPI4L27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue1_B': {'SWBin': 54, 'HWBin': 6, 'Priority': 401,
                                                  'IsNan': 'MIPI4L27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue2_R': {'SWBin': 54, 'HWBin': 6, 'Priority': 402,
                                                  'IsNan': 'MIPI4L27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue2_Gr': {'SWBin': 54, 'HWBin': 6, 'Priority': 403,
                                                   'IsNan': 'MIPI4L27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue2_Gb': {'SWBin': 54, 'HWBin': 6, 'Priority': 404,
                                                   'IsNan': 'MIPI4L27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue2_B': {'SWBin': 54, 'HWBin': 6, 'Priority': 405,
                                                  'IsNan': 'MIPI4L27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue3_R': {'SWBin': 54, 'HWBin': 6, 'Priority': 406,
                                                  'IsNan': 'MIPI4L27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue3_Gr': {'SWBin': 54, 'HWBin': 6, 'Priority': 407,
                                                   'IsNan': 'MIPI4L27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue3_Gb': {'SWBin': 54, 'HWBin': 6, 'Priority': 408,
                                                   'IsNan': 'MIPI4L27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue3_B': {'SWBin': 54, 'HWBin': 6, 'Priority': 409,
                                                  'IsNan': 'MIPI4L27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue4_R': {'SWBin': 54, 'HWBin': 6, 'Priority': 410,
                                                  'IsNan': 'MIPI4L27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue4_Gr': {'SWBin': 54, 'HWBin': 6, 'Priority': 411,
                                                   'IsNan': 'MIPI4L27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue4_Gb': {'SWBin': 54, 'HWBin': 6, 'Priority': 412,
                                                   'IsNan': 'MIPI4L27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue4_B': {'SWBin': 54, 'HWBin': 6, 'Priority': 413,
                                                  'IsNan': 'MIPI4L27M30F_BK32X_Cap'},
            'BK32X_BK_Mean_R': {'SWBin': 53, 'HWBin': 6, 'Priority': 415, 'IsNan': 'MIPI4L27M30F_BK32X_Cap'},
            'BK32X_BK_Mean_Gr': {'SWBin': 53, 'HWBin': 6, 'Priority': 416, 'IsNan': 'MIPI4L27M30F_BK32X_Cap'},
            'BK32X_BK_Mean_Gb': {'SWBin': 53, 'HWBin': 6, 'Priority': 417, 'IsNan': 'MIPI4L27M30F_BK32X_Cap'},
            'BK32X_BK_Mean_B': {'SWBin': 53, 'HWBin': 6, 'Priority': 418, 'IsNan': 'MIPI4L27M30F_BK32X_Cap'},
            'BK32X_BK_StdDEV_R': {'SWBin': 53, 'HWBin': 6, 'Priority': 419, 'IsNan': 'MIPI4L27M30F_BK32X_Cap'},
            'BK32X_BK_StdDEV_Gr': {'SWBin': 53, 'HWBin': 6, 'Priority': 420, 'IsNan': 'MIPI4L27M30F_BK32X_Cap'},
            'BK32X_BK_StdDEV_Gb': {'SWBin': 53, 'HWBin': 6, 'Priority': 421, 'IsNan': 'MIPI4L27M30F_BK32X_Cap'},
            'BK32X_BK_StdDEV_B': {'SWBin': 53, 'HWBin': 6, 'Priority': 422, 'IsNan': 'MIPI4L27M30F_BK32X_Cap','IsLast': 'MIPI4L27M30F_BK32X_Cap'},
            'Binning': {'SWBin': 2, 'HWBin': 3, 'Priority': 423},
            'All Pass': {'SWBin': 255, 'HWBin': 3, 'Priority': 424}
        },
    'JX832':
        {
            'VSYNC_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 1},
            'HSYNC_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 2},
            'PCLK_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 3},
            'EXCLK_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 4},
            'RSTB_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 5},
            'PWDN_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 6},
            'SDA_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 7},
            'SCL_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 8},
            'D0_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 9},
            'D1_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 10},
            'D2_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 11},
            'D3_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 12},
            'D4_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 13},
            'D5_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 14},
            'D6_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 15},
            'D7_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 16},
            'D8_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 17},
            'D9_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 18},
            'AVDD_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 19, 'Limit': {'L': -2, 'H': 2}},
            'DOVDD_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 20},
            'VRamp_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 21},
            'VH_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 22},
            'VN1_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 23},
            'VSYNC_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 24},
            'HSYNC_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 25},
            'PCLK_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 26},
            'EXCLK_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 27},
            'RSTB_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 28},
            'PWDN_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 29},
            'SDA_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 30},
            'SCL_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 31},
            'D0_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 32},
            'D1_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 33},
            'D2_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 34},
            'D3_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 35},
            'D4_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 36},
            'D5_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 37},
            'D6_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 38},
            'D7_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 39},
            'D8_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 40},
            'D9_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 41},
            'VSYNC_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 42},
            'HSYNC_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 43},
            'PCLK_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 44},
            'EXCLK_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 45},
            'RSTB_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 46},
            'PWDN_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 47},
            'SDA_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 48},
            'SCL_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 49},
            'D0_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 50},
            'D1_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 51},
            'D2_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 52},
            'D3_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 53},
            'D4_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 54},
            'D5_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 55},
            'D6_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 56},
            'D7_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 57},
            'D8_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 58},
            'D9_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 59},
            'iic_test': {'SWBin': 7, 'HWBin': 5, 'Priority': 60},
            'ChipVer': {'SWBin': 7, 'HWBin': 5, 'Priority': 61},
            'DVDD_voltage': {'SWBin': 9, 'HWBin': 5, 'Priority': 62},
            'VH_voltage': {'SWBin': 9, 'HWBin': 5, 'Priority': 63},
            'VN1_voltage': {'SWBin': 9, 'HWBin': 5, 'Priority': 64},
            'Active_AVDD': {'SWBin': 12, 'HWBin': 5, 'Priority': 65},
            'Active_DOVDD': {'SWBin': 12, 'HWBin': 5, 'Priority': 66},
            'PWDN_AVDD': {'SWBin': 8, 'HWBin': 5, 'Priority': 67},
            'PWDN_DOVDD': {'SWBin': 8, 'HWBin': 5, 'Priority': 68},
            'PWDN_Total': {'SWBin': 8, 'HWBin': 5, 'Priority': 69},
            'DVP27M30F_BK_Cap': {'SWBin': 98, 'HWBin': 5, 'Priority': 70, 'Status': []},
            'DVP27M30F_BK_Calc': {'SWBin': 97, 'HWBin': 5, 'Priority': 71},
            'BLC_R': {'SWBin': 7, 'HWBin': 5, 'Priority': 72},
            'BLC_Gr': {'SWBin': 7, 'HWBin': 5, 'Priority': 73},
            'BLC_Gb': {'SWBin': 7, 'HWBin': 5, 'Priority': 74},
            'BLC_B': {'SWBin': 7, 'HWBin': 5, 'Priority': 75},
            'BK_DeadRowExBPix_R': {'SWBin': 15, 'HWBin': 6, 'Priority': 76, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_DeadRowExBPix_Gr': {'SWBin': 15, 'HWBin': 6, 'Priority': 77, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_DeadRowExBPix_Gb': {'SWBin': 15, 'HWBin': 6, 'Priority': 78, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_DeadRowExBPix_B': {'SWBin': 15, 'HWBin': 6, 'Priority': 79, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_DeadColExBPix_R': {'SWBin': 14, 'HWBin': 6, 'Priority': 80, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_DeadColExBPix_Gr': {'SWBin': 14, 'HWBin': 6, 'Priority': 81, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_DeadColExBPix_Gb': {'SWBin': 14, 'HWBin': 6, 'Priority': 82, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_DeadColExBPix_B': {'SWBin': 14, 'HWBin': 6, 'Priority': 83, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VfpnQty_R': {'SWBin': 14, 'HWBin': 6, 'Priority': 84, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VfpnQty_Gr': {'SWBin': 14, 'HWBin': 6, 'Priority': 85, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VfpnQty_Gb': {'SWBin': 14, 'HWBin': 6, 'Priority': 86, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VfpnQty_B': {'SWBin': 14, 'HWBin': 6, 'Priority': 87, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue0_R': {'SWBin': 14, 'HWBin': 6, 'Priority': 88, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue0_Gr': {'SWBin': 14, 'HWBin': 6, 'Priority': 89, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue0_Gb': {'SWBin': 14, 'HWBin': 6, 'Priority': 90, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue0_B': {'SWBin': 14, 'HWBin': 6, 'Priority': 91, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue1_R': {'SWBin': 14, 'HWBin': 6, 'Priority': 92, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue1_Gr': {'SWBin': 14, 'HWBin': 6, 'Priority': 93, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue1_Gb': {'SWBin': 14, 'HWBin': 6, 'Priority': 94, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue1_B': {'SWBin': 14, 'HWBin': 6, 'Priority': 95, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue2_R': {'SWBin': 14, 'HWBin': 6, 'Priority': 96, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue2_Gr': {'SWBin': 14, 'HWBin': 6, 'Priority': 97, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue2_Gb': {'SWBin': 14, 'HWBin': 6, 'Priority': 98, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue2_B': {'SWBin': 14, 'HWBin': 6, 'Priority': 99, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue3_R': {'SWBin': 14, 'HWBin': 6, 'Priority': 100, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue3_Gr': {'SWBin': 14, 'HWBin': 6, 'Priority': 101, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue3_Gb': {'SWBin': 14, 'HWBin': 6, 'Priority': 102, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue3_B': {'SWBin': 14, 'HWBin': 6, 'Priority': 103, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue4_R': {'SWBin': 14, 'HWBin': 6, 'Priority': 104, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue4_Gr': {'SWBin': 14, 'HWBin': 6, 'Priority': 105, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue4_Gb': {'SWBin': 14, 'HWBin': 6, 'Priority': 106, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_MixDCol_VFPN_MaxValue4_B': {'SWBin': 14, 'HWBin': 6, 'Priority': 107, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_Mean_R': {'SWBin': 13, 'HWBin': 6, 'Priority': 108, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_Mean_Gr': {'SWBin': 13, 'HWBin': 6, 'Priority': 109, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_Mean_Gb': {'SWBin': 13, 'HWBin': 6, 'Priority': 110, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_Mean_B': {'SWBin': 13, 'HWBin': 6, 'Priority': 111, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_StdDEV_R': {'SWBin': 13, 'HWBin': 6, 'Priority': 112, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_StdDEV_Gr': {'SWBin': 13, 'HWBin': 6, 'Priority': 113, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_StdDEV_Gb': {'SWBin': 13, 'HWBin': 6, 'Priority': 114, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_StdDEV_B': {'SWBin': 13, 'HWBin': 6, 'Priority': 115, 'IsNan': 'DVP27M30F_BK_Cap',
                            'IsLast': 'DVP27M30F_BK_Calc'},
            'DVP27M30F_LT_Cap': {'SWBin': 99, 'HWBin': 5, 'Priority': 116, 'Status': []},
            'DVP27M30F_LT_Calc': {'SWBin': 97, 'HWBin': 5, 'Priority': 117},
            'PLCK_Freq': {'SWBin': 93, 'HWBin': 5, 'Priority': 118},
            'LT_DRow_R': {'SWBin': 25, 'HWBin': 4, 'Priority': 119, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_DRow_Gr': {'SWBin': 25, 'HWBin': 4, 'Priority': 120, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_DRow_Gb': {'SWBin': 25, 'HWBin': 4, 'Priority': 121, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_DRow_B': {'SWBin': 25, 'HWBin': 4, 'Priority': 122, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_DCol_R': {'SWBin': 24, 'HWBin': 4, 'Priority': 123, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_DCol_Gr': {'SWBin': 24, 'HWBin': 4, 'Priority': 124, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_DCol_Gb': {'SWBin': 24, 'HWBin': 4, 'Priority': 125, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_DCol_B': {'SWBin': 24, 'HWBin': 4, 'Priority': 126, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_DRow_Color': {'SWBin': 25, 'HWBin': 4, 'Priority': 127, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_DCol_Color': {'SWBin': 24, 'HWBin': 4, 'Priority': 128, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_WeakLineRow_R': {'SWBin': 25, 'HWBin': 4, 'Priority': 129, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_WeakLineRow_Gr': {'SWBin': 25, 'HWBin': 4, 'Priority': 130, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_WeakLineRow_Gb': {'SWBin': 25, 'HWBin': 4, 'Priority': 131, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_WeakLineRow_B': {'SWBin': 25, 'HWBin': 4, 'Priority': 132, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_WeakLineCol_R': {'SWBin': 24, 'HWBin': 4, 'Priority': 133, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_WeakLineCol_Gr': {'SWBin': 24, 'HWBin': 4, 'Priority': 134, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_WeakLineCol_Gb': {'SWBin': 24, 'HWBin': 4, 'Priority': 135, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_WeakLineCol_B': {'SWBin': 24, 'HWBin': 4, 'Priority': 136, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_Mean_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 137, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_Mean_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 138, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_Mean_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 139, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_Mean_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 140, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_StdDEV_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 141, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_StdDEV_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 142, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_StdDEV_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 143, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_StdDEV_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 144, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_RI_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 145, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_RI_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 146, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_RI_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 147, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_RI_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 148, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_Ratio_GrR': {'SWBin': 23, 'HWBin': 4, 'Priority': 149, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_Ratio_GbR': {'SWBin': 23, 'HWBin': 4, 'Priority': 150, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_Ratio_GrB': {'SWBin': 23, 'HWBin': 4, 'Priority': 151, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_Ratio_GbB': {'SWBin': 23, 'HWBin': 4, 'Priority': 152, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_Ratio_GbGr': {'SWBin': 23, 'HWBin': 4, 'Priority': 153, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_LostBit': {'SWBin': 23, 'HWBin': 4, 'Priority': 154, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_LostBitSNR_SumDIFF_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 155, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_LostBitSNR_SumDIFF_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 156, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_LostBitSNR_SumDIFF_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 157, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_LostBitSNR_SumDIFF_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 158, 'IsNan': 'DVP27M30F_LT_Cap',
                                        'IsLast': 'DVP27M30F_LT_Calc'},
            'DVP27M30F_LBIT_Cap': {'SWBin': 99, 'HWBin': 5, 'Priority': 159, 'Status': []},
            'DVP27M30F_LBIT_Calc': {'SWBin': 97, 'HWBin': 5, 'Priority': 160},
            'LBIT_LT_Mean_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 161, 'IsNan': 'DVP27M30F_LBIT_Cap'},
            'LBIT_LT_Mean_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 162, 'IsNan': 'DVP27M30F_LBIT_Cap'},
            'LBIT_LT_Mean_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 163, 'IsNan': 'DVP27M30F_LBIT_Cap'},
            'LBIT_LT_Mean_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 164, 'IsNan': 'DVP27M30F_LBIT_Cap'},
            'LBIT_LT_LostBit': {'SWBin': 23, 'HWBin': 4, 'Priority': 165, 'IsNan': 'DVP27M30F_LBIT_Cap'},
            'LBIT_LT_LostBitSNR_SumDIFF_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 166, 'IsNan': 'DVP27M30F_LBIT_Cap'},
            'LBIT_LT_LostBitSNR_SumDIFF_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 167, 'IsNan': 'DVP27M30F_LBIT_Cap'},
            'LBIT_LT_LostBitSNR_SumDIFF_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 168, 'IsNan': 'DVP27M30F_LBIT_Cap'},
            'LBIT_LT_LostBitSNR_SumDIFF_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 169, 'IsNan': 'DVP27M30F_LBIT_Cap',
                                             'IsLast': 'DVP27M30F_LBIT_Calc'},
            'DVP27M30F_FW_Cap': {'SWBin': 99, 'HWBin': 5, 'Priority': 170, 'Status': []},
            'DVP27M30F_FW_Calc': {'SWBin': 97, 'HWBin': 5, 'Priority': 171},
            'FW_LT_DRow_R': {'SWBin': 36, 'HWBin': 4, 'Priority': 172, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_DRow_Gr': {'SWBin': 36, 'HWBin': 4, 'Priority': 173, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_DRow_Gb': {'SWBin': 36, 'HWBin': 4, 'Priority': 174, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_DRow_B': {'SWBin': 36, 'HWBin': 4, 'Priority': 175, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_DCol_R': {'SWBin': 36, 'HWBin': 4, 'Priority': 176, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_DCol_Gr': {'SWBin': 36, 'HWBin': 4, 'Priority': 177, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_DCol_Gb': {'SWBin': 36, 'HWBin': 4, 'Priority': 178, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_DCol_B': {'SWBin': 36, 'HWBin': 4, 'Priority': 179, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_DRow_Color': {'SWBin': 36, 'HWBin': 4, 'Priority': 180, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_DCol_Color': {'SWBin': 36, 'HWBin': 4, 'Priority': 181, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_Mean_R': {'SWBin': 36, 'HWBin': 4, 'Priority': 182, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_Mean_Gr': {'SWBin': 36, 'HWBin': 4, 'Priority': 183, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_Mean_Gb': {'SWBin': 36, 'HWBin': 4, 'Priority': 184, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_Mean_B': {'SWBin': 36, 'HWBin': 4, 'Priority': 185, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_StdDEV_R': {'SWBin': 36, 'HWBin': 4, 'Priority': 186, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_StdDEV_Gr': {'SWBin': 36, 'HWBin': 4, 'Priority': 187, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_StdDEV_Gb': {'SWBin': 36, 'HWBin': 4, 'Priority': 188, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_StdDEV_B': {'SWBin': 36, 'HWBin': 4, 'Priority': 189, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_RI_R': {'SWBin': 36, 'HWBin': 4, 'Priority': 190, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_RI_Gr': {'SWBin': 36, 'HWBin': 4, 'Priority': 191, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_RI_Gb': {'SWBin': 36, 'HWBin': 4, 'Priority': 192, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_RI_B': {'SWBin': 36, 'HWBin': 4, 'Priority': 193, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_Ratio_GrR': {'SWBin': 36, 'HWBin': 4, 'Priority': 194, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_Ratio_GbR': {'SWBin': 36, 'HWBin': 4, 'Priority': 195, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_Ratio_GrB': {'SWBin': 36, 'HWBin': 4, 'Priority': 196, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_Ratio_GbB': {'SWBin': 36, 'HWBin': 4, 'Priority': 197, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_Ratio_GbGr': {'SWBin': 36, 'HWBin': 4, 'Priority': 198, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_LostBit': {'SWBin': 36, 'HWBin': 4, 'Priority': 199, 'IsNan': 'DVP27M30F_FW_Cap',
                              'IsLast': 'DVP27M30F_FW_Calc'},
            'DVP27M30F_BSun_Cap': {'SWBin': 99, 'HWBin': 5, 'Priority': 200, 'Status': []},
            'DVP27M30F_BSun_Calc': {'SWBin': 97, 'HWBin': 5, 'Priority': 201},
            'One_BKSN_DROW_R': {'SWBin': 48, 'HWBin': 6, 'Priority': 202, 'IsNan': 'DVP27M30F_BSun_Cap'},
            'One_BKSN_DROW_Gr': {'SWBin': 48, 'HWBin': 6, 'Priority': 203, 'IsNan': 'DVP27M30F_BSun_Cap'},
            'One_BKSN_DROW_Gb': {'SWBin': 48, 'HWBin': 6, 'Priority': 204, 'IsNan': 'DVP27M30F_BSun_Cap'},
            'One_BKSN_DROW_B': {'SWBin': 48, 'HWBin': 6, 'Priority': 205, 'IsNan': 'DVP27M30F_BSun_Cap'},
            'One_BKSN_DCOL_R': {'SWBin': 47, 'HWBin': 6, 'Priority': 206, 'IsNan': 'DVP27M30F_BSun_Cap'},
            'One_BKSN_DCOL_Gr': {'SWBin': 47, 'HWBin': 6, 'Priority': 207, 'IsNan': 'DVP27M30F_BSun_Cap'},
            'One_BKSN_DCOL_Gb': {'SWBin': 47, 'HWBin': 6, 'Priority': 208, 'IsNan': 'DVP27M30F_BSun_Cap'},
            'One_BKSN_DCOL_B': {'SWBin': 47, 'HWBin': 6, 'Priority': 209, 'IsNan': 'DVP27M30F_BSun_Cap'},
            'One_BKSN_MEAN_R': {'SWBin': 46, 'HWBin': 6, 'Priority': 210, 'IsNan': 'DVP27M30F_BSun_Cap'},
            'One_BKSN_MEAN_Gr': {'SWBin': 46, 'HWBin': 6, 'Priority': 211, 'IsNan': 'DVP27M30F_BSun_Cap'},
            'One_BKSN_MEAN_Gb': {'SWBin': 46, 'HWBin': 6, 'Priority': 212, 'IsNan': 'DVP27M30F_BSun_Cap'},
            'One_BKSN_MEAN_B': {'SWBin': 46, 'HWBin': 6, 'Priority': 213, 'IsNan': 'DVP27M30F_BSun_Cap'},
            'One_BKSN_STDDEV_R': {'SWBin': 46, 'HWBin': 6, 'Priority': 214, 'IsNan': 'DVP27M30F_BSun_Cap'},
            'One_BKSN_STDDEV_Gr': {'SWBin': 46, 'HWBin': 6, 'Priority': 215, 'IsNan': 'DVP27M30F_BSun_Cap'},
            'One_BKSN_STDDEV_Gb': {'SWBin': 46, 'HWBin': 6, 'Priority': 216, 'IsNan': 'DVP27M30F_BSun_Cap'},
            'One_BKSN_STDDEV_B': {'SWBin': 46, 'HWBin': 6, 'Priority': 217, 'IsNan': 'DVP27M30F_BSun_Cap',
                                  'IsLast': 'DVP27M30F_BSun_Calc'},
            'LT_CornerLine': {'SWBin': 26, 'HWBin': 4, 'Priority': 218},
            'LT_ScratchLine': {'SWBin': 26, 'HWBin': 4, 'Priority': 219},
            'LT_Blemish': {'SWBin': 31, 'HWBin': 4, 'Priority': 220},
            'LT_LineStripe': {'SWBin': 27, 'HWBin': 4, 'Priority': 221},
            'LT_Particle': {'SWBin': 32, 'HWBin': 4, 'Priority': 222},
            'BK_Cluster2': {'SWBin': 30, 'HWBin': 4, 'Priority': 223},
            'LT_Cluster2': {'SWBin': 30, 'HWBin': 4, 'Priority': 224},
            'BK_Cluster1': {'SWBin': 29, 'HWBin': 4, 'Priority': 225},
            'LT_Cluster1': {'SWBin': 29, 'HWBin': 4, 'Priority': 226},
            'BK_Cluster3GrGb': {'SWBin': 39, 'HWBin': 4, 'Priority': 227},
            'LT_Cluster3GrGb': {'SWBin': 40, 'HWBin': 4, 'Priority': 228},
            'BK_Cluster3SubtractGrGb': {'SWBin': 33, 'HWBin': 4, 'Priority': 229},
            'LT_Cluster3SubtractGrGb': {'SWBin': 34, 'HWBin': 4, 'Priority': 230},
            'WP_Count': {'SWBin': 35, 'HWBin': 6, 'Priority': 231},
            'BK_Z1_BT_DP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 232},
            'BK_Z1_BT_DP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 233},
            'BK_Z1_BT_DP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 234},
            'BK_Z1_BT_DP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 235},
            'BK_Z2_BT_DP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 236},
            'BK_Z2_BT_DP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 237},
            'BK_Z2_BT_DP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 238},
            'BK_Z2_BT_DP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 239},
            'BK_Z1_BT_WP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 240},
            'BK_Z1_BT_WP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 241},
            'BK_Z1_BT_WP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 242},
            'BK_Z1_BT_WP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 243},
            'BK_Z2_BT_WP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 244},
            'BK_Z2_BT_WP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 245},
            'BK_Z2_BT_WP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 246},
            'BK_Z2_BT_WP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 247},
            'LT_Z1_BT_DP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 248},
            'LT_Z1_BT_DP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 249},
            'LT_Z1_BT_DP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 250},
            'LT_Z1_BT_DP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 251},
            'LT_Z2_BT_DP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 252},
            'LT_Z2_BT_DP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 253},
            'LT_Z2_BT_DP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 254},
            'LT_Z2_BT_DP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 255},
            'LT_Z1_BT_WP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 256},
            'LT_Z1_BT_WP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 257},
            'LT_Z1_BT_WP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 258},
            'LT_Z1_BT_WP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 259},
            'LT_Z2_BT_WP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 260},
            'LT_Z2_BT_WP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 261},
            'LT_Z2_BT_WP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 262},
            'LT_Z2_BT_WP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 263},
            'LT_Z1_DK_WP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 264},
            'LT_Z1_DK_WP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 265},
            'LT_Z1_DK_WP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 266},
            'LT_Z1_DK_WP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 267},
            'LT_Z2_DK_WP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 268},
            'LT_Z2_DK_WP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 269},
            'LT_Z2_DK_WP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 270},
            'LT_Z2_DK_WP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 271},
            'LT_Z1_DK_DP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 272},
            'LT_Z1_DK_DP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 273},
            'LT_Z1_DK_DP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 274},
            'LT_Z1_DK_DP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 275},
            'LT_Z2_DK_DP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 276},
            'LT_Z2_DK_DP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 277},
            'LT_Z2_DK_DP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 278},
            'LT_Z2_DK_DP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 279},
            'BK_Z1_BT_DP': {'SWBin': 60, 'HWBin': 6, 'Priority': 280},
            'BK_Z2_BT_DP': {'SWBin': 60, 'HWBin': 6, 'Priority': 281},
            'BK_Z1_BT_WP': {'SWBin': 60, 'HWBin': 6, 'Priority': 282},
            'BK_Z2_BT_WP': {'SWBin': 60, 'HWBin': 6, 'Priority': 283},
            'LT_Z1_BT_DP': {'SWBin': 60, 'HWBin': 6, 'Priority': 284},
            'LT_Z2_BT_DP': {'SWBin': 60, 'HWBin': 6, 'Priority': 285},
            'LT_Z1_BT_WP': {'SWBin': 60, 'HWBin': 6, 'Priority': 286},
            'LT_Z2_BT_WP': {'SWBin': 60, 'HWBin': 6, 'Priority': 287},
            'LT_Z1_DK_DP': {'SWBin': 60, 'HWBin': 6, 'Priority': 288},
            'LT_Z2_DK_DP': {'SWBin': 60, 'HWBin': 6, 'Priority': 289},
            'LT_Z1_DK_WP': {'SWBin': 60, 'HWBin': 6, 'Priority': 290},
            'LT_Z2_DK_WP': {'SWBin': 60, 'HWBin': 6, 'Priority': 291},
            'DVP27M30F_BK32X_Cap': {'SWBin': 98, 'HWBin': 5, 'Priority': 292, 'Status': []},
            'DVP27M30F_BK32X_Calc': {'SWBin': 97, 'HWBin': 5, 'Priority': 293},
            'BK32X_BK_BadPixel_Area': {'SWBin': 78, 'HWBin': 4, 'Priority': 294, 'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DashQty_R': {'SWBin': 51, 'HWBin': 6, 'Priority': 295, 'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DashQty_Gr': {'SWBin': 51, 'HWBin': 6, 'Priority': 296, 'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DashQty_Gb': {'SWBin': 51, 'HWBin': 6, 'Priority': 297, 'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DashQty_B': {'SWBin': 51, 'HWBin': 6, 'Priority': 298, 'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_SP1_BK_MixDCol_DashQty_R': {'SWBin': 54, 'HWBin': 2, 'Priority': 299,
                                               'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_SP1_BK_MixDCol_DashQty_Gr': {'SWBin': 54, 'HWBin': 2, 'Priority': 300,
                                                'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_SP1_BK_MixDCol_DashQty_Gb': {'SWBin': 54, 'HWBin': 2, 'Priority': 301,
                                                'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_SP1_BK_MixDCol_DashQty_B': {'SWBin': 54, 'HWBin': 2, 'Priority': 302,
                                               'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_DeltaAvg_R': {'SWBin': 54, 'HWBin': 2, 'Priority': 303,
                                                 'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_DeltaAvg_Gr': {'SWBin': 54, 'HWBin': 2, 'Priority': 304,
                                                  'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_DeltaAvg_Gb': {'SWBin': 54, 'HWBin': 2, 'Priority': 305,
                                                  'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_DeltaAvg_B': {'SWBin': 54, 'HWBin': 2, 'Priority': 306,
                                                 'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue0_R': {'SWBin': 54, 'HWBin': 2, 'Priority': 307,
                                                  'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue0_Gr': {'SWBin': 54, 'HWBin': 2, 'Priority': 308,
                                                   'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue0_Gb': {'SWBin': 54, 'HWBin': 2, 'Priority': 309,
                                                   'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue0_B': {'SWBin': 54, 'HWBin': 2, 'Priority': 310,
                                                  'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue1_R': {'SWBin': 54, 'HWBin': 2, 'Priority': 311,
                                                  'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue1_Gr': {'SWBin': 54, 'HWBin': 2, 'Priority': 312,
                                                   'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue1_Gb': {'SWBin': 54, 'HWBin': 2, 'Priority': 313,
                                                   'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue1_B': {'SWBin': 54, 'HWBin': 2, 'Priority': 314,
                                                  'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue2_R': {'SWBin': 54, 'HWBin': 2, 'Priority': 315,
                                                  'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue2_Gr': {'SWBin': 54, 'HWBin': 2, 'Priority': 316,
                                                   'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue2_Gb': {'SWBin': 54, 'HWBin': 2, 'Priority': 317,
                                                   'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue2_B': {'SWBin': 54, 'HWBin': 2, 'Priority': 318,
                                                  'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue3_R': {'SWBin': 54, 'HWBin': 2, 'Priority': 319,
                                                  'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue3_Gr': {'SWBin': 54, 'HWBin': 2, 'Priority': 320,
                                                   'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue3_Gb': {'SWBin': 54, 'HWBin': 2, 'Priority': 321,
                                                   'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue3_B': {'SWBin': 54, 'HWBin': 2, 'Priority': 322,
                                                  'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue4_R': {'SWBin': 54, 'HWBin': 2, 'Priority': 323,
                                                  'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue4_Gr': {'SWBin': 54, 'HWBin': 2, 'Priority': 324,
                                                   'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue4_Gb': {'SWBin': 54, 'HWBin': 2, 'Priority': 325,
                                                   'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DASH_MaxValue4_B': {'SWBin': 54, 'HWBin': 2, 'Priority': 326,
                                                  'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_Mean_R': {'SWBin': 53, 'HWBin': 2, 'Priority': 327, 'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_Mean_Gr': {'SWBin': 53, 'HWBin': 2, 'Priority': 328, 'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_Mean_Gb': {'SWBin': 53, 'HWBin': 2, 'Priority': 329, 'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_Mean_B': {'SWBin': 53, 'HWBin': 2, 'Priority': 330, 'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_StdDEV_R': {'SWBin': 53, 'HWBin': 2, 'Priority': 331, 'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_StdDEV_Gr': {'SWBin': 53, 'HWBin': 2, 'Priority': 332, 'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_StdDEV_Gb': {'SWBin': 53, 'HWBin': 2, 'Priority': 333, 'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_StdDEV_B': {'SWBin': 53, 'HWBin': 2, 'Priority': 334, 'IsNan': 'DVP27M30F_BK32X_Cap',
                                  'IsLast': 'DVP27M30F_BK32X_Calc'},
            'BK_BadPixel_Area': {'SWBin': 73, 'HWBin': 2, 'Priority': 335},
            'LT_BadPixel_Area': {'SWBin': 74, 'HWBin': 2, 'Priority': 336},
            'MIPI1L27M30F_WK1_Cap': {'SWBin': 89, 'HWBin': 2, 'Priority': 337, 'Status': []},
            'MIPI1L27M30F_WK1_Calc': {'SWBin': 97, 'HWBin': 5, 'Priority': 338},
            'MIPI1L27M30F_One_W1_FixedPattern': {'SWBin': 94, 'HWBin': 2, 'Priority': 339,
                                                 'IsNan': 'MIPI1L27M30F_WK1_Cap', 'IsLast': 'MIPI1L27M30F_WK1_Calc'},
            'MIPI1L27M30F_LT_Cap': {'SWBin': 90, 'HWBin': 2, 'Priority': 340, 'Status': []},
            'MIPI1L27M30F_LT_Calc': {'SWBin': 97, 'HWBin': 5, 'Priority': 341},
            'MIPI1L27M30F_LT_DRow_R': {'SWBin': 65, 'HWBin': 2, 'Priority': 342, 'IsNan': 'MIPI1L27M30F_LT_Cap'},
            'MIPI1L27M30F_LT_DRow_Gr': {'SWBin': 65, 'HWBin': 2, 'Priority': 343, 'IsNan': 'MIPI1L27M30F_LT_Cap'},
            'MIPI1L27M30F_LT_DRow_Gb': {'SWBin': 65, 'HWBin': 2, 'Priority': 344, 'IsNan': 'MIPI1L27M30F_LT_Cap'},
            'MIPI1L27M30F_LT_DRow_B': {'SWBin': 65, 'HWBin': 2, 'Priority': 345, 'IsNan': 'MIPI1L27M30F_LT_Cap'},
            'MIPI1L27M30F_LT_DCol_R': {'SWBin': 64, 'HWBin': 2, 'Priority': 346, 'IsNan': 'MIPI1L27M30F_LT_Cap'},
            'MIPI1L27M30F_LT_DCol_Gr': {'SWBin': 64, 'HWBin': 2, 'Priority': 347, 'IsNan': 'MIPI1L27M30F_LT_Cap'},
            'MIPI1L27M30F_LT_DCol_Gb': {'SWBin': 64, 'HWBin': 2, 'Priority': 348, 'IsNan': 'MIPI1L27M30F_LT_Cap'},
            'MIPI1L27M30F_LT_DCol_B': {'SWBin': 64, 'HWBin': 2, 'Priority': 349, 'IsNan': 'MIPI1L27M30F_LT_Cap'},
            'MIPI1L27M30F_LT_DRow_Color': {'SWBin': 65, 'HWBin': 2, 'Priority': 350, 'IsNan': 'MIPI1L27M30F_LT_Cap'},
            'MIPI1L27M30F_LT_DCol_Color': {'SWBin': 64, 'HWBin': 2, 'Priority': 351, 'IsNan': 'MIPI1L27M30F_LT_Cap'},
            'MIPI1L27M30F_LT_WeakLineRow_R': {'SWBin': 65, 'HWBin': 2, 'Priority': 352, 'IsNan': 'MIPI1L27M30F_LT_Cap'},
            'MIPI1L27M30F_LT_WeakLineRow_Gr': {'SWBin': 65, 'HWBin': 2, 'Priority': 353,
                                               'IsNan': 'MIPI1L27M30F_LT_Cap'},
            'MIPI1L27M30F_LT_WeakLineRow_Gb': {'SWBin': 65, 'HWBin': 2, 'Priority': 354,
                                               'IsNan': 'MIPI1L27M30F_LT_Cap'},
            'MIPI1L27M30F_LT_WeakLineRow_B': {'SWBin': 65, 'HWBin': 2, 'Priority': 355, 'IsNan': 'MIPI1L27M30F_LT_Cap'},
            'MIPI1L27M30F_LT_WeakLineCol_R': {'SWBin': 64, 'HWBin': 2, 'Priority': 356, 'IsNan': 'MIPI1L27M30F_LT_Cap'},
            'MIPI1L27M30F_LT_WeakLineCol_Gr': {'SWBin': 64, 'HWBin': 2, 'Priority': 357,
                                               'IsNan': 'MIPI1L27M30F_LT_Cap'},
            'MIPI1L27M30F_LT_WeakLineCol_Gb': {'SWBin': 64, 'HWBin': 2, 'Priority': 358,
                                               'IsNan': 'MIPI1L27M30F_LT_Cap'},
            'MIPI1L27M30F_LT_WeakLineCol_B': {'SWBin': 64, 'HWBin': 2, 'Priority': 359, 'IsNan': 'MIPI1L27M30F_LT_Cap'},
            'MIPI1L27M30F_LT_Mean_R': {'SWBin': 63, 'HWBin': 2, 'Priority': 360, 'IsNan': 'MIPI1L27M30F_LT_Cap'},
            'MIPI1L27M30F_LT_Mean_Gr': {'SWBin': 63, 'HWBin': 2, 'Priority': 361, 'IsNan': 'MIPI1L27M30F_LT_Cap'},
            'MIPI1L27M30F_LT_Mean_Gb': {'SWBin': 63, 'HWBin': 2, 'Priority': 362, 'IsNan': 'MIPI1L27M30F_LT_Cap'},
            'MIPI1L27M30F_LT_Mean_B': {'SWBin': 63, 'HWBin': 2, 'Priority': 363, 'IsNan': 'MIPI1L27M30F_LT_Cap'},
            'MIPI1L27M30F_LT_StdDEV_R': {'SWBin': 63, 'HWBin': 2, 'Priority': 364, 'IsNan': 'MIPI1L27M30F_LT_Cap'},
            'MIPI1L27M30F_LT_StdDEV_Gr': {'SWBin': 63, 'HWBin': 2, 'Priority': 365, 'IsNan': 'MIPI1L27M30F_LT_Cap'},
            'MIPI1L27M30F_LT_StdDEV_Gb': {'SWBin': 63, 'HWBin': 2, 'Priority': 366, 'IsNan': 'MIPI1L27M30F_LT_Cap'},
            'MIPI1L27M30F_LT_StdDEV_B': {'SWBin': 63, 'HWBin': 2, 'Priority': 367, 'IsNan': 'MIPI1L27M30F_LT_Cap'},
            'MIPI1L27M30F_LT_RI_R': {'SWBin': 63, 'HWBin': 2, 'Priority': 368, 'IsNan': 'MIPI1L27M30F_LT_Cap'},
            'MIPI1L27M30F_LT_RI_Gr': {'SWBin': 63, 'HWBin': 2, 'Priority': 369, 'IsNan': 'MIPI1L27M30F_LT_Cap'},
            'MIPI1L27M30F_LT_RI_Gb': {'SWBin': 63, 'HWBin': 2, 'Priority': 370, 'IsNan': 'MIPI1L27M30F_LT_Cap'},
            'MIPI1L27M30F_LT_RI_B': {'SWBin': 63, 'HWBin': 2, 'Priority': 371, 'IsNan': 'MIPI1L27M30F_LT_Cap'},
            'MIPI1L27M30F_LT_Ratio_GrR': {'SWBin': 63, 'HWBin': 2, 'Priority': 372, 'IsNan': 'MIPI1L27M30F_LT_Cap'},
            'MIPI1L27M30F_LT_Ratio_GbR': {'SWBin': 63, 'HWBin': 2, 'Priority': 373, 'IsNan': 'MIPI1L27M30F_LT_Cap'},
            'MIPI1L27M30F_LT_Ratio_GrB': {'SWBin': 63, 'HWBin': 2, 'Priority': 374, 'IsNan': 'MIPI1L27M30F_LT_Cap'},
            'MIPI1L27M30F_LT_Ratio_GbB': {'SWBin': 63, 'HWBin': 2, 'Priority': 375, 'IsNan': 'MIPI1L27M30F_LT_Cap'},
            'MIPI1L27M30F_LT_Ratio_GbGr': {'SWBin': 63, 'HWBin': 2, 'Priority': 376, 'IsNan': 'MIPI1L27M30F_LT_Cap'},
            'MIPI1L27M30F_LT_LostBit': {'SWBin': 63, 'HWBin': 2, 'Priority': 377, 'IsNan': 'MIPI1L27M30F_LT_Cap'},
            'MIPI1L27M30F_LT_LostBitSNR_SumDIFF_R': {'SWBin': 63, 'HWBin': 2, 'Priority': 378,
                                                     'IsNan': 'MIPI1L27M30F_LT_Cap'},
            'MIPI1L27M30F_LT_LostBitSNR_SumDIFF_Gr': {'SWBin': 63, 'HWBin': 2, 'Priority': 379,
                                                      'IsNan': 'MIPI1L27M30F_LT_Cap'},
            'MIPI1L27M30F_LT_LostBitSNR_SumDIFF_Gb': {'SWBin': 63, 'HWBin': 2, 'Priority': 380,
                                                      'IsNan': 'MIPI1L27M30F_LT_Cap'},
            'MIPI1L27M30F_LT_LostBitSNR_SumDIFF_B': {'SWBin': 63, 'HWBin': 2, 'Priority': 381,
                                                     'IsNan': 'MIPI1L27M30F_LT_Cap', 'IsLast': 'MIPI1L27M30F_LT_Calc'},
            'DVP27M60F_LT_Cap': {'SWBin': 96, 'HWBin': 1, 'Priority': 382, 'Status': []},
            'DVP27M60F_LT_Calc': {'SWBin': 97, 'HWBin': 5, 'Priority': 383},
            'DVP27M60F_LT_DRow_R': {'SWBin': 58, 'HWBin': 1, 'Priority': 384, 'IsNan': 'DVP27M60F_LT_Cap'},
            'DVP27M60F_LT_DRow_Gr': {'SWBin': 58, 'HWBin': 1, 'Priority': 385, 'IsNan': 'DVP27M60F_LT_Cap'},
            'DVP27M60F_LT_DRow_Gb': {'SWBin': 58, 'HWBin': 1, 'Priority': 386, 'IsNan': 'DVP27M60F_LT_Cap'},
            'DVP27M60F_LT_DRow_B': {'SWBin': 58, 'HWBin': 1, 'Priority': 387, 'IsNan': 'DVP27M60F_LT_Cap'},
            'DVP27M60F_LT_DCol_R': {'SWBin': 57, 'HWBin': 1, 'Priority': 388, 'IsNan': 'DVP27M60F_LT_Cap'},
            'DVP27M60F_LT_DCol_Gr': {'SWBin': 57, 'HWBin': 1, 'Priority': 389, 'IsNan': 'DVP27M60F_LT_Cap'},
            'DVP27M60F_LT_DCol_Gb': {'SWBin': 57, 'HWBin': 1, 'Priority': 390, 'IsNan': 'DVP27M60F_LT_Cap'},
            'DVP27M60F_LT_DCol_B': {'SWBin': 57, 'HWBin': 1, 'Priority': 391, 'IsNan': 'DVP27M60F_LT_Cap'},
            'DVP27M60F_LT_DRow_Color': {'SWBin': 58, 'HWBin': 1, 'Priority': 392, 'IsNan': 'DVP27M60F_LT_Cap'},
            'DVP27M60F_LT_DCol_Color': {'SWBin': 57, 'HWBin': 1, 'Priority': 393, 'IsNan': 'DVP27M60F_LT_Cap'},
            'DVP27M60F_LT_WeakLineRow_R': {'SWBin': 58, 'HWBin': 1, 'Priority': 394, 'IsNan': 'DVP27M60F_LT_Cap'},
            'DVP27M60F_LT_WeakLineRow_Gr': {'SWBin': 58, 'HWBin': 1, 'Priority': 395, 'IsNan': 'DVP27M60F_LT_Cap'},
            'DVP27M60F_LT_WeakLineRow_Gb': {'SWBin': 58, 'HWBin': 1, 'Priority': 396, 'IsNan': 'DVP27M60F_LT_Cap'},
            'DVP27M60F_LT_WeakLineRow_B': {'SWBin': 58, 'HWBin': 1, 'Priority': 397, 'IsNan': 'DVP27M60F_LT_Cap'},
            'DVP27M60F_LT_WeakLineCol_R': {'SWBin': 57, 'HWBin': 1, 'Priority': 398, 'IsNan': 'DVP27M60F_LT_Cap'},
            'DVP27M60F_LT_WeakLineCol_Gr': {'SWBin': 57, 'HWBin': 1, 'Priority': 399, 'IsNan': 'DVP27M60F_LT_Cap'},
            'DVP27M60F_LT_WeakLineCol_Gb': {'SWBin': 57, 'HWBin': 1, 'Priority': 400, 'IsNan': 'DVP27M60F_LT_Cap'},
            'DVP27M60F_LT_WeakLineCol_B': {'SWBin': 57, 'HWBin': 1, 'Priority': 401, 'IsNan': 'DVP27M60F_LT_Cap'},
            'DVP27M60F_LT_Mean_R': {'SWBin': 56, 'HWBin': 1, 'Priority': 402, 'IsNan': 'DVP27M60F_LT_Cap'},
            'DVP27M60F_LT_Mean_Gr': {'SWBin': 56, 'HWBin': 1, 'Priority': 403, 'IsNan': 'DVP27M60F_LT_Cap'},
            'DVP27M60F_LT_Mean_Gb': {'SWBin': 56, 'HWBin': 1, 'Priority': 404, 'IsNan': 'DVP27M60F_LT_Cap'},
            'DVP27M60F_LT_Mean_B': {'SWBin': 56, 'HWBin': 1, 'Priority': 405, 'IsNan': 'DVP27M60F_LT_Cap'},
            'DVP27M60F_LT_StdDEV_R': {'SWBin': 56, 'HWBin': 1, 'Priority': 406, 'IsNan': 'DVP27M60F_LT_Cap'},
            'DVP27M60F_LT_StdDEV_Gr': {'SWBin': 56, 'HWBin': 1, 'Priority': 407, 'IsNan': 'DVP27M60F_LT_Cap'},
            'DVP27M60F_LT_StdDEV_Gb': {'SWBin': 56, 'HWBin': 1, 'Priority': 408, 'IsNan': 'DVP27M60F_LT_Cap'},
            'DVP27M60F_LT_StdDEV_B': {'SWBin': 56, 'HWBin': 1, 'Priority': 409, 'IsNan': 'DVP27M60F_LT_Cap'},
            'DVP27M60F_LT_RI_R': {'SWBin': 56, 'HWBin': 1, 'Priority': 410, 'IsNan': 'DVP27M60F_LT_Cap'},
            'DVP27M60F_LT_RI_Gr': {'SWBin': 56, 'HWBin': 1, 'Priority': 411, 'IsNan': 'DVP27M60F_LT_Cap'},
            'DVP27M60F_LT_RI_Gb': {'SWBin': 56, 'HWBin': 1, 'Priority': 412, 'IsNan': 'DVP27M60F_LT_Cap'},
            'DVP27M60F_LT_RI_B': {'SWBin': 56, 'HWBin': 1, 'Priority': 413, 'IsNan': 'DVP27M60F_LT_Cap'},
            'DVP27M60F_LT_Ratio_GrR': {'SWBin': 56, 'HWBin': 1, 'Priority': 414, 'IsNan': 'DVP27M60F_LT_Cap'},
            'DVP27M60F_LT_Ratio_GbR': {'SWBin': 56, 'HWBin': 1, 'Priority': 415, 'IsNan': 'DVP27M60F_LT_Cap'},
            'DVP27M60F_LT_Ratio_GrB': {'SWBin': 56, 'HWBin': 1, 'Priority': 416, 'IsNan': 'DVP27M60F_LT_Cap'},
            'DVP27M60F_LT_Ratio_GbB': {'SWBin': 56, 'HWBin': 1, 'Priority': 417, 'IsNan': 'DVP27M60F_LT_Cap'},
            'DVP27M60F_LT_Ratio_GbGr': {'SWBin': 56, 'HWBin': 1, 'Priority': 418, 'IsNan': 'DVP27M60F_LT_Cap'},
            'DVP27M60F_LT_LostBit': {'SWBin': 56, 'HWBin': 1, 'Priority': 419, 'IsNan': 'DVP27M60F_LT_Cap'},
            'DVP27M60F_LT_LostBitSNR_SumDIFF_R': {'SWBin': 56, 'HWBin': 1, 'Priority': 420,
                                                  'IsNan': 'DVP27M60F_LT_Cap'},
            'DVP27M60F_LT_LostBitSNR_SumDIFF_Gr': {'SWBin': 56, 'HWBin': 1, 'Priority': 421,
                                                   'IsNan': 'DVP27M60F_LT_Cap'},
            'DVP27M60F_LT_LostBitSNR_SumDIFF_Gb': {'SWBin': 56, 'HWBin': 1, 'Priority': 422,
                                                   'IsNan': 'DVP27M60F_LT_Cap'},
            'DVP27M60F_LT_LostBitSNR_SumDIFF_B': {'SWBin': 56, 'HWBin': 1, 'Priority': 423,
                                                  'IsNan': 'DVP27M60F_LT_Cap', 'IsLast': 'DVP27M60F_LT_Calc'},
            'MIPI1L27M60F_LT_Cap': {'SWBin': 96, 'HWBin': 1, 'Priority': 424, 'Status': []},
            'MIPI1L27M60F_LT_Calc': {'SWBin': 97, 'HWBin': 5, 'Priority': 425},
            'MIPI1L27M60F_LT_DRow_R': {'SWBin': 68, 'HWBin': 1, 'Priority': 426, 'IsNan': 'MIPI1L27M60F_LT_Cap'},
            'MIPI1L27M60F_LT_DRow_Gr': {'SWBin': 68, 'HWBin': 1, 'Priority': 427, 'IsNan': 'MIPI1L27M60F_LT_Cap'},
            'MIPI1L27M60F_LT_DRow_Gb': {'SWBin': 68, 'HWBin': 1, 'Priority': 428, 'IsNan': 'MIPI1L27M60F_LT_Cap'},
            'MIPI1L27M60F_LT_DRow_B': {'SWBin': 68, 'HWBin': 1, 'Priority': 429, 'IsNan': 'MIPI1L27M60F_LT_Cap'},
            'MIPI1L27M60F_LT_DCol_R': {'SWBin': 67, 'HWBin': 1, 'Priority': 430, 'IsNan': 'MIPI1L27M60F_LT_Cap'},
            'MIPI1L27M60F_LT_DCol_Gr': {'SWBin': 67, 'HWBin': 1, 'Priority': 431, 'IsNan': 'MIPI1L27M60F_LT_Cap'},
            'MIPI1L27M60F_LT_DCol_Gb': {'SWBin': 67, 'HWBin': 1, 'Priority': 432, 'IsNan': 'MIPI1L27M60F_LT_Cap'},
            'MIPI1L27M60F_LT_DCol_B': {'SWBin': 67, 'HWBin': 1, 'Priority': 433, 'IsNan': 'MIPI1L27M60F_LT_Cap'},
            'MIPI1L27M60F_LT_DRow_Color': {'SWBin': 68, 'HWBin': 1, 'Priority': 434, 'IsNan': 'MIPI1L27M60F_LT_Cap'},
            'MIPI1L27M60F_LT_DCol_Color': {'SWBin': 67, 'HWBin': 1, 'Priority': 435, 'IsNan': 'MIPI1L27M60F_LT_Cap'},
            'MIPI1L27M60F_LT_WeakLineRow_R': {'SWBin': 68, 'HWBin': 1, 'Priority': 436, 'IsNan': 'MIPI1L27M60F_LT_Cap'},
            'MIPI1L27M60F_LT_WeakLineRow_Gr': {'SWBin': 68, 'HWBin': 1, 'Priority': 437,
                                               'IsNan': 'MIPI1L27M60F_LT_Cap'},
            'MIPI1L27M60F_LT_WeakLineRow_Gb': {'SWBin': 68, 'HWBin': 1, 'Priority': 438,
                                               'IsNan': 'MIPI1L27M60F_LT_Cap'},
            'MIPI1L27M60F_LT_WeakLineRow_B': {'SWBin': 68, 'HWBin': 1, 'Priority': 439, 'IsNan': 'MIPI1L27M60F_LT_Cap'},
            'MIPI1L27M60F_LT_WeakLineCol_R': {'SWBin': 67, 'HWBin': 1, 'Priority': 440, 'IsNan': 'MIPI1L27M60F_LT_Cap'},
            'MIPI1L27M60F_LT_WeakLineCol_Gr': {'SWBin': 67, 'HWBin': 1, 'Priority': 441,
                                               'IsNan': 'MIPI1L27M60F_LT_Cap'},
            'MIPI1L27M60F_LT_WeakLineCol_Gb': {'SWBin': 67, 'HWBin': 1, 'Priority': 442,
                                               'IsNan': 'MIPI1L27M60F_LT_Cap'},
            'MIPI1L27M60F_LT_WeakLineCol_B': {'SWBin': 67, 'HWBin': 1, 'Priority': 443, 'IsNan': 'MIPI1L27M60F_LT_Cap'},
            'MIPI1L27M60F_LT_Mean_R': {'SWBin': 66, 'HWBin': 1, 'Priority': 444, 'IsNan': 'MIPI1L27M60F_LT_Cap'},
            'MIPI1L27M60F_LT_Mean_Gr': {'SWBin': 66, 'HWBin': 1, 'Priority': 445, 'IsNan': 'MIPI1L27M60F_LT_Cap'},
            'MIPI1L27M60F_LT_Mean_Gb': {'SWBin': 66, 'HWBin': 1, 'Priority': 446, 'IsNan': 'MIPI1L27M60F_LT_Cap'},
            'MIPI1L27M60F_LT_Mean_B': {'SWBin': 66, 'HWBin': 1, 'Priority': 447, 'IsNan': 'MIPI1L27M60F_LT_Cap'},
            'MIPI1L27M60F_LT_StdDEV_R': {'SWBin': 66, 'HWBin': 1, 'Priority': 448, 'IsNan': 'MIPI1L27M60F_LT_Cap'},
            'MIPI1L27M60F_LT_StdDEV_Gr': {'SWBin': 66, 'HWBin': 1, 'Priority': 449, 'IsNan': 'MIPI1L27M60F_LT_Cap'},
            'MIPI1L27M60F_LT_StdDEV_Gb': {'SWBin': 66, 'HWBin': 1, 'Priority': 450, 'IsNan': 'MIPI1L27M60F_LT_Cap'},
            'MIPI1L27M60F_LT_StdDEV_B': {'SWBin': 66, 'HWBin': 1, 'Priority': 451, 'IsNan': 'MIPI1L27M60F_LT_Cap'},
            'MIPI1L27M60F_LT_RI_R': {'SWBin': 66, 'HWBin': 1, 'Priority': 452, 'IsNan': 'MIPI1L27M60F_LT_Cap'},
            'MIPI1L27M60F_LT_RI_Gr': {'SWBin': 66, 'HWBin': 1, 'Priority': 453, 'IsNan': 'MIPI1L27M60F_LT_Cap'},
            'MIPI1L27M60F_LT_RI_Gb': {'SWBin': 66, 'HWBin': 1, 'Priority': 454, 'IsNan': 'MIPI1L27M60F_LT_Cap'},
            'MIPI1L27M60F_LT_RI_B': {'SWBin': 66, 'HWBin': 1, 'Priority': 455, 'IsNan': 'MIPI1L27M60F_LT_Cap'},
            'MIPI1L27M60F_LT_Ratio_GrR': {'SWBin': 66, 'HWBin': 1, 'Priority': 456, 'IsNan': 'MIPI1L27M60F_LT_Cap'},
            'MIPI1L27M60F_LT_Ratio_GbR': {'SWBin': 66, 'HWBin': 1, 'Priority': 457, 'IsNan': 'MIPI1L27M60F_LT_Cap'},
            'MIPI1L27M60F_LT_Ratio_GrB': {'SWBin': 66, 'HWBin': 1, 'Priority': 458, 'IsNan': 'MIPI1L27M60F_LT_Cap'},
            'MIPI1L27M60F_LT_Ratio_GbB': {'SWBin': 66, 'HWBin': 1, 'Priority': 459, 'IsNan': 'MIPI1L27M60F_LT_Cap'},
            'MIPI1L27M60F_LT_Ratio_GbGr': {'SWBin': 66, 'HWBin': 1, 'Priority': 460, 'IsNan': 'MIPI1L27M60F_LT_Cap'},
            'MIPI1L27M60F_LT_LostBit': {'SWBin': 66, 'HWBin': 1, 'Priority': 461, 'IsNan': 'MIPI1L27M60F_LT_Cap'},
            'MIPI1L27M60F_LT_LostBitSNR_SumDIFF_R': {'SWBin': 66, 'HWBin': 1, 'Priority': 462,
                                                     'IsNan': 'MIPI1L27M60F_LT_Cap'},
            'MIPI1L27M60F_LT_LostBitSNR_SumDIFF_Gr': {'SWBin': 66, 'HWBin': 1, 'Priority': 463,
                                                      'IsNan': 'MIPI1L27M60F_LT_Cap'},
            'MIPI1L27M60F_LT_LostBitSNR_SumDIFF_Gb': {'SWBin': 66, 'HWBin': 1, 'Priority': 464,
                                                      'IsNan': 'MIPI1L27M60F_LT_Cap'},
            'MIPI1L27M60F_LT_LostBitSNR_SumDIFF_B': {'SWBin': 66, 'HWBin': 1, 'Priority': 465,
                                                     'IsNan': 'MIPI1L27M60F_LT_Cap', 'IsLast': 'MIPI1L27M60F_LT_Calc'},
            'Binning': {'SWBin': 2, 'HWBin': 3, 'Priority': 466},
            'All Pass': {'SWBin': 1, 'HWBin': 3, 'Priority': 467}
        }
}


@time_calc
def parse_file(file, _signal, analysis_folder):
    """
    parse file
    get basic information,datacheck result,binningcheck result,nanchipno result and summary result
    """
    global current_row_count
    parse_result = {}
    global high_limit_row_num
    data = []
    # get file data
    with open(file) as f:
        csv_reader = reader(f)
        for row in csv_reader:
            data.append(row)

    # get row and column of last test item
    if project == 'F28':
        search_last_test_item = search_string(data, 'Full_Error')
    elif project == 'JX828':
        search_last_test_item = search_string(data, 'ChipVer')
    elif project == 'JX825':
        search_last_test_item = search_string(data, 'SRAM_0x56_Read')
    elif project == 'JX832':
        search_last_test_item = search_string(data, 'Full_Error')
    if not search_last_test_item:
        exit()
    else:
        test_item_row_num = search_last_test_item[0]
        last_test_item_column_num = search_last_test_item[1]
    high_limit_row_num = test_item_row_num + 2
    low_limit_row_num = test_item_row_num + 3

    # get row and column of last DC test item
    search_pwdn_total = search_string(data, 'PWDN_Total')
    if not search_pwdn_total:
        exit()
    else:
        pwdn_total_column_num = search_pwdn_total[1]

    # get row and column of Binning
    search_binning = search_string(data, 'Binning')
    if not search_binning:
        exit()
    else:
        binning_column_num = search_binning[1]

    # get column of iic_test
    search_iic_test = search_string(data, 'iic_test')
    if not search_iic_test:
        exit()
    else:
        iic_test_column_num = search_iic_test[1]

    # get row of first register
    first_register_row_num = len(data)
    for i in range(test_item_row_num + row_offset, len(data)):
        try:
            int(data[i][0])
        except:
            first_register_row_num = i
            break

    # fill '' to align with last test item col num
    for i in range(0, len(data)):
        add_count = last_test_item_column_num + 1 - len(data[i])
        if add_count > 0:
            for j in range(add_count):
                data[i].append('')

    # parse file name
    file_name = basename(file).split('.')[0]
    # calculate chip count
    chip_count = first_register_row_num - test_item_row_num - row_offset
    # get lot_no
    lot_no = data[5][1]

    if analysis_item in ('All', 'Data Check'):
        row_count = len(data)

        # get row number with null value
        exist_nan_row_list = []
        for i in range(test_item_row_num + row_offset, first_register_row_num):
            for j in range(binning_column_num + 1):
                if len(data[i][j].strip()) == 0:
                    exist_nan_row_list.append(i)
                    break

        data_check_result = []
        for row_num in range(row_count):
            row_data = []
            for column_num in range(last_test_item_column_num + 1):
                test_item = data[test_item_row_num][column_num].strip()
                value = data[row_num][column_num]
                if column_offset <= column_num <= binning_column_num and high_limit_row_num <= row_num <= low_limit_row_num:
                    # fill color of limit value is green
                    row_data.append((value, '008000'))
                elif column_num == 0 and row_num in exist_nan_row_list:
                    # fill color of first value of exist nan row is purple
                    row_data.append((value, 'A020F0'))
                elif column_offset <= column_num <= binning_column_num and test_item_row_num + row_offset <= row_num < first_register_row_num:
                    try:
                        value_convert = float(value)
                        high_limit_data = data[high_limit_row_num][column_num]
                        try:
                            high_limit = float(high_limit_data)
                        except:
                            high_limit = high_limit_data
                        low_limit_data = data[low_limit_row_num][column_num]
                        try:
                            low_limit = float(low_limit_data)
                        except:
                            low_limit = low_limit_data
                        if value_convert == int(value_convert):
                            value_convert = int(value_convert)
                        if high_limit == 'N' or (high_limit != 'N' and low_limit <= value_convert <= high_limit):
                            # Fill color of following scenario value is white
                            # 1.Limit is N
                            # 2.Limit is not N and data wihtin limit
                            row_data.append((value_convert, 'FFFFFF'))
                        elif 'Limit' in BIN_DEFINITION[project][test_item].keys():
                            if BIN_DEFINITION[project][test_item]['Limit']['L'] <= value_convert <= \
                                    BIN_DEFINITION[project][test_item]['Limit']['H']:
                                # Fill color is white
                                # bin_definitionList has limit key and data within limt
                                row_data.append((value_convert, 'FFFFFF'))
                            else:
                                # fill color of over limit is yellow
                                row_data.append((value_convert, 'FFFF00'))
                        else:
                            # fill color of over limit is red
                            row_data.append((value_convert, 'FF0000'))
                    except:
                        if not value.strip():
                            # fill color of '' is purple
                            row_data.append((value, 'A020F0'))
                        elif column_num == iic_test_column_num:
                            if value == '0':
                                # fill color of iic_test 0 is white
                                row_data.append((value_convert, 'FFFFFF'))
                            else:
                                # fill color of iic_test not 0 is yellow
                                row_data.append((value, 'FFFF00'))
                else:
                    # set '' when value is nan
                    if isinstance(value, float) and isnan(value):
                        value = ''
                    # fill color is white
                    row_data.append((value, 'FFFFFF'))
            data_check_result.append(row_data)
            current_row_count += 1
            if current_row_count != total_row_count:
                _signal.emit(str(current_row_count * 100 // total_row_count))

        data_check_file = join(analysis_folder, file_name + '_DataCheck_analysis' + now_time + '.xlsx')
        data_check_wb = Workbook()  # create file object
        data_check_sheet = data_check_wb.active  # get first sheet
        data_check_sheet.freeze_panes = 'E17'  # set freeze panes
        irow = 1
        for i in range(len(data_check_result)):
            for j in range(len(data_check_result[i])):
                data_check_sheet.cell(row=irow, column=j + 1).value = data_check_result[i][j][0]
                data_check_sheet.cell(row=irow, column=j + 1).fill = PatternFill(fill_type='solid',
                                                                                 fgColor=data_check_result[i][j][1])
                data_check_sheet.cell(row=irow, column=j + 1).border = border
            irow += 1
            current_row_count += 1
            if current_row_count != total_row_count:
                _signal.emit(str(current_row_count * 100 // total_row_count))
        data_check_wb.save(data_check_file)
    if analysis_item in ('All', 'Binning Check'):
        binning_check_result = []
        for row_num in range(test_item_row_num + row_offset, first_register_row_num):
            # set current_test_item is All Pass
            current_test_item = 'All Pass'
            bin_definition = deepcopy(BIN_DEFINITION[project])
            for column_num in range(column_offset, binning_column_num + 1):
                if row_num==559 and column_num==242:
                    print('')
                test_item = data[test_item_row_num][column_num].strip()
                high_limit_data = data[high_limit_row_num][column_num]
                low_limit_data = data[low_limit_row_num][column_num]
                try:
                    high_limit = float(high_limit_data)
                    low_limit = float(low_limit_data)
                    try:
                        value = data[row_num][column_num]
                        value_convert = float(value)
                        if value_convert == int(value_convert):
                            value_convert = int(value_convert)
                        if 'IsNan' in bin_definition[test_item].keys():
                            bin_definition[bin_definition[test_item]['IsNan']]['Status'].append('1')
                        if 'Limit' in bin_definition[test_item].keys():
                            if (value_convert < bin_definition[test_item]['Limit']['L'] or value_convert >
                                bin_definition[test_item]['Limit']['H']) and bin_definition[test_item]['Priority'] < \
                                    bin_definition[current_test_item]['Priority']:
                                # set current_test_item is test_item when test_item's priority less than current_test_item's priority
                                current_test_item = test_item
                        elif value_convert < low_limit or high_limit < value_convert:
                            # over limit
                            if bin_definition[test_item]['Priority'] < bin_definition[current_test_item]['Priority']:
                                # set current_test_item is test_item when test_item's priority less than current_test_item's priority
                                current_test_item = test_item
                    except:
                        # value is ''
                        if 'IsNan' in bin_definition[test_item].keys():
                            bin_definition[bin_definition[test_item]['IsNan']]['Status'].append('0')
                        if test_item in bin_definition.keys() and 'IsNan' not in bin_definition[
                            test_item].keys() and bin_definition[test_item]['Priority'] < \
                                bin_definition[current_test_item]['Priority']:
                            # set current_test_item is test_item
                            # when test_item in keys and InNan not in keys
                            # and test_item's priority less than current_test_item's priority
                            current_test_item = test_item
                except:
                    # limit is N or nan
                    try:
                        value = data[row_num][column_num]
                        value_convert = float(value)
                        if test_item in bin_definition.keys():
                            if 'IsNan' in bin_definition[test_item].keys():
                                bin_definition[bin_definition[test_item]['IsNan']]['Status'].append('1')
                            if test_item == 'iic_test' and value == '1':
                                if bin_definition[test_item]['Priority'] < bin_definition[current_test_item][
                                    'Priority']:
                                    # set current_test_item is bin_definition[test_item]['IsNan']
                                    # when bin_definition[test_item]['IsNan']'s priority less than current_test_item's priority
                                    current_test_item = test_item
                            elif 'Limit' in bin_definition[test_item].keys():
                                if (value_convert < bin_definition[test_item]['Limit']['L'] or value_convert >
                                    bin_definition[test_item]['Limit']['H']) and bin_definition[test_item]['Priority'] < \
                                        bin_definition[current_test_item]['Priority']:
                                    # set current_test_item is test_item when test_item's priority less than current_test_item's priority
                                    current_test_item = test_item
                    except:
                        # value is ''
                        if test_item in bin_definition.keys():
                            if 'IsNan' in bin_definition[test_item].keys():
                                bin_definition[bin_definition[test_item]['IsNan']]['Status'].append('0')
                            if test_item in bin_definition.keys() and 'IsNan' not in bin_definition[
                                test_item].keys() and \
                                            bin_definition[test_item]['Priority'] < bin_definition[current_test_item][
                                        'Priority']:
                                # set current_test_item is test_item
                                # when test_item in keys and InNan not in keys
                                # and test_item's priority less than current_test_item's priority
                                current_test_item = test_item
                finally:
                    if test_item in bin_definition.keys():
                        if 'IsLast' in bin_definition[test_item].keys():
                            if ('0' in bin_definition[bin_definition[test_item]['IsNan']]['Status'] and '1' in
                                bin_definition[bin_definition[test_item]['IsNan']]['Status']) and \
                                            bin_definition[bin_definition[test_item]['IsLast']]['Priority'] < \
                                            bin_definition[current_test_item]['Priority']:
                                current_test_item = bin_definition[test_item]['IsLast']
                            elif ('0' in bin_definition[bin_definition[test_item]['IsNan']]['Status'] and '1' not in
                                bin_definition[bin_definition[test_item]['IsNan']]['Status']) and \
                                            bin_definition[bin_definition[test_item]['IsNan']]['Priority'] < \
                                            bin_definition[current_test_item]['Priority']:
                                current_test_item = bin_definition[test_item]['IsNan']
            if bin_definition[current_test_item]['SWBin'] != int(data[row_num][2]):
                # current_test_item's SWBin is not equal to SW_BIN in csv file
                binning_check_result.append(
                    "              SB_BIN error of chipNo " + data[row_num][
                        0] + " : according to the priority,swbin should be " + str(
                        bin_definition[current_test_item]['SWBin']) + " ,but in CSV it's " + data[row_num][2] + "\n")
            if bin_definition[current_test_item]['HWBin'] != int(data[row_num][3]):
                # current_test_item's HWBin is not equal to hW_BIN in csv file
                binning_check_result.append("              hW_BIN error of chipNo " + data[row_num][
                    0] + " : according to the priority,hwbin should be " + str(
                    bin_definition[current_test_item]['HWBin']) + " ,but in CSV it's " + data[row_num][3] + "\n")
            current_row_count += 1
            if current_row_count != total_row_count:
                _signal.emit(str(current_row_count * 100 // total_row_count))
        binning_check_file = join(analysis_folder, lot_no + '_BinningCheck_analysis' + now_time + '.txt')
        with open(binning_check_file, 'a') as f:
            f.write("            " + file_name + "：\n")
            if len(binning_check_result) > 0:
                for item in binning_check_result:
                    f.write(item)
            else:
                f.write('              No problem.\n')
    if analysis_item in ('All', 'Nan ChipNo'):
        dc_nan_chip_no = []
        image_nan_chip_no = []
        for i in range(test_item_row_num + row_offset, first_register_row_num):
            for j in range(column_offset, pwdn_total_column_num + 1):
                if len(data[i][j].strip()) == 0:
                    # append value which is '' in columns colOffset to pwdn_total_column_num
                    dc_nan_chip_no.append(data[i][0])
                    break
        current_row_count += 1
        if current_row_count != total_row_count:
            _signal.emit(str(current_row_count * 100 // total_row_count))
        for i in range(test_item_row_num + row_offset, first_register_row_num):
            for j in range(pwdn_total_column_num + 1, binning_column_num + 1):
                if len(data[i][j].strip()) == 0:
                    # append value which is '' in columns pwdn_total_column_num+1 to binningcolumn_num
                    image_nan_chip_no.append(data[i][0])
                    break
        current_row_count += 1
        if current_row_count != total_row_count:
            _signal.emit(str(current_row_count * 100 // total_row_count))
        nan_chip_no_file = join(analysis_folder, lot_no + '_NanChipNo_analysis' + now_time + '.txt')
        dc_count = len(dc_nan_chip_no)
        image_count = len(image_nan_chip_no)
        with open(nan_chip_no_file, 'a') as f:
            f.write("            " + file_name + "：\n")
            if dc_count > 0 or image_count > 0:
                if dc_count == 0:
                    f.write("              (1) DC : no nan value.\n")
                else:
                    f.write("              (1) DC : " + str(dc_count) + " in total.They are : " + str(
                        dc_nan_chip_no) + "\n")
                if image_count == 0:
                    f.write("              (2) IMAGE : no nan value\n")
                else:
                    f.write("              (2) IMAGE : " + str(image_count) + " in total.They are : " + str(
                        image_nan_chip_no) + '\n')
            else:
                f.write("              No nan value.\n")
        current_row_count += 1
        if current_row_count != total_row_count:
            _signal.emit(str(current_row_count * 100 // total_row_count))
    if analysis_item in ('All', 'Summary'):
        summary_result = {}
        for group_by_id in range(1, 4):
            # 1:Site    2:SW_BIN    3:hW_BIN
            groupName = data[test_item_row_num][group_by_id]
            group_list = []
            for i in range(test_item_row_num + row_offset, first_register_row_num):
                # get value
                group_list.append(int(data[i][group_by_id]))
            # remove duplicate value
            format_group_list = list(set(group_list))
            # sort list
            format_group_list.sort()

            group_index = {}
            for i in format_group_list:
                temp = []
                # find i in groupList
                data_list = find_item(group_list, i)
                if group_by_id == 1:
                    temp_list = []
                    for j in data_list:
                        # get the corresponding SW_BIN of Site
                        temp_list.append(data[test_item_row_num + row_offset + j][2])
                    # remove duplicate value
                    format_temp_list = list(set(temp_list))
                    temp_index = {}
                    for m in format_temp_list:
                        # find m in temp_list
                        temp_data_list = find_item(temp_list, m)
                        temp_index[m] = temp_data_list
                    temp.append(temp_index)
                else:
                    temp.append(data_list)
                test_item_fail_count = []
                for column_num in range(column_offset, last_test_item_column_num + 1):
                    test_item = data[test_item_row_num][column_num].strip()
                    if test_item == 'AVDD_O/S' and project == 'JX828':
                        # set AVDD_O/S's limit in JX828 project
                        high_limit_data = -0.2
                        low_limit_data = -0.6
                    else:
                        high_limit_data = data[high_limit_row_num][column_num]
                        low_limit_data = data[low_limit_row_num][column_num]
                    temp_list = [data[test_item_row_num][column_num], 0]
                    try:
                        high_limit = float(high_limit_data)
                        low_limit = float(low_limit_data)
                        for j in data_list:
                            try:
                                value = data[test_item_row_num + row_offset + j][column_num]
                                value_convert = float(value)
                                if value_convert == int(value_convert):
                                    value_convert = int(value_convert)
                                if value_convert < low_limit or high_limit < value_convert:
                                    # count+1 when value over limit
                                    temp_list[1] += 1
                            except:
                                # count+1 when value is ''
                                temp_list[1] += 1
                    except:
                        # limit is N or nan
                        continue
                    test_item_fail_count.append(temp_list)
                temp.append(test_item_fail_count)
                group_index[i] = temp
            summary_result[groupName] = group_index
            parse_result['summary'] = summary_result
            current_row_count += 1
            if current_row_count != total_row_count:
                _signal.emit(str(current_row_count * 100 // total_row_count))
    parse_result['chip count'] = chip_count
    parse_result['lot_no'] = lot_no
    return parse_result


@time_calc
def save_data(analysis_folder, _signal, parse_data):
    """
    save data
    save data check,binning check,nan chipno and summary to files
    """
    global current_row_count
    site_data = []
    softbin_data = []
    hardbin_data = []
    lot_count = parse_data[0]['chip count']
    lot_no = parse_data[0]['lot_no']
    for data in parse_data:
        if analysis_item in ('All', 'Summary'):
            site_data.append(data['summary']['Site'])
            # sort softbin
            softbin_data.append(
                sorted(data['summary']['SW_BIN'].items(), key=lambda item: len(item[1][0]), reverse=True))
            hardbin_data.append(data['summary']['hW_BIN'])
    if analysis_item in ('All', 'Summary'):
        summary_file = join(analysis_folder, lot_no + '_Summary_analysis_' + now_time + '.xlsx')
        summary_wb = Workbook()

        ok_hwbin_count = 0
        begin_fail_swbin = 0
        for hwbin_key in HWBIN_TO_SWBIN[project].keys():
            if HWBIN_TO_SWBIN[project][hwbin_key]['isPassBin']:
                ok_hwbin_count += 1
                begin_fail_swbin += len(HWBIN_TO_SWBIN[project][hwbin_key]['SWBin'])

        color_list = ['99FFFF', '33FF00', 'FFFFCC', 'FFFF33', 'FF9900', 'FF0099', 'FF0000']
        swbin_list = []
        key_index = 0
        for hwbin_key in HWBIN_TO_SWBIN[project].keys():
            for swbin in HWBIN_TO_SWBIN[project][hwbin_key]['SWBin']:
                swbin_list.append([swbin, color_list[key_index]])
            key_index += 1

        hardbin_sheet = summary_wb.create_sheet('HWBin')
        hardbin_sheet.freeze_panes = 'B2'
        summary_data = []
        for i in range(len(hardbin_data)):
            temp_list = []
            if i == 0:
                temp_list.append('FT')
            else:
                temp_list.append('RT' + str(i))
            test_count = 0
            pass_count = 0
            for hwbin_key in HWBIN_TO_SWBIN[project].keys():
                if hwbin_key in hardbin_data[i].keys():
                    bin_count = len(hardbin_data[i][hwbin_key][0])
                else:
                    bin_count = 0
                if HWBIN_TO_SWBIN[project][hwbin_key]['isPassBin']:
                    pass_count += bin_count
                test_count += bin_count
                temp_list.append(bin_count)
            temp_list.append(test_count)
            pass_percent = '{:.2%}'.format(pass_count / test_count)
            temp_list.append(pass_percent)
            summary_data.append(temp_list)
        irow = 1
        icol = 1
        hardbin_sheet.cell(row=irow, column=icol).value = lot_no
        icol += 1
        for hwbin_key in HWBIN_TO_SWBIN[project].keys():
            hardbin_sheet.cell(row=irow, column=icol).value = 'HWBin' + str(hwbin_key)
            icol += 1
        hardbin_sheet.cell(row=irow, column=icol).value = 'TestCount'
        icol += 1
        hardbin_sheet.cell(row=irow, column=icol).value = 'PassPercent'
        irow += 1
        for i in range(len(summary_data)):
            for j in range(len(summary_data[i])):
                hardbin_sheet.cell(row=irow, column=(j + 1)).value = summary_data[i][j]
                if j == 0:
                    hardbin_sheet.cell(row=irow, column=(j + 1)).fill = PatternFill(fill_type='solid', fgColor=GREEN)
                elif 1 <= j < len(summary_data[i]) - 2:
                    hardbin_sheet.cell(row=irow + 1, column=(j + 1)).value = '{:.2%}'.format(
                        summary_data[i][j] / summary_data[i][-2])
                    hardbin_sheet.cell(row=irow + 1, column=(j + 1)).fill = PatternFill(fill_type='solid',
                                                                                        fgColor=GREEN)
                elif j == len(summary_data[i]) - 2 and i > 0:
                    compareCount = 0
                    for m in range(ok_hwbin_count + 1, len(summary_data[i]) - 2):
                        compareCount += summary_data[i - 1][m]
                    if summary_data[i][j] != compareCount:
                        # the number of test is not equal to the total number of chips failed in the previous test
                        if summary_data[i][j] > compareCount:
                            flag = '↓'
                        else:
                            flag = '↑'
                        hardbin_sheet.cell(row=irow, column=(j + 1)).fill = PatternFill(fill_type='solid', fgColor=RED)
                        hardbin_sheet.cell(row=irow + 1, column=(j + 1)).value = flag + str(compareCount)
                        hardbin_sheet.cell(row=irow + 1, column=(j + 1)).font = Font(color=RED, bold=True)
            irow += 2
        hardbin_sheet.cell(row=irow, column=1).value = 'Summary'
        hardbin_sheet.cell(row=irow, column=1).fill = PatternFill(fill_type='solid', fgColor=GREEN)
        # summary pass hwbin count : add all lots pass hwbin count
        pass_hwbin_count = [0] * ok_hwbin_count
        for i in range(len(summary_data)):
            for j in range(len(pass_hwbin_count)):
                pass_hwbin_count[j] += summary_data[i][j + 1]
        # summary total pass hwbin count : add all pass hwbin count
        total_pass_hwbin_count = 0
        for i in range(len(pass_hwbin_count)):
            hardbin_sheet.cell(row=irow, column=2 + i).value = pass_hwbin_count[i]
            total_pass_hwbin_count += pass_hwbin_count[i]
        # summary fail hwbin count : last lot's fail hwbin count
        for i in range(len(pass_hwbin_count) + 1, len(summary_data[-1]) - 2):
            hardbin_sheet.cell(row=irow, column=(i + 1)).value = summary_data[-1][i]
        hardbin_sheet.cell(row=irow, column=icol - 1).value = summary_data[0][-2]
        hardbin_sheet.cell(row=irow, column=icol).value = '{:.2%}'.format(total_pass_hwbin_count / summary_data[0][-2])

        for row in hardbin_sheet.rows:
            for cell in row:
                cell.border = border
                if cell.row == 1:
                    cell.fill = PatternFill(fill_type='solid', fgColor=YELLOW)
                    cell.font = Font(bold=True)
                    cell.alignment = alignment
                if cell.row == 2 and cell.column == hardbin_sheet.max_column:
                    cell.fill = PatternFill(fill_type='solid', fgColor=GREEN)
                if cell.row == hardbin_sheet.max_row and cell.column == hardbin_sheet.max_column:
                    cell.fill = PatternFill(fill_type='solid', fgColor=GREEN)

        current_row_count += 1
        if current_row_count != total_row_count:
            _signal.emit(str(current_row_count * 100 // total_row_count))

        hardsoftbin_sheet = summary_wb.create_sheet('HWBin-SWBin')
        hardsoftbin_sheet.freeze_panes = 'C2'
        irow = 1
        hardsoftbin_sheet.cell(row=irow, column=1).value = lot_no
        hardsoftbin_sheet.cell(row=irow, column=1).fill = PatternFill(fill_type='solid', fgColor=YELLOW)
        hardsoftbin_sheet.cell(row=irow, column=3).value = 'FT'
        hardsoftbin_sheet.cell(row=irow, column=3).fill = PatternFill(fill_type='solid', fgColor=YELLOW)
        hardsoftbin_sheet.merge_cells(start_row=irow, end_row=irow, start_column=3, end_column=4)
        for i in range(1, len(softbin_data)):
            hardsoftbin_sheet.cell(row=irow, column=3 + 2 * i).value = 'RT' + str(i)
            hardsoftbin_sheet.cell(row=irow, column=3 + 2 * i).fill = PatternFill(fill_type='solid', fgColor=YELLOW)
            hardsoftbin_sheet.merge_cells(start_row=irow, end_row=irow, start_column=3 + 2 * i, end_column=4 + 2 * i)
        irow += 1

        # sort swbin in hwbin from FT by swbin count from more to less
        swbin_sort_by_FT_list = []
        for hwbin_key in HWBIN_TO_SWBIN[project].keys():
            temp_list = []
            for swbin in HWBIN_TO_SWBIN[project][hwbin_key]['SWBin']:
                find_softbin = False
                for x in range(len(softbin_data[0])):
                    if swbin == softbin_data[0][x][0]:
                        find_softbin = True
                        swbin_count = len(softbin_data[0][x][1][0])
                        temp_list.append((swbin, swbin_count))
                        break
                if not find_softbin:
                    temp_list.append((swbin, 0))
            for m in range(len(temp_list) - 1):
                for n in range(m + 1, len(temp_list)):
                    if temp_list[m][1] < temp_list[n][1]:
                        temp_list[m], temp_list[n] = temp_list[n], temp_list[m]
            temp_count_list = [hwbin_key]
            for y in range(len(temp_list)):
                temp_count_list.append(temp_list[y][0])
            swbin_sort_by_FT_list.append(temp_count_list)

        summary_soft_count = []
        for i in range(len(swbin_sort_by_FT_list)):
            hardsoftbin_sheet.cell(row=irow, column=1).value = 'HWBin' + str(swbin_sort_by_FT_list[i][0])
            hardsoftbin_sheet.cell(row=irow, column=1).fill = PatternFill(fill_type='solid', fgColor=GREEN)
            hardsoftbin_sheet.cell(row=irow, column=1).font = Font(bold=True)
            for j in range(1, len(swbin_sort_by_FT_list[i])):
                hardsoftbin_sheet.cell(row=irow, column=2).value = 'SWBin' + str(swbin_sort_by_FT_list[i][j])
                temp_count = 0
                for m in range(len(softbin_data)):
                    find_softbin = False
                    for n in range(len(softbin_data[m])):
                        if swbin_sort_by_FT_list[i][j] == softbin_data[m][n][0]:
                            find_softbin = True
                            swbin_count = len(softbin_data[m][n][1][0])
                            break
                    if not find_softbin:
                        swbin_count = 0
                    if i in range(ok_hwbin_count):
                        # add the number of pass hwbin's swbin of all lots
                        temp_count += swbin_count
                    elif m == len(softbin_data) - 1:
                        # set the number of fail hwbin's swbin of last lot
                        temp_count = swbin_count
                    hardsoftbin_sheet.cell(row=irow, column=(2 * m + 3)).value = swbin_count
                    hardsoftbin_sheet.cell(row=irow, column=(2 * m + 4)).value = '{:.2%}'.format(
                        swbin_count / summary_data[0][-2])
                    hardsoftbin_sheet.cell(row=irow, column=(2 * m + 4)).fill = PatternFill(fill_type='solid',
                                                                                            fgColor=GREEN)
                irow += 1
                summary_soft_count.append(temp_count)
            summary_soft_count.append('')
            irow += 1

        irow = 1
        icol = 4 + 2 * len(softbin_data)
        hardsoftbin_sheet.cell(row=irow, column=icol).value = 'Summary'
        hardsoftbin_sheet.cell(row=irow, column=icol).fill = PatternFill(fill_type='solid', fgColor='FFA500')
        hardsoftbin_sheet.merge_cells(start_row=irow, end_row=irow, start_column=icol, end_column=icol + 1)
        irow += 1
        for i in range(len(summary_soft_count)):
            if isinstance(summary_soft_count[i], int):
                hardsoftbin_sheet.cell(row=irow, column=icol).value = summary_soft_count[i]
                hardsoftbin_sheet.cell(row=irow, column=icol + 1).value = '{:.2%}'.format(
                    summary_soft_count[i] / summary_data[0][-2])
                hardsoftbin_sheet.cell(row=irow, column=icol + 1).fill = PatternFill(fill_type='solid',
                                                                                     fgColor='FFA500')
            irow += 1

        for row in hardsoftbin_sheet.rows:
            for cell in row:
                cell.border = border
                if cell.row == 1:
                    cell.font = Font(bold=True)
                    cell.alignment = alignment

        current_row_count += 1
        if current_row_count != total_row_count:
            _signal.emit(str(current_row_count * 100 // total_row_count))

        site_softbin_sheet = summary_wb.create_sheet('Site-SWBin')
        site_softbin_sheet.freeze_panes = 'B2'
        irow = 1
        site_softbin_sheet.cell(row=irow, column=1).value = lot_no
        site_softbin_sheet.cell(row=irow, column=1).fill = PatternFill(fill_type='solid', fgColor=YELLOW)
        for i in range(16):
            site_softbin_sheet.cell(row=irow, column=2 + i).value = 'Site' + str(i)
            site_softbin_sheet.cell(row=irow, column=2 + i).fill = PatternFill(fill_type='solid', fgColor=YELLOW)
        site_softbin_sheet.merge_cells(start_row=irow, end_row=irow, start_column=19, end_column=20)
        site_softbin_sheet.cell(row=irow, column=19).value = 'Summary'
        site_softbin_sheet.cell(row=irow, column=19).fill = PatternFill(fill_type='solid', fgColor='FFA500')
        irow += 1

        lot_no_site_swbin_count = []
        for x in range(len(swbin_list)):
            site_count_list = [[0, 0], [1, 0], [2, 0], [3, 0], [4, 0], [5, 0], [6, 0], [7, 0], [8, 0], [9, 0], [10, 0],
                               [11, 0], [12, 0], [13, 0], [14, 0], [15, 0]]
            swbin = str(swbin_list[x][0])
            for site in range(16):
                if x < begin_fail_swbin:
                    for i in range(len(site_data)):
                        if (site in site_data[i].keys()) and (swbin in site_data[i][site][0].keys()):
                            site_count_list[site][1] += len(site_data[i][site][0][swbin])
                else:
                    if (site in site_data[-1].keys()) and (swbin in site_data[-1][site][0].keys()):
                        site_count_list[site][1] += len(site_data[-1][site][0][swbin])
            lot_no_site_swbin_count.append(site_count_list)
        site_swbin_count = []
        for i in range(len(lot_no_site_swbin_count)):
            temp_total_count = 0
            temp_list = [0] * 16
            temp_swbin_count = []
            for j in range(len(temp_list)):
                temp_list[j] += lot_no_site_swbin_count[i][j][1]
                temp_swbin_count.append([temp_list[j], WHITE])
                temp_total_count += temp_list[j]
                if j == len(temp_list) - 1:
                    temp_swbin_count.append([temp_total_count, 'FFA500'])
            site_swbin_count.append(temp_swbin_count)
        # add the number of fail swbin by all site
        # fill color of max value in fail swbin is green(all values are 0 do not fill green)
        # fill color of min value in fail swbin is red(multiple values of 0 do not fill red)
        site_pass_total_list = [0] * 16
        for i in range(begin_fail_swbin):
            for j in range(len(site_swbin_count[i]) - 1):
                site_pass_total_list[j] += site_swbin_count[i][j][0]
        site_fail_total_list = [0] * 16
        for i in range(begin_fail_swbin, len(site_swbin_count)):
            min_value = site_swbin_count[i][0][0]
            max_value = site_swbin_count[i][0][0]
            for m in range(1, len(site_swbin_count[i]) - 1):
                if min_value > site_swbin_count[i][m][0]:
                    min_value = site_swbin_count[i][m][0]
                if max_value < site_swbin_count[i][m][0]:
                    max_value = site_swbin_count[i][m][0]
            min_index = []
            max_index = []
            for j in range(len(site_swbin_count[i]) - 1):
                site_fail_total_list[j] += site_swbin_count[i][j][0]
                if site_swbin_count[i][j][0] == min_value:
                    min_index.append(j)
                if site_swbin_count[i][j][0] == max_value:
                    max_index.append(j)
            if min_value == 0 and len(min_index) == 1:
                site_swbin_count[i][min_index[0]][1] = GREEN
                for y in range(len(max_index)):
                    site_swbin_count[i][max_index[y]][1] = RED
            elif min_value == 0 and max_value > 0:
                for y in range(len(max_index)):
                    site_swbin_count[i][max_index[y]][1] = RED
            elif min_value > 0:
                for x in range(len(min_index)):
                    site_swbin_count[i][min_index[x]][1] = GREEN
                for y in range(len(max_index)):
                    site_swbin_count[i][max_index[y]][1] = RED

        for x in range(len(swbin_list)):
            site_softbin_sheet.cell(row=irow, column=1).value = 'SWBin' + str(swbin_list[x][0])
            site_softbin_sheet.cell(row=irow, column=1).fill = PatternFill(fill_type='solid', fgColor=swbin_list[x][1])
            for y in range(len(site_swbin_count[x])):
                if y < len(site_swbin_count[x]) - 1:
                    if site_swbin_count[x][y][0] > 0:
                        site_softbin_sheet.cell(row=irow, column=2 + y).value = '{:.4%}'.format(
                            site_swbin_count[x][y][0] / summary_data[0][-2])
                    site_softbin_sheet.cell(row=irow, column=2 + y).fill = PatternFill(fill_type='solid',
                                                                                       fgColor=site_swbin_count[x][y][
                                                                                           1])
                else:
                    site_softbin_sheet.cell(row=irow, column=3 + y).value = site_swbin_count[x][y][0]
                    site_softbin_sheet.cell(row=irow, column=4 + y).value = '{:.4%}'.format(
                        site_swbin_count[x][y][0] / summary_data[0][-2])
                    site_softbin_sheet.cell(row=irow, column=4 + y).fill = PatternFill(fill_type='solid',
                                                                                       fgColor=site_swbin_count[x][y][
                                                                                           1])
            irow += 1
        irow += 1
        site_softbin_sheet.merge_cells(start_row=irow, end_row=irow + 1, start_column=1, end_column=1)
        site_softbin_sheet.cell(row=irow, column=1).value = 'FailPercent'
        site_softbin_sheet.cell(row=irow, column=1).font = Font(bold=True)
        site_softbin_sheet.cell(row=irow, column=1).alignment = alignment
        site_softbin_sheet.cell(row=irow, column=1).fill = PatternFill(fill_type='solid', fgColor='FFA500')
        for i in range(len(site_fail_total_list)):
            site_softbin_sheet.cell(row=irow, column=2 + i).value = site_fail_total_list[i]
            site_softbin_sheet.cell(row=irow + 1, column=2 + i).value = '{:.4%}'.format(
                site_fail_total_list[i] / summary_data[0][-2])
            site_softbin_sheet.cell(row=irow + 1, column=2 + i).fill = PatternFill(fill_type='solid', fgColor=RED)
        site_softbin_sheet.cell(row=irow, column=3 + len(site_fail_total_list)).value = sum(site_fail_total_list)
        site_softbin_sheet.cell(row=irow, column=3 + len(site_fail_total_list)).fill = PatternFill(fill_type='solid',
                                                                                                   fgColor=RED)
        site_softbin_sheet.cell(row=irow + 1, column=3 + len(site_fail_total_list)).value = '{:.4%}'.format(
            sum(site_fail_total_list) / summary_data[0][-2])
        site_softbin_sheet.cell(row=irow + 1, column=3 + len(site_fail_total_list)).fill = PatternFill(
            fill_type='solid',
            fgColor=RED)
        irow += 2
        site_softbin_sheet.merge_cells(start_row=irow, end_row=irow + 1, start_column=1, end_column=1)
        site_softbin_sheet.cell(row=irow, column=1).value = 'PassPercent'
        site_softbin_sheet.cell(row=irow, column=1).font = Font(bold=True)
        site_softbin_sheet.cell(row=irow, column=1).alignment = alignment
        site_softbin_sheet.cell(row=irow, column=1).fill = PatternFill(fill_type='solid', fgColor=GREEN)
        for i in range(len(site_pass_total_list)):
            site_softbin_sheet.cell(row=irow, column=2 + i).value = site_pass_total_list[i]
            site_softbin_sheet.cell(row=irow + 1, column=2 + i).value = '{:.4%}'.format(
                site_pass_total_list[i] / summary_data[0][-2])
            site_softbin_sheet.cell(row=irow + 1, column=2 + i).fill = PatternFill(fill_type='solid', fgColor=GREEN)
        site_softbin_sheet.cell(row=irow, column=3 + len(site_pass_total_list)).value = sum(site_pass_total_list)
        site_softbin_sheet.cell(row=irow, column=3 + len(site_pass_total_list)).fill = PatternFill(fill_type='solid',
                                                                                                   fgColor=GREEN)
        site_softbin_sheet.cell(row=irow + 1, column=3 + len(site_pass_total_list)).value = '{:.4%}'.format(
            sum(site_pass_total_list) / summary_data[0][-2])
        site_softbin_sheet.cell(row=irow + 1, column=3 + len(site_pass_total_list)).fill = PatternFill(
            fill_type='solid',
            fgColor=GREEN)
        for row in site_softbin_sheet.rows:
            for cell in row:
                cell.border = border
                if cell.row == 1:
                    cell.font = Font(bold=True)
                    cell.alignment = alignment

        current_row_count += 1
        if current_row_count != total_row_count:
            _signal.emit(str(current_row_count * 100 // total_row_count))

        hwbin_test_item_sheet = summary_wb.create_sheet('HWBin-test_item')
        hwbin_test_item_sheet.freeze_panes = 'B2'
        irow = 1
        hwbin_test_item_sheet.cell(row=irow, column=1).value = lot_count
        for hwbin_key in hardbin_data[0].keys():
            for i in range(len(hardbin_data[0][hwbin_key][1])):
                hwbin_test_item_sheet.cell(row=irow, column=i + 2).value = hardbin_data[0][hwbin_key][1][i][0]
            break
        irow += 1
        for hwbin_key in hardbin_data[0].keys():
            hwbin_test_item_sheet.cell(row=irow, column=1).value = 'HWBin' + str(hwbin_key)
            hwbin_test_item_sheet.cell(row=irow, column=1).fill = PatternFill(fill_type='solid', fgColor=GREEN)
            for i in range(len(hardbin_data[0][hwbin_key][1])):
                hwbin_test_item_sheet.cell(row=irow, column=2 + i).value = hardbin_data[0][hwbin_key][1][i][1]
                hwbin_test_item_sheet.cell(row=irow + 1, column=2 + i).value = '{:.2%}'.format(
                    hardbin_data[0][hwbin_key][1][i][1] / lot_count)
                if hardbin_data[0][hwbin_key][1][i][1] > 0:
                    # percent value greater than 0 are filled in red
                    hwbin_test_item_sheet.cell(row=irow + 1, column=2 + i).fill = PatternFill(fill_type='solid',
                                                                                              fgColor=RED)
            irow += 2

        for row in hwbin_test_item_sheet.rows:
            for cell in row:
                cell.border = border
                if cell.row == 1 and cell.column > 1:
                    cell.fill = PatternFill(fill_type='solid', fgColor=YELLOW)
                    cell.font = Font(bold=True)
                    cell.alignment = alignment

        current_row_count += 1
        if current_row_count != total_row_count:
            _signal.emit(str(current_row_count * 100 // total_row_count))

        swbin_test_item_sheet = summary_wb.create_sheet('SWBin-test_item')
        swbin_test_item_sheet.freeze_panes = 'B2'
        irow = 1
        swbin_test_item_sheet.cell(row=irow, column=1).value = lot_count
        for i in range(len(softbin_data[0][0][1][1])):
            swbin_test_item_sheet.cell(row=irow, column=i + 2).value = softbin_data[0][0][1][1][i][0]
        irow += 1
        for x in range(len(swbin_list)):
            swbin_test_item_sheet.cell(row=irow, column=1).value = 'SWBin' + str(swbin_list[x][0])
            swbin_test_item_sheet.cell(row=irow, column=1).fill = PatternFill(fill_type='solid',
                                                                              fgColor=swbin_list[x][1])
            for i in range(len(softbin_data[0])):
                if swbin_list[x][0] == softbin_data[0][i][0]:
                    for j in range(len(softbin_data[0][i][1][1])):
                        swbin_test_item_sheet.cell(row=irow, column=2 + j).value = softbin_data[0][i][1][1][j][1]
                        swbin_test_item_sheet.cell(row=irow + 1, column=2 + j).value = '{:.2%}'.format(
                            softbin_data[0][i][1][1][j][1] / lot_count)
                        if softbin_data[0][i][1][1][j][1] > 0:
                            # percent value greater than 0 are filled in red
                            swbin_test_item_sheet.cell(row=irow + 1, column=2 + j).fill = PatternFill(fill_type='solid',
                                                                                                      fgColor=RED)
            irow += 2

        for row in swbin_test_item_sheet.rows:
            for cell in row:
                cell.border = border
                if cell.row == 1 and cell.column > 1:
                    cell.fill = PatternFill(fill_type='solid', fgColor=YELLOW)
                    cell.font = Font(bold=True)
                    cell.alignment = alignment

        current_row_count += 1
        if current_row_count != total_row_count:
            _signal.emit(str(current_row_count * 100 // total_row_count))

        site_test_item_sheet = summary_wb.create_sheet('Site-test_item')
        site_test_item_sheet.freeze_panes = 'B2'
        irow = 1
        site_test_item_sheet.cell(row=irow, column=1).value = lot_count
        for siteKey in site_data[0].keys():
            for i in range(len(site_data[0][siteKey][1])):
                site_test_item_sheet.cell(row=irow, column=i + 2).value = site_data[0][siteKey][1][i][0]
            break
        irow += 1
        for siteKey in site_data[0].keys():
            site_test_item_sheet.cell(row=irow, column=1).value = 'Site' + str(siteKey)
            site_test_item_sheet.cell(row=irow, column=1).fill = PatternFill(fill_type='solid',
                                                                             fgColor=GREEN)
            for i in range(len(site_data[0][siteKey][1])):
                site_test_item_sheet.cell(row=irow, column=2 + i).value = site_data[0][siteKey][1][i][1]
                site_test_item_sheet.cell(row=irow + 1, column=2 + i).value = '{:.2%}'.format(
                    site_data[0][siteKey][1][i][1] / lot_count)
                if site_data[0][siteKey][1][i][1] > 0:
                    # percent value greater than 0 are filled in red
                    site_test_item_sheet.cell(row=irow + 1, column=2 + i).fill = PatternFill(fill_type='solid',
                                                                                             fgColor=RED)
            irow += 2

        for row in site_test_item_sheet.rows:
            for cell in row:
                cell.border = border
                if cell.row == 1 and cell.column > 1:
                    cell.fill = PatternFill(fill_type='solid', fgColor=YELLOW)
                    cell.font = Font(bold=True)
                    cell.alignment = alignment

        for sheet_name in summary_wb.sheetnames:
            if sheet_name == 'Sheet':
                del summary_wb[sheet_name]
            else:
                set_column_width(summary_wb[sheet_name])
        summary_wb.save(summary_file)
        current_row_count += 1
        if current_row_count != total_row_count:
            _signal.emit(str(current_row_count * 100 // total_row_count))


class RunThread(QThread):
    _signal = pyqtSignal(str)

    def __init__(self, choose_radio):
        super(RunThread, self).__init__()
        self.choose_radio = choose_radio

    def __del__(self):
        self.wait()

    def run(self):
        if self.choose_radio == 'DateFolder':
            lot_no_names = listdir(open_path)
            for name in lot_no_names:
                folder = join(open_path, name)
                if isdir(folder):
                    # get CSV file under the folder
                    file_list = get_file_list(folder, '.csv')
                    if not file_list:
                        exit()

                    # analysis folder path
                    if argv.count('-a') == 0:
                        # default analysis folder
                        analysis_folder = folder + '\analysis'
                    else:
                        analysis_folder = argv[argv.index('-a') + 1]

                    # create analysis folder
                    create_directory(analysis_folder)

                    parse_data = []
                    for file in file_list:
                        # parse file
                        parse_data.append(parse_file(file, self._signal, analysis_folder))
                    # save data
                    save_data(analysis_folder, self._signal, parse_data)
        else:
            file_list = get_file_list(open_path, '.csv')
            if not file_list:
                exit()

            # analysis folder path
            if argv.count('-a') == 0:
                # default analysis folder
                analysis_folder = open_path + '\Analysis'
            else:
                analysis_folder = argv[argv.index('-a') + 1]

            # create analysis folder
            create_directory(analysis_folder)

            parse_data = []
            for file in file_list:
                # parse file
                parse_data.append(parse_file(file, self._signal, analysis_folder))
            # save data
            save_data(analysis_folder, self._signal, parse_data)
        self._signal.emit(str(100))


class MainWindow(QMainWindow, Datalog_DateLot_Analysis_UI.Ui_MainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        self.setupUi(self)
        self.cwd = getcwd()
        self.qe = QErrorMessage(self)
        self.pale = QPalette()
        # self.pale.setBrush(self.backgroundRole(), QBrush(QPixmap('./images/kobe3.jpg')))
        self.setPalette(self.pale)
        self.DateFolder_radioButton.setChecked(True)
        # click button call openfolder
        self.Open_pushButton.clicked.connect(self.open)
        self.Project_comboBox.insertItem(0, self.tr('F28'))
        self.Project_comboBox.insertItem(1, self.tr('JX828'))
        self.Project_comboBox.insertItem(2, self.tr('JX825'))
        self.Project_comboBox.insertItem(2, self.tr('JX832'))
        self.AnalysisItem_comboBox.insertItem(0, self.tr('All'))
        self.AnalysisItem_comboBox.insertItem(1, self.tr('Summary'))
        self.AnalysisItem_comboBox.insertItem(2, self.tr('Data Check'))
        self.AnalysisItem_comboBox.insertItem(3, self.tr('Binning Check'))
        self.AnalysisItem_comboBox.insertItem(4, self.tr('Nan ChipNo'))
        self.Analysis_pushButton.clicked.connect(self.analysis)

    def open(self):
        if self.DateFolder_radioButton.isChecked():
            dir_choose = QFileDialog.getExistingDirectory(self, 'Select DateFolder Directory', self.cwd)
        else:
            dir_choose = QFileDialog.getExistingDirectory(self, 'Select lot_folder Directory', self.cwd)
        if not dir_choose:
            return
        self.Open_lineEdit.setText(dir_choose)

    def call_back_log(self, msg):
        self.progressBar.setValue(int(msg))  # pass the thread's parameters to progressBar
        if msg == '100':
            self.Analysis_pushButton.setEnabled(True)

    def analysis(self):
        if not self.Open_lineEdit.text():
            self.qe.showMessage('Path cannot be empty!')
            return
        else:
            global total_row_count, analysis_item, open_path
            total_row_count = 0
            open_path = self.Open_lineEdit.text()
            analysis_item = self.AnalysisItem_comboBox.currentText()
            if self.DateFolder_radioButton.isChecked():
                lot_no_names = listdir(open_path)
                for lot_name in lot_no_names:
                    lot_folder = join(open_path, lot_name)
                    if isdir(lot_folder):
                        file_list = get_file_list(lot_folder, '.csv')
                        for file in file_list:
                            with open(file, encoding='unicode_escape') as f:
                                csv_reader = reader(f)
                                if analysis_item == 'All':
                                    total_row_count += (array(list(csv_reader)).shape[0] * 3)
                                    total_row_count += 12
                                elif analysis_item == 'Data Check':
                                    total_row_count += (array(list(csv_reader)).shape[0] * 2)
                                elif analysis_item == 'Binning Check':
                                    total_row_count += (array(list(csv_reader)).shape[0])
                                elif analysis_item == 'Nan ChipNo':
                                    total_row_count += 3
                                elif analysis_item == 'Summary':
                                    total_row_count += 9
            else:
                file_list = get_file_list(open_path, '.csv')
                for file in file_list:
                    with open(file, encoding='unicode_escape') as f:
                        csv_reader = reader(f)
                        if analysis_item == 'All':
                            total_row_count += (array(list(csv_reader)).shape[0] * 3)
                            total_row_count += 12
                        elif analysis_item == 'Data Check':
                            total_row_count += (array(list(csv_reader)).shape[0] * 2)
                        elif analysis_item == 'Binning Check':
                            total_row_count += (array(list(csv_reader)).shape[0])
                        elif analysis_item == 'Nan ChipNo':
                            total_row_count += 3
                        elif analysis_item == 'Summary':
                            total_row_count += 9
            if total_row_count == 0:
                self.qe.showMessage('No files available!')
                return
            global now_time, current_row_count, project
            now_time = datetime.now().strftime("%Y%m%d%H%M%S")
            current_row_count = 0
            self.progressBar.setValue(0)
            project = self.Project_comboBox.currentText()
            if self.DateFolder_radioButton.isChecked():
                choose_radio = self.DateFolder_radioButton.text()
            else:
                choose_radio = self.LotFolder_radioButton.text()
            # create thread
            self.Analysis_pushButton.setEnabled(False)
            self.thread = RunThread(choose_radio)
            # connect signal
            self.thread._signal.connect(self.call_back_log)
            self.thread.start()


if __name__ == '__main__':
    # _csv.Error:field larger than field limit(131072)
    maxInt = maxsize
    while True:
        try:
            field_size_limit(maxInt)
            break
        except OverflowError:
            maxInt = int(maxInt / 10)
    app = QApplication(argv)
    myMainWindow = MainWindow()
    myMainWindow.show()
    exit(app.exec_())
