# encoding:utf-8
# @Time     : 2019/9/11
# @Author   : Jerry Chou
# @File     :
# @Function : 1、多文件分析  2、根据需要的分组方式自动识别文件中分组

from csv import reader
from os.path import dirname, join, abspath, basename
from os import getcwd
from math import isnan
from numpy import std, mean
from datetime import datetime
from sys import argv, path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.styles.colors import BLACK,YELLOW,DARKYELLOW,RED,GREEN

path.append(abspath(join(getcwd(), '..')))
import Common

alignment = Alignment(horizontal='center', vertical='center')
thin = Side(border_style='thin', color=BLACK)
border = Border(top=thin, left=thin, right=thin, bottom=thin)

row_offset = 5
col_offset = 4


def parse_file(file, group_by_id):
    """
    解析文件
    :param file: 文件
    :param group_by_id: 分组id
    :return: 
    """
    data = []
    with open(file) as f:
        csv_reader = reader(f)
        for row in csv_reader:
            data.append(row)

    search_binning = Common.search_string(data, 'Binning')
    if not search_binning:
        exit()
    else:
        binning_row_num = search_binning[0]
        binning_col_num = search_binning[1]

    col_count = binning_col_num + 1

    group_name = data[binning_row_num][group_by_id]

    first_register_row_num = len(data) - 1
    for i in range(binning_row_num + row_offset, len(data)):
        try:
            int(data[i][0])
        except:
            first_register_row_num = i
            break

    for i in range(0, len(data)):
        add_count = col_count - len(data[i])
        if add_count > 0:
            for j in range(add_count):
                data[i].append('')

    result = []

    exist_nan_row_list = []

    for i in range(binning_row_num + row_offset, first_register_row_num):
        for j in range(col_count):
            if len(data[i][j].strip()) == 0:
                exist_nan_row_list.append(i)
                break

    data_row_list = []
    for i in range(binning_row_num + row_offset, first_register_row_num):
        data_row_list.append(i)

    ok_row_list = list(set(data_row_list) - set(exist_nan_row_list))

    ok_group_index = {}
    ok_group_list = []
    for row in ok_row_list:
        ok_group_list.append(data[row][group_by_id])
    ok_group_int_list = [int(i) for i in ok_group_list]
    format_ok_group_list = list(set(ok_group_int_list))
    format_ok_group_list.sort()
    for i in format_ok_group_list:
        ok_group_index[i] = Common.find_item(ok_group_int_list, i)
    # 遍历所有测试项
    for col_num in range(col_offset, col_count):
        test_name = data[binning_row_num][col_num]
        unit = data[binning_row_num + 1][col_num]
        if isinstance(unit, float) and isnan(unit):
            unit = ''
        hi_limit = data[binning_row_num + 2][col_num]
        lo_limit = data[binning_row_num + 3][col_num]
        # 取得数据
        test_item_data = []
        for row in ok_row_list:
            test_item_data.append(data[row][col_num])
        temp = [test_name, unit, hi_limit, lo_limit]
        temp.append(cal_data(test_item_data, ok_group_index))
        # 计算数据
        result.append(temp)
    return result, format_ok_group_list, ok_group_index, group_name


def calc(data):
    """
    计算数据
    :param data: 数据
    :return: 最小值，中位数，最大值，标准方差
    """
    min_value = min(data)
    max_value = max(data)
    std_value = std(data)
    mean_value = mean(data)

    # 付博观察值，>5%为异常值
    if mean_value != 0:
        std_div_mean = std_value / mean_value  # '{:.2%}'.format(std_value / mean_value)
    else:
        std_div_mean = 0
    return round(min_value, 6), round(mean_value, 6), round(max_value, 6), round(std_value, 6), round(std_div_mean, 6)


def cal_data(data, group_index):
    """
    计算数据
    :param data:数据 
    :param group_index:分组索引 
    :return: 
    """
    data_val = []
    for num in data:
        data_val.append(float(num))
    calc_data = []
    for i in group_index:
        temp = []
        for j in group_index[i]:
            temp.append(data_val[j])
        calc_data.append(calc(temp))
    return calc_data

def save_data(save_file, data):
    """
    保存数据
    :param save_file: 保存的文件路径
    :param data: 数据
    :return: 
    """
    file_name = save_file.split('.')[0]
    date = datetime.now().strftime("%Y%m%d%H%M")
    test_file_name = file_name + '_Analysis_by' + data[3] + '_' + date + '.xlsx'
    print(test_file_name)

    wb = Workbook()

    line_title = ['TestName', 'Unit', 'LoLimit', 'HiLimit', 'Min', 'Mean', 'Max', 'std', 'StdDivMeanPercent']
    std_style = ['(1,10]', '(10,100]', '(100,+∞)']
    StdDivMeanPercent_style = ['[-50%,-5%),(5%,50%]', '[-500%,-50%),(50%,500%]', '(-∞,-500%),(500%,+∞)']

    for i in range(len(data[1])):
        ws = wb.create_sheet(data[3] + str(data[1][i]))
        ws.freeze_panes = 'B7'

        irow = 1
        ws.cell(row=irow, column=8).value = line_title[7]
        ws.cell(row=irow, column=9).value = line_title[8]

        irow += 1
        for j in range(len(std_style)):
            ws.cell(row=irow, column=8).value = std_style[j]
            ws.cell(row=irow, column=9).value = StdDivMeanPercent_style[j]
            irow += 1
        ws.cell(row=2, column=8).fill = PatternFill(fill_type='solid', fgColor=YELLOW)
        ws.cell(row=2, column=9).fill = PatternFill(fill_type='solid', fgColor=YELLOW)
        ws.cell(row=3, column=8).fill = PatternFill(fill_type='solid', fgColor=DARKYELLOW)
        ws.cell(row=3, column=9).fill = PatternFill(fill_type='solid', fgColor=DARKYELLOW)
        ws.cell(row=4, column=8).fill = PatternFill(fill_type='solid', fgColor=RED)
        ws.cell(row=4, column=9).fill = PatternFill(fill_type='solid', fgColor=RED)

        line_loop_times = ['Test times', len(data[2][data[1][i]])]
        for index in range(len(line_loop_times)):
            ws.cell(row=irow, column=index + 1).value = line_loop_times[index]

        irow += 1
        for index in range(len(line_title)):
            ws.cell(row=irow, column=index + 1).value = line_title[index]
            ws.cell(row=irow, column=index + 1).fill = PatternFill(fill_type='solid', fgColor=GREEN)

        irow += 1
        for j in range(len(data[0])):
            line = [data[0][j][0], data[0][j][1], data[0][j][3], data[0][j][2], data[0][j][4][i][0],
                    data[0][j][4][i][1], data[0][j][4][i][2], data[0][j][4][i][3], data[0][j][4][i][4]]
            for index in range(len(line)):
                if index == 7:
                    ws.cell(row=irow, column=index + 1).value = line[index]
                    if 1 < line[index] <= 10:
                        ws.cell(row=irow, column=index + 1).fill = PatternFill(fill_type='solid', fgColor=YELLOW)
                    elif 10 < line[index] <= 100:
                        ws.cell(row=irow, column=index + 1).fill = PatternFill(fill_type='solid', fgColor=DARKYELLOW)
                    elif 100 < line[index]:
                        ws.cell(row=irow, column=index + 1).fill = PatternFill(fill_type='solid', fgColor=RED)
                elif index == 8:
                    ws.cell(row=irow, column=index + 1).value = '{:.2%}'.format(line[index])
                    if -0.5 <= line[index] < -0.05 or 0.05 < line[index] <= 0.5:
                        ws.cell(row=irow, column=index + 1).fill = PatternFill(fill_type='solid', fgColor=YELLOW)
                    elif -5 <= line[index] < -0.5 or 0.5 < line[index] <= 5:
                        ws.cell(row=irow, column=index + 1).fill = PatternFill(fill_type='solid', fgColor=DARKYELLOW)
                    elif line[index] < -5 or 5 < line[index]:
                        ws.cell(row=irow, column=index + 1).fill = PatternFill(fill_type='solid', fgColor=RED)
                else:
                    ws.cell(row=irow, column=index + 1).value = line[index]
            irow += 1
        for row in ws.rows:
            for cell in row:
                cell.border = border
    for sheet_name in wb.sheetnames:
        if sheet_name == 'Sheet':
            del wb[sheet_name]
        else:
            Common.set_column_width(wb[sheet_name])
    print("保存数据开始-----------------")
    wb.save(test_file_name)
    print("保存数据结束-----------------")


def main():
    # Sourcefile folder path
    if argv.count('-s') == 0 and argv.count('-f') == 0:
        print(
            "Error：Sourcefile folder path 或 single file path为必填项，格式：“-s D:\sourcefile” 或 “-f D:\sourcefile\ST5-440mm-out1.csv”。")
        exit()

    if argv.count('-s') != 0:
        sourcefile_folder = argv[argv.index('-s') + 1]
        file_list = Common.get_filelist(sourcefile_folder, '.csv')
        if not file_list:
            exit()

    if argv.count('-f') != 0:
        single_file = argv[argv.index('-f') + 1]
        sourcefile_folder = dirname(single_file)
        file_list = [single_file]

    # Analysis folder path
    if argv.count('-a') == 0:
        analysis_folder = sourcefile_folder + '\Analysis'
    else:
        analysis_folder = argv[argv.index('-a') + 1]

    if argv.count('-b') == 0:
        group_by_id = 0  # 默认根据ChipNo分组
    else:
        group_by_id = int(argv[argv.index('-b') + 1])

    Common.mkdir(analysis_folder)

    for file in file_list:
        # parse file
        data = parse_file(file, group_by_id)

        save_file = join(analysis_folder, basename(file))
        # save data
        save_data(save_file, data)


if __name__ == '__main__':
    main()
