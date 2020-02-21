# encoding:utf-8
# @Time     : 2019/9/9 13:32
# @Author   : Jerry Chou
# @File     :
# @Function : Date data analysis

from csv import reader
from os.path import basename, isdir, join, exists
from os import listdir, makedirs
from datetime import datetime
from math import isnan
from sys import argv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.styles.colors import YELLOW, GREEN, BLACK, WHITE, RED
from progressbar import *
from openpyxl.utils import get_column_letter


def find_item(item_list, value):
    """
    find value in item list
    """
    return [i for i, v in enumerate(item_list) if v == value]


def search_string(data, target):
    """
    find out if target exists in data
    """
    for i in range(len(data)):
        try:
            col_num = data[i].index(target)
            row_num = i
            return row_num, col_num
        except:
            pass
    print("Can't find " + target + " !")
    return False


def set_column_width(sheet):
    """
    set column width
    """
    # get the maximum width of each column
    col_width = [0.5] * sheet.max_column
    for row in range(sheet.max_row):
        for col in range(sheet.max_column):
            value = sheet.cell(row=row + 1, column=col + 1).value
            if value:
                width = len(str(value))
                if width > col_width[col]:
                    col_width[col] = width
    # set column width
    for i in range(len(col_width)):
        col_lettert = get_column_letter(i + 1)
        if col_width[i] > 100:
            # set to 100 if col_width greater than 100
            sheet.column_dimensions[col_lettert].width = 100
        else:
            sheet.column_dimensions[col_lettert].width = col_width[i] + 4


def get_filelist(folder, postfix=None):
    """
    find a list of files with a postfix
    """
    fullname_list = []
    if isdir(folder):
        files = listdir(folder)
        for filename in files:
            fullname_list.append(join(folder, filename))
        if postfix:
            target_file_list = []
            for fullname in fullname_list:
                if fullname.endswith(postfix):
                    target_file_list.append(fullname)
            return target_file_list
        else:
            return fullname_list
    else:
        print("Error：Not a folder!")
        return False


def mkdir(path):
    """
    create a folder
    """
    path = path.strip()
    path = path.rstrip("\\")
    is_exists = exists(path)
    if not is_exists:
        makedirs(path)
        return True
    else:
        return False


row_offset = 5
col_offset = 4
high_limit_row_num = 0
# default project
project = 'F28'
# get current time
now_time = datetime.now().strftime("%Y%m%d%H%M")
alignment = Alignment(horizontal='center', vertical='center')
thin = Side(border_style='thin', color=BLACK)
border = Border(top=thin, left=thin, right=thin, bottom=thin)

hwbin_to_swbin = {
    'F28': {
        1: {'SWBin': (1, 2), 'isPassBin': True},
        2: {'SWBin': (41, 42, 43, 44, 45, 53, 54, 55), 'isPassBin': True},
        4: {'SWBin': (23, 24, 25, 29, 30, 33, 34, 36, 39, 40, 70, 71, 72), 'isPassBin': True},
        5: {'SWBin': (5, 6, 7, 8, 9, 12, 96, 97, 98, 99), 'isPassBin': False},
        6: {'SWBin': (13, 14, 15, 35), 'isPassBin': False},
        8: {'SWBin': (26, 27, 31, 32), 'isPassBin': False}  # separate them out from HWBin4
    },
    'JX828': {
        3: {'SWBin': (1, 2, 3), 'isPassBin': True},
        1: {'SWBin': (63, 64, 65, 89, 90, 94), 'isPassBin': True},
        2: {'SWBin': (53, 54, 73, 74), 'isPassBin': True},
        4: {'SWBin': (23, 24, 25, 26, 27, 29, 30, 31, 32, 33, 34, 36, 39, 40, 56, 57, 58, 75), 'isPassBin': False},
        5: {'SWBin': (5, 6, 7, 8, 9, 12, 93, 96, 98, 99), 'isPassBin': False},
        6: {'SWBin': (13, 14, 15, 35, 46, 47, 48, 51, 60), 'isPassBin': False}
    },
    'JX825': {
        3: {'SWBin': (2, 255), 'isPassBin': True},
        4: {'SWBin': (23, 24, 25, 26, 27, 29, 30, 31, 32, 33, 34, 36, 39, 40, 63, 64, 65), 'isPassBin': False},
        5: {'SWBin': (5, 6, 7, 8, 9, 12, 89), 'isPassBin': False},
        6: {'SWBin': (13, 14, 15, 35, 46, 48, 53, 54, 60), 'isPassBin': False},
        8: {'SWBin': (94, 96, 97, 98, 99), 'isPassBin': False}
    }
}

bin_definition_list = {
    'F28':
        {
            'PCLK_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 1},
            'HSYNC_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 1},
            'VSYNC_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 1},
            'D9_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 1},
            'D8_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 1},
            'D7_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 1},
            'D6_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 1},
            'D5_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 1},
            'D4_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 1},
            'D3_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 1},
            'D2_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 1},
            'D1_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 1},
            'D0_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 1},
            'SCL_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 1},
            'SDA_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 1},
            'RSTB_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 1},
            'PWDN_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 1},
            'DVDD_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 1},
            'VRamp_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 1},
            'VH_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 1},
            'VN1_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 1},
            'EXCLK_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 1},
            'PCLK_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 2},
            'HSYNC_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 2},
            'VSYNC_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 2},
            'D9_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 2},
            'D8_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 2},
            'D7_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 2},
            'D6_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 2},
            'D5_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 2},
            'D4_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 2},
            'D3_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 2},
            'D2_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 2},
            'D1_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 2},
            'D0_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 2},
            'SCL_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 2},
            'SDA_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 2},
            'RSTB_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 2},
            'PWDN_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 2},
            'EXCLK_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 2},
            'PCLK_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 2},
            'HSYNC_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 2},
            'VSYNC_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 2},
            'D9_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 2},
            'D8_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 2},
            'D7_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 2},
            'D6_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 2},
            'D5_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 2},
            'D4_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 2},
            'D3_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 2},
            'D2_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 2},
            'D1_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 2},
            'D0_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 2},
            'SCL_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 2},
            'SDA_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 2},
            'RSTB_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 2},
            'PWDN_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 2},
            'EXCLK_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 2},
            'iic_test': {'SWBin': 7, 'HWBin': 5, 'Priority': 3},
            'DVDD_voltage': {'SWBin': 9, 'HWBin': 5, 'Priority': 4},
            'VH_voltage': {'SWBin': 9, 'HWBin': 5, 'Priority': 5},
            'VN1_voltage': {'SWBin': 9, 'HWBin': 5, 'Priority': 6},
            'Active_AVDD': {'SWBin': 12, 'HWBin': 5, 'Priority': 7},
            'Active_DOVDD': {'SWBin': 12, 'HWBin': 5, 'Priority': 8},
            'PWDN_AVDD': {'SWBin': 8, 'HWBin': 5, 'Priority': 9},
            'PWDN_DOVDD': {'SWBin': 8, 'HWBin': 5, 'Priority': 10},
            'PWDN_Total': {'SWBin': 8, 'HWBin': 5, 'Priority': 11},
            'PLCK_Freq': {'SWBin': 9, 'HWBin': 5, 'Priority': 12},
            'BK_ImageCapture': {'SWBin': 98, 'HWBin': 5, 'Priority': 13},
            'BK_ImageCalculation': {'SWBin': 97, 'HWBin': 5, 'Priority': 14},
            'BLC_R': {'SWBin': 7, 'HWBin': 5, 'Priority': 15},
            'BLC_Gr': {'SWBin': 7, 'HWBin': 5, 'Priority': 15},
            'BLC_Gb': {'SWBin': 7, 'HWBin': 5, 'Priority': 15},
            'BLC_B': {'SWBin': 7, 'HWBin': 5, 'Priority': 15},
            'BK_DeadRowExBPix_R': {'SWBin': 15, 'HWBin': 6, 'Priority': 16, 'IsNan': 'BK_ImageCapture'},
            'BK_DeadRowExBPix_Gr': {'SWBin': 15, 'HWBin': 6, 'Priority': 16},
            'BK_DeadRowExBPix_Gb': {'SWBin': 15, 'HWBin': 6, 'Priority': 16},
            'BK_DeadRowExBPix_B': {'SWBin': 15, 'HWBin': 6, 'Priority': 16},
            'BK_DeadColExBPix_R': {'SWBin': 14, 'HWBin': 6, 'Priority': 17},
            'BK_DeadColExBPix_Gr': {'SWBin': 14, 'HWBin': 6, 'Priority': 17},
            'BK_DeadColExBPix_Gb': {'SWBin': 14, 'HWBin': 6, 'Priority': 17},
            'BK_DeadColExBPix_B': {'SWBin': 14, 'HWBin': 6, 'Priority': 17},
            'BK_Mean_R': {'SWBin': 13, 'HWBin': 6, 'Priority': 18},
            'BK_Mean_Gr': {'SWBin': 13, 'HWBin': 6, 'Priority': 19},
            'BK_Mean_Gb': {'SWBin': 13, 'HWBin': 6, 'Priority': 20},
            'BK_Mean_B': {'SWBin': 13, 'HWBin': 6, 'Priority': 21},
            'BK_StdDEV_R': {'SWBin': 13, 'HWBin': 6, 'Priority': 22},
            'BK_StdDEV_Gr': {'SWBin': 13, 'HWBin': 6, 'Priority': 23},
            'BK_StdDEV_Gb': {'SWBin': 13, 'HWBin': 6, 'Priority': 24},
            'BK_StdDEV_B': {'SWBin': 13, 'HWBin': 6, 'Priority': 25},
            'LT_ImageCapture': {'SWBin': 99, 'HWBin': 5, 'Priority': 26},
            'LT_ImageCalculation': {'SWBin': 97, 'HWBin': 5, 'Priority': 27},
            'LT_DRow_R': {'SWBin': 25, 'HWBin': 4, 'Priority': 28, 'IsNan': 'LT_ImageCapture'},
            'LT_DRow_Gr': {'SWBin': 25, 'HWBin': 4, 'Priority': 28},
            'LT_DRow_Gb': {'SWBin': 25, 'HWBin': 4, 'Priority': 28},
            'LT_DRow_B': {'SWBin': 25, 'HWBin': 4, 'Priority': 28},
            'LT_DCol_R': {'SWBin': 24, 'HWBin': 4, 'Priority': 29},
            'LT_DCol_Gr': {'SWBin': 24, 'HWBin': 4, 'Priority': 29},
            'LT_DCol_Gb': {'SWBin': 24, 'HWBin': 4, 'Priority': 29},
            'LT_DCol_B': {'SWBin': 24, 'HWBin': 4, 'Priority': 29},
            'LT_DRow_Color': {'SWBin': 25, 'HWBin': 4, 'Priority': 30},
            'LT_DCol_Color': {'SWBin': 24, 'HWBin': 4, 'Priority': 31},
            'LT_Mean_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 32},
            'LT_Mean_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 33},
            'LT_Mean_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 34},
            'LT_Mean_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 35},
            'LT_StdDEV_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 36},
            'LT_StdDEV_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 37},
            'LT_StdDEV_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 38},
            'LT_StdDEV_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 39},
            'LT_RI_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 40},
            'LT_RI_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 41},
            'LT_RI_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 42},
            'LT_RI_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 43},
            'LT_Ratio_GrR': {'SWBin': 23, 'HWBin': 4, 'Priority': 44},
            'LT_Ratio_GbR': {'SWBin': 23, 'HWBin': 4, 'Priority': 45},
            'LT_Ratio_GrB': {'SWBin': 23, 'HWBin': 4, 'Priority': 46},
            'LT_Ratio_GbB': {'SWBin': 23, 'HWBin': 4, 'Priority': 47},
            'LT_Ratio_GbGr': {'SWBin': 23, 'HWBin': 4, 'Priority': 48},
            'LT_LostBit': {'SWBin': 23, 'HWBin': 4, 'Priority': 49},
            'LT_LostBitSNR_SumDIFF_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 50},
            'LT_LostBitSNR_SumDIFF_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 51},
            'LT_LostBitSNR_SumDIFF_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 52},
            'LT_LostBitSNR_SumDIFF_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 53},
            'LB_LT_ImageCapture': {'SWBin': 99, 'HWBin': 5, 'Priority': 54},
            'LB_LT_ImageCalculation': {'SWBin': 97, 'HWBin': 5, 'Priority': 55},
            'LB_LT_Mean_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 56, 'IsNan': 'LB_LT_ImageCapture'},
            'LB_LT_Mean_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 57},
            'LB_LT_Mean_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 58},
            'LB_LT_Mean_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 59},
            'LB_LT_LostBit': {'SWBin': 23, 'HWBin': 4, 'Priority': 60},
            'LB_LT_LostBitSNR_SumDIFF_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 61},
            'LB_LT_LostBitSNR_SumDIFF_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 62},
            'LB_LT_LostBitSNR_SumDIFF_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 63},
            'LB_LT_LostBitSNR_SumDIFF_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 64},
            'FW_LT_ImageCapture': {'SWBin': 99, 'HWBin': 5, 'Priority': 65},
            'FW_LT_Calculation': {'SWBin': 97, 'HWBin': 5, 'Priority': 66},
            'FW_LT_DRow_R': {'SWBin': 36, 'HWBin': 4, 'Priority': 67, 'IsNan': 'FW_LT_ImageCapture'},
            'FW_LT_DRow_Gr': {'SWBin': 36, 'HWBin': 4, 'Priority': 67},
            'FW_LT_DRow_Gb': {'SWBin': 36, 'HWBin': 4, 'Priority': 67},
            'FW_LT_DRow_B': {'SWBin': 36, 'HWBin': 4, 'Priority': 67},
            'FW_LT_DCol_R': {'SWBin': 36, 'HWBin': 4, 'Priority': 68},
            'FW_LT_DCol_Gr': {'SWBin': 36, 'HWBin': 4, 'Priority': 68},
            'FW_LT_DCol_Gb': {'SWBin': 36, 'HWBin': 4, 'Priority': 68},
            'FW_LT_DCol_B': {'SWBin': 36, 'HWBin': 4, 'Priority': 68},
            'FW_LT_DRow_Color': {'SWBin': 36, 'HWBin': 4, 'Priority': 69},
            'FW_LT_DCol_Color': {'SWBin': 36, 'HWBin': 4, 'Priority': 70},
            'FW_LT_Mean_R': {'SWBin': 36, 'HWBin': 4, 'Priority': 77},
            'FW_LT_Mean_Gr': {'SWBin': 36, 'HWBin': 4, 'Priority': 78},
            'FW_LT_Mean_Gb': {'SWBin': 36, 'HWBin': 4, 'Priority': 79},
            'FW_LT_Mean_B': {'SWBin': 36, 'HWBin': 4, 'Priority': 80},
            'FW_LT_StdDEV_R': {'SWBin': 36, 'HWBin': 4, 'Priority': 81},
            'FW_LT_StdDEV_Gr': {'SWBin': 36, 'HWBin': 4, 'Priority': 82},
            'FW_LT_StdDEV_Gb': {'SWBin': 36, 'HWBin': 4, 'Priority': 83},
            'FW_LT_StdDEV_B': {'SWBin': 36, 'HWBin': 4, 'Priority': 84},
            'FW_LT_RI_R': {'SWBin': 36, 'HWBin': 4, 'Priority': 85},
            'FW_LT_RI_Gr': {'SWBin': 36, 'HWBin': 4, 'Priority': 86},
            'FW_LT_RI_Gb': {'SWBin': 36, 'HWBin': 4, 'Priority': 87},
            'FW_LT_RI_B': {'SWBin': 36, 'HWBin': 4, 'Priority': 88},
            'FW_LT_Ratio_GrR': {'SWBin': 36, 'HWBin': 4, 'Priority': 89},
            'FW_LT_Ratio_GbR': {'SWBin': 36, 'HWBin': 4, 'Priority': 90},
            'FW_LT_Ratio_GrB': {'SWBin': 36, 'HWBin': 4, 'Priority': 91},
            'FW_LT_Ratio_GbB': {'SWBin': 36, 'HWBin': 4, 'Priority': 92},
            'FW_LT_Ratio_GbGr': {'SWBin': 36, 'HWBin': 4, 'Priority': 93},
            'LT_CornerLine': {'SWBin': 26, 'HWBin': 8, 'Priority': 94},  # 4
            'LT_ScratchLine': {'SWBin': 26, 'HWBin': 8, 'Priority': 95},  # 4
            'LT_Blemish': {'SWBin': 31, 'HWBin': 8, 'Priority': 96},  # 4
            'LT_LineStripe': {'SWBin': 27, 'HWBin': 8, 'Priority': 97},  # 4
            'LT_Particle': {'SWBin': 32, 'HWBin': 8, 'Priority': 98},  # 4
            'BK_Cluster2': {'SWBin': 30, 'HWBin': 4, 'Priority': 99},
            'LT_Cluster2': {'SWBin': 30, 'HWBin': 4, 'Priority': 100},
            'BK_Cluster1': {'SWBin': 29, 'HWBin': 4, 'Priority': 101},
            'LT_Cluster1': {'SWBin': 29, 'HWBin': 4, 'Priority': 102},
            'WP_Count': {'SWBin': 35, 'HWBin': 6, 'Priority': 103},
            'BK_Cluster3GrGb': {'SWBin': 39, 'HWBin': 4, 'Priority': 104},
            'LT_Cluster3GrGb': {'SWBin': 40, 'HWBin': 4, 'Priority': 105},
            'BK_Cluster3SubtractGrGb': {'SWBin': 33, 'HWBin': 4, 'Priority': 106},
            'LT_Cluster3SubtractGrGb': {'SWBin': 34, 'HWBin': 4, 'Priority': 107},
            'LB_BK_DeadRowExBPix_R': {'SWBin': 45, 'HWBin': 2, 'Priority': 108},
            'LB_BK_DeadRowExBPix_Gr': {'SWBin': 45, 'HWBin': 2, 'Priority': 108},
            'LB_BK_DeadRowExBPix_Gb': {'SWBin': 45, 'HWBin': 2, 'Priority': 108},
            'LB_BK_DeadRowExBPix_B': {'SWBin': 45, 'HWBin': 2, 'Priority': 108},
            'LB_BK_DeadColExBPix_R': {'SWBin': 44, 'HWBin': 2, 'Priority': 109},
            'LB_BK_DeadColExBPix_Gr': {'SWBin': 44, 'HWBin': 2, 'Priority': 109},
            'LB_BK_DeadColExBPix_Gb': {'SWBin': 44, 'HWBin': 2, 'Priority': 109},
            'LB_BK_DeadColExBPix_B': {'SWBin': 44, 'HWBin': 2, 'Priority': 109},
            'SR_LT_ImageCapture': {'SWBin': 96, 'HWBin': 5, 'Priority': 110},
            'SR_LT_Calculation': {'SWBin': 97, 'HWBin': 5, 'Priority': 110},
            'SR_LT_DRow_R': {'SWBin': 55, 'HWBin': 2, 'Priority': 111, 'IsNan': 'SR_LT_ImageCapture'},
            'SR_LT_DRow_Gr': {'SWBin': 55, 'HWBin': 2, 'Priority': 111},
            'SR_LT_DRow_Gb': {'SWBin': 55, 'HWBin': 2, 'Priority': 111},
            'SR_LT_DRow_B': {'SWBin': 55, 'HWBin': 2, 'Priority': 111},
            'SR_LT_DCol_R': {'SWBin': 54, 'HWBin': 2, 'Priority': 112},
            'SR_LT_DCol_Gr': {'SWBin': 54, 'HWBin': 2, 'Priority': 112},
            'SR_LT_DCol_Gb': {'SWBin': 54, 'HWBin': 2, 'Priority': 112},
            'SR_LT_DCol_B': {'SWBin': 54, 'HWBin': 2, 'Priority': 112},
            'SR_LT_DRow_Color': {'SWBin': 55, 'HWBin': 2, 'Priority': 113},
            'SR_LT_DCol_Color': {'SWBin': 54, 'HWBin': 2, 'Priority': 114},
            'SR_LT_Mean_R': {'SWBin': 53, 'HWBin': 2, 'Priority': 115},
            'SR_LT_Mean_Gr': {'SWBin': 53, 'HWBin': 2, 'Priority': 116},
            'SR_LT_Mean_Gb': {'SWBin': 53, 'HWBin': 2, 'Priority': 117},
            'SR_LT_Mean_B': {'SWBin': 53, 'HWBin': 2, 'Priority': 118},
            'SR_LT_StdDEV_R': {'SWBin': 53, 'HWBin': 2, 'Priority': 119},
            'SR_LT_StdDEV_Gr': {'SWBin': 53, 'HWBin': 2, 'Priority': 120},
            'SR_LT_StdDEV_Gb': {'SWBin': 53, 'HWBin': 2, 'Priority': 121},
            'SR_LT_StdDEV_B': {'SWBin': 53, 'HWBin': 2, 'Priority': 122},
            'SR_LT_RI_R': {'SWBin': 53, 'HWBin': 2, 'Priority': 123},
            'SR_LT_RI_Gr': {'SWBin': 53, 'HWBin': 2, 'Priority': 124},
            'SR_LT_RI_Gb': {'SWBin': 53, 'HWBin': 2, 'Priority': 125},
            'SR_LT_RI_B': {'SWBin': 53, 'HWBin': 2, 'Priority': 126},
            'SR_LT_Ratio_GrR': {'SWBin': 53, 'HWBin': 2, 'Priority': 127},
            'SR_LT_Ratio_GbR': {'SWBin': 53, 'HWBin': 2, 'Priority': 128},
            'SR_LT_Ratio_GrB': {'SWBin': 53, 'HWBin': 2, 'Priority': 129},
            'SR_LT_Ratio_GbB': {'SWBin': 53, 'HWBin': 2, 'Priority': 130},
            'SR_LT_Ratio_GbGr': {'SWBin': 53, 'HWBin': 2, 'Priority': 131},
            'SR_LT_LostBit': {'SWBin': 53, 'HWBin': 2, 'Priority': 132},
            'SR_BK_ImageCapture': {'SWBin': 96, 'HWBin': 5, 'Priority': 133},
            'SR_BK_Calculation': {'SWBin': 97, 'HWBin': 5, 'Priority': 134},
            'SR_BK_DeadRowExBPix_R': {'SWBin': 42, 'HWBin': 2, 'Priority': 135, 'IsNan': 'SR_BK_ImageCapture'},
            'SR_BK_DeadRowExBPix_Gr': {'SWBin': 42, 'HWBin': 2, 'Priority': 135},
            'SR_BK_DeadRowExBPix_Gb': {'SWBin': 42, 'HWBin': 2, 'Priority': 135},
            'SR_BK_DeadRowExBPix_B': {'SWBin': 42, 'HWBin': 2, 'Priority': 135},
            'SR_BK_DeadColExBPix_R': {'SWBin': 41, 'HWBin': 2, 'Priority': 136},
            'SR_BK_DeadColExBPix_Gr': {'SWBin': 41, 'HWBin': 2, 'Priority': 136},
            'SR_BK_DeadColExBPix_Gb': {'SWBin': 41, 'HWBin': 2, 'Priority': 136},
            'SR_BK_DeadColExBPix_B': {'SWBin': 41, 'HWBin': 2, 'Priority': 136},
            'SR_BK_Mean_R': {'SWBin': 43, 'HWBin': 2, 'Priority': 137},
            'SR_BK_Mean_Gr': {'SWBin': 43, 'HWBin': 2, 'Priority': 138},
            'SR_BK_Mean_Gb': {'SWBin': 43, 'HWBin': 2, 'Priority': 139},
            'SR_BK_Mean_B': {'SWBin': 43, 'HWBin': 2, 'Priority': 140},
            'SR_BK_StdDEV_R': {'SWBin': 43, 'HWBin': 2, 'Priority': 141},
            'SR_BK_StdDEV_Gr': {'SWBin': 43, 'HWBin': 2, 'Priority': 142},
            'SR_BK_StdDEV_Gb': {'SWBin': 43, 'HWBin': 2, 'Priority': 143},
            'SR_BK_StdDEV_B': {'SWBin': 43, 'HWBin': 2, 'Priority': 144},
            'Regular Line Pattern Change Fail': {'SWBin': 70, 'HWBin': 4, 'Priority': 145},  # not use
            'Diffuser2_LT_ImageCapture': {'SWBin': 99, 'HWBin': 5, 'Priority': 146},
            'Diffuser2_LT_Calculation': {'SWBin': 97, 'HWBin': 5, 'Priority': 147},
            'Diffuser2_LT_WeakLineRow_R': {'SWBin': 72, 'HWBin': 4, 'Priority': 148,
                                           'IsNan': 'Diffuser2_LT_ImageCapture'},
            'Diffuser2_LT_WeakLineRow_Gr': {'SWBin': 72, 'HWBin': 4, 'Priority': 148},
            'Diffuser2_LT_WeakLineRow_Gb': {'SWBin': 72, 'HWBin': 4, 'Priority': 148},
            'Diffuser2_LT_WeakLineRow_B': {'SWBin': 72, 'HWBin': 4, 'Priority': 148},
            'Diffuser2_LT_WeakLineCol_R': {'SWBin': 71, 'HWBin': 4, 'Priority': 149},
            'Diffuser2_LT_WeakLineCol_Gr': {'SWBin': 71, 'HWBin': 4, 'Priority': 149},
            'Diffuser2_LT_WeakLineCol_Gb': {'SWBin': 71, 'HWBin': 4, 'Priority': 149},
            'Diffuser2_LT_WeakLineCol_B': {'SWBin': 71, 'HWBin': 4, 'Priority': 149},
            'Binning': {'SWBin': 2, 'HWBin': 1, 'Priority': 145},
            'All Pass': {'SWBin': 1, 'HWBin': 1, 'Priority': 146}
        },
    'JX828':
        {
            'VSYNC_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 1},
            'HSYNC_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 2},
            'PCLK_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 3},
            'EXCLK_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 4},
            'RSTB_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 5},
            'PWDN_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 6},
            'SDA_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 7},
            'SCL_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 8},
            'D0_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 9},
            'D1_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 10},
            'D2_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 11},
            'D3_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 12},
            'D4_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 13},
            'D5_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 14},
            'D6_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 15},
            'D7_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 16},
            'D8_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 17},
            'D9_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 18},
            'MCP_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 19},
            'MCN_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 20},
            'MDP0_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 21},
            'MDN0_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 22},
            'MDP1_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 23},
            'MDN1_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 24},
            'AVDD_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 25},
            'DOVDD_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 26},
            'VRamp_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 27},
            'VH_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 28},
            'VN1_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 29},
            'VSYNC_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 30},
            'HSYNC_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 31},
            'PCLK_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 32},
            'EXCLK_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 33},
            'RSTB_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 34},
            'PWDN_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 35},
            'SDA_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 36},
            'SCL_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 37},
            'D0_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 38},
            'D1_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 39},
            'D2_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 40},
            'D4_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 41},
            'D5_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 42},
            'D3_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 43},
            'D6_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 44},
            'D7_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 45},
            'D8_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 46},
            'D9_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 47},
            'MCP_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 48},
            'MCN_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 49},
            'MDP0_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 50},
            'MDN0_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 51},
            'MDP1_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 52},
            'MDN1_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 53},
            'VSYNC_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 54},
            'HSYNC_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 55},
            'PCLK_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 56},
            'EXCLK_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 57},
            'RSTB_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 58},
            'PWDN_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 59},
            'SDA_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 60},
            'SCL_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 61},
            'D0_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 62},
            'D1_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 63},
            'D2_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 64},
            'D4_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 65},
            'D5_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 66},
            'D3_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 67},
            'D6_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 68},
            'D7_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 69},
            'D8_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 70},
            'D9_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 71},
            'MCP_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 72},
            'MCN_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 73},
            'MDP0_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 74},
            'MDN0_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 75},
            'MDP1_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 76},
            'MDN1_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 77},
            'iic_test': {'SWBin': 7, 'HWBin': 5, 'Priority': 78},
            'DVDD_voltage': {'SWBin': 9, 'HWBin': 5, 'Priority': 79},
            'VH_voltage': {'SWBin': 9, 'HWBin': 5, 'Priority': 80},
            'VN1_voltage': {'SWBin': 9, 'HWBin': 5, 'Priority': 81},
            'Active_AVDD': {'SWBin': 12, 'HWBin': 5, 'Priority': 82},
            'Active_DOVDD': {'SWBin': 12, 'HWBin': 5, 'Priority': 83},
            'PWDN_AVDD': {'SWBin': 8, 'HWBin': 5, 'Priority': 84},
            'PWDN_DOVDD': {'SWBin': 8, 'HWBin': 5, 'Priority': 85},
            'PWDN_Total': {'SWBin': 8, 'HWBin': 5, 'Priority': 86},
            'DVP27M30F_BK_Cap': {'SWBin': 98, 'HWBin': 5, 'Priority': 87},
            'DVP27M30F_BK_Calc': {'SWBin': 97, 'HWBin': 5, 'Priority': 88},
            'DVP27M30F_BK_REG_VRamp': {'SWBin': 7, 'HWBin': 5, 'Priority': 89},
            'BLC_R': {'SWBin': 7, 'HWBin': 5, 'Priority': 90},
            'BLC_Gr': {'SWBin': 7, 'HWBin': 5, 'Priority': 91},
            'BLC_Gb': {'SWBin': 7, 'HWBin': 5, 'Priority': 92},
            'BLC_B': {'SWBin': 7, 'HWBin': 5, 'Priority': 93},
            'BK_DeadRowExBPix_R': {'SWBin': 15, 'HWBin': 6, 'Priority': 94, 'IsNan': 'DVP27M30F_BK_Cap'},
            'BK_DeadRowExBPix_Gr': {'SWBin': 15, 'HWBin': 6, 'Priority': 95},
            'BK_DeadRowExBPix_Gb': {'SWBin': 15, 'HWBin': 6, 'Priority': 96},
            'BK_DeadRowExBPix_B': {'SWBin': 15, 'HWBin': 6, 'Priority': 97},
            'BK_DeadColExBPix_R': {'SWBin': 14, 'HWBin': 6, 'Priority': 98},
            'BK_DeadColExBPix_Gr': {'SWBin': 14, 'HWBin': 6, 'Priority': 99},
            'BK_DeadColExBPix_Gb': {'SWBin': 14, 'HWBin': 6, 'Priority': 100},
            'BK_DeadColExBPix_B': {'SWBin': 14, 'HWBin': 6, 'Priority': 101},
            'BK_MixDCol_VfpnQty_R': {'SWBin': 14, 'HWBin': 6, 'Priority': 102},
            'BK_MixDCol_VfpnQty_Gr': {'SWBin': 14, 'HWBin': 6, 'Priority': 103},
            'BK_MixDCol_VfpnQty_Gb': {'SWBin': 14, 'HWBin': 6, 'Priority': 104},
            'BK_MixDCol_VfpnQty_B': {'SWBin': 14, 'HWBin': 6, 'Priority': 105},
            'BK_MixDCol_VFPN_MaxValue0_R': {'SWBin': 14, 'HWBin': 6, 'Priority': 106},
            'BK_MixDCol_VFPN_MaxValue0_Gr': {'SWBin': 14, 'HWBin': 6, 'Priority': 107},
            'BK_MixDCol_VFPN_MaxValue0_Gb': {'SWBin': 14, 'HWBin': 6, 'Priority': 108},
            'BK_MixDCol_VFPN_MaxValue0_B': {'SWBin': 14, 'HWBin': 6, 'Priority': 109},
            'BK_MixDCol_VFPN_MaxValue1_R': {'SWBin': 14, 'HWBin': 6, 'Priority': 110},
            'BK_MixDCol_VFPN_MaxValue1_Gr': {'SWBin': 14, 'HWBin': 6, 'Priority': 111},
            'BK_MixDCol_VFPN_MaxValue1_Gb': {'SWBin': 14, 'HWBin': 6, 'Priority': 112},
            'BK_MixDCol_VFPN_MaxValue1_B': {'SWBin': 14, 'HWBin': 6, 'Priority': 113},
            'BK_MixDCol_VFPN_MaxValue2_R': {'SWBin': 14, 'HWBin': 6, 'Priority': 114},
            'BK_MixDCol_VFPN_MaxValue2_Gr': {'SWBin': 14, 'HWBin': 6, 'Priority': 115},
            'BK_MixDCol_VFPN_MaxValue2_Gb': {'SWBin': 14, 'HWBin': 6, 'Priority': 116},
            'BK_MixDCol_VFPN_MaxValue2_B': {'SWBin': 14, 'HWBin': 6, 'Priority': 117},
            'BK_MixDCol_VFPN_MaxValue3_R': {'SWBin': 14, 'HWBin': 6, 'Priority': 118},
            'BK_MixDCol_VFPN_MaxValue3_Gr': {'SWBin': 14, 'HWBin': 6, 'Priority': 119},
            'BK_MixDCol_VFPN_MaxValue3_Gb': {'SWBin': 14, 'HWBin': 6, 'Priority': 120},
            'BK_MixDCol_VFPN_MaxValue3_B': {'SWBin': 14, 'HWBin': 6, 'Priority': 121},
            'BK_MixDCol_VFPN_MaxValue4_R': {'SWBin': 14, 'HWBin': 6, 'Priority': 122},
            'BK_MixDCol_VFPN_MaxValue4_Gr': {'SWBin': 14, 'HWBin': 6, 'Priority': 123},
            'BK_MixDCol_VFPN_MaxValue4_Gb': {'SWBin': 14, 'HWBin': 6, 'Priority': 124},
            'BK_MixDCol_VFPN_MaxValue4_B': {'SWBin': 14, 'HWBin': 6, 'Priority': 125},
            'BK_Mean_R': {'SWBin': 13, 'HWBin': 6, 'Priority': 126},
            'BK_Mean_Gr': {'SWBin': 13, 'HWBin': 6, 'Priority': 127},
            'BK_Mean_Gb': {'SWBin': 13, 'HWBin': 6, 'Priority': 128},
            'BK_Mean_B': {'SWBin': 13, 'HWBin': 6, 'Priority': 129},
            'BK_StdDEV_R': {'SWBin': 13, 'HWBin': 6, 'Priority': 130},
            'BK_StdDEV_Gr': {'SWBin': 13, 'HWBin': 6, 'Priority': 131},
            'BK_StdDEV_Gb': {'SWBin': 13, 'HWBin': 6, 'Priority': 132},
            'BK_StdDEV_B': {'SWBin': 13, 'HWBin': 6, 'Priority': 133},
            'DVP27M30F_LT_Cap': {'SWBin': 99, 'HWBin': 5, 'Priority': 134},
            'DVP27M30F_LT_Calc': {'SWBin': 97, 'HWBin': 5, 'Priority': 135},
            'PLCK_Freq': {'SWBin': 93, 'HWBin': 5, 'Priority': 136},
            'LT_DRow_R': {'SWBin': 25, 'HWBin': 4, 'Priority': 137, 'IsNan': 'DVP27M30F_LT_Cap'},
            'LT_DRow_Gr': {'SWBin': 25, 'HWBin': 4, 'Priority': 138},
            'LT_DRow_Gb': {'SWBin': 25, 'HWBin': 4, 'Priority': 139},
            'LT_DRow_B': {'SWBin': 25, 'HWBin': 4, 'Priority': 140},
            'LT_DCol_R': {'SWBin': 24, 'HWBin': 4, 'Priority': 141},
            'LT_DCol_Gr': {'SWBin': 24, 'HWBin': 4, 'Priority': 142},
            'LT_DCol_Gb': {'SWBin': 24, 'HWBin': 4, 'Priority': 143},
            'LT_DCol_B': {'SWBin': 24, 'HWBin': 4, 'Priority': 144},
            'LT_DRow_Color': {'SWBin': 25, 'HWBin': 4, 'Priority': 145},
            'LT_DCol_Color': {'SWBin': 24, 'HWBin': 4, 'Priority': 146},
            'LT_WeakLineRow_R': {'SWBin': 25, 'HWBin': 4, 'Priority': 147},
            'LT_WeakLineRow_Gr': {'SWBin': 25, 'HWBin': 4, 'Priority': 148},
            'LT_WeakLineRow_Gb': {'SWBin': 25, 'HWBin': 4, 'Priority': 149},
            'LT_WeakLineRow_B': {'SWBin': 25, 'HWBin': 4, 'Priority': 150},
            'LT_WeakLineCol_R': {'SWBin': 24, 'HWBin': 4, 'Priority': 151},
            'LT_WeakLineCol_Gr': {'SWBin': 24, 'HWBin': 4, 'Priority': 152},
            'LT_WeakLineCol_Gb': {'SWBin': 24, 'HWBin': 4, 'Priority': 153},
            'LT_WeakLineCol_B': {'SWBin': 24, 'HWBin': 4, 'Priority': 154},
            'LT_Mean_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 155},
            'LT_Mean_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 156},
            'LT_Mean_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 157},
            'LT_Mean_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 158},
            'LT_StdDEV_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 159},
            'LT_StdDEV_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 160},
            'LT_StdDEV_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 161},
            'LT_StdDEV_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 162},
            'LT_RI_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 163},
            'LT_RI_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 164},
            'LT_RI_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 165},
            'LT_RI_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 166},
            'LT_Ratio_GrR': {'SWBin': 23, 'HWBin': 4, 'Priority': 167},
            'LT_Ratio_GbR': {'SWBin': 23, 'HWBin': 4, 'Priority': 168},
            'LT_Ratio_GrB': {'SWBin': 23, 'HWBin': 4, 'Priority': 169},
            'LT_Ratio_GbB': {'SWBin': 23, 'HWBin': 4, 'Priority': 170},
            'LT_Ratio_GbGr': {'SWBin': 23, 'HWBin': 4, 'Priority': 171},
            'LT_LostBit': {'SWBin': 23, 'HWBin': 4, 'Priority': 172},
            'LT_LostBitSNR_SumDIFF_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 173},
            'LT_LostBitSNR_SumDIFF_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 174},
            'LT_LostBitSNR_SumDIFF_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 175},
            'LT_LostBitSNR_SumDIFF_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 176},
            'DVP27M30F_LBIT_Cap': {'SWBin': 99, 'HWBin': 5, 'Priority': 177},
            'DVP27M30F_LBIT_Calc': {'SWBin': 97, 'HWBin': 5, 'Priority': 178},
            'LBIT_LT_Mean_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 179, 'IsNan': 'DVP27M30F_LBIT_Cap'},
            'LBIT_LT_Mean_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 180},
            'LBIT_LT_Mean_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 181},
            'LBIT_LT_Mean_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 182},
            'LBIT_LT_LostBit': {'SWBin': 23, 'HWBin': 4, 'Priority': 183},
            'LBIT_LT_LostBitSNR_SumDIFF_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 184},
            'LBIT_LT_LostBitSNR_SumDIFF_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 185},
            'LBIT_LT_LostBitSNR_SumDIFF_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 186},
            'LBIT_LT_LostBitSNR_SumDIFF_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 187},
            'DVP27M30F_FW_Cap': {'SWBin': 99, 'HWBin': 5, 'Priority': 188},
            'DVP27M30F_FW_Calc': {'SWBin': 97, 'HWBin': 5, 'Priority': 189},
            'FW_LT_DRow_R': {'SWBin': 36, 'HWBin': 4, 'Priority': 190, 'IsNan': 'DVP27M30F_FW_Cap'},
            'FW_LT_DRow_Gr': {'SWBin': 36, 'HWBin': 4, 'Priority': 191},
            'FW_LT_DRow_Gb': {'SWBin': 36, 'HWBin': 4, 'Priority': 192},
            'FW_LT_DRow_B': {'SWBin': 36, 'HWBin': 4, 'Priority': 193},
            'FW_LT_DCol_R': {'SWBin': 36, 'HWBin': 4, 'Priority': 194},
            'FW_LT_DCol_Gr': {'SWBin': 36, 'HWBin': 4, 'Priority': 195},
            'FW_LT_DCol_Gb': {'SWBin': 36, 'HWBin': 4, 'Priority': 196},
            'FW_LT_DCol_B': {'SWBin': 36, 'HWBin': 4, 'Priority': 197},
            'FW_LT_DRow_Color': {'SWBin': 36, 'HWBin': 4, 'Priority': 198},
            'FW_LT_DCol_Color': {'SWBin': 36, 'HWBin': 4, 'Priority': 199},
            'FW_LT_Mean_R': {'SWBin': 36, 'HWBin': 4, 'Priority': 200},
            'FW_LT_Mean_Gr': {'SWBin': 36, 'HWBin': 4, 'Priority': 201},
            'FW_LT_Mean_Gb': {'SWBin': 36, 'HWBin': 4, 'Priority': 202},
            'FW_LT_Mean_B': {'SWBin': 36, 'HWBin': 4, 'Priority': 203},
            'FW_LT_StdDEV_R': {'SWBin': 36, 'HWBin': 4, 'Priority': 204},
            'FW_LT_StdDEV_Gr': {'SWBin': 36, 'HWBin': 4, 'Priority': 205},
            'FW_LT_StdDEV_Gb': {'SWBin': 36, 'HWBin': 4, 'Priority': 206},
            'FW_LT_StdDEV_B': {'SWBin': 36, 'HWBin': 4, 'Priority': 207},
            'FW_LT_RI_R': {'SWBin': 36, 'HWBin': 4, 'Priority': 208},
            'FW_LT_RI_Gr': {'SWBin': 36, 'HWBin': 4, 'Priority': 209},
            'FW_LT_RI_Gb': {'SWBin': 36, 'HWBin': 4, 'Priority': 210},
            'FW_LT_RI_B': {'SWBin': 36, 'HWBin': 4, 'Priority': 211},
            'FW_LT_Ratio_GrR': {'SWBin': 36, 'HWBin': 4, 'Priority': 212},
            'FW_LT_Ratio_GbR': {'SWBin': 36, 'HWBin': 4, 'Priority': 213},
            'FW_LT_Ratio_GrB': {'SWBin': 36, 'HWBin': 4, 'Priority': 214},
            'FW_LT_Ratio_GbB': {'SWBin': 36, 'HWBin': 4, 'Priority': 215},
            'FW_LT_Ratio_GbGr': {'SWBin': 36, 'HWBin': 4, 'Priority': 216},
            'FW_LT_LostBit': {'SWBin': 36, 'HWBin': 4, 'Priority': 217},
            'DVP27M30F_BSun_Cap': {'SWBin': 96, 'HWBin': 5, 'Priority': 218},
            'DVP27M30F_BSun_Calc': {'SWBin': 97, 'HWBin': 5, 'Priority': 219},
            'BSun_BK_DeadRowExBPix_R': {'SWBin': 48, 'HWBin': 6, 'Priority': 220, 'IsNan': 'DVP27M30F_BSun_Cap'},
            'BSun_BK_DeadRowExBPix_Gr': {'SWBin': 48, 'HWBin': 6, 'Priority': 221},
            'BSun_BK_DeadRowExBPix_Gb': {'SWBin': 48, 'HWBin': 6, 'Priority': 222},
            'BSun_BK_DeadRowExBPix_B': {'SWBin': 48, 'HWBin': 6, 'Priority': 223},
            'BSun_BK_DeadColExBPix_R': {'SWBin': 47, 'HWBin': 6, 'Priority': 224},
            'BSun_BK_DeadColExBPix_Gr': {'SWBin': 47, 'HWBin': 6, 'Priority': 225},
            'BSun_BK_DeadColExBPix_Gb': {'SWBin': 47, 'HWBin': 6, 'Priority': 226},
            'BSun_BK_DeadColExBPix_B': {'SWBin': 47, 'HWBin': 6, 'Priority': 227},
            'BSun_BK_Mean_R': {'SWBin': 46, 'HWBin': 6, 'Priority': 228},
            'BSun_BK_Mean_Gr': {'SWBin': 46, 'HWBin': 6, 'Priority': 229},
            'BSun_BK_Mean_Gb': {'SWBin': 46, 'HWBin': 6, 'Priority': 230},
            'BSun_BK_Mean_B': {'SWBin': 46, 'HWBin': 6, 'Priority': 231},
            'BSun_BK_StdDEV_R': {'SWBin': 46, 'HWBin': 6, 'Priority': 232},
            'BSun_BK_StdDEV_Gr': {'SWBin': 46, 'HWBin': 6, 'Priority': 233},
            'BSun_BK_StdDEV_Gb': {'SWBin': 46, 'HWBin': 6, 'Priority': 234},
            'BSun_BK_StdDEV_B': {'SWBin': 46, 'HWBin': 6, 'Priority': 235},
            'DVP27M30F_LMFlip_Cap': {'SWBin': 96, 'HWBin': 5, 'Priority': 236},
            'DVP27M30F_LMFlip_Calc': {'SWBin': 97, 'HWBin': 5, 'Priority': 237},
            'LMFlip_LT_DRow_R': {'SWBin': 58, 'HWBin': 4, 'Priority': 238, 'IsNan': 'DVP27M30F_LMFlip_Cap'},
            'LMFlip_LT_DRow_Gr': {'SWBin': 58, 'HWBin': 4, 'Priority': 239},
            'LMFlip_LT_DRow_Gb': {'SWBin': 58, 'HWBin': 4, 'Priority': 240},
            'LMFlip_LT_DRow_B': {'SWBin': 58, 'HWBin': 4, 'Priority': 241},
            'LMFlip_LT_DCol_R': {'SWBin': 57, 'HWBin': 4, 'Priority': 242},
            'LMFlip_LT_DCol_Gr': {'SWBin': 57, 'HWBin': 4, 'Priority': 243},
            'LMFlip_LT_DCol_Gb': {'SWBin': 57, 'HWBin': 4, 'Priority': 244},
            'LMFlip_LT_DCol_B': {'SWBin': 57, 'HWBin': 4, 'Priority': 245},
            'LMFlip_LT_DRow_Color': {'SWBin': 58, 'HWBin': 4, 'Priority': 246},
            'LMFlip_LT_DCol_Color': {'SWBin': 57, 'HWBin': 4, 'Priority': 247},
            'LMFlip_LT_WeakLineRow_R': {'SWBin': 58, 'HWBin': 4, 'Priority': 248},
            'LMFlip_LT_WeakLineRow_Gr': {'SWBin': 58, 'HWBin': 4, 'Priority': 249},
            'LMFlip_LT_WeakLineRow_Gb': {'SWBin': 58, 'HWBin': 4, 'Priority': 250},
            'LMFlip_LT_WeakLineRow_B': {'SWBin': 58, 'HWBin': 4, 'Priority': 251},
            'LMFlip_LT_WeakLineCol_R': {'SWBin': 57, 'HWBin': 4, 'Priority': 252},
            'LMFlip_LT_WeakLineCol_Gr': {'SWBin': 57, 'HWBin': 4, 'Priority': 253},
            'LMFlip_LT_WeakLineCol_Gb': {'SWBin': 57, 'HWBin': 4, 'Priority': 254},
            'LMFlip_LT_WeakLineCol_B': {'SWBin': 57, 'HWBin': 4, 'Priority': 255},
            'LMFlip_LT_Mean_R': {'SWBin': 56, 'HWBin': 4, 'Priority': 256},
            'LMFlip_LT_Mean_Gr': {'SWBin': 56, 'HWBin': 4, 'Priority': 257},
            'LMFlip_LT_Mean_Gb': {'SWBin': 56, 'HWBin': 4, 'Priority': 258},
            'LMFlip_LT_Mean_B': {'SWBin': 56, 'HWBin': 4, 'Priority': 259},
            'LMFlip_LT_StdDEV_R': {'SWBin': 56, 'HWBin': 4, 'Priority': 260},
            'LMFlip_LT_StdDEV_Gr': {'SWBin': 56, 'HWBin': 4, 'Priority': 261},
            'LMFlip_LT_StdDEV_Gb': {'SWBin': 56, 'HWBin': 4, 'Priority': 262},
            'LMFlip_LT_StdDEV_B': {'SWBin': 56, 'HWBin': 4, 'Priority': 263},
            'LMFlip_LT_RI_R': {'SWBin': 56, 'HWBin': 4, 'Priority': 264},
            'LMFlip_LT_RI_Gr': {'SWBin': 56, 'HWBin': 4, 'Priority': 265},
            'LMFlip_LT_RI_Gb': {'SWBin': 56, 'HWBin': 4, 'Priority': 266},
            'LMFlip_LT_RI_B': {'SWBin': 56, 'HWBin': 4, 'Priority': 267},
            'LMFlip_LT_Ratio_GrR': {'SWBin': 56, 'HWBin': 4, 'Priority': 268},
            'LMFlip_LT_Ratio_GbR': {'SWBin': 56, 'HWBin': 4, 'Priority': 269},
            'LMFlip_LT_Ratio_GrB': {'SWBin': 56, 'HWBin': 4, 'Priority': 270},
            'LMFlip_LT_Ratio_GbB': {'SWBin': 56, 'HWBin': 4, 'Priority': 271},
            'LMFlip_LT_Ratio_GbGr': {'SWBin': 56, 'HWBin': 4, 'Priority': 272},
            'LMFlip_LT_LostBit': {'SWBin': 56, 'HWBin': 4, 'Priority': 273},
            'LMFlip_LT_LostBitSNR_SumDIFF_R': {'SWBin': 56, 'HWBin': 4, 'Priority': 274},
            'LMFlip_LT_LostBitSNR_SumDIFF_Gr': {'SWBin': 56, 'HWBin': 4, 'Priority': 275},
            'LMFlip_LT_LostBitSNR_SumDIFF_Gb': {'SWBin': 56, 'HWBin': 4, 'Priority': 276},
            'LMFlip_LT_LostBitSNR_SumDIFF_B': {'SWBin': 56, 'HWBin': 4, 'Priority': 277},
            'LT_CornerLine': {'SWBin': 26, 'HWBin': 4, 'Priority': 278},
            'LT_ScratchLine': {'SWBin': 26, 'HWBin': 4, 'Priority': 279},
            'LT_Blemish': {'SWBin': 31, 'HWBin': 4, 'Priority': 280},
            'LT_LineStripe': {'SWBin': 27, 'HWBin': 4, 'Priority': 281},
            'LT_Particle': {'SWBin': 32, 'HWBin': 4, 'Priority': 282},
            'BK_Cluster2': {'SWBin': 30, 'HWBin': 4, 'Priority': 283},  # 4/8
            'LT_Cluster2': {'SWBin': 30, 'HWBin': 4, 'Priority': 284},  # 4/8
            'BK_Cluster1': {'SWBin': 29, 'HWBin': 4, 'Priority': 285},  # 4/8
            'LT_Cluster1': {'SWBin': 29, 'HWBin': 4, 'Priority': 286},  # 4/8
            'BK_Cluster3GrGb': {'SWBin': 39, 'HWBin': 4, 'Priority': 287},  # 4/8
            'LT_Cluster3GrGb': {'SWBin': 40, 'HWBin': 4, 'Priority': 288},  # 4/8
            'BK_Cluster3SubtractGrGb': {'SWBin': 33, 'HWBin': 4, 'Priority': 289},  # 4/8
            'LT_Cluster3SubtractGrGb': {'SWBin': 34, 'HWBin': 4, 'Priority': 290},  # 4/8
            'WP_Count': {'SWBin': 35, 'HWBin': 4, 'Priority': 291},  # 6/8
            'BK_Z1_BT_DP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 292},
            'BK_Z1_BT_DP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 293},
            'BK_Z1_BT_DP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 294},
            'BK_Z1_BT_DP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 295},
            'BK_Z2_BT_DP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 296},
            'BK_Z2_BT_DP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 297},
            'BK_Z2_BT_DP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 298},
            'BK_Z2_BT_DP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 299},
            'BK_Z1_BT_WP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 300},
            'BK_Z1_BT_WP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 301},
            'BK_Z1_BT_WP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 302},
            'BK_Z1_BT_WP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 303},
            'BK_Z2_BT_WP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 304},
            'BK_Z2_BT_WP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 305},
            'BK_Z2_BT_WP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 306},
            'BK_Z2_BT_WP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 307},
            'LT_Z1_BT_DP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 308},
            'LT_Z1_BT_DP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 309},
            'LT_Z1_BT_DP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 310},
            'LT_Z1_BT_DP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 311},
            'LT_Z2_BT_DP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 312},
            'LT_Z2_BT_DP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 313},
            'LT_Z2_BT_DP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 314},
            'LT_Z2_BT_DP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 315},
            'LT_Z1_BT_WP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 316},
            'LT_Z1_BT_WP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 317},
            'LT_Z1_BT_WP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 318},
            'LT_Z1_BT_WP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 319},
            'LT_Z2_BT_WP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 320},
            'LT_Z2_BT_WP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 321},
            'LT_Z2_BT_WP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 322},
            'LT_Z2_BT_WP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 323},
            'LT_Z1_DK_WP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 324},
            'LT_Z1_DK_WP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 325},
            'LT_Z1_DK_WP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 326},
            'LT_Z1_DK_WP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 327},
            'LT_Z2_DK_WP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 328},
            'LT_Z2_DK_WP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 329},
            'LT_Z2_DK_WP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 330},
            'LT_Z2_DK_WP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 331},
            'LT_Z1_DK_DP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 332},
            'LT_Z1_DK_DP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 333},
            'LT_Z1_DK_DP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 334},
            'LT_Z1_DK_DP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 335},
            'LT_Z2_DK_DP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 336},
            'LT_Z2_DK_DP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 337},
            'LT_Z2_DK_DP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 338},
            'LT_Z2_DK_DP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 339},
            'BK_Z1_BT_DP': {'SWBin': 60, 'HWBin': 6, 'Priority': 340},
            'BK_Z2_BT_DP': {'SWBin': 60, 'HWBin': 6, 'Priority': 341},
            'BK_Z1_BT_WP': {'SWBin': 60, 'HWBin': 6, 'Priority': 342},
            'BK_Z2_BT_WP': {'SWBin': 60, 'HWBin': 6, 'Priority': 343},
            'LT_Z1_BT_DP': {'SWBin': 60, 'HWBin': 6, 'Priority': 344},
            'LT_Z2_BT_DP': {'SWBin': 60, 'HWBin': 6, 'Priority': 345},
            'LT_Z1_BT_WP': {'SWBin': 60, 'HWBin': 6, 'Priority': 346},
            'LT_Z2_BT_WP': {'SWBin': 60, 'HWBin': 6, 'Priority': 347},
            'LT_Z1_DK_DP': {'SWBin': 60, 'HWBin': 6, 'Priority': 348},
            'LT_Z2_DK_DP': {'SWBin': 60, 'HWBin': 6, 'Priority': 349},
            'LT_Z1_DK_WP': {'SWBin': 60, 'HWBin': 6, 'Priority': 350},
            'LT_Z2_DK_WP': {'SWBin': 60, 'HWBin': 6, 'Priority': 351},
            'DVP27M30F_BK32X_Cap': {'SWBin': 98, 'HWBin': 5, 'Priority': 352},
            'DVP27M30F_BK32X_Calc': {'SWBin': 97, 'HWBin': 5, 'Priority': 353},
            'BK32X_BK_BadPixel_Area': {'SWBin': 75, 'HWBin': 4, 'Priority': 354, 'IsNan': 'DVP27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DashQty_R': {'SWBin': 51, 'HWBin': 6, 'Priority': 355},
            'BK32X_BK_MixDCol_DashQty_Gr': {'SWBin': 51, 'HWBin': 6, 'Priority': 356},
            'BK32X_BK_MixDCol_DashQty_Gb': {'SWBin': 51, 'HWBin': 6, 'Priority': 357},
            'BK32X_BK_MixDCol_DashQty_B': {'SWBin': 51, 'HWBin': 6, 'Priority': 358},
            'BK32X_SP_BK_MixDCol_DashQty_R': {'SWBin': 54, 'HWBin': 2, 'Priority': 359},
            'BK32X_SP_BK_MixDCol_DashQty_Gr': {'SWBin': 54, 'HWBin': 2, 'Priority': 360},
            'BK32X_SP_BK_MixDCol_DashQty_Gb': {'SWBin': 54, 'HWBin': 2, 'Priority': 361},
            'BK32X_SP_BK_MixDCol_DashQty_B': {'SWBin': 54, 'HWBin': 2, 'Priority': 362},
            'BK32X_BK_MixDCol_DASH_DeltaAvg_R': {'SWBin': 54, 'HWBin': 2, 'Priority': 363},
            'BK32X_BK_MixDCol_DASH_DeltaAvg_Gr': {'SWBin': 54, 'HWBin': 2, 'Priority': 364},
            'BK32X_BK_MixDCol_DASH_DeltaAvg_Gb': {'SWBin': 54, 'HWBin': 2, 'Priority': 365},
            'BK32X_BK_MixDCol_DASH_DeltaAvg_B': {'SWBin': 54, 'HWBin': 2, 'Priority': 366},
            'BK32X_BK_MixDCol_DASH_MaxValue0_R': {'SWBin': 54, 'HWBin': 2, 'Priority': 367},
            'BK32X_BK_MixDCol_DASH_MaxValue0_Gr': {'SWBin': 54, 'HWBin': 2, 'Priority': 368},
            'BK32X_BK_MixDCol_DASH_MaxValue0_Gb': {'SWBin': 54, 'HWBin': 2, 'Priority': 369},
            'BK32X_BK_MixDCol_DASH_MaxValue0_B': {'SWBin': 54, 'HWBin': 2, 'Priority': 370},
            'BK32X_BK_MixDCol_DASH_MaxValue1_R': {'SWBin': 54, 'HWBin': 2, 'Priority': 371},
            'BK32X_BK_MixDCol_DASH_MaxValue1_Gr': {'SWBin': 54, 'HWBin': 2, 'Priority': 372},
            'BK32X_BK_MixDCol_DASH_MaxValue1_Gb': {'SWBin': 54, 'HWBin': 2, 'Priority': 373},
            'BK32X_BK_MixDCol_DASH_MaxValue1_B': {'SWBin': 54, 'HWBin': 2, 'Priority': 374},
            'BK32X_BK_MixDCol_DASH_MaxValue2_R': {'SWBin': 54, 'HWBin': 2, 'Priority': 375},
            'BK32X_BK_MixDCol_DASH_MaxValue2_Gr': {'SWBin': 54, 'HWBin': 2, 'Priority': 376},
            'BK32X_BK_MixDCol_DASH_MaxValue2_Gb': {'SWBin': 54, 'HWBin': 2, 'Priority': 377},
            'BK32X_BK_MixDCol_DASH_MaxValue2_B': {'SWBin': 54, 'HWBin': 2, 'Priority': 378},
            'BK32X_BK_MixDCol_DASH_MaxValue3_R': {'SWBin': 54, 'HWBin': 2, 'Priority': 379},
            'BK32X_BK_MixDCol_DASH_MaxValue3_Gr': {'SWBin': 54, 'HWBin': 2, 'Priority': 380},
            'BK32X_BK_MixDCol_DASH_MaxValue3_Gb': {'SWBin': 54, 'HWBin': 2, 'Priority': 381},
            'BK32X_BK_MixDCol_DASH_MaxValue3_B': {'SWBin': 54, 'HWBin': 2, 'Priority': 382},
            'BK32X_BK_MixDCol_DASH_MaxValue4_R': {'SWBin': 54, 'HWBin': 2, 'Priority': 383},
            'BK32X_BK_MixDCol_DASH_MaxValue4_Gr': {'SWBin': 54, 'HWBin': 2, 'Priority': 384},
            'BK32X_BK_MixDCol_DASH_MaxValue4_Gb': {'SWBin': 54, 'HWBin': 2, 'Priority': 385},
            'BK32X_BK_MixDCol_DASH_MaxValue4_B': {'SWBin': 54, 'HWBin': 2, 'Priority': 386},
            'BK32X_BK_Mean_R': {'SWBin': 53, 'HWBin': 2, 'Priority': 387},
            'BK32X_BK_Mean_Gr': {'SWBin': 53, 'HWBin': 2, 'Priority': 388},
            'BK32X_BK_Mean_Gb': {'SWBin': 53, 'HWBin': 2, 'Priority': 389},
            'BK32X_BK_Mean_B': {'SWBin': 53, 'HWBin': 2, 'Priority': 390},
            'BK32X_BK_StdDEV_R': {'SWBin': 53, 'HWBin': 2, 'Priority': 391},
            'BK32X_BK_StdDEV_Gr': {'SWBin': 53, 'HWBin': 2, 'Priority': 392},
            'BK32X_BK_StdDEV_Gb': {'SWBin': 53, 'HWBin': 2, 'Priority': 393},
            'BK32X_BK_StdDEV_B': {'SWBin': 53, 'HWBin': 2, 'Priority': 394},
            'BK_Cluster3_2': {'SWBin': 73, 'HWBin': 2, 'Priority': 395},
            'LT_Cluster3_2': {'SWBin': 74, 'HWBin': 2, 'Priority': 396},
            'MIPI2L24M30F_WK1_Cap': {'SWBin': 89, 'HWBin': 1, 'Priority': 397},
            'MIPI2L24M30F_WK1_Calc': {'SWBin': 97, 'HWBin': 5, 'Priority': 398},
            'MIPI_One_W1_FixedPattern': {'SWBin': 94, 'HWBin': 1, 'Priority': 399, 'IsNan': 'MIPI2L24M30F_WK1_Cap'},
            'MIPI2L24M30F_LT_Cap': {'SWBin': 90, 'HWBin': 1, 'Priority': 400},
            'MIPI2L24M30F_LT_Calc': {'SWBin': 97, 'HWBin': 5, 'Priority': 401},
            'MIPI_LT_DRow_R': {'SWBin': 65, 'HWBin': 1, 'Priority': 402, 'IsNan': 'MIPI2L24M30F_LT_Cap'},
            'MIPI_LT_DRow_Gr': {'SWBin': 65, 'HWBin': 1, 'Priority': 403},
            'MIPI_LT_DRow_Gb': {'SWBin': 65, 'HWBin': 1, 'Priority': 404},
            'MIPI_LT_DRow_B': {'SWBin': 65, 'HWBin': 1, 'Priority': 405},
            'MIPI_LT_DCol_R': {'SWBin': 64, 'HWBin': 1, 'Priority': 406},
            'MIPI_LT_DCol_Gr': {'SWBin': 64, 'HWBin': 1, 'Priority': 407},
            'MIPI_LT_DCol_Gb': {'SWBin': 64, 'HWBin': 1, 'Priority': 408},
            'MIPI_LT_DCol_B': {'SWBin': 64, 'HWBin': 1, 'Priority': 409},
            'MIPI_LT_DRow_Color': {'SWBin': 65, 'HWBin': 1, 'Priority': 410},
            'MIPI_LT_DCol_Color': {'SWBin': 64, 'HWBin': 1, 'Priority': 411},
            'MIPI_LT_WeakLineRow_R': {'SWBin': 65, 'HWBin': 1, 'Priority': 412},
            'MIPI_LT_WeakLineRow_Gr': {'SWBin': 65, 'HWBin': 1, 'Priority': 413},
            'MIPI_LT_WeakLineRow_Gb': {'SWBin': 65, 'HWBin': 1, 'Priority': 414},
            'MIPI_LT_WeakLineRow_B': {'SWBin': 65, 'HWBin': 1, 'Priority': 415},
            'MIPI_LT_WeakLineCol_R': {'SWBin': 64, 'HWBin': 1, 'Priority': 416},
            'MIPI_LT_WeakLineCol_Gr': {'SWBin': 64, 'HWBin': 1, 'Priority': 417},
            'MIPI_LT_WeakLineCol_Gb': {'SWBin': 64, 'HWBin': 1, 'Priority': 418},
            'MIPI_LT_WeakLineCol_B': {'SWBin': 64, 'HWBin': 1, 'Priority': 419},
            'MIPI_LT_Mean_R': {'SWBin': 63, 'HWBin': 1, 'Priority': 420},
            'MIPI_LT_Mean_Gr': {'SWBin': 63, 'HWBin': 1, 'Priority': 421},
            'MIPI_LT_Mean_Gb': {'SWBin': 63, 'HWBin': 1, 'Priority': 422},
            'MIPI_LT_Mean_B': {'SWBin': 63, 'HWBin': 1, 'Priority': 423},
            'MIPI_LT_StdDEV_R': {'SWBin': 63, 'HWBin': 1, 'Priority': 424},
            'MIPI_LT_StdDEV_Gr': {'SWBin': 63, 'HWBin': 1, 'Priority': 425},
            'MIPI_LT_StdDEV_Gb': {'SWBin': 63, 'HWBin': 1, 'Priority': 426},
            'MIPI_LT_StdDEV_B': {'SWBin': 63, 'HWBin': 1, 'Priority': 427},
            'MIPI_LT_RI_R': {'SWBin': 63, 'HWBin': 1, 'Priority': 428},
            'MIPI_LT_RI_Gr': {'SWBin': 63, 'HWBin': 1, 'Priority': 429},
            'MIPI_LT_RI_Gb': {'SWBin': 63, 'HWBin': 1, 'Priority': 430},
            'MIPI_LT_RI_B': {'SWBin': 63, 'HWBin': 1, 'Priority': 431},
            'MIPI_LT_Ratio_GrR': {'SWBin': 63, 'HWBin': 1, 'Priority': 432},
            'MIPI_LT_Ratio_GbR': {'SWBin': 63, 'HWBin': 1, 'Priority': 433},
            'MIPI_LT_Ratio_GrB': {'SWBin': 63, 'HWBin': 1, 'Priority': 434},
            'MIPI_LT_Ratio_GbB': {'SWBin': 63, 'HWBin': 1, 'Priority': 435},
            'MIPI_LT_Ratio_GbGr': {'SWBin': 63, 'HWBin': 1, 'Priority': 436},
            'MIPI_LT_LostBit': {'SWBin': 63, 'HWBin': 1, 'Priority': 437},
            'MIPI_LT_LostBitSNR_SumDIFF_R': {'SWBin': 63, 'HWBin': 1, 'Priority': 438},
            'MIPI_LT_LostBitSNR_SumDIFF_Gr': {'SWBin': 63, 'HWBin': 1, 'Priority': 439},
            'MIPI_LT_LostBitSNR_SumDIFF_Gb': {'SWBin': 63, 'HWBin': 1, 'Priority': 440},
            'MIPI_LT_LostBitSNR_SumDIFF_B': {'SWBin': 63, 'HWBin': 1, 'Priority': 441},
            'MIPI2L24M30F_LBIT_Cap': {'SWBin': 90, 'HWBin': 1, 'Priority': 442},
            'MIPI2L24M30F_LBIT_Calc': {'SWBin': 97, 'HWBin': 5, 'Priority': 443},
            'MIPI_LBIT_LT_Mean_R': {'SWBin': 63, 'HWBin': 1, 'Priority': 444, 'IsNan': 'MIPI2L24M30F_LBIT_Cap'},
            'MIPI_LBIT_LT_Mean_Gr': {'SWBin': 63, 'HWBin': 1, 'Priority': 445},
            'MIPI_LBIT_LT_Mean_Gb': {'SWBin': 63, 'HWBin': 1, 'Priority': 446},
            'MIPI_LBIT_LT_Mean_B': {'SWBin': 63, 'HWBin': 1, 'Priority': 447},
            'MIPI_LBIT_LT_LostBit': {'SWBin': 63, 'HWBin': 1, 'Priority': 448},
            'MIPI_LBIT_LT_LostBitSNR_SumDIFF_R': {'SWBin': 63, 'HWBin': 1, 'Priority': 449},
            'MIPI_LBIT_LT_LostBitSNR_SumDIFF_Gr': {'SWBin': 63, 'HWBin': 1, 'Priority': 450},
            'MIPI_LBIT_LT_LostBitSNR_SumDIFF_Gb': {'SWBin': 63, 'HWBin': 1, 'Priority': 451},
            'MIPI_LBIT_LT_LostBitSNR_SumDIFF_B': {'SWBin': 63, 'HWBin': 1, 'Priority': 452},
            'Binning': {'SWBin': 2, 'HWBin': 3, 'Priority': 453},
            'All Pass': {'SWBin': 1, 'HWBin': 3, 'Priority': 454}
        },
    'JX825':
        {
            'VSYNC_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 1},
            'HSYNC_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 2},
            'PCLK_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 3},
            'EXCLK_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 4},
            'RSTB_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 5},
            'PWDN_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 6},
            'SDA_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 7},
            'SCL_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 8},
            'VH_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 9},
            'VN1_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 10},
            'VN2_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 11},
            'D0_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 12},
            'D1_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 13},
            'D2_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 14},
            'D3_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 15},
            'D4_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 16},
            'D5_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 17},
            'D6_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 18},
            'D7_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 19},
            'D8_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 20},
            'D9_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 21},
            'MCP_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 22},
            'MCN_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 23},
            'MDP0_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 24},
            'MDN0_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 25},
            'MDP1_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 26},
            'MDN1_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 27},
            'MDP2_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 28},
            'MDN2_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 29},
            'MDP3_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 30},
            'MDN3_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 31},
            'DVDD_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 32},
            'AVDD_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 33},
            'DOVDD_O/S': {'SWBin': 5, 'HWBin': 5, 'Priority': 34},
            'VSYNC_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 35},
            'HSYNC_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 36},
            'PCLK_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 37},
            'EXCLK_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 38},
            'RSTB_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 39},
            'PWDN_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 40},
            'SDA_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 41},
            'SCL_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 42},
            'D0_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 43},
            'D1_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 44},
            'D2_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 45},
            'D3_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 46},
            'D4_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 47},
            'D5_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 48},
            'D6_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 49},
            'D7_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 50},
            'D8_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 51},
            'D9_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 52},
            'MCP_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 53},
            'MCN_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 54},
            'MDP0_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 55},
            'MDN0_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 56},
            'MDP1_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 57},
            'MDN1_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 58},
            'MDP2_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 59},
            'MDN2_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 60},
            'MDP3_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 61},
            'MDN3_Leakage/iiL': {'SWBin': 6, 'HWBin': 5, 'Priority': 62},
            'VSYNC_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 63},
            'HSYNC_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 64},
            'PCLK_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 65},
            'EXCLK_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 66},
            'RSTB_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 67},
            'PWDN_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 68},
            'SDA_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 69},
            'SCL_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 70},
            'D0_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 71},
            'D1_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 72},
            'D2_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 73},
            'D3_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 74},
            'D4_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 75},
            'D5_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 76},
            'D6_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 77},
            'D7_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 78},
            'D8_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 79},
            'D9_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 80},
            'MCP_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 81},
            'MCN_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 82},
            'MDP0_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 83},
            'MDN0_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 84},
            'MDP1_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 85},
            'MDN1_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 86},
            'MDP2_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 87},
            'MDN2_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 88},
            'MDP3_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 89},
            'MDN3_Leakage/iiH': {'SWBin': 6, 'HWBin': 5, 'Priority': 90},
            'iic_test': {'SWBin': 7, 'HWBin': 5, 'Priority': 91},
            'MIPI4L27M30F_BK_Cap': {'SWBin': 98, 'HWBin': 8, 'Priority': 92},
            'MIPI4L27M30F_BK_Calc': {'SWBin': 97, 'HWBin': 8, 'Priority': 93},
            'BLC_R': {'SWBin': 7, 'HWBin': 5, 'Priority': 94},
            'BLC_Gr': {'SWBin': 7, 'HWBin': 5, 'Priority': 95},
            'BLC_Gb': {'SWBin': 7, 'HWBin': 5, 'Priority': 96},
            'BLC_B': {'SWBin': 7, 'HWBin': 5, 'Priority': 97},
            'BK_DeadRowExBPix_R': {'SWBin': 15, 'HWBin': 6, 'Priority': 98},
            'BK_DeadRowExBPix_Gr': {'SWBin': 15, 'HWBin': 6, 'Priority': 99},
            'BK_DeadRowExBPix_Gb': {'SWBin': 15, 'HWBin': 6, 'Priority': 100},
            'BK_DeadRowExBPix_B': {'SWBin': 15, 'HWBin': 6, 'Priority': 101},
            'BK_DeadColExBPix_R': {'SWBin': 14, 'HWBin': 6, 'Priority': 102},
            'BK_DeadColExBPix_Gr': {'SWBin': 14, 'HWBin': 6, 'Priority': 103},
            'BK_DeadColExBPix_Gb': {'SWBin': 14, 'HWBin': 6, 'Priority': 104},
            'BK_DeadColExBPix_B': {'SWBin': 14, 'HWBin': 6, 'Priority': 105},
            'BK_MixDCol_VfpnQty_R': {'SWBin': 14, 'HWBin': 6, 'Priority': 106},
            'BK_MixDCol_VfpnQty_Gr': {'SWBin': 14, 'HWBin': 6, 'Priority': 107},
            'BK_MixDCol_VfpnQty_Gb': {'SWBin': 14, 'HWBin': 6, 'Priority': 108},
            'BK_MixDCol_VfpnQty_B': {'SWBin': 14, 'HWBin': 6, 'Priority': 109},
            'BK_MixDCol_VFPN_MaxValue0_R': {'SWBin': 13, 'HWBin': 6, 'Priority': 110},
            'BK_MixDCol_VFPN_MaxValue0_Gr': {'SWBin': 13, 'HWBin': 6, 'Priority': 111},
            'BK_MixDCol_VFPN_MaxValue0_Gb': {'SWBin': 13, 'HWBin': 6, 'Priority': 112},
            'BK_MixDCol_VFPN_MaxValue0_B': {'SWBin': 13, 'HWBin': 6, 'Priority': 113},
            'BK_MixDCol_VFPN_MaxValue1_R': {'SWBin': 13, 'HWBin': 6, 'Priority': 114},
            'BK_MixDCol_VFPN_MaxValue1_Gr': {'SWBin': 13, 'HWBin': 6, 'Priority': 115},
            'BK_MixDCol_VFPN_MaxValue1_Gb': {'SWBin': 13, 'HWBin': 6, 'Priority': 116},
            'BK_MixDCol_VFPN_MaxValue1_B': {'SWBin': 13, 'HWBin': 6, 'Priority': 117},
            'BK_MixDCol_VFPN_MaxValue2_R': {'SWBin': 13, 'HWBin': 6, 'Priority': 118},
            'BK_MixDCol_VFPN_MaxValue2_Gr': {'SWBin': 13, 'HWBin': 6, 'Priority': 119},
            'BK_MixDCol_VFPN_MaxValue2_Gb': {'SWBin': 13, 'HWBin': 6, 'Priority': 120},
            'BK_MixDCol_VFPN_MaxValue2_B': {'SWBin': 13, 'HWBin': 6, 'Priority': 121},
            'BK_MixDCol_VFPN_MaxValue3_R': {'SWBin': 13, 'HWBin': 6, 'Priority': 122},
            'BK_MixDCol_VFPN_MaxValue3_Gr': {'SWBin': 13, 'HWBin': 6, 'Priority': 123},
            'BK_MixDCol_VFPN_MaxValue3_Gb': {'SWBin': 13, 'HWBin': 6, 'Priority': 124},
            'BK_MixDCol_VFPN_MaxValue3_B': {'SWBin': 13, 'HWBin': 6, 'Priority': 125},
            'BK_MixDCol_VFPN_MaxValue4_R': {'SWBin': 13, 'HWBin': 6, 'Priority': 126},
            'BK_MixDCol_VFPN_MaxValue4_Gr': {'SWBin': 13, 'HWBin': 6, 'Priority': 127},
            'BK_MixDCol_VFPN_MaxValue4_Gb': {'SWBin': 13, 'HWBin': 6, 'Priority': 128},
            'BK_MixDCol_VFPN_MaxValue4_B': {'SWBin': 13, 'HWBin': 6, 'Priority': 129},
            'BK_Mean_R': {'SWBin': 13, 'HWBin': 6, 'Priority': 130},
            'BK_Mean_Gr': {'SWBin': 13, 'HWBin': 6, 'Priority': 131},
            'BK_Mean_Gb': {'SWBin': 13, 'HWBin': 6, 'Priority': 132},
            'BK_Mean_B': {'SWBin': 13, 'HWBin': 6, 'Priority': 133},
            'BK_StdDEV_R': {'SWBin': 13, 'HWBin': 6, 'Priority': 134},
            'BK_StdDEV_Gr': {'SWBin': 13, 'HWBin': 6, 'Priority': 135},
            'BK_StdDEV_Gb': {'SWBin': 13, 'HWBin': 6, 'Priority': 136},
            'BK_StdDEV_B': {'SWBin': 13, 'HWBin': 6, 'Priority': 137},
            'VH_voltage': {'SWBin': 9, 'HWBin': 5, 'Priority': 138},
            'VN1_voltage': {'SWBin': 9, 'HWBin': 5, 'Priority': 139},
            'VN2_voltage': {'SWBin': 9, 'HWBin': 5, 'Priority': 140},
            'DVDD_voltage': {'SWBin': 9, 'HWBin': 5, 'Priority': 141},
            'MIPI4L27M30F_LT_Cap': {'SWBin': 99, 'HWBin': 8, 'Priority': 142},
            'MIPI4L27M30F_LT_Calc': {'SWBin': 99, 'HWBin': 8, 'Priority': 143},
            'LT_DRow_R': {'SWBin': 25, 'HWBin': 4, 'Priority': 144, 'IsNan': 'MIPI4L27M30F_LT_Cap'},
            'LT_DRow_Gr': {'SWBin': 25, 'HWBin': 4, 'Priority': 145},
            'LT_DRow_Gb': {'SWBin': 25, 'HWBin': 4, 'Priority': 146},
            'LT_DRow_B': {'SWBin': 25, 'HWBin': 4, 'Priority': 147},
            'LT_DCol_R': {'SWBin': 24, 'HWBin': 4, 'Priority': 148},
            'LT_DCol_Gr': {'SWBin': 24, 'HWBin': 4, 'Priority': 149},
            'LT_DCol_Gb': {'SWBin': 24, 'HWBin': 4, 'Priority': 150},
            'LT_DCol_B': {'SWBin': 24, 'HWBin': 4, 'Priority': 151},
            'LT_DRow_Color': {'SWBin': 25, 'HWBin': 4, 'Priority': 152},
            'LT_DCol_Color': {'SWBin': 24, 'HWBin': 4, 'Priority': 153},
            'LT_WeakLineRow_R': {'SWBin': 25, 'HWBin': 4, 'Priority': 154},
            'LT_WeakLineRow_Gr': {'SWBin': 25, 'HWBin': 4, 'Priority': 155},
            'LT_WeakLineRow_Gb': {'SWBin': 25, 'HWBin': 4, 'Priority': 156},
            'LT_WeakLineRow_B': {'SWBin': 25, 'HWBin': 4, 'Priority': 157},
            'LT_WeakLineCol_R': {'SWBin': 24, 'HWBin': 4, 'Priority': 158},
            'LT_WeakLineCol_Gr': {'SWBin': 24, 'HWBin': 4, 'Priority': 159},
            'LT_WeakLineCol_Gb': {'SWBin': 24, 'HWBin': 4, 'Priority': 160},
            'LT_WeakLineCol_B': {'SWBin': 24, 'HWBin': 4, 'Priority': 161},
            'LT_Mean_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 162},
            'LT_Mean_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 163},
            'LT_Mean_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 164},
            'LT_Mean_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 165},
            'LT_StdDEV_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 166},
            'LT_StdDEV_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 167},
            'LT_StdDEV_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 168},
            'LT_StdDEV_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 169},
            'LT_RI_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 170},
            'LT_RI_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 171},
            'LT_RI_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 172},
            'LT_RI_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 173},
            'LT_Ratio_GrR': {'SWBin': 23, 'HWBin': 4, 'Priority': 174},
            'LT_Ratio_GbR': {'SWBin': 23, 'HWBin': 4, 'Priority': 175},
            'LT_Ratio_GrB': {'SWBin': 23, 'HWBin': 4, 'Priority': 176},
            'LT_Ratio_GbB': {'SWBin': 23, 'HWBin': 4, 'Priority': 177},
            'LT_Ratio_GbGr': {'SWBin': 23, 'HWBin': 4, 'Priority': 178},
            'LT_LostBit': {'SWBin': 23, 'HWBin': 4, 'Priority': 179},
            'LT_LostBitSNR_SumDIFF_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 180},
            'LT_LostBitSNR_SumDIFF_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 181},
            'LT_LostBitSNR_SumDIFF_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 182},
            'LT_LostBitSNR_SumDIFF_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 183},
            'MIPI4L27M30F_LBIT_Cap': {'SWBin': 99, 'HWBin': 8, 'Priority': 184},
            'MIPI4L27M30F_LBIT_Calc': {'SWBin': 97, 'HWBin': 8, 'Priority': 185},
            'LBIT_LT_Mean_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 186, 'IsNan': 'MIPI4L27M30F_LBIT_Cap'},
            'LBIT_LT_Mean_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 187},
            'LBIT_LT_Mean_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 188},
            'LBIT_LT_Mean_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 189},
            'LBIT_LT_DRow_R': {'SWBin': 25, 'HWBin': 4, 'Priority': 190},
            'LBIT_LT_DRow_Gr': {'SWBin': 25, 'HWBin': 4, 'Priority': 191},
            'LBIT_LT_DRow_Gb': {'SWBin': 25, 'HWBin': 4, 'Priority': 192},
            'LBIT_LT_DRow_B': {'SWBin': 25, 'HWBin': 4, 'Priority': 193},
            'LBIT_LT_DCol_R': {'SWBin': 24, 'HWBin': 4, 'Priority': 194},
            'LBIT_LT_DCol_Gr': {'SWBin': 24, 'HWBin': 4, 'Priority': 195},
            'LBIT_LT_DCol_Gb': {'SWBin': 24, 'HWBin': 4, 'Priority': 196},
            'LBIT_LT_DCol_B': {'SWBin': 24, 'HWBin': 4, 'Priority': 197},
            'LBIT_LT_DRow_Color': {'SWBin': 25, 'HWBin': 4, 'Priority': 198},
            'LBIT_LT_DCol_Color': {'SWBin': 24, 'HWBin': 4, 'Priority': 199},
            'LBIT_LT_WeakLineRow_R': {'SWBin': 25, 'HWBin': 4, 'Priority': 200},
            'LBIT_LT_WeakLineRow_Gr': {'SWBin': 25, 'HWBin': 4, 'Priority': 201},
            'LBIT_LT_WeakLineRow_Gb': {'SWBin': 25, 'HWBin': 4, 'Priority': 202},
            'LBIT_LT_WeakLineRow_B': {'SWBin': 25, 'HWBin': 4, 'Priority': 203},
            'LBIT_LT_WeakLineCol_R': {'SWBin': 24, 'HWBin': 4, 'Priority': 204},
            'LBIT_LT_WeakLineCol_Gr': {'SWBin': 24, 'HWBin': 4, 'Priority': 205},
            'LBIT_LT_WeakLineCol_Gb': {'SWBin': 24, 'HWBin': 4, 'Priority': 206},
            'LBIT_LT_WeakLineCol_B': {'SWBin': 24, 'HWBin': 4, 'Priority': 207},
            'LBIT_LT_Cluster2': {'SWBin': 23, 'HWBin': 4, 'Priority': 208},
            'LBIT_LT_Cluster1': {'SWBin': 23, 'HWBin': 4, 'Priority': 209},
            'LBIT_LT_LostBit': {'SWBin': 23, 'HWBin': 4, 'Priority': 210},
            'LBIT_LT_LostBitSNR_SumDIFF_R': {'SWBin': 23, 'HWBin': 4, 'Priority': 211},
            'LBIT_LT_LostBitSNR_SumDIFF_Gr': {'SWBin': 23, 'HWBin': 4, 'Priority': 212},
            'LBIT_LT_LostBitSNR_SumDIFF_Gb': {'SWBin': 23, 'HWBin': 4, 'Priority': 213},
            'LBIT_LT_LostBitSNR_SumDIFF_B': {'SWBin': 23, 'HWBin': 4, 'Priority': 214},
            'LT_CornerLine': {'SWBin': 26, 'HWBin': 4, 'Priority': 215},
            'LT_ScratchLine': {'SWBin': 26, 'HWBin': 4, 'Priority': 216},
            'LT_Blemish': {'SWBin': 31, 'HWBin': 4, 'Priority': 217},
            'LT_LineStripe': {'SWBin': 27, 'HWBin': 4, 'Priority': 218},
            'LT_Particle': {'SWBin': 32, 'HWBin': 4, 'Priority': 219},
            'BK_Cluster2': {'SWBin': 30, 'HWBin': 4, 'Priority': 220},
            'LT_Cluster2': {'SWBin': 30, 'HWBin': 4, 'Priority': 221},
            'BK_Cluster1': {'SWBin': 29, 'HWBin': 4, 'Priority': 222},
            'LT_Cluster1': {'SWBin': 29, 'HWBin': 4, 'Priority': 223},
            'BK_BadPixel_Area': {'SWBin': 33, 'HWBin': 4, 'Priority': 224},
            'LT_BadPixel_Area': {'SWBin': 34, 'HWBin': 4, 'Priority': 225},
            'WP_Count': {'SWBin': 35, 'HWBin': 6, 'Priority': 226},
            'BK_Z1_BT_DP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 227},
            'BK_Z1_BT_DP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 228},
            'BK_Z1_BT_DP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 229},
            'BK_Z1_BT_DP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 230},
            'BK_Z2_BT_DP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 231},
            'BK_Z2_BT_DP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 232},
            'BK_Z2_BT_DP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 233},
            'BK_Z2_BT_DP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 234},
            'BK_Z1_BT_WP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 235},
            'BK_Z1_BT_WP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 236},
            'BK_Z1_BT_WP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 237},
            'BK_Z1_BT_WP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 238},
            'BK_Z2_BT_WP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 239},
            'BK_Z2_BT_WP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 240},
            'BK_Z2_BT_WP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 241},
            'BK_Z2_BT_WP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 242},
            'LT_Z1_BT_DP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 243},
            'LT_Z1_BT_DP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 244},
            'LT_Z1_BT_DP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 245},
            'LT_Z1_BT_DP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 246},
            'LT_Z2_BT_DP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 247},
            'LT_Z2_BT_DP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 248},
            'LT_Z2_BT_DP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 249},
            'LT_Z2_BT_DP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 250},
            'LT_Z1_BT_WP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 251},
            'LT_Z1_BT_WP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 252},
            'LT_Z1_BT_WP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 253},
            'LT_Z1_BT_WP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 254},
            'LT_Z2_BT_WP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 255},
            'LT_Z2_BT_WP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 256},
            'LT_Z2_BT_WP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 257},
            'LT_Z2_BT_WP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 258},
            'LT_Z1_DK_WP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 259},
            'LT_Z1_DK_WP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 260},
            'LT_Z1_DK_WP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 261},
            'LT_Z1_DK_WP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 262},
            'LT_Z2_DK_WP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 263},
            'LT_Z2_DK_WP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 264},
            'LT_Z2_DK_WP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 265},
            'LT_Z2_DK_WP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 266},
            'LT_Z1_DK_DP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 267},
            'LT_Z1_DK_DP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 268},
            'LT_Z1_DK_DP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 269},
            'LT_Z1_DK_DP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 270},
            'LT_Z2_DK_DP_R': {'SWBin': 60, 'HWBin': 6, 'Priority': 271},
            'LT_Z2_DK_DP_Gr': {'SWBin': 60, 'HWBin': 6, 'Priority': 272},
            'LT_Z2_DK_DP_Gb': {'SWBin': 60, 'HWBin': 6, 'Priority': 273},
            'LT_Z2_DK_DP_B': {'SWBin': 60, 'HWBin': 6, 'Priority': 274},
            'BK_Z1_BT_DP': {'SWBin': 60, 'HWBin': 6, 'Priority': 275},
            'BK_Z2_BT_DP': {'SWBin': 60, 'HWBin': 6, 'Priority': 276},
            'BK_Z1_BT_WP': {'SWBin': 60, 'HWBin': 6, 'Priority': 277},
            'BK_Z2_BT_WP': {'SWBin': 60, 'HWBin': 6, 'Priority': 278},
            'LT_Z1_BT_DP': {'SWBin': 60, 'HWBin': 6, 'Priority': 279},
            'LT_Z2_BT_DP': {'SWBin': 60, 'HWBin': 6, 'Priority': 280},
            'LT_Z1_BT_WP': {'SWBin': 60, 'HWBin': 6, 'Priority': 281},
            'LT_Z2_BT_WP': {'SWBin': 60, 'HWBin': 6, 'Priority': 282},
            'LT_Z1_DK_DP': {'SWBin': 60, 'HWBin': 6, 'Priority': 283},
            'LT_Z2_DK_DP': {'SWBin': 60, 'HWBin': 6, 'Priority': 284},
            'LT_Z1_DK_WP': {'SWBin': 60, 'HWBin': 6, 'Priority': 285},
            'LT_Z2_DK_WP': {'SWBin': 60, 'HWBin': 6, 'Priority': 286},
            'Active_AVDD': {'SWBin': 12, 'HWBin': 5, 'Priority': 287},
            'Active_DOVDD': {'SWBin': 12, 'HWBin': 5, 'Priority': 288},
            'PWDN_AVDD': {'SWBin': 8, 'HWBin': 5, 'Priority': 289},
            'PWDN_DOVDD': {'SWBin': 8, 'HWBin': 5, 'Priority': 290},
            'PWDN_Total': {'SWBin': 8, 'HWBin': 5, 'Priority': 291},
            'BK_Cluster3GrGb': {'SWBin': 39, 'HWBin': 4, 'Priority': 292},
            'LT_Cluster3GrGb': {'SWBin': 40, 'HWBin': 4, 'Priority': 293},
            'MIPI4L27M30F_FW_Cap': {'SWBin': 99, 'HWBin': 8, 'Priority': 294},
            'MIPI4L27M30F_FW_Calc': {'SWBin': 97, 'HWBin': 8, 'Priority': 295},
            'FW_LT_DRow_R': {'SWBin': 36, 'HWBin': 4, 'Priority': 296, 'IsNan': 'MIPI4L27M30F_FW_Cap'},
            'FW_LT_DRow_Gr': {'SWBin': 36, 'HWBin': 4, 'Priority': 297},
            'FW_LT_DRow_Gb': {'SWBin': 36, 'HWBin': 4, 'Priority': 298},
            'FW_LT_DRow_B': {'SWBin': 36, 'HWBin': 4, 'Priority': 299},
            'FW_LT_DCol_R': {'SWBin': 36, 'HWBin': 4, 'Priority': 300},
            'FW_LT_DCol_Gr': {'SWBin': 36, 'HWBin': 4, 'Priority': 301},
            'FW_LT_DCol_Gb': {'SWBin': 36, 'HWBin': 4, 'Priority': 302},
            'FW_LT_DCol_B': {'SWBin': 36, 'HWBin': 4, 'Priority': 303},
            'FW_LT_DRow_Color': {'SWBin': 36, 'HWBin': 4, 'Priority': 304},
            'FW_LT_DCol_Color': {'SWBin': 36, 'HWBin': 4, 'Priority': 305},
            'FW_LT_Mean_R': {'SWBin': 36, 'HWBin': 4, 'Priority': 306},
            'FW_LT_Mean_Gr': {'SWBin': 36, 'HWBin': 4, 'Priority': 307},
            'FW_LT_Mean_Gb': {'SWBin': 36, 'HWBin': 4, 'Priority': 308},
            'FW_LT_Mean_B': {'SWBin': 36, 'HWBin': 4, 'Priority': 309},
            'FW_LT_StdDEV_R': {'SWBin': 36, 'HWBin': 4, 'Priority': 310},
            'FW_LT_StdDEV_Gr': {'SWBin': 36, 'HWBin': 4, 'Priority': 311},
            'FW_LT_StdDEV_Gb': {'SWBin': 36, 'HWBin': 4, 'Priority': 312},
            'FW_LT_StdDEV_B': {'SWBin': 36, 'HWBin': 4, 'Priority': 313},
            'FW_LT_RI_R': {'SWBin': 36, 'HWBin': 4, 'Priority': 314},
            'FW_LT_RI_Gr': {'SWBin': 36, 'HWBin': 4, 'Priority': 315},
            'FW_LT_RI_Gb': {'SWBin': 36, 'HWBin': 4, 'Priority': 316},
            'FW_LT_RI_B': {'SWBin': 36, 'HWBin': 4, 'Priority': 317},
            'FW_LT_Ratio_GrR': {'SWBin': 36, 'HWBin': 4, 'Priority': 318},
            'FW_LT_Ratio_GbR': {'SWBin': 36, 'HWBin': 4, 'Priority': 319},
            'FW_LT_Ratio_GrB': {'SWBin': 36, 'HWBin': 4, 'Priority': 320},
            'FW_LT_Ratio_GbB': {'SWBin': 36, 'HWBin': 4, 'Priority': 321},
            'FW_LT_Ratio_GbGr': {'SWBin': 36, 'HWBin': 4, 'Priority': 322},
            'FW_LT_LostBit': {'SWBin': 36, 'HWBin': 4, 'Priority': 323},
            'MIPI2L_LT_Cap': {'SWBin': 90, 'HWBin': 8, 'Priority': 324},
            'MIPI2L_LT_Calc': {'SWBin': 97, 'HWBin': 8, 'Priority': 325},
            'MIPI2L_LT_DRow_R': {'SWBin': 65, 'HWBin': 4, 'Priority': 326, 'IsNan': 'MIPI2L_LT_Cap'},
            'MIPI2L_LT_DRow_Gr': {'SWBin': 65, 'HWBin': 4, 'Priority': 327},
            'MIPI2L_LT_DRow_Gb': {'SWBin': 65, 'HWBin': 4, 'Priority': 328},
            'MIPI2L_LT_DRow_B': {'SWBin': 65, 'HWBin': 4, 'Priority': 329},
            'MIPI2L_LT_DCol_R': {'SWBin': 64, 'HWBin': 4, 'Priority': 330},
            'MIPI2L_LT_DCol_Gr': {'SWBin': 64, 'HWBin': 4, 'Priority': 331},
            'MIPI2L_LT_DCol_Gb': {'SWBin': 64, 'HWBin': 4, 'Priority': 332},
            'MIPI2L_LT_DCol_B': {'SWBin': 64, 'HWBin': 4, 'Priority': 333},
            'MIPI2L_LT_DRow_Color': {'SWBin': 65, 'HWBin': 4, 'Priority': 334},
            'MIPI2L_LT_DCol_Color': {'SWBin': 64, 'HWBin': 4, 'Priority': 335},
            'MIPI2L_LT_WeakLineRow_R': {'SWBin': 65, 'HWBin': 4, 'Priority': 336},
            'MIPI2L_LT_WeakLineRow_Gr': {'SWBin': 65, 'HWBin': 4, 'Priority': 337},
            'MIPI2L_LT_WeakLineRow_Gb': {'SWBin': 65, 'HWBin': 4, 'Priority': 338},
            'MIPI2L_LT_WeakLineRow_B': {'SWBin': 65, 'HWBin': 4, 'Priority': 339},
            'MIPI2L_LT_WeakLineCol_R': {'SWBin': 64, 'HWBin': 4, 'Priority': 340},
            'MIPI2L_LT_WeakLineCol_Gr': {'SWBin': 64, 'HWBin': 4, 'Priority': 341},
            'MIPI2L_LT_WeakLineCol_Gb': {'SWBin': 64, 'HWBin': 4, 'Priority': 342},
            'MIPI2L_LT_WeakLineCol_B': {'SWBin': 64, 'HWBin': 4, 'Priority': 343},
            'MIPI2L_LT_Mean_R': {'SWBin': 63, 'HWBin': 4, 'Priority': 344},
            'MIPI2L_LT_Mean_Gr': {'SWBin': 63, 'HWBin': 4, 'Priority': 345},
            'MIPI2L_LT_Mean_Gb': {'SWBin': 63, 'HWBin': 4, 'Priority': 346},
            'MIPI2L_LT_Mean_B': {'SWBin': 63, 'HWBin': 4, 'Priority': 347},
            'MIPI2L_LT_StdDEV_R': {'SWBin': 63, 'HWBin': 4, 'Priority': 348},
            'MIPI2L_LT_StdDEV_Gr': {'SWBin': 63, 'HWBin': 4, 'Priority': 349},
            'MIPI2L_LT_StdDEV_Gb': {'SWBin': 63, 'HWBin': 4, 'Priority': 350},
            'MIPI2L_LT_StdDEV_B': {'SWBin': 63, 'HWBin': 4, 'Priority': 351},
            'MIPI2L_LT_RI_R': {'SWBin': 63, 'HWBin': 4, 'Priority': 352},
            'MIPI2L_LT_RI_Gr': {'SWBin': 63, 'HWBin': 4, 'Priority': 353},
            'MIPI2L_LT_RI_Gb': {'SWBin': 63, 'HWBin': 4, 'Priority': 354},
            'MIPI2L_LT_RI_B': {'SWBin': 63, 'HWBin': 4, 'Priority': 355},
            'MIPI2L_LT_Ratio_GrR': {'SWBin': 63, 'HWBin': 4, 'Priority': 356},
            'MIPI2L_LT_Ratio_GbR': {'SWBin': 63, 'HWBin': 4, 'Priority': 357},
            'MIPI2L_LT_Ratio_GrB': {'SWBin': 63, 'HWBin': 4, 'Priority': 358},
            'MIPI2L_LT_Ratio_GbB': {'SWBin': 63, 'HWBin': 4, 'Priority': 359},
            'MIPI2L_LT_Ratio_GbGr': {'SWBin': 63, 'HWBin': 4, 'Priority': 360},
            'MIPI2L_LT_LostBit': {'SWBin': 63, 'HWBin': 4, 'Priority': 361},
            'MIPI2L_LT_LostBitSNR_SumDIFF_R': {'SWBin': 63, 'HWBin': 4, 'Priority': 362},
            'MIPI2L_LT_LostBitSNR_SumDIFF_Gr': {'SWBin': 63, 'HWBin': 4, 'Priority': 363},
            'MIPI2L_LT_LostBitSNR_SumDIFF_Gb': {'SWBin': 63, 'HWBin': 4, 'Priority': 364},
            'MIPI2L_LT_LostBitSNR_SumDIFF_B': {'SWBin': 63, 'HWBin': 4, 'Priority': 365},
            'MIPI4L27M30F_BK1X_Cap': {'SWBin': 96, 'HWBin': 8, 'Priority': 366},
            'MIPI4L27M30F_BK1X_Calc': {'SWBin': 97, 'HWBin': 8, 'Priority': 367},
            'BK1X_BK_DeadRowExBPix_R': {'SWBin': 48, 'HWBin': 6, 'Priority': 368, 'IsNan': 'MIPI4L27M30F_BK1X_Cap'},
            'BK1X_BK_DeadRowExBPix_Gr': {'SWBin': 48, 'HWBin': 6, 'Priority': 369},
            'BK1X_BK_DeadRowExBPix_Gb': {'SWBin': 48, 'HWBin': 6, 'Priority': 370},
            'BK1X_BK_DeadRowExBPix_B': {'SWBin': 48, 'HWBin': 6, 'Priority': 371},
            'BK1X_BK_DeadColExBPix_R': {'SWBin': 47, 'HWBin': 6, 'Priority': 372},
            'BK1X_BK_DeadColExBPix_Gr': {'SWBin': 47, 'HWBin': 6, 'Priority': 373},
            'BK1X_BK_DeadColExBPix_Gb': {'SWBin': 47, 'HWBin': 6, 'Priority': 374},
            'BK1X_BK_DeadColExBPix_B': {'SWBin': 47, 'HWBin': 6, 'Priority': 375},
            'BK1X_BK_Mean_R': {'SWBin': 46, 'HWBin': 6, 'Priority': 376},
            'BK1X_BK_Mean_Gr': {'SWBin': 46, 'HWBin': 6, 'Priority': 377},
            'BK1X_BK_Mean_Gb': {'SWBin': 46, 'HWBin': 6, 'Priority': 378},
            'BK1X_BK_Mean_B': {'SWBin': 46, 'HWBin': 6, 'Priority': 379},
            'BK1X_BK_StdDEV_R': {'SWBin': 46, 'HWBin': 6, 'Priority': 380},
            'BK1X_BK_StdDEV_Gr': {'SWBin': 46, 'HWBin': 6, 'Priority': 381},
            'BK1X_BK_StdDEV_Gb': {'SWBin': 46, 'HWBin': 6, 'Priority': 382},
            'BK1X_BK_StdDEV_B': {'SWBin': 46, 'HWBin': 6, 'Priority': 383},
            'MIPI4L27M30F_BK32X_Cap': {'SWBin': 96, 'HWBin': 8, 'Priority': 384},
            'MIPI4L27M30F_BK32X_Calc': {'SWBin': 97, 'HWBin': 8, 'Priority': 385},
            'BK32X_BK_MixDCol_DashQty_R': {'SWBin': 54, 'HWBin': 6, 'Priority': 386, 'IsNan': 'MIPI4L27M30F_BK32X_Cap'},
            'BK32X_BK_MixDCol_DashQty_Gr': {'SWBin': 54, 'HWBin': 6, 'Priority': 387},
            'BK32X_BK_MixDCol_DashQty_Gb': {'SWBin': 54, 'HWBin': 6, 'Priority': 388},
            'BK32X_BK_MixDCol_DashQty_B': {'SWBin': 54, 'HWBin': 6, 'Priority': 389},
            'BK32X_BK_MixDCol_DASH_DeltaAvg_R': {'SWBin': 54, 'HWBin': 6, 'Priority': 390},
            'BK32X_BK_MixDCol_DASH_DeltaAvg_Gr': {'SWBin': 54, 'HWBin': 6, 'Priority': 391},
            'BK32X_BK_MixDCol_DASH_DeltaAvg_Gb': {'SWBin': 54, 'HWBin': 6, 'Priority': 392},
            'BK32X_BK_MixDCol_DASH_DeltaAvg_B': {'SWBin': 54, 'HWBin': 6, 'Priority': 393},
            'BK32X_BK_MixDCol_DASH_MaxValue0_R': {'SWBin': 54, 'HWBin': 6, 'Priority': 394},
            'BK32X_BK_MixDCol_DASH_MaxValue0_Gr': {'SWBin': 54, 'HWBin': 6, 'Priority': 395},
            'BK32X_BK_MixDCol_DASH_MaxValue0_Gb': {'SWBin': 54, 'HWBin': 6, 'Priority': 396},
            'BK32X_BK_MixDCol_DASH_MaxValue0_B': {'SWBin': 54, 'HWBin': 6, 'Priority': 397},
            'BK32X_BK_MixDCol_DASH_MaxValue1_R': {'SWBin': 54, 'HWBin': 6, 'Priority': 398},
            'BK32X_BK_MixDCol_DASH_MaxValue1_Gr': {'SWBin': 54, 'HWBin': 6, 'Priority': 399},
            'BK32X_BK_MixDCol_DASH_MaxValue1_Gb': {'SWBin': 54, 'HWBin': 6, 'Priority': 400},
            'BK32X_BK_MixDCol_DASH_MaxValue1_B': {'SWBin': 54, 'HWBin': 6, 'Priority': 401},
            'BK32X_BK_MixDCol_DASH_MaxValue2_R': {'SWBin': 54, 'HWBin': 6, 'Priority': 402},
            'BK32X_BK_MixDCol_DASH_MaxValue2_Gr': {'SWBin': 54, 'HWBin': 6, 'Priority': 403},
            'BK32X_BK_MixDCol_DASH_MaxValue2_Gb': {'SWBin': 54, 'HWBin': 6, 'Priority': 404},
            'BK32X_BK_MixDCol_DASH_MaxValue2_B': {'SWBin': 54, 'HWBin': 6, 'Priority': 405},
            'BK32X_BK_MixDCol_DASH_MaxValue3_R': {'SWBin': 54, 'HWBin': 6, 'Priority': 406},
            'BK32X_BK_MixDCol_DASH_MaxValue3_Gr': {'SWBin': 54, 'HWBin': 6, 'Priority': 407},
            'BK32X_BK_MixDCol_DASH_MaxValue3_Gb': {'SWBin': 54, 'HWBin': 6, 'Priority': 408},
            'BK32X_BK_MixDCol_DASH_MaxValue3_B': {'SWBin': 54, 'HWBin': 6, 'Priority': 409},
            'BK32X_BK_MixDCol_DASH_MaxValue4_R': {'SWBin': 54, 'HWBin': 6, 'Priority': 410},
            'BK32X_BK_MixDCol_DASH_MaxValue4_Gr': {'SWBin': 54, 'HWBin': 6, 'Priority': 411},
            'BK32X_BK_MixDCol_DASH_MaxValue4_Gb': {'SWBin': 54, 'HWBin': 6, 'Priority': 412},
            'BK32X_BK_MixDCol_DASH_MaxValue4_B': {'SWBin': 54, 'HWBin': 6, 'Priority': 413},
            'BK32X_BK_Mean_R': {'SWBin': 53, 'HWBin': 6, 'Priority': 415},
            'BK32X_BK_Mean_Gr': {'SWBin': 53, 'HWBin': 6, 'Priority': 416},
            'BK32X_BK_Mean_Gb': {'SWBin': 53, 'HWBin': 6, 'Priority': 417},
            'BK32X_BK_Mean_B': {'SWBin': 53, 'HWBin': 6, 'Priority': 418},
            'BK32X_BK_StdDEV_R': {'SWBin': 53, 'HWBin': 6, 'Priority': 419},
            'BK32X_BK_StdDEV_Gr': {'SWBin': 53, 'HWBin': 6, 'Priority': 420},
            'BK32X_BK_StdDEV_Gb': {'SWBin': 53, 'HWBin': 6, 'Priority': 421},
            'BK32X_BK_StdDEV_B': {'SWBin': 53, 'HWBin': 6, 'Priority': 422},
            'Binning': {'SWBin': 2, 'HWBin': 3, 'Priority': 423},
            'All Pass': {'SWBin': 255, 'HWBin': 3, 'Priority': 424}
        }
}


def parse_file(file):
    """
    parse file
    get basic information,datacheck result,binningcheck result,nanchipno result and summary result
    """
    parse_result = {}
    global high_limit_row_num
    data = []
    # get file data
    with open(file) as f:
        csv_reader = reader(f)
        for row in csv_reader:
            data.append(row)

    # get row and column of last test item
    if project == 'F28':
        search_last_test_item = search_string(data, 'Full_Error')
        search_first_image_test_item = search_string(data, 'BK_DeadRowExBPix_R ')
    elif project == 'JX828':
        search_last_test_item = search_string(data, 'ChipVer')
        search_first_image_test_item = search_string(data, 'BK_DeadRowExBPix_R ')
    elif project == 'JX825':
        search_last_test_item = search_string(data, 'SRAM_0x56_Read')
        search_first_image_test_item = search_string(data, 'BK_DeadRowExBPix_R ')
    if not search_last_test_item:
        exit()
    else:
        test_item_row_num = search_last_test_item[0]
        last_test_item_col_num = search_last_test_item[1]
    if not search_first_image_test_item:
        exit()
    else:
        first_image_test_item_col_num = search_first_image_test_item[1]
    high_limit_row_num = test_item_row_num + 2
    low_limit_row_num = test_item_row_num + 3

    # get row and column of last DC test item
    search_pwdn_total = search_string(data, 'PWDN_Total')
    if not search_pwdn_total:
        exit()
    else:
        pwdn_total_col_num = search_pwdn_total[1]

    # get row and column of Binning
    search_binning = search_string(data, 'Binning')
    if not search_binning:
        exit()
    else:
        binning_col_num = search_binning[1]

    # get column of iic_test
    search_iic_test = search_string(data, 'iic_test')
    if not search_iic_test:
        exit()
    else:
        iic_test_col_num = search_iic_test[1]

    # get row of first register
    first_register_row_num = len(data)
    for i in range(test_item_row_num + row_offset, len(data)):
        try:
            int(data[i][0])
        except:
            first_register_row_num = i
            break

    # fill '' to align with last test item col num
    for i in range(0, len(data)):
        add_count = last_test_item_col_num + 1 - len(data[i])
        if add_count > 0:
            for j in range(add_count):
                data[i].append('')

    # data check----------------------------
    row_count = len(data)

    # get row number with null value
    exist_nan_row_list = []
    for i in range(test_item_row_num + row_offset, first_register_row_num):
        for j in range(binning_col_num + 1):
            if len(data[i][j].strip()) == 0:
                exist_nan_row_list.append(i)
                break

    datacheck_result = []
    data_check_widgets = ['ParseFile-DataCheck: ', Percentage(), ' ', Bar('#'), ' ', Timer(), ' ', ETA(), ' ',
                          FileTransferSpeed()]
    data_check_pbar = ProgressBar(widgets=data_check_widgets, maxval=row_count).start()
    for row_num in range(row_count):
        row_data = []
        for col_num in range(last_test_item_col_num + 1):
            test_item_name = data[test_item_row_num][col_num].strip()
            value = data[row_num][col_num]
            if col_offset <= col_num <= binning_col_num and high_limit_row_num <= row_num <= low_limit_row_num:
                # fill color of limit value is green
                row_data.append((value, '008000'))
            elif col_num == 0 and row_num in exist_nan_row_list:
                # fill color of first value of exist nan row is purple
                row_data.append((value, 'A020F0'))
            elif col_offset <= col_num <= binning_col_num and test_item_row_num + row_offset <= row_num < first_register_row_num:
                try:
                    value_convert = float(value)
                    high_limit_data = data[high_limit_row_num][col_num]
                    try:
                        high_limit = float(high_limit_data)
                    except:
                        high_limit = high_limit_data
                    low_limit_data = data[low_limit_row_num][col_num]
                    try:
                        low_limit = float(low_limit_data)
                    except:
                        low_limit = low_limit_data
                    if value_convert == int(value_convert):
                        value_convert = int(value_convert)
                    if (test_item_name == 'AVDD_O/S' and project == 'JX828') or high_limit == 'N' or (
                                    high_limit != 'N' and low_limit <= value_convert <= high_limit):
                        # fill color of following scenario value is white
                        # 1.JX828 ignore AVDD_O/S
                        # 2.limit is N
                        # 3.limit is not N and data wihtin limit
                        row_data.append((value_convert, 'FFFFFF'))
                    else:
                        # fill color of over limit is red
                        row_data.append((value_convert, 'FF0000'))
                except:
                    if not value.strip():
                        # fill color of '' is purple
                        row_data.append((value, 'A020F0'))
                    elif col_num == iic_test_col_num:
                        if value == '1':
                            # fill color of iic_test 1 is white
                            row_data.append((value_convert, 'FFFFFF'))
                        else:
                            # fill color of iic_test not 1 is yellow
                            row_data.append((value, 'FFFF00'))
            else:
                # set '' when value is nan
                if isinstance(value, float) and isnan(value):
                    value = ''
                # fill color is white
                row_data.append((value, 'FFFFFF'))
        datacheck_result.append(row_data)
        if row_num % 1000 == 0 or row_num == row_count - 1:
            # update bar when row num divisible by 1000 or is last row count
            data_check_pbar.update(row_num + 1)
    data_check_pbar.finish()

    # binning check----------------------------
    bin_definition = bin_definition_list[project]
    binning_check_result = []
    binning_check_widgets = ['ParseFile-BinningCheck: ', Percentage(), ' ', Bar('#'), ' ', Timer(), ' ', ETA(), ' ',
                             FileTransferSpeed()]
    binning_check_pbar = ProgressBar(widgets=binning_check_widgets, maxval=first_register_row_num).start()
    for row_num in range(test_item_row_num + row_offset, first_register_row_num):
        # set current_test_item is All Pass
        current_test_item = 'All Pass'
        for col_num in range(col_offset, binning_col_num + 1):
            test_item = data[test_item_row_num][col_num].strip()
            if test_item == 'AVDD_O/S' and project == 'JX828':
                # JX828 skip AVDD_O/S
                continue
            else:
                high_limit_data = data[high_limit_row_num][col_num]
                low_limit_data = data[low_limit_row_num][col_num]
            try:
                high_limit = float(high_limit_data)
                low_limit = float(low_limit_data)
                try:
                    value = data[row_num][col_num]
                    value_convert = float(value)
                    if value_convert == int(value_convert):
                        value_convert = int(value_convert)
                    if value_convert < low_limit or high_limit < value_convert:
                        # over limit
                        if bin_definition[test_item]['Priority'] < bin_definition[current_test_item]['Priority']:
                            # set current_test_item is test_item when test_item's priority less than current_test_item's priority
                            current_test_item = test_item
                            break
                except:
                    # value is ''
                    if col_num >= first_image_test_item_col_num and \
                                    bin_definition[bin_definition[test_item]['IsNan']]['Priority'] < \
                                    bin_definition[current_test_item]['Priority']:
                        # set current_test_item is bin_definition[test_item]['IsNan']
                        # when column number greater than or euqal to first image test item
                        # and bin_definition[test_item]['IsNan']'s priority less than current_test_item's priority
                        current_test_item = bin_definition[test_item]['IsNan']
                        break
                    elif col_num < first_image_test_item_col_num and bin_definition[test_item]['Priority'] < \
                            bin_definition[current_test_item]['Priority']:
                        # set current_test_item is test_item
                        # when when when column number less than first image test item
                        # and test_item's priority less than current_test_item's priority
                        current_test_item = test_item
                        break
            except:
                # limit is N or nan
                try:
                    value = data[row_num][col_num]
                    float(value)
                    if test_item == 'iic_test' and value == '0':
                        if bin_definition[test_item]['Priority'] < bin_definition[current_test_item]['Priority']:
                            # set current_test_item is bin_definition[test_item]['IsNan']
                            # when bin_definition[test_item]['IsNan']'s priority less than current_test_item's priority
                            current_test_item = test_item
                            break
                except:
                    # value is ''
                    if col_num >= first_image_test_item_col_num and \
                                    bin_definition[bin_definition[test_item]['IsNan']]['Priority'] < \
                                    bin_definition[current_test_item]['Priority']:
                        # set current_test_item is bin_definition[test_item]['IsNan']
                        # when column number greater than or euqal to first image test item
                        # and bin_definition[test_item]['IsNan']'s priority less than current_test_item's priority
                        current_test_item = bin_definition[test_item]['IsNan']
                        break
                    elif col_num < first_image_test_item_col_num and bin_definition[test_item]['Priority'] < \
                            bin_definition[current_test_item]['Priority']:
                        # set current_test_item is test_item
                        # when when when column number less than first image test item
                        # and test_item's priority less than current_test_item's priority
                        current_test_item = test_item
                        break

        if bin_definition[current_test_item]['SWBin'] != int(data[row_num][2]):
            # current_test_item's SWBin is not equal to SW_BIN in csv file
            binning_check_result.append(
                "              SB_BIN error of chipNo " + data[row_num][
                    0] + " : according to the priority,swbin should be " + str(
                    bin_definition[current_test_item]['SWBin']) + " ,but in CSV it's " + data[row_num][2] + "\n")
        if bin_definition[current_test_item]['HWBin'] != int(data[row_num][3]):
            # current_test_item's HWBin is not equal to hW_BIN in csv file
            binning_check_result.append(
                "              hW_BIN error of chipNo " + data[row_num][
                    0] + " : according to the priority,hwbin should be " + str(
                    bin_definition[current_test_item]['SWBin']) + " ,but in CSV it's " + data[row_num][3] + "\n")
        if row_num % 1000 == 0 or row_num == first_register_row_num - 1:
            # update bar when row num divisible by 1000 or is last chipno row
            binning_check_pbar.update(row_num + 1)
    binning_check_pbar.finish()

    # nan chipno----------------------------
    dc_nan_chipno = []
    image_nan_chipno = []
    for i in range(test_item_row_num + row_offset, first_register_row_num):
        for j in range(col_offset, pwdn_total_col_num + 1):
            if len(data[i][j].strip()) == 0:
                # append value which is '' in columns col_offset to pwdn_total_col_num
                dc_nan_chipno.append(data[i][0])
                break
    for i in range(test_item_row_num + row_offset, first_register_row_num):
        for j in range(pwdn_total_col_num + 1, binning_col_num + 1):
            if len(data[i][j].strip()) == 0:
                # append value which is '' in columns pwdn_total_col_num+1 to binning_col_num
                image_nan_chipno.append(data[i][0])
                break
    nan_chipno_result = {'DC': dc_nan_chipno, 'IMAGE': image_nan_chipno}

    # summary----------------------------
    summary_result = {}
    for group_by_id in range(1, 4):
        # 1:Site    2:SW_BIN    3:hW_BIN
        group_name = data[test_item_row_num][group_by_id]
        group_list = []
        for i in range(test_item_row_num + row_offset, first_register_row_num):
            # get value
            group_list.append(int(data[i][group_by_id]))
        # remove duplicate value
        format_group_list = list(set(group_list))
        # sort list
        format_group_list.sort()

        group_index = {}
        for i in format_group_list:
            temp = []
            # find i in group_list
            data_list = find_item(group_list, i)
            if group_by_id == 1:
                temp_list = []
                for j in data_list:
                    # get the corresponding SW_BIN of Site
                    temp_list.append(data[test_item_row_num + row_offset + j][2])
                # remove duplicate value
                format_temp_list = list(set(temp_list))
                temp_index = {}
                for m in format_temp_list:
                    # find m in temp_list
                    temp_data_list = find_item(temp_list, m)
                    temp_index[m] = temp_data_list
                temp.append(temp_index)
            else:
                temp.append(data_list)
            testitem_fail_count = []
            for col_num in range(col_offset, last_test_item_col_num + 1):
                test_item = data[test_item_row_num][col_num].strip()
                if test_item == 'AVDD_O/S' and project == 'JX828':
                    # set AVDD_O/S's limit in JX828 project
                    high_limit_data = -0.2
                    low_limit_data = -0.6
                else:
                    high_limit_data = data[high_limit_row_num][col_num]
                    low_limit_data = data[low_limit_row_num][col_num]
                temp_list = [data[test_item_row_num][col_num], 0]
                try:
                    high_limit = float(high_limit_data)
                    low_limit = float(low_limit_data)
                    for j in data_list:
                        try:
                            value = data[test_item_row_num + row_offset + j][col_num]
                            value_convert = float(value)
                            if value_convert == int(value_convert):
                                value_convert = int(value_convert)
                            if value_convert < low_limit or high_limit < value_convert:
                                # count+1 when value over limit
                                temp_list[1] += 1
                        except:
                            # count+1 when value is ''
                            temp_list[1] += 1
                except:
                    # limit is N or nan
                    continue
                testitem_fail_count.append(temp_list)
            temp.append(testitem_fail_count)
            group_index[i] = temp
        summary_result[group_name] = group_index
    # parse file name
    file_name = basename(file).split('.')[0]
    # calculate chip count
    chip_count = first_register_row_num - test_item_row_num - row_offset
    # get lotno
    lotno = data[5][1]
    parse_result['file name'] = file_name
    parse_result['chip count'] = chip_count
    parse_result['lotno'] = lotno
    parse_result['data check'] = datacheck_result
    parse_result['binning check'] = binning_check_result
    parse_result['nan chipno'] = nan_chipno_result
    parse_result['summary'] = summary_result
    return parse_result


def save_data(analysis_folder, parse_data):
    """
    save data
    save data check,binning check,nan chipno and summary to files
    """
    site_data = []
    softbin_data = []
    hardbin_data = []
    lot_count = parse_data[0]['chip count']
    lotno = parse_data[0]['lotno']
    for data in parse_data:
        file_name = data['file name']
        datacheck_data = data['data check']
        binning_check_parse_data = data['binning check']
        nan_chipno_parse_data = data['nan chipno']
        site_data.append(data['summary']['Site'])
        # sort softbin
        softbin_data.append(sorted(data['summary']['SW_BIN'].items(), key=lambda item: len(item[1][0]), reverse=True))
        hardbin_data.append(data['summary']['hW_BIN'])

        # data check----------------------------
        data_check_file = join(analysis_folder, file_name + '_DataCheck_Analysis' + now_time + '.xlsx')
        print(data_check_file)
        data_check_wb = Workbook()  # create file object
        data_check_sheet = data_check_wb.active  # get first sheet
        data_check_sheet.freeze_panes = 'E17'  # set freeze panes

        irow = 1
        gather_data_widgets = ['GatherData-DataCheck: ', Percentage(), ' ', Bar('#'), ' ', Timer(), ' ', ETA(), ' ',
                               FileTransferSpeed()]
        gather_data_pbar = ProgressBar(widgets=gather_data_widgets, maxval=len(datacheck_data)).start()
        for i in range(len(datacheck_data)):
            for j in range(len(datacheck_data[i])):
                data_check_sheet.cell(row=irow, column=j + 1).value = datacheck_data[i][j][0]
                data_check_sheet.cell(row=irow, column=j + 1).fill = PatternFill(fill_type='solid',
                                                                                 fgColor=datacheck_data[i][j][1])
                data_check_sheet.cell(row=irow, column=j + 1).border = border
            if i % 100 == 0 or i == len(datacheck_data) - 1:
                gather_data_pbar.update(i + 1)
            irow += 1
        gather_data_pbar.finish()
        print("save data check begin>>>>>>")
        data_check_wb.save(data_check_file)
        print("save data check finished<<<<<<")

        # binning check----------------------------
        binning_check_file = join(analysis_folder, lotno + '_BinningCheck_Analysis' + now_time + '.txt')
        print("save binning check data begin>>>>>>")
        with open(binning_check_file, 'a') as f:
            f.write("            " + file_name + "：\n")
            if len(binning_check_parse_data) > 0:
                for item in binning_check_parse_data:
                    f.write(item)
            else:
                f.write('              No problem.\n')
        print("save binning check data finished<<<<<<")

        # nan chipno----------------------------
        nan_chipno_file = join(analysis_folder, lotno + '_NanChipno_Analysis' + now_time + '.txt')
        dc_count = len(nan_chipno_parse_data['DC'])
        image_count = len(nan_chipno_parse_data['IMAGE'])
        print("save nan chipno data begin>>>>>>")
        with open(nan_chipno_file, 'a') as f:
            f.write("            " + file_name + "：\n")
            if dc_count > 0 or image_count > 0:
                if dc_count == 0:
                    f.write("              (1) DC : no nan value.\n")
                else:
                    f.write("              (1) DC : " + str(dc_count) + " in total.They are : " + str(
                        nan_chipno_parse_data['DC']) + "\n")
                if image_count == 0:
                    f.write("              (2) IMAGE : no nan value\n")
                else:
                    f.write("              (2) IMAGE : " + str(image_count) + " in total.They are : " + str(
                        nan_chipno_parse_data['IMAGE']) + '\n')
            else:
                f.write("              No nan value.\n")
        print("save nan chipno data finished<<<<<<")

    # Summary----------------------------
    summary_file = join(analysis_folder, lotno + '_Summary_Analysis_' + now_time + '.xlsx')
    summary_wb = Workbook()

    ok_hwbin_count = 0
    begin_fail_swbin = 0
    for hw_bin_key in hwbin_to_swbin[project].keys():
        if hwbin_to_swbin[project][hw_bin_key]['isPassBin']:
            ok_hwbin_count += 1
            begin_fail_swbin += len(hwbin_to_swbin[project][hw_bin_key]['SWBin'])

    color_list = ['99FFFF', '33FF00', 'FFFFCC', 'FFFF33', 'FF9900', 'FF0099', 'FF0000']
    swbin_list = []
    key_index = 0
    for hwbin_key in hwbin_to_swbin[project].keys():
        for swbin in hwbin_to_swbin[project][hwbin_key]['SWBin']:
            swbin_list.append([swbin, color_list[key_index]])
        key_index += 1

    hardbin_sheet = summary_wb.create_sheet('HWBin')
    hardbin_sheet.freeze_panes = 'B2'
    summary_data = []
    for i in range(len(hardbin_data)):
        temp_list = []
        if i == 0:
            temp_list.append('FT')
        else:
            temp_list.append('RT' + str(i))
        test_count = 0
        pass_count = 0
        for hw_bin_key in hwbin_to_swbin[project].keys():
            if hw_bin_key in hardbin_data[i].keys():
                bin_count = len(hardbin_data[i][hw_bin_key][0])
            else:
                bin_count = 0
            if hwbin_to_swbin[project][hw_bin_key]['isPassBin']:
                pass_count += bin_count
            test_count += bin_count
            temp_list.append(bin_count)
        temp_list.append(test_count)
        pass_percent = '{:.2%}'.format(pass_count / test_count)
        temp_list.append(pass_percent)
        summary_data.append(temp_list)
    irow = 1
    icol = 1
    hardbin_sheet.cell(row=irow, column=icol).value = lotno
    icol += 1
    for hw_bin_key in hwbin_to_swbin[project].keys():
        hardbin_sheet.cell(row=irow, column=icol).value = 'HWBin' + str(hw_bin_key)
        icol += 1
    hardbin_sheet.cell(row=irow, column=icol).value = 'TestCount'
    icol += 1
    hardbin_sheet.cell(row=irow, column=icol).value = 'PassPercent'
    irow += 1
    for i in range(len(summary_data)):
        for j in range(len(summary_data[i])):
            hardbin_sheet.cell(row=irow, column=(j + 1)).value = summary_data[i][j]
            if j == 0:
                hardbin_sheet.cell(row=irow, column=(j + 1)).fill = PatternFill(fill_type='solid', fgColor=GREEN)
            elif 1 <= j < len(summary_data[i]) - 2:
                hardbin_sheet.cell(row=irow + 1, column=(j + 1)).value = '{:.2%}'.format(
                    summary_data[i][j] / summary_data[i][-2])
                hardbin_sheet.cell(row=irow + 1, column=(j + 1)).fill = PatternFill(fill_type='solid', fgColor=GREEN)
            elif j == len(summary_data[i]) - 2 and i > 0:
                compare_count = 0
                for m in range(ok_hwbin_count + 1, len(summary_data[i]) - 2):
                    compare_count += summary_data[i - 1][m]
                if summary_data[i][j] != compare_count:
                    # the number of test is not equal to the total number of chips failed in the previous test
                    if summary_data[i][j] > compare_count:
                        flag = '↓'
                    else:
                        flag = '↑'
                    hardbin_sheet.cell(row=irow, column=(j + 1)).fill = PatternFill(fill_type='solid', fgColor=RED)
                    hardbin_sheet.cell(row=irow + 1, column=(j + 1)).value = flag + str(compare_count)
                    hardbin_sheet.cell(row=irow + 1, column=(j + 1)).font = Font(color=RED, bold=True)
        irow += 2
    hardbin_sheet.cell(row=irow, column=1).value = 'Summary'
    hardbin_sheet.cell(row=irow, column=1).fill = PatternFill(fill_type='solid', fgColor=GREEN)
    # summary pass hwbin count : add all lots pass hwbin count
    pass_hwbin_count = [0] * ok_hwbin_count
    for i in range(len(summary_data)):
        for j in range(len(pass_hwbin_count)):
            pass_hwbin_count[j] += summary_data[i][j + 1]
    # summary total pass hwbin count : add all pass hwbin count
    total_pass_hwbin_count = 0
    for i in range(len(pass_hwbin_count)):
        hardbin_sheet.cell(row=irow, column=2 + i).value = pass_hwbin_count[i]
        total_pass_hwbin_count += pass_hwbin_count[i]
    # summary fail hwbin count : last lot's fail hwbin count
    for i in range(len(pass_hwbin_count) + 1, len(summary_data[-1]) - 2):
        hardbin_sheet.cell(row=irow, column=(i + 1)).value = summary_data[-1][i]
    hardbin_sheet.cell(row=irow, column=icol - 1).value = summary_data[0][-2]
    hardbin_sheet.cell(row=irow, column=icol).value = '{:.2%}'.format(total_pass_hwbin_count / summary_data[0][-2])

    for row in hardbin_sheet.rows:
        for cell in row:
            cell.border = border
            if cell.row == 1:
                cell.fill = PatternFill(fill_type='solid', fgColor=YELLOW)
                cell.font = Font(bold=True)
                cell.alignment = alignment
            if cell.row == 2 and cell.column == hardbin_sheet.max_column:
                cell.fill = PatternFill(fill_type='solid', fgColor=GREEN)
            if cell.row == hardbin_sheet.max_row and cell.column == hardbin_sheet.max_column:
                cell.fill = PatternFill(fill_type='solid', fgColor=GREEN)

    hardsoftbin_sheet = summary_wb.create_sheet('HWBin-SWBin')
    hardsoftbin_sheet.freeze_panes = 'C2'
    irow = 1
    hardsoftbin_sheet.cell(row=irow, column=1).value = lotno
    hardsoftbin_sheet.cell(row=irow, column=1).fill = PatternFill(fill_type='solid', fgColor=YELLOW)
    hardsoftbin_sheet.cell(row=irow, column=3).value = 'FT'
    hardsoftbin_sheet.cell(row=irow, column=3).fill = PatternFill(fill_type='solid', fgColor=YELLOW)
    hardsoftbin_sheet.merge_cells(start_row=irow, end_row=irow, start_column=3, end_column=4)
    for i in range(1, len(softbin_data)):
        hardsoftbin_sheet.cell(row=irow, column=3 + 2 * i).value = 'RT' + str(i)
        hardsoftbin_sheet.cell(row=irow, column=3 + 2 * i).fill = PatternFill(fill_type='solid', fgColor=YELLOW)
        hardsoftbin_sheet.merge_cells(start_row=irow, end_row=irow, start_column=3 + 2 * i, end_column=4 + 2 * i)
    irow += 1

    # sort swbin in hwbin from FT by swbin count from more to less
    swbin_sort_by_FT_list = []
    for hw_bin_key in hwbin_to_swbin[project].keys():
        temp_list = []
        for swbin in hwbin_to_swbin[project][hw_bin_key]['SWBin']:
            find_softbin = False
            for x in range(len(softbin_data[0])):
                if swbin == softbin_data[0][x][0]:
                    find_softbin = True
                    swbin_count = len(softbin_data[0][x][1][0])
                    temp_list.append((swbin, swbin_count))
                    break
            if not find_softbin:
                temp_list.append((swbin, 0))
        for m in range(len(temp_list) - 1):
            for n in range(m + 1, len(temp_list)):
                if temp_list[m][1] < temp_list[n][1]:
                    temp_list[m], temp_list[n] = temp_list[n], temp_list[m]
        temp_count_list = [hw_bin_key]
        for y in range(len(temp_list)):
            temp_count_list.append(temp_list[y][0])
        swbin_sort_by_FT_list.append(temp_count_list)

    summary_soft_count = []
    for i in range(len(swbin_sort_by_FT_list)):
        hardsoftbin_sheet.cell(row=irow, column=1).value = 'HWBin' + str(swbin_sort_by_FT_list[i][0])
        hardsoftbin_sheet.cell(row=irow, column=1).fill = PatternFill(fill_type='solid', fgColor=GREEN)
        hardsoftbin_sheet.cell(row=irow, column=1).font = Font(bold=True)
        for j in range(1, len(swbin_sort_by_FT_list[i])):
            hardsoftbin_sheet.cell(row=irow, column=2).value = 'SWBin' + str(swbin_sort_by_FT_list[i][j])
            temp_count = 0
            for m in range(len(softbin_data)):
                find_softbin = False
                for n in range(len(softbin_data[m])):
                    if swbin_sort_by_FT_list[i][j] == softbin_data[m][n][0]:
                        find_softbin = True
                        swbin_count = len(softbin_data[m][n][1][0])
                        break
                if not find_softbin:
                    swbin_count = 0
                if i in range(ok_hwbin_count):
                    # add the number of pass hwbin's swbin of all lots
                    temp_count += swbin_count
                elif m == len(softbin_data) - 1:
                    # set the number of fail hwbin's swbin of last lot
                    temp_count = swbin_count
                hardsoftbin_sheet.cell(row=irow, column=(2 * m + 3)).value = swbin_count
                hardsoftbin_sheet.cell(row=irow, column=(2 * m + 4)).value = '{:.2%}'.format(
                    swbin_count / summary_data[0][-2])
                hardsoftbin_sheet.cell(row=irow, column=(2 * m + 4)).fill = PatternFill(fill_type='solid',
                                                                                        fgColor=GREEN)
            irow += 1
            summary_soft_count.append(temp_count)
        summary_soft_count.append('')
        irow += 1

    irow = 1
    icol = 4 + 2 * len(softbin_data)
    hardsoftbin_sheet.cell(row=irow, column=icol).value = 'Summary'
    hardsoftbin_sheet.cell(row=irow, column=icol).fill = PatternFill(fill_type='solid', fgColor='FFA500')
    hardsoftbin_sheet.merge_cells(start_row=irow, end_row=irow, start_column=icol, end_column=icol + 1)
    irow += 1
    for i in range(len(summary_soft_count)):
        if isinstance(summary_soft_count[i], int):
            hardsoftbin_sheet.cell(row=irow, column=icol).value = summary_soft_count[i]
            hardsoftbin_sheet.cell(row=irow, column=icol + 1).value = '{:.2%}'.format(
                summary_soft_count[i] / summary_data[0][-2])
            hardsoftbin_sheet.cell(row=irow, column=icol + 1).fill = PatternFill(fill_type='solid', fgColor='FFA500')
        irow += 1

    for row in hardsoftbin_sheet.rows:
        for cell in row:
            cell.border = border
            if cell.row == 1:
                cell.font = Font(bold=True)
                cell.alignment = alignment

    sitesoftbin_sheet = summary_wb.create_sheet('Site-SWBin')
    sitesoftbin_sheet.freeze_panes = 'B2'
    irow = 1
    sitesoftbin_sheet.cell(row=irow, column=1).value = lotno
    sitesoftbin_sheet.cell(row=irow, column=1).fill = PatternFill(fill_type='solid', fgColor=YELLOW)
    for i in range(16):
        sitesoftbin_sheet.cell(row=irow, column=2 + i).value = 'Site' + str(i)
        sitesoftbin_sheet.cell(row=irow, column=2 + i).fill = PatternFill(fill_type='solid', fgColor=YELLOW)
    sitesoftbin_sheet.merge_cells(start_row=irow, end_row=irow, start_column=19, end_column=20)
    sitesoftbin_sheet.cell(row=irow, column=19).value = 'Summary'
    sitesoftbin_sheet.cell(row=irow, column=19).fill = PatternFill(fill_type='solid', fgColor='FFA500')
    irow += 1

    lotno_site_swbin_count = []
    for x in range(len(swbin_list)):
        site_count_list = [[0, 0], [1, 0], [2, 0], [3, 0], [4, 0], [5, 0], [6, 0], [7, 0], [8, 0], [9, 0], [10, 0],
                           [11, 0], [12, 0], [13, 0], [14, 0], [15, 0]]
        swbin = str(swbin_list[x][0])
        for site in range(16):
            if x < begin_fail_swbin:
                for i in range(len(site_data)):
                    if (site in site_data[i].keys()) and (swbin in site_data[i][site][0].keys()):
                        site_count_list[site][1] += len(site_data[i][site][0][swbin])
            else:
                if (site in site_data[-1].keys()) and (swbin in site_data[-1][site][0].keys()):
                    site_count_list[site][1] += len(site_data[-1][site][0][swbin])
        lotno_site_swbin_count.append(site_count_list)
    site_swbin_count = []
    for i in range(len(lotno_site_swbin_count)):
        temp_total_count = 0
        temp_list = [0] * 16
        temp_swbin_count = []
        for j in range(len(temp_list)):
            temp_list[j] += lotno_site_swbin_count[i][j][1]
            temp_swbin_count.append([temp_list[j], WHITE])
            temp_total_count += temp_list[j]
            if j == len(temp_list) - 1:
                temp_swbin_count.append([temp_total_count, 'FFA500'])
        site_swbin_count.append(temp_swbin_count)
    # add the number of fail swbin by all site
    # fill color of max value in fail swbin is green(all values are 0 do not fill green)
    # fill color of min value in fail swbin is red(multiple values of 0 do not fill red)
    site_pass_total_list = [0] * 16
    for i in range(begin_fail_swbin):
        for j in range(len(site_swbin_count[i]) - 1):
            site_pass_total_list[j] += site_swbin_count[i][j][0]
    site_fail_total_list = [0] * 16
    for i in range(begin_fail_swbin, len(site_swbin_count)):
        min_value = site_swbin_count[i][0][0]
        max_value = site_swbin_count[i][0][0]
        for m in range(1, len(site_swbin_count[i]) - 1):
            if min_value > site_swbin_count[i][m][0]:
                min_value = site_swbin_count[i][m][0]
            if max_value < site_swbin_count[i][m][0]:
                max_value = site_swbin_count[i][m][0]
        min_index = []
        max_index = []
        for j in range(len(site_swbin_count[i]) - 1):
            site_fail_total_list[j] += site_swbin_count[i][j][0]
            if site_swbin_count[i][j][0] == min_value:
                min_index.append(j)
            if site_swbin_count[i][j][0] == max_value:
                max_index.append(j)
        if min_value == 0 and len(min_index) == 1:
            site_swbin_count[i][min_index[0]][1] = GREEN
            for y in range(len(max_index)):
                site_swbin_count[i][max_index[y]][1] = RED
        elif min_value == 0 and max_value > 0:
            for y in range(len(max_index)):
                site_swbin_count[i][max_index[y]][1] = RED
        elif min_value > 0:
            for x in range(len(min_index)):
                site_swbin_count[i][min_index[x]][1] = GREEN
            for y in range(len(max_index)):
                site_swbin_count[i][max_index[y]][1] = RED

    for x in range(len(swbin_list)):
        sitesoftbin_sheet.cell(row=irow, column=1).value = 'SWBin' + str(swbin_list[x][0])
        sitesoftbin_sheet.cell(row=irow, column=1).fill = PatternFill(fill_type='solid', fgColor=swbin_list[x][1])
        for y in range(len(site_swbin_count[x])):
            if y < len(site_swbin_count[x]) - 1:
                if site_swbin_count[x][y][0] > 0:
                    sitesoftbin_sheet.cell(row=irow, column=2 + y).value = '{:.4%}'.format(
                        site_swbin_count[x][y][0] / summary_data[0][-2])
                sitesoftbin_sheet.cell(row=irow, column=2 + y).fill = PatternFill(fill_type='solid',
                                                                                  fgColor=site_swbin_count[x][y][1])
            else:
                sitesoftbin_sheet.cell(row=irow, column=3 + y).value = site_swbin_count[x][y][0]
                sitesoftbin_sheet.cell(row=irow, column=4 + y).value = '{:.4%}'.format(
                    site_swbin_count[x][y][0] / summary_data[0][-2])
                sitesoftbin_sheet.cell(row=irow, column=4 + y).fill = PatternFill(fill_type='solid',
                                                                                  fgColor=site_swbin_count[x][y][1])
        irow += 1
    irow += 1
    sitesoftbin_sheet.merge_cells(start_row=irow, end_row=irow + 1, start_column=1, end_column=1)
    sitesoftbin_sheet.cell(row=irow, column=1).value = 'FailPercent'
    sitesoftbin_sheet.cell(row=irow, column=1).font = Font(bold=True)
    sitesoftbin_sheet.cell(row=irow, column=1).alignment = alignment
    sitesoftbin_sheet.cell(row=irow, column=1).fill = PatternFill(fill_type='solid', fgColor='FFA500')
    for i in range(len(site_fail_total_list)):
        sitesoftbin_sheet.cell(row=irow, column=2 + i).value = site_fail_total_list[i]
        sitesoftbin_sheet.cell(row=irow + 1, column=2 + i).value = '{:.4%}'.format(
            site_fail_total_list[i] / summary_data[0][-2])
        sitesoftbin_sheet.cell(row=irow + 1, column=2 + i).fill = PatternFill(fill_type='solid', fgColor=RED)
    sitesoftbin_sheet.cell(row=irow, column=3 + len(site_fail_total_list)).value = sum(site_fail_total_list)
    sitesoftbin_sheet.cell(row=irow, column=3 + len(site_fail_total_list)).fill = PatternFill(fill_type='solid',
                                                                                              fgColor=RED)
    sitesoftbin_sheet.cell(row=irow + 1, column=3 + len(site_fail_total_list)).value = '{:.4%}'.format(
        sum(site_fail_total_list) / summary_data[0][-2])
    sitesoftbin_sheet.cell(row=irow + 1, column=3 + len(site_fail_total_list)).fill = PatternFill(fill_type='solid',
                                                                                                  fgColor=RED)
    irow += 2
    sitesoftbin_sheet.merge_cells(start_row=irow, end_row=irow + 1, start_column=1, end_column=1)
    sitesoftbin_sheet.cell(row=irow, column=1).value = 'PassPercent'
    sitesoftbin_sheet.cell(row=irow, column=1).font = Font(bold=True)
    sitesoftbin_sheet.cell(row=irow, column=1).alignment = alignment
    sitesoftbin_sheet.cell(row=irow, column=1).fill = PatternFill(fill_type='solid', fgColor=GREEN)
    for i in range(len(site_pass_total_list)):
        sitesoftbin_sheet.cell(row=irow, column=2 + i).value = site_pass_total_list[i]
        sitesoftbin_sheet.cell(row=irow + 1, column=2 + i).value = '{:.4%}'.format(
            site_pass_total_list[i] / summary_data[0][-2])
        sitesoftbin_sheet.cell(row=irow + 1, column=2 + i).fill = PatternFill(fill_type='solid', fgColor=GREEN)
    sitesoftbin_sheet.cell(row=irow, column=3 + len(site_pass_total_list)).value = sum(site_pass_total_list)
    sitesoftbin_sheet.cell(row=irow, column=3 + len(site_pass_total_list)).fill = PatternFill(fill_type='solid',
                                                                                              fgColor=GREEN)
    sitesoftbin_sheet.cell(row=irow + 1, column=3 + len(site_pass_total_list)).value = '{:.4%}'.format(
        sum(site_pass_total_list) / summary_data[0][-2])
    sitesoftbin_sheet.cell(row=irow + 1, column=3 + len(site_pass_total_list)).fill = PatternFill(fill_type='solid',
                                                                                                  fgColor=GREEN)
    for row in sitesoftbin_sheet.rows:
        for cell in row:
            cell.border = border
            if cell.row == 1:
                cell.font = Font(bold=True)
                cell.alignment = alignment

    hwbin_testitem_sheet = summary_wb.create_sheet('HWBin-TestItem')
    hwbin_testitem_sheet.freeze_panes = 'B2'
    irow = 1
    hwbin_testitem_sheet.cell(row=irow, column=1).value = lot_count
    for hwbin_key in hardbin_data[0].keys():
        for i in range(len(hardbin_data[0][hwbin_key][1])):
            hwbin_testitem_sheet.cell(row=irow, column=i + 2).value = hardbin_data[0][hwbin_key][1][i][0]
        break
    irow += 1
    for hwbin_key in hardbin_data[0].keys():
        hwbin_testitem_sheet.cell(row=irow, column=1).value = 'HWBin' + str(hwbin_key)
        hwbin_testitem_sheet.cell(row=irow, column=1).fill = PatternFill(fill_type='solid', fgColor=GREEN)
        for i in range(len(hardbin_data[0][hwbin_key][1])):
            hwbin_testitem_sheet.cell(row=irow, column=2 + i).value = hardbin_data[0][hwbin_key][1][i][1]
            hwbin_testitem_sheet.cell(row=irow + 1, column=2 + i).value = '{:.2%}'.format(
                hardbin_data[0][hwbin_key][1][i][1] / lot_count)
            if hardbin_data[0][hwbin_key][1][i][1] > 0:
                # percent value greater than 0 are filled in red
                hwbin_testitem_sheet.cell(row=irow + 1, column=2 + i).fill = PatternFill(fill_type='solid',
                                                                                         fgColor=RED)
        irow += 2

    for row in hwbin_testitem_sheet.rows:
        for cell in row:
            cell.border = border
            if cell.row == 1 and cell.column > 1:
                cell.fill = PatternFill(fill_type='solid', fgColor=YELLOW)
                cell.font = Font(bold=True)
                cell.alignment = alignment

    swbin_testitem_sheet = summary_wb.create_sheet('SWBin-TestItem')
    swbin_testitem_sheet.freeze_panes = 'B2'
    irow = 1
    swbin_testitem_sheet.cell(row=irow, column=1).value = lot_count
    for i in range(len(softbin_data[0][0][1][1])):
        swbin_testitem_sheet.cell(row=irow, column=i + 2).value = softbin_data[0][0][1][1][i][0]
    irow += 1
    for x in range(len(swbin_list)):
        swbin_testitem_sheet.cell(row=irow, column=1).value = 'SWBin' + str(swbin_list[x][0])
        swbin_testitem_sheet.cell(row=irow, column=1).fill = PatternFill(fill_type='solid',
                                                                         fgColor=swbin_list[x][1])
        for i in range(len(softbin_data[0])):
            if swbin_list[x][0] == softbin_data[0][i][0]:
                for j in range(len(softbin_data[0][i][1][1])):
                    swbin_testitem_sheet.cell(row=irow, column=2 + j).value = softbin_data[0][i][1][1][j][1]
                    swbin_testitem_sheet.cell(row=irow + 1, column=2 + j).value = '{:.2%}'.format(
                        softbin_data[0][i][1][1][j][1] / lot_count)
                    if softbin_data[0][i][1][1][j][1] > 0:
                        # percent value greater than 0 are filled in red
                        swbin_testitem_sheet.cell(row=irow + 1, column=2 + j).fill = PatternFill(fill_type='solid',
                                                                                                 fgColor=RED)
        irow += 2

    for row in swbin_testitem_sheet.rows:
        for cell in row:
            cell.border = border
            if cell.row == 1 and cell.column > 1:
                cell.fill = PatternFill(fill_type='solid', fgColor=YELLOW)
                cell.font = Font(bold=True)
                cell.alignment = alignment

    site_testitem_sheet = summary_wb.create_sheet('Site-TestItem')
    site_testitem_sheet.freeze_panes = 'B2'
    irow = 1
    site_testitem_sheet.cell(row=irow, column=1).value = lot_count
    for site_key in site_data[0].keys():
        for i in range(len(site_data[0][site_key][1])):
            site_testitem_sheet.cell(row=irow, column=i + 2).value = site_data[0][site_key][1][i][0]
        break
    irow += 1
    for site_key in site_data[0].keys():
        site_testitem_sheet.cell(row=irow, column=1).value = 'Site' + str(site_key)
        site_testitem_sheet.cell(row=irow, column=1).fill = PatternFill(fill_type='solid',
                                                                        fgColor=GREEN)
        for i in range(len(site_data[0][site_key][1])):
            site_testitem_sheet.cell(row=irow, column=2 + i).value = site_data[0][site_key][1][i][1]
            site_testitem_sheet.cell(row=irow + 1, column=2 + i).value = '{:.2%}'.format(
                site_data[0][site_key][1][i][1] / lot_count)
            if site_data[0][site_key][1][i][1] > 0:
                # percent value greater than 0 are filled in red
                site_testitem_sheet.cell(row=irow + 1, column=2 + i).fill = PatternFill(fill_type='solid',
                                                                                        fgColor=RED)
        irow += 2

    for row in site_testitem_sheet.rows:
        for cell in row:
            cell.border = border
            if cell.row == 1 and cell.column > 1:
                cell.fill = PatternFill(fill_type='solid', fgColor=YELLOW)
                cell.font = Font(bold=True)
                cell.alignment = alignment

    for sheet_name in summary_wb.sheetnames:
        if sheet_name == 'Sheet':
            del summary_wb[sheet_name]
        else:
            set_column_width(summary_wb[sheet_name])
    print("save summary data begin>>>>>>")
    summary_wb.save(summary_file)
    print("save summary data finished<<<<<<")


def main():
    global project

    # date folder path
    if argv.count('-d') == 0:
        print("Error：Date folder path is required.Format:-d D:\Date folder.")
        exit()
    else:
        date_folder = argv[argv.index('-d') + 1]
        for i in range(argv.index('-d') + 2, len(argv)):
            if not argv[i].startswith('-'):
                date_folder += (' ' + argv[i])
            else:
                break

    # project F28,JX828,JX825
    if argv.count('-p') != 0:
        project = argv[argv.index('-p') + 1]

    lotno_names = listdir(date_folder)
    for name in lotno_names:
        folder = join(date_folder, name)
        if isdir(folder):
            # get CSV file under the folder
            file_list = get_filelist(folder, '.csv')
            if not file_list:
                exit()

            # analysis folder path
            if argv.count('-a') == 0:
                # default analysis folder
                analysis_folder = folder + '\Analysis'
            else:
                analysis_folder = argv[argv.index('-a') + 1]

            # create analysis folder
            mkdir(analysis_folder)

            parse_data = []
            for file in file_list:
                print(file)
                # parse file
                parse_data.append(parse_file(file))
            # save data
            save_data(analysis_folder, parse_data)


if __name__ == '__main__':
    main()
