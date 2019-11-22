# encoding:utf-8
# @Time     : 2019/11/21
# @Author   : Jerry Chou
# @File     :
# @Function : 公共函数

from os.path import join, exists, isdir
from os import remove, walk, makedirs, listdir
from openpyxl.utils import get_column_letter
from psutil import pids, Process


def find_item(item_list, value):
    """
    找到value在item_list中所有的index
    :param item_list: 查询列表
    :param value: 查询值
    :return: 查询值的index
    """
    return [i for i, v in enumerate(item_list) if v == value]


def sort_data(data):
    """
    根据列表子节点的元素个数从多到少排序
    :param data: 列表名
    :return: 
    """
    for i in range(len(data) - 1):
        for j in range(i + 1, len(data)):
            if len(data[i][1]) < len(data[j][1]):
                data[i], data[j] = data[j], data[i]


def set_column_width(sheet):
    """
    设置列宽
    :param sheet: sheet名
    :return: 
    """
    # 获取每一列的内容的最大宽度
    col_width = [0.5] * sheet.max_column
    for row in range(sheet.max_row):
        for col in range(sheet.max_column):
            value = sheet.cell(row=row + 1, column=col + 1).value
            if value:
                width = len(str(value))
                if width > col_width[col]:
                    col_width[col] = width
    # 设置列宽
    for i in range(len(col_width)):
        col_lettert = get_column_letter(i + 1)
        if col_width[i] > 100:
            sheet.column_dimensions[col_lettert].width = 100
        else:
            sheet.column_dimensions[col_lettert].width = col_width[i] + 4


def mkdir(path):
    """
    创建文件夹目录
    :param path: 文件夹目录
    :return: 
    """
    path = path.strip()
    path = path.rstrip("\\")
    is_exists = exists(path)
    if not is_exists:
        makedirs(path)
        return True
    else:
        return False


def get_filelist(folder, postfix=None):
    """
    获取某个后缀的文件列表
    :param postfix: 后缀，默认为None
    :param folder: 文件夹路径
    :return: 文件列表
    """
    fullname_list = []
    if isdir(folder):
        files = listdir(folder)
        for filename in files:
            fullname_list.append(join(folder, filename))
        if postfix:
            target_file_list = []
            for fullname in fullname_list:
                if fullname.endswith(postfix):
                    target_file_list.append(fullname)
            return target_file_list
        else:
            return fullname_list
    else:
        print("Error：不是文件夹！")
        return False


def delete_files(path, postfix):
    """
    删除一个目录文件中的某些特定后缀名文件
    :param path: 文件夹路径
    :param postfix: 后缀
    :return: 
    """
    for root, dirs, files in walk(path):
        for name in files:
            if name.endswith(postfix):
                remove(join(root, name))
        return


def search_string(data, target):
    """
    查找字符串
    :param data: 数据
    :param target: 待查找字符串
    :return: 行、列
    """
    for i in range(len(data)):
        try:
            col_num = data[i].index(target)
            row_num = i
            return row_num, col_num
        except:
            pass
    print("没有找到 " + target + " ！")
    return False


def judge_process(process_name):
    """
    判断进程是否存在
    :param process_name: 进程名
    :return: True 存在  False 不存在
    """
    pl = pids()
    for pid in pl:
        if Process(pid).name() == process_name:
            return True
    else:
        return False


def read_output(p, match_times):
    """
    读取子进程的返回值
    :param p: 子进程
    :param match_times: 匹配次数
    :return: 返回字符串
    """
    match = ' >'
    count = 0
    buf = ''
    while True:
        # 逐个获取字符串
        out = p.stdout.read(1)
        buf = buf + out
        # 比较最后两个字符串是否与匹配值一致
        if buf[-2:] == match:
            count += 1
        # 达到匹配次数
        if count == match_times:
            return buf


def debug_mode(p, buf):
    """
    断点模式：是否产生断点
    :param p: 子进程
    :param buf: 返回字符串
    :return: 没有断点，返回
    """
    buf_list = buf.split("\n")
    # 当最后一行不匹配
    if buf_list[-1] != 'Everest >':
        # 打印返回字符串，获取用户输入
        command = input(buf)
        # 将用户输入写进子进程的输入中
        p.stdin.write(command + '\n')
        # 获取子进程返回
        buf2 = read_output(p, 1)
        # 断点模式判断
        debug_mode(p, buf2)
    else:
        return

def set_command(p, buf, cmd_str):
    """
    发送命令
    :param p: 子进程
    :param buf: 返回字符串
    :param cmd_str: 命令字符串
    :return: 
    """
    buf_list = buf.split("\n")
    # 当最后一行不匹配
    if buf_list[-1] == 'Everest >':
        # 将cmd_str写进子进程的输入中
        p.stdin.write(cmd_str)

def get_attriValue(element, attribute):
    """
    获得属性值
    :param element: 元素节点
    :param attribute: 属性名
    :return: 属性名,属性值
    """
    value = element.getAttribute(attribute)
    return attribute, value


def get_tree(element):
    """
    获得树节点
    :param element: 元素节点
    :return: 树
    """
    tree = []
    if element.hasAttribute('name'):
        tree.append(get_attriValue(element, 'name'))
    if element.hasAttribute('sigRef'):
        tree.append(get_attriValue(element, 'sigRef'))
    if element.hasAttribute('minorTestNum'):
        tree.append(get_attriValue(element, 'minorTestNum'))
    if element.hasAttribute('testNum'):
        tree.append(get_attriValue(element, 'testNum'))
    if element.hasAttribute('type'):
        tree.append(get_attriValue(element, 'type'))
    if len(element.childNodes) == 1:
        tag_name = element.tagName
        tag_content = element.childNodes[0].data
        tree.append((tag_name, tag_content))
    elif len(element.childNodes) > 1:
        temp = []
        for childNode in element.childNodes:
            if len(childNode.childNodes) == 1:
                tag_name = childNode.tagName
                tag_content = childNode.childNodes[0].data
                temp.append((tag_name, tag_content))
            elif len(childNode.childNodes) > 1:
                tree.append(get_tree(childNode))
        if len(temp) > 0:
            tree.append(temp)
    return tree

def num_to_str(num):
    """
    数字转字符串，如果个位数前面加0
    :param num: 
    :return: 
    """
    if num < 10:
        return '0' + str(num)
    else:
        return str(num)

def cal_factor(num):
    """
    计算一个数的因数
    :param num: 
    :return: 
    """
    factor_list = []
    for i in range(1, num + 1):
        if num % i == 0:
            factor_list.append(i)
    return factor_list